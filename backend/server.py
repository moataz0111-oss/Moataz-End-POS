from fastapi import FastAPI, APIRouter, HTTPException, Depends, Request, BackgroundTasks, UploadFile, File, Form, Body
from fastapi.security import HTTPBearer, HTTPAuthorizationCredentials
from fastapi.staticfiles import StaticFiles
from fastapi.middleware.gzip import GZipMiddleware
from dotenv import load_dotenv
from starlette.middleware.cors import CORSMiddleware
from motor.motor_asyncio import AsyncIOMotorClient
import os
import logging
from pathlib import Path
from pydantic import BaseModel, Field, ConfigDict, EmailStr
from typing import List, Optional, Dict, Any
from enum import Enum
import uuid
from datetime import datetime, timezone, timedelta
import bcrypt
import jwt
from sendgrid import SendGridAPIClient
from sendgrid.helpers.mail import Mail
from PIL import Image
import io
import base64
import aiofiles
import asyncio

ROOT_DIR = Path(__file__).parent
load_dotenv(ROOT_DIR / '.env')

# MongoDB connection
mongo_url = os.environ['MONGO_URL']
client = AsyncIOMotorClient(mongo_url)
db = client[os.environ['DB_NAME']]

# JWT Configuration
JWT_SECRET = os.environ.get('JWT_SECRET')
if not JWT_SECRET:
    raise ValueError("JWT_SECRET environment variable is required")
JWT_ALGORITHM = 'HS256'
JWT_EXPIRATION_HOURS = 24

# SendGrid Configuration
SENDGRID_API_KEY = os.environ.get('SENDGRID_API_KEY', '')
SENDER_EMAIL = os.environ.get('SENDER_EMAIL', 'noreply@maestroegp.com')

# Static Files Configuration
UPLOAD_DIR = ROOT_DIR / "uploads"
UPLOAD_DIR.mkdir(exist_ok=True)
BACKGROUNDS_DIR = UPLOAD_DIR / "backgrounds"
BACKGROUNDS_DIR.mkdir(exist_ok=True)
LOGOS_DIR = UPLOAD_DIR / "logos"
LOGOS_DIR.mkdir(exist_ok=True)
IMAGES_DIR = UPLOAD_DIR / "images"
IMAGES_DIR.mkdir(exist_ok=True)
PRODUCTS_DIR = IMAGES_DIR / "products"
PRODUCTS_DIR.mkdir(exist_ok=True)
CATEGORIES_DIR = IMAGES_DIR / "categories"
CATEGORIES_DIR.mkdir(exist_ok=True)

app = FastAPI(title="Maestro EGP API", version="2.0.0")
api_router = APIRouter(prefix="/api")
security = HTTPBearer()

# إضافة GZip compression لتسريع نقل البيانات
app.add_middleware(GZipMiddleware, minimum_size=1000)

# Mount static files directory
app.mount("/uploads", StaticFiles(directory=str(UPLOAD_DIR)), name="uploads")
app.mount("/api/uploads", StaticFiles(directory=str(UPLOAD_DIR)), name="api_uploads")

# Helper function to check user roles
def has_role(user: dict, roles: list) -> bool:
    """التحقق من صلاحية المستخدم"""
    user_role = user.get("role", "")
    return user_role in roles or "super_admin" in roles and user_role == "super_admin"

# ==================== DATABASE INITIALIZATION ====================

async def create_indexes():
    """إنشاء indexes لتسريع الاستعلامات"""
    try:
        # Users indexes
        await db.users.create_index("id", unique=True)
        await db.users.create_index("email", unique=True)
        await db.users.create_index("tenant_id")
        await db.users.create_index("role")
        
        # Orders indexes - الأكثر أهمية للأداء
        await db.orders.create_index("id", unique=True)
        await db.orders.create_index("tenant_id")
        await db.orders.create_index("branch_id")
        await db.orders.create_index("status")
        await db.orders.create_index("created_at")
        await db.orders.create_index([("tenant_id", 1), ("status", 1)])
        await db.orders.create_index([("tenant_id", 1), ("created_at", -1)])
        await db.orders.create_index([("cashier_id", 1), ("created_at", -1)])
        
        # Products indexes
        await db.products.create_index("id", unique=True)
        await db.products.create_index("tenant_id")
        await db.products.create_index("category_id")
        await db.products.create_index([("tenant_id", 1), ("is_active", 1)])
        
        # Categories indexes
        await db.categories.create_index("id", unique=True)
        await db.categories.create_index("tenant_id")
        
        # Drivers indexes
        await db.drivers.create_index("id", unique=True)
        await db.drivers.create_index("tenant_id")
        await db.drivers.create_index("branch_id")
        
        # Employees indexes
        await db.employees.create_index("id", unique=True)
        await db.employees.create_index("tenant_id")
        await db.employees.create_index("branch_id")
        
        # Shifts indexes
        await db.shifts.create_index("id", unique=True)
        await db.shifts.create_index([("cashier_id", 1), ("status", 1)])
        await db.shifts.create_index("tenant_id")
        
        # Expenses indexes
        await db.expenses.create_index("id", unique=True)
        await db.expenses.create_index("tenant_id")
        await db.expenses.create_index([("branch_id", 1), ("created_at", -1)])
        
        # Inventory indexes
        await db.inventory.create_index("id", unique=True)
        await db.inventory.create_index("tenant_id")
        await db.inventory.create_index("branch_id")
        
        logger.info("✅ Database indexes created successfully")
    except Exception as e:
        logger.warning(f"⚠️ Some indexes may already exist: {e}")

async def init_database():
    """تهيئة قاعدة البيانات بالبيانات الأساسية عند بدء التطبيق"""
    try:
        logger.info("🔍 Checking database initialization...")
        
        # إنشاء indexes لتسريع الاستعلامات
        await create_indexes()
        
        # التحقق من الاتصال بقاعدة البيانات
        await db.command('ping')
        logger.info("✅ Database connection successful")
        
        # التحقق من وجود Super Admin
        super_admin = await db.users.find_one({"role": "super_admin"})
        if not super_admin:
            logger.info("🔧 Initializing database with default data...")
            
            # إنشاء Super Admin
            super_admin_password = bcrypt.hashpw("owner123".encode('utf-8'), bcrypt.gensalt()).decode('utf-8')
            super_admin_doc = {
                "id": str(uuid.uuid4()),
                "username": "super_admin",
                "email": "owner@maestroegp.com",
                "password": super_admin_password,
                "full_name": "مالك النظام",
                "role": "super_admin",
                "branch_id": None,
                "tenant_id": "system",
                "permissions": ["all", "super_admin"],
                "is_active": True,
                "created_at": datetime.now(timezone.utc).isoformat()
            }
            await db.users.insert_one(super_admin_doc)
            logger.info("✅ Super Admin created: owner@maestroegp.com / owner123")
            
            # إنشاء مدير النظام الرئيسي
            admin_password = bcrypt.hashpw("admin123".encode('utf-8'), bcrypt.gensalt()).decode('utf-8')
            admin_doc = {
                "id": str(uuid.uuid4()),
                "username": "admin",
                "email": "admin@maestroegp.com",
                "password": admin_password,
                "full_name": "مدير النظام",
                "role": "admin",
                "branch_id": None,
                "tenant_id": "default",
                "permissions": ["all"],
                "is_active": True,
                "created_at": datetime.now(timezone.utc).isoformat()
            }
            await db.users.insert_one(admin_doc)
            logger.info("✅ Main Admin created: admin@maestroegp.com / admin123")
            
            # إنشاء الفرع الرئيسي
            branch_doc = {
                "id": str(uuid.uuid4()),
                "name": "الفرع الرئيسي",
                "code": "MAIN",
                "address": "العنوان الرئيسي",
                "phone": "",
                "is_main": True,
                "is_active": True,
                "tenant_id": "default",
                "created_at": datetime.now(timezone.utc).isoformat()
            }
            await db.branches.insert_one(branch_doc)
            logger.info("✅ Main Branch created")
            
            # إعدادات النظام
            branding_doc = {
                "type": "system_branding",
                "value": {
                    "name": "Maestro",
                    "name_ar": "Maestro",
                    "name_en": "Maestro",
                    "logo_url": None
                }
            }
            await db.settings.insert_one(branding_doc)
            logger.info("✅ System branding created")
            
            # خلفيات تسجيل الدخول الافتراضية - صور متعددة
            bg_doc = {
                "type": "login_backgrounds",
                "backgrounds": [
                    {
                        "id": str(uuid.uuid4()),
                        "image_url": "https://images.unsplash.com/photo-1517248135467-4c7edcad34c4?w=1920",
                        "title": "مطعم فاخر",
                        "is_active": True
                    },
                    {
                        "id": str(uuid.uuid4()),
                        "image_url": "https://images.unsplash.com/photo-1552566626-52f8b828add9?w=1920",
                        "title": "مطعم حديث",
                        "is_active": True
                    },
                    {
                        "id": str(uuid.uuid4()),
                        "image_url": "https://images.unsplash.com/photo-1414235077428-338989a2e8c0?w=1920",
                        "title": "طعام شهي",
                        "is_active": True
                    },
                    {
                        "id": str(uuid.uuid4()),
                        "image_url": "https://images.unsplash.com/photo-1555396273-367ea4eb4db5?w=1920",
                        "title": "مطعم أنيق",
                        "is_active": True
                    },
                    {
                        "id": str(uuid.uuid4()),
                        "image_url": "https://images.unsplash.com/photo-1559339352-11d035aa65de?w=1920",
                        "title": "كافيه عصري",
                        "is_active": True
                    },
                    {
                        "id": str(uuid.uuid4()),
                        "image_url": "https://images.unsplash.com/photo-1466978913421-dad2ebd01d17?w=1920",
                        "title": "مطبخ احترافي",
                        "is_active": True
                    }
                ],
                "settings": {
                    "transition_effect": "fade",
                    "transition_speed": 5,
                    "overlay_color": "rgba(0,0,0,0.5)",
                    "text_color": "#ffffff"
                }
            }
            await db.settings.insert_one(bg_doc)
            logger.info("✅ Login backgrounds created (6 images)")
            
            # إنشاء الفئات الافتراضية مع الصور
            default_categories = [
                {"id": str(uuid.uuid4()), "name": "برغر", "name_ar": "برغر", "sort_order": 1, "tenant_id": "default", "image": "https://images.unsplash.com/photo-1635275650933-7b0911815a2e?w=400", "created_at": datetime.now(timezone.utc).isoformat()},
                {"id": str(uuid.uuid4()), "name": "بيتزا", "name_ar": "بيتزا", "sort_order": 2, "tenant_id": "default", "image": "https://images.unsplash.com/photo-1703073186021-021fb5a0bde1?w=400", "created_at": datetime.now(timezone.utc).isoformat()},
                {"id": str(uuid.uuid4()), "name": "مشروبات", "name_ar": "مشروبات", "sort_order": 3, "tenant_id": "default", "image": "https://images.unsplash.com/photo-1657958977261-d75e81b4713f?w=400", "created_at": datetime.now(timezone.utc).isoformat()},
                {"id": str(uuid.uuid4()), "name": "حلويات", "name_ar": "حلويات", "sort_order": 4, "tenant_id": "default", "image": "https://images.unsplash.com/photo-1546902189-eaaf09f8e38f?w=400", "created_at": datetime.now(timezone.utc).isoformat()},
                {"id": str(uuid.uuid4()), "name": "سلطات", "name_ar": "سلطات", "sort_order": 5, "tenant_id": "default", "image": "https://images.unsplash.com/photo-1677653805080-59c57727c84e?w=400", "created_at": datetime.now(timezone.utc).isoformat()},
            ]
            await db.categories.insert_many(default_categories)
            logger.info("✅ Default categories created (5) with images")
            
            # إنشاء منتجات افتراضية مع الصور
            burger_cat = default_categories[0]["id"]
            pizza_cat = default_categories[1]["id"]
            drinks_cat = default_categories[2]["id"]
            desserts_cat = default_categories[3]["id"]
            salads_cat = default_categories[4]["id"]
            default_products = [
                {"id": str(uuid.uuid4()), "name": "برغر كلاسيك", "price": 5000, "cost": 2000, "category_id": burger_cat, "is_available": True, "tenant_id": "default", "image": "https://images.unsplash.com/photo-1656439659132-24c68e36b553?w=400", "created_at": datetime.now(timezone.utc).isoformat()},
                {"id": str(uuid.uuid4()), "name": "برغر دبل", "price": 7500, "cost": 3000, "category_id": burger_cat, "is_available": True, "tenant_id": "default", "image": "https://images.unsplash.com/photo-1635275650933-7b0911815a2e?w=400", "created_at": datetime.now(timezone.utc).isoformat()},
                {"id": str(uuid.uuid4()), "name": "بيتزا مارغريتا", "price": 10000, "cost": 4000, "category_id": pizza_cat, "is_available": True, "tenant_id": "default", "image": "https://images.unsplash.com/photo-1681567604770-0dc826c870ae?w=400", "created_at": datetime.now(timezone.utc).isoformat()},
                {"id": str(uuid.uuid4()), "name": "بيتزا خضار", "price": 12000, "cost": 5000, "category_id": pizza_cat, "is_available": True, "tenant_id": "default", "image": "https://images.unsplash.com/photo-1602104980741-b87a33837f9f?w=400", "created_at": datetime.now(timezone.utc).isoformat()},
                {"id": str(uuid.uuid4()), "name": "كولا", "price": 1500, "cost": 500, "category_id": drinks_cat, "is_available": True, "tenant_id": "default", "image": "https://images.unsplash.com/photo-1657958977261-d75e81b4713f?w=400", "created_at": datetime.now(timezone.utc).isoformat()},
                {"id": str(uuid.uuid4()), "name": "عصير برتقال", "price": 2500, "cost": 1000, "category_id": drinks_cat, "is_available": True, "tenant_id": "default", "image": "https://images.unsplash.com/photo-1716925539259-ce0115263d37?w=400", "created_at": datetime.now(timezone.utc).isoformat()},
                {"id": str(uuid.uuid4()), "name": "كيكة شوكولاته", "price": 3500, "cost": 1500, "category_id": desserts_cat, "is_available": True, "tenant_id": "default", "image": "https://images.unsplash.com/photo-1546902189-eaaf09f8e38f?w=400", "created_at": datetime.now(timezone.utc).isoformat()},
                {"id": str(uuid.uuid4()), "name": "سلطة خضراء", "price": 4000, "cost": 1200, "category_id": salads_cat, "is_available": True, "tenant_id": "default", "image": "https://images.unsplash.com/photo-1677653805080-59c57727c84e?w=400", "created_at": datetime.now(timezone.utc).isoformat()},
            ]
            await db.products.insert_many(default_products)
            logger.info("✅ Default products created (8) with images")
            
            # إنشاء سائقين افتراضيين
            default_drivers = [
                {"id": str(uuid.uuid4()), "name": "سائق 1", "phone": "07801111111", "is_active": True, "tenant_id": "default", "created_at": datetime.now(timezone.utc).isoformat()},
                {"id": str(uuid.uuid4()), "name": "سائق 2", "phone": "07802222222", "is_active": True, "tenant_id": "default", "created_at": datetime.now(timezone.utc).isoformat()},
            ]
            await db.drivers.insert_many(default_drivers)
            logger.info("✅ Default drivers created (2)")
            
            # إنشاء موظفين افتراضيين
            default_employees = [
                {"id": str(uuid.uuid4()), "name": "موظف 1", "position": "كاشير", "phone": "07803333333", "salary": 500000, "is_active": True, "tenant_id": "default", "created_at": datetime.now(timezone.utc).isoformat()},
                {"id": str(uuid.uuid4()), "name": "موظف 2", "position": "طباخ", "phone": "07804444444", "salary": 600000, "is_active": True, "tenant_id": "default", "created_at": datetime.now(timezone.utc).isoformat()},
            ]
            await db.employees.insert_many(default_employees)
            logger.info("✅ Default employees created (2)")
            
            logger.info("=" * 50)
            logger.info("🎉 DATABASE INITIALIZATION COMPLETE!")
            logger.info("=" * 50)
            logger.info("📋 LOGIN CREDENTIALS:")
            logger.info("   Super Admin: owner@maestroegp.com / owner123")
            logger.info("   Secret Key: 271018")
            logger.info("   Main Admin: admin@maestroegp.com / admin123")
            logger.info("=" * 50)
        else:
            logger.info("ℹ️ Database already initialized - Super Admin exists")
            # لا نحتاج لإنشاء بيانات default لأن كل عميل سيكون له بياناته الخاصة
        
        # التحقق من وجود خلفيات تسجيل الدخول
        login_bg = await db.settings.find_one({"type": "login_backgrounds"})
        if not login_bg:
            logger.info("🔧 Adding default login backgrounds...")
            bg_doc = {
                "type": "login_backgrounds",
                "backgrounds": [
                    {
                        "id": str(uuid.uuid4()),
                        "image_url": "https://images.unsplash.com/photo-1517248135467-4c7edcad34c4?w=1920",
                        "title": "مطعم فاخر",
                        "is_active": True
                    },
                    {
                        "id": str(uuid.uuid4()),
                        "image_url": "https://images.unsplash.com/photo-1552566626-52f8b828add9?w=1920",
                        "title": "مطعم حديث",
                        "is_active": True
                    },
                    {
                        "id": str(uuid.uuid4()),
                        "image_url": "https://images.unsplash.com/photo-1414235077428-338989a2e8c0?w=1920",
                        "title": "طعام شهي",
                        "is_active": True
                    },
                    {
                        "id": str(uuid.uuid4()),
                        "image_url": "https://images.unsplash.com/photo-1555396273-367ea4eb4db5?w=1920",
                        "title": "مطعم أنيق",
                        "is_active": True
                    },
                    {
                        "id": str(uuid.uuid4()),
                        "image_url": "https://images.unsplash.com/photo-1559339352-11d035aa65de?w=1920",
                        "title": "كافيه عصري",
                        "is_active": True
                    },
                    {
                        "id": str(uuid.uuid4()),
                        "image_url": "https://images.unsplash.com/photo-1466978913421-dad2ebd01d17?w=1920",
                        "title": "مطبخ احترافي",
                        "is_active": True
                    }
                ],
                "settings": {
                    "transition_effect": "fade",
                    "transition_speed": 5,
                    "overlay_color": "rgba(0,0,0,0.5)",
                    "text_color": "#ffffff"
                }
            }
            await db.settings.insert_one(bg_doc)
            logger.info("✅ Login backgrounds added (6 images)")
            
    except Exception as e:
        logger.error(f"❌ Database initialization error: {str(e)}")
        logger.error(f"❌ Error type: {type(e).__name__}")
        import traceback
        logger.error(f"❌ Traceback: {traceback.format_exc()}")

@app.on_event("startup")
async def startup_event():
    """يتم تشغيله عند بدء التطبيق"""
    logger.info("🚀 Starting Maestro EGP API...")
    await init_database()
    
    # تشغيل التحديثات التلقائية لجميع العملاء
    await apply_automatic_updates()
    
    logger.info("✅ Application started successfully")

async def apply_automatic_updates():
    """تطبيق التحديثات التلقائية على جميع العملاء عند كل بدء تشغيل"""
    logger.info("🔄 Applying automatic updates to all tenants...")
    
    try:
        # 0. تحديث البيانات القديمة التي ليس لها tenant_id لتصبح "default"
        # تحديث المستخدمين الرئيسيين
        await db.users.update_many(
            {"$or": [{"tenant_id": {"$exists": False}}, {"tenant_id": None}, {"tenant_id": ""}]},
            {"$set": {"tenant_id": "default"}}
        )
        logger.info("   ✅ Updated users without tenant_id")
        
        # تحديث السائقين
        drivers_tenant_result = await db.drivers.update_many(
            {"$or": [{"tenant_id": {"$exists": False}}, {"tenant_id": None}, {"tenant_id": ""}]},
            {"$set": {"tenant_id": "default"}}
        )
        if drivers_tenant_result.modified_count > 0:
            logger.info(f"   ✅ Updated {drivers_tenant_result.modified_count} drivers with default tenant_id")
        
        # تحديث الموظفين
        employees_tenant_result = await db.employees.update_many(
            {"$or": [{"tenant_id": {"$exists": False}}, {"tenant_id": None}, {"tenant_id": ""}]},
            {"$set": {"tenant_id": "default"}}
        )
        if employees_tenant_result.modified_count > 0:
            logger.info(f"   ✅ Updated {employees_tenant_result.modified_count} employees with default tenant_id")
        
        # تحديث الفروع
        branches_tenant_result = await db.branches.update_many(
            {"$or": [{"tenant_id": {"$exists": False}}, {"tenant_id": None}, {"tenant_id": ""}]},
            {"$set": {"tenant_id": "default"}}
        )
        if branches_tenant_result.modified_count > 0:
            logger.info(f"   ✅ Updated {branches_tenant_result.modified_count} branches with default tenant_id")
        
        # تحديث الفئات
        categories_tenant_result = await db.categories.update_many(
            {"$or": [{"tenant_id": {"$exists": False}}, {"tenant_id": None}, {"tenant_id": ""}]},
            {"$set": {"tenant_id": "default"}}
        )
        if categories_tenant_result.modified_count > 0:
            logger.info(f"   ✅ Updated {categories_tenant_result.modified_count} categories with default tenant_id")
        
        # تحديث المنتجات
        products_tenant_result = await db.products.update_many(
            {"$or": [{"tenant_id": {"$exists": False}}, {"tenant_id": None}, {"tenant_id": ""}]},
            {"$set": {"tenant_id": "default"}}
        )
        if products_tenant_result.modified_count > 0:
            logger.info(f"   ✅ Updated {products_tenant_result.modified_count} products with default tenant_id")
        
        # تحديث الطلبات
        orders_tenant_result = await db.orders.update_many(
            {"$or": [{"tenant_id": {"$exists": False}}, {"tenant_id": None}, {"tenant_id": ""}]},
            {"$set": {"tenant_id": "default"}}
        )
        if orders_tenant_result.modified_count > 0:
            logger.info(f"   ✅ Updated {orders_tenant_result.modified_count} orders with default tenant_id")
        
        # تحديث المصاريف
        expenses_tenant_result = await db.expenses.update_many(
            {"$or": [{"tenant_id": {"$exists": False}}, {"tenant_id": None}, {"tenant_id": ""}]},
            {"$set": {"tenant_id": "default"}}
        )
        if expenses_tenant_result.modified_count > 0:
            logger.info(f"   ✅ Updated {expenses_tenant_result.modified_count} expenses with default tenant_id")
        
        # 1. تفعيل جميع السائقين
        drivers_result = await db.drivers.update_many(
            {},
            {"$set": {"is_active": True}}
        )
        if drivers_result.modified_count > 0:
            logger.info(f"   ✅ Activated {drivers_result.modified_count} drivers")
        
        # 2. إضافة is_available للسائقين
        await db.drivers.update_many(
            {"is_available": {"$exists": False}},
            {"$set": {"is_available": True}}
        )
        
        # 2.5 إصلاح السائقين المرتبطين بفروع غير موجودة
        default_branch = await db.branches.find_one({"tenant_id": "default"})
        if default_branch:
            # جلب جميع الفروع الصالحة (مع حد أقصى لتجنب مشاكل الأداء)
            valid_branch_ids = [b["id"] async for b in db.branches.find({}, {"id": 1}).limit(1000)]
            
            # تحديث السائقين بفروع غير موجودة
            drivers_fixed = await db.drivers.update_many(
                {"branch_id": {"$nin": valid_branch_ids}},
                {"$set": {"branch_id": default_branch["id"]}}
            )
            if drivers_fixed.modified_count > 0:
                logger.info(f"   ✅ Fixed {drivers_fixed.modified_count} drivers with invalid branch_id")
        
        # 3. تفعيل جميع الفروع
        await db.branches.update_many(
            {"is_active": {"$exists": False}},
            {"$set": {"is_active": True}}
        )
        
        # 4. إنشاء فرع افتراضي لكل عميل ليس لديه فرع
        tenants = await db.tenants.find({}, {"_id": 0, "id": 1, "name": 1}).to_list(1000)
        for tenant in tenants:
            tenant_branch = await db.branches.find_one({"tenant_id": tenant["id"]})
            if not tenant_branch:
                default_branch = {
                    "id": str(uuid.uuid4()),
                    "name": "الفرع الرئيسي",
                    "address": "",
                    "phone": "",
                    "is_active": True,
                    "is_main": True,
                    "tenant_id": tenant["id"],
                    "created_at": datetime.now(timezone.utc).isoformat()
                }
                await db.branches.insert_one(default_branch)
                logger.info(f"   ✅ Created branch for: {tenant.get('name', tenant['id'][:8])}")
        
        # 5. إغلاق الورديات القديمة (أكثر من 24 ساعة)
        old_cutoff = (datetime.now(timezone.utc) - timedelta(hours=24)).isoformat()
        old_shifts = await db.shifts.update_many(
            {"status": "open", "started_at": {"$lt": old_cutoff}},
            {"$set": {
                "status": "auto_closed",
                "ended_at": datetime.now(timezone.utc).isoformat(),
                "auto_closed_reason": "تم الإغلاق تلقائياً بعد 24 ساعة"
            }}
        )
        if old_shifts.modified_count > 0:
            logger.info(f"   ✅ Auto-closed {old_shifts.modified_count} old shifts")
        
        # 6. تحديث الطاولات القديمة التي ليس لها tenant_id
        tables_tenant_result = await db.tables.update_many(
            {"$or": [{"tenant_id": {"$exists": False}}, {"tenant_id": None}, {"tenant_id": ""}]},
            {"$set": {"tenant_id": "default"}}
        )
        if tables_tenant_result.modified_count > 0:
            logger.info(f"   ✅ Updated {tables_tenant_result.modified_count} tables with default tenant_id")
        
        # 7. إنشاء طاولات افتراضية لكل عميل ليس لديه طاولات
        for tenant in tenants:
            tenant_tables = await db.tables.count_documents({"tenant_id": tenant["id"]})
            if tenant_tables == 0:
                # جلب فرع العميل
                tenant_branch = await db.branches.find_one({"tenant_id": tenant["id"]})
                branch_id = tenant_branch["id"] if tenant_branch else None
                
                # إنشاء 5 طاولات افتراضية
                default_tables = []
                for i in range(1, 6):
                    default_tables.append({
                        "id": str(uuid.uuid4()),
                        "number": i,
                        "capacity": 4,
                        "section": "القاعة الرئيسية",
                        "status": "available",
                        "current_order_id": None,
                        "branch_id": branch_id,
                        "tenant_id": tenant["id"]
                    })
                await db.tables.insert_many(default_tables)
                logger.info(f"   ✅ Created 5 default tables for: {tenant.get('name', tenant['id'][:8])}")
        
        # 8. تحديث صور الفئات والمنتجات الافتراضية للنظام الرئيسي (إذا لم تكن موجودة)
        category_images = {
            "برغر": "https://images.unsplash.com/photo-1635275650933-7b0911815a2e?w=400",
            "بيتزا": "https://images.unsplash.com/photo-1703073186021-021fb5a0bde1?w=400",
            "مشروبات": "https://images.unsplash.com/photo-1657958977261-d75e81b4713f?w=400",
            "حلويات": "https://images.unsplash.com/photo-1546902189-eaaf09f8e38f?w=400",
            "سلطات": "https://images.unsplash.com/photo-1677653805080-59c57727c84e?w=400",
        }
        for cat_name, cat_image in category_images.items():
            await db.categories.update_many(
                {"name": cat_name, "tenant_id": "default", "$or": [{"image": {"$exists": False}}, {"image": None}, {"image": ""}]},
                {"$set": {"image": cat_image}}
            )
        
        product_images = {
            "برغر كلاسيك": "https://images.unsplash.com/photo-1656439659132-24c68e36b553?w=400",
            "برغر دبل": "https://images.unsplash.com/photo-1635275650933-7b0911815a2e?w=400",
            "بيتزا مارغريتا": "https://images.unsplash.com/photo-1681567604770-0dc826c870ae?w=400",
            "بيتزا خضار": "https://images.unsplash.com/photo-1602104980741-b87a33837f9f?w=400",
            "كولا": "https://images.unsplash.com/photo-1657958977261-d75e81b4713f?w=400",
            "عصير برتقال": "https://images.unsplash.com/photo-1716925539259-ce0115263d37?w=400",
        }
        for prod_name, prod_image in product_images.items():
            await db.products.update_many(
                {"name": prod_name, "tenant_id": "default", "$or": [{"image": {"$exists": False}}, {"image": None}, {"image": ""}]},
                {"$set": {"image": prod_image}}
            )
        logger.info("   ✅ Updated default category and product images")
        
        logger.info("✅ Automatic updates applied successfully")
    except Exception as e:
        logger.error(f"❌ Error applying automatic updates: {e}")

# ==================== HEALTH CHECK ====================

@app.get("/")
def read_root():
    return {"status": "Server is running successfully 🚀", "app": "Maestro EGP", "version": "2.0.0"}

@app.get("/health")
def health_check():
    return {"status": "ok"}

@api_router.get("/health")
def api_health_check():
    return {"status": "ok", "api": "Maestro EGP API"}

# ==================== DATABASE INITIALIZATION ENDPOINT ====================

@api_router.get("/init-db")
async def initialize_database_endpoint():
    """
    Endpoint لتهيئة قاعدة البيانات يدوياً
    يمكن استدعاؤه عبر: GET /api/init-db
    """
    try:
        # التحقق من وجود Super Admin
        super_admin = await db.users.find_one({"role": "super_admin"})
        
        if super_admin:
            # التحقق من وجود خلفيات
            login_bg = await db.settings.find_one({"type": "login_backgrounds"})
            if not login_bg:
                bg_doc = {
                    "type": "login_backgrounds",
                    "backgrounds": [
                        {
                            "id": str(uuid.uuid4()),
                            "image_url": "https://images.unsplash.com/photo-1517248135467-4c7edcad34c4?w=1920",
                            "title": "مطعم فاخر",
                            "is_active": True
                        },
                        {
                            "id": str(uuid.uuid4()),
                            "image_url": "https://images.unsplash.com/photo-1552566626-52f8b828add9?w=1920",
                            "title": "مطعم حديث",
                            "is_active": True
                        }
                    ],
                    "settings": {
                        "transition_effect": "fade",
                        "transition_speed": 5,
                        "overlay_color": "rgba(0,0,0,0.5)",
                        "text_color": "#ffffff"
                    }
                }
                await db.settings.insert_one(bg_doc)
            
            return {
                "status": "already_initialized",
                "message": "قاعدة البيانات مهيأة مسبقاً"
            }
        
        # إنشاء Super Admin
        super_admin_password = bcrypt.hashpw("owner123".encode('utf-8'), bcrypt.gensalt()).decode('utf-8')
        super_admin_doc = {
            "id": str(uuid.uuid4()),
            "username": "super_admin",
            "email": "owner@maestroegp.com",
            "password": super_admin_password,
            "full_name": "Owner",
            "role": "super_admin",
            "branch_id": None,
            "tenant_id": None,
            "permissions": ["all", "super_admin"],
            "is_active": True,
            "created_at": datetime.now(timezone.utc).isoformat()
        }
        await db.users.insert_one(super_admin_doc)
        
        # إنشاء مدير النظام الرئيسي
        admin_password = bcrypt.hashpw("admin123".encode('utf-8'), bcrypt.gensalt()).decode('utf-8')
        admin_doc = {
            "id": str(uuid.uuid4()),
            "username": "admin",
            "email": "admin@maestroegp.com",
            "password": admin_password,
            "full_name": "مدير النظام",
            "role": "admin",
            "branch_id": None,
            "tenant_id": None,
            "permissions": ["all"],
            "is_active": True,
            "created_at": datetime.now(timezone.utc).isoformat()
        }
        await db.users.insert_one(admin_doc)
        
        # إنشاء الفرع الرئيسي
        branch_doc = {
            "id": str(uuid.uuid4()),
            "name": "الفرع الرئيسي",
            "code": "MAIN",
            "address": "العنوان الرئيسي",
            "phone": "",
            "is_main": True,
            "is_active": True,
            "tenant_id": None,
            "created_at": datetime.now(timezone.utc).isoformat()
        }
        await db.branches.insert_one(branch_doc)
        
        # إعدادات النظام
        branding_doc = {
            "type": "system_branding",
            "value": {
                "name": "Maestro",
                "name_ar": "Maestro",
                "name_en": "Maestro",
                "logo_url": None
            }
        }
        await db.settings.insert_one(branding_doc)
        
        # خلفيات تسجيل الدخول - 6 خلفيات
        bg_doc = {
            "type": "login_backgrounds",
            "backgrounds": [
                {
                    "id": str(uuid.uuid4()),
                    "image_url": "https://images.unsplash.com/photo-1517248135467-4c7edcad34c4?w=1920",
                    "title": "مطعم فاخر",
                    "is_active": True
                },
                {
                    "id": str(uuid.uuid4()),
                    "image_url": "https://images.unsplash.com/photo-1552566626-52f8b828add9?w=1920",
                    "title": "مطعم حديث",
                    "is_active": True
                },
                {
                    "id": str(uuid.uuid4()),
                    "image_url": "https://images.unsplash.com/photo-1414235077428-338989a2e8c0?w=1920",
                    "title": "طعام شهي",
                    "is_active": True
                },
                {
                    "id": str(uuid.uuid4()),
                    "image_url": "https://images.unsplash.com/photo-1555396273-367ea4eb4db5?w=1920",
                    "title": "مطعم أنيق",
                    "is_active": True
                },
                {
                    "id": str(uuid.uuid4()),
                    "image_url": "https://images.unsplash.com/photo-1559339352-11d035aa65de?w=1920",
                    "title": "كافيه عصري",
                    "is_active": True
                },
                {
                    "id": str(uuid.uuid4()),
                    "image_url": "https://images.unsplash.com/photo-1466978913421-dad2ebd01d17?w=1920",
                    "title": "مطبخ احترافي",
                    "is_active": True
                }
            ],
            "settings": {
                "transition_effect": "fade",
                "transition_speed": 5,
                "overlay_color": "rgba(0,0,0,0.5)",
                "text_color": "#ffffff"
            }
        }
        await db.settings.insert_one(bg_doc)
        
        # Log credentials to server logs only (not in response)
        logger.info("=" * 50)
        logger.info("🎉 DATABASE INITIALIZED - CREDENTIALS (check server logs):")
        logger.info("   Super Admin: owner@maestroegp.com / owner123")
        logger.info("   Secret Key: 271018")
        logger.info("   Main Admin: admin@maestroegp.com / admin123")
        logger.info("=" * 50)
        
        return {
            "status": "success",
            "message": "تم تهيئة قاعدة البيانات بنجاح! تحقق من البريد الإلكتروني أو تواصل مع مزود الخدمة للحصول على بيانات الدخول."
        }
        
    except Exception as e:
        return {
            "status": "error",
            "message": f"حدث خطأ: {str(e)}"
        }

# Configure logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(name)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

# ==================== MODELS ====================

class UserRole:
    SUPER_ADMIN = "super_admin"  # مالك النظام الرئيسي
    ADMIN = "admin"
    MANAGER = "manager"
    SUPERVISOR = "supervisor"
    CASHIER = "cashier"
    DELIVERY = "delivery"  # دور جديد للسائقين

class OrderType:
    DINE_IN = "dine_in"
    TAKEAWAY = "takeaway"
    DELIVERY = "delivery"

class OrderStatus:
    PENDING = "pending"
    PREPARING = "preparing"
    READY = "ready"
    DELIVERED = "delivered"
    CANCELLED = "cancelled"

class PaymentMethod:
    CASH = "cash"
    CARD = "card"
    CREDIT = "credit"
    PENDING = "pending"

# ==================== TENANT MODELS (Multi-tenant) ====================

class TenantCreate(BaseModel):
    name: str  # اسم المطعم/الكافيه
    name_ar: Optional[str] = None  # اسم المطعم بالعربي
    name_en: Optional[str] = None  # اسم المطعم بالإنجليزي
    slug: str  # رابط فريد (مثل: my-restaurant)
    owner_name: str  # اسم المالك
    owner_email: EmailStr
    owner_phone: str
    subscription_type: str = "trial"  # trial, bronze, silver, gold, basic, premium, demo
    subscription_duration: int = 1  # مدة الاشتراك بالأشهر (1, 3, 6, 12)
    max_branches: int = 1
    max_users: int = 5
    logo_url: Optional[str] = None  # شعار المطعم
    is_demo: bool = False  # هل هو حساب تجريبي

class TenantFeatures(BaseModel):
    """ميزات العميل المتاحة"""
    showPOS: bool = True
    showTables: bool = True
    showOrders: bool = True
    showExpenses: bool = True
    showInventory: bool = True
    showDelivery: bool = True
    showReports: bool = True
    showSettings: bool = True
    showHR: bool = False
    showWarehouse: bool = False
    showCallLogs: bool = False
    showCallCenter: bool = False
    showKitchen: bool = False
    showLoyalty: bool = True
    showCoupons: bool = True
    showRecipes: bool = True
    showReservations: bool = True
    showReviews: bool = True
    showRatings: bool = True
    showSmartReports: bool = True

class TenantResponse(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str
    name: str
    name_ar: Optional[str] = None
    name_en: Optional[str] = None
    slug: str
    owner_name: str
    owner_email: str
    owner_phone: str
    subscription_type: str
    max_branches: int
    max_users: int
    is_active: bool
    created_at: str
    expires_at: Optional[str] = None
    logo_url: Optional[str] = None
    subscription_duration: Optional[int] = None  # مدة الاشتراك بالأشهر

# ==================== نظام الإشعارات ====================

class NotificationType(str, Enum):
    NEW_TENANT = "new_tenant"  # عميل جديد
    SUBSCRIPTION_EXPIRING = "subscription_expiring"  # اشتراك قارب على الانتهاء
    SUBSCRIPTION_EXPIRED = "subscription_expired"  # اشتراك انتهى
    TENANT_ACTIVATED = "tenant_activated"  # تفعيل عميل
    TENANT_DEACTIVATED = "tenant_deactivated"  # تعطيل عميل
    SYSTEM = "system"  # إشعار نظام عام

class NotificationCreate(BaseModel):
    type: str
    title: str
    message: str
    tenant_id: Optional[str] = None
    data: Optional[dict] = None

class NotificationResponse(BaseModel):
    id: str
    type: str
    title: str
    message: str
    tenant_id: Optional[str] = None
    data: Optional[dict] = None
    is_read: bool = False
    created_at: str

class NotificationSettings(BaseModel):
    """إعدادات الإشعارات للمالك"""
    days_before_expiry: int = 15  # عدد الأيام قبل انتهاء الاشتراك للتنبيه (الافتراضي 15 يوم)
    email_notifications: bool = False  # إرسال بريد إلكتروني (معطل افتراضياً)
    push_notifications: bool = True  # إشعارات المتصفح
    notify_new_tenant: bool = True  # إشعار عند إضافة عميل جديد
    notify_tenant_status: bool = True  # إشعار عند تفعيل/تعطيل عميل

# User Models
class UserCreate(BaseModel):
    username: str
    email: EmailStr
    password: str
    full_name: str
    role: str = UserRole.CASHIER
    branch_id: Optional[str] = None
    permissions: List[str] = []
    tenant_id: Optional[str] = None  # للنظام متعدد المستأجرين

class UserLogin(BaseModel):
    email: str
    password: str

class UserResponse(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str
    username: Optional[str] = ""
    email: str
    full_name: Optional[str] = ""
    role: str
    branch_id: Optional[str] = None
    permissions: List[str] = []
    is_active: bool = True
    created_at: str
    tenant_id: Optional[str] = None

class UserUpdate(BaseModel):
    username: Optional[str] = None
    email: Optional[EmailStr] = None
    full_name: Optional[str] = None
    role: Optional[str] = None
    branch_id: Optional[str] = None
    permissions: Optional[List[str]] = None
    is_active: Optional[bool] = None

# Branch Models
class BranchCreate(BaseModel):
    name: str
    address: str
    phone: str
    email: Optional[str] = None

class BranchResponse(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str
    name: str
    address: str
    phone: str
    email: Optional[str] = None
    is_active: bool = True
    created_at: str

# Category Models
class CategoryCreate(BaseModel):
    name: str
    name_en: Optional[str] = None
    icon: Optional[str] = None
    image: Optional[str] = None
    color: Optional[str] = None
    sort_order: int = 0

class CategoryResponse(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str
    name: str
    name_en: Optional[str] = None
    icon: Optional[str] = None
    image: Optional[str] = None
    color: Optional[str] = None
    sort_order: int = 0
    is_active: bool = True

# Product Models with Pre-Manufacturing Cost
class ProductCreate(BaseModel):
    name: str
    name_en: Optional[str] = None
    category_id: str
    price: float
    cost: float = 0.0  # تكلفة ما قبل التصنيع
    operating_cost: float = 0.0  # تكلفة تشغيلية
    packaging_cost: float = 0.0  # تكلفة التغليف (للسفري/التوصيل)
    image: Optional[str] = None
    description: Optional[str] = None
    is_available: bool = True
    ingredients: List[Dict[str, Any]] = []
    barcode: Optional[str] = None
    finished_product_id: Optional[str] = None  # ربط بالمنتج النهائي للخصم التلقائي من الوصفة
    manufactured_product_id: Optional[str] = None  # ربط بالمنتج المصنع من النظام الجديد
    printer_ids: List[str] = []  # الطابعات المرتبطة بالمنتج

class ProductResponse(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str
    name: str
    name_en: Optional[str] = None
    category_id: str
    price: float
    cost: float = 0.0
    operating_cost: float = 0.0
    packaging_cost: float = 0.0  # تكلفة التغليف
    profit: float = 0.0  # حقل محسوب
    image: Optional[str] = None
    description: Optional[str] = None
    is_available: bool = True
    ingredients: List[Dict[str, Any]] = []
    barcode: Optional[str] = None
    finished_product_id: Optional[str] = None  # ربط بالمنتج النهائي
    manufactured_product_id: Optional[str] = None  # ربط بالمنتج المصنع من النظام الجديد
    printer_ids: List[str] = []  # الطابعات المرتبطة بالمنتج

# Inventory Models
class InventoryItemCreate(BaseModel):
    name: str
    name_en: Optional[str] = None
    unit: str
    quantity: float = 0.0
    min_quantity: float = 0.0
    cost_per_unit: float = 0.0
    branch_id: str
    item_type: str = "raw"  # raw or finished

class InventoryResponse(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str
    name: str
    name_en: Optional[str] = None
    unit: str
    quantity: float
    min_quantity: float
    cost_per_unit: float
    branch_id: str
    item_type: str
    last_updated: str

class InventoryTransaction(BaseModel):
    inventory_id: str
    transaction_type: str  # in or out
    quantity: float
    notes: Optional[str] = None

# Purchase Models - المشتريات
class PurchaseCreate(BaseModel):
    supplier_name: str
    invoice_number: Optional[str] = None
    items: List[Dict[str, Any]]  # [{inventory_id, quantity, cost_per_unit}]
    total_amount: float
    payment_method: str = "cash"
    payment_status: str = "paid"  # paid, pending, partial
    notes: Optional[str] = None
    branch_id: str

class PurchaseResponse(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str
    supplier_name: str
    invoice_number: Optional[str] = None
    items: List[Dict[str, Any]]
    total_amount: float
    payment_method: str
    payment_status: str
    notes: Optional[str] = None
    branch_id: str
    created_by: str
    created_at: str

# Expense Models - المصاريف اليومية
class ExpenseCreate(BaseModel):
    category: str  # rent, utilities, salaries, maintenance, supplies, other
    description: str
    amount: float
    payment_method: str = "cash"
    reference_number: Optional[str] = None
    branch_id: str
    date: Optional[str] = None

class ExpenseResponse(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str
    category: str
    description: str
    amount: float
    payment_method: str
    reference_number: Optional[str] = None
    branch_id: str
    created_by: str
    date: str
    created_at: str

# Operating Cost Models - التكاليف التشغيلية
class OperatingCostCreate(BaseModel):
    name: str
    cost_type: str  # fixed or variable
    amount: float
    frequency: str  # daily, weekly, monthly
    branch_id: str

# Table Models
class TableCreate(BaseModel):
    number: int
    capacity: int
    branch_id: str
    section: Optional[str] = None

class TableResponse(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str
    number: int
    capacity: int
    branch_id: str
    section: Optional[str] = None
    status: str = "available"
    current_order_id: Optional[str] = None

# Customer Models - إدارة العملاء
class CustomerCreate(BaseModel):
    name: str
    phone: str
    phone2: Optional[str] = None  # رقم إضافي
    address: Optional[str] = None
    area: Optional[str] = None  # المنطقة
    notes: Optional[str] = None
    is_blocked: bool = False  # حظر العميل

class CustomerResponse(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str
    name: str
    phone: str
    phone2: Optional[str] = None
    address: Optional[str] = None
    area: Optional[str] = None
    notes: Optional[str] = None
    is_blocked: bool = False
    total_orders: int = 0
    total_spent: float = 0.0
    last_order_date: Optional[str] = None
    created_at: str

# Order Models
class OrderItemCreate(BaseModel):
    product_id: str
    product_name: str
    quantity: int
    price: float
    cost: float = 0.0
    notes: Optional[str] = None

class OrderCreate(BaseModel):
    order_type: str = OrderType.DINE_IN
    table_id: Optional[str] = None
    customer_name: Optional[str] = None
    customer_phone: Optional[str] = None
    delivery_address: Optional[str] = None
    buzzer_number: Optional[str] = None  # رقم جهاز التنبيه للسفري
    items: List[OrderItemCreate]
    branch_id: str
    payment_method: str = PaymentMethod.CASH
    discount: float = 0.0
    notes: Optional[str] = None
    delivery_app: Optional[str] = None
    driver_id: Optional[str] = None
    auto_ready: bool = False  # الطلب جاهز تلقائياً

class OrderResponse(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str
    order_number: int
    order_type: str
    table_id: Optional[str] = None
    customer_name: Optional[str] = None
    customer_phone: Optional[str] = None
    delivery_address: Optional[str] = None
    buzzer_number: Optional[str] = None  # رقم جهاز التنبيه
    items: List[Dict[str, Any]]
    subtotal: float
    discount: float = 0.0  # Default for legacy orders
    tax: float = 0.0  # Default for legacy orders
    total: float
    total_cost: float = 0.0
    profit: float = 0.0
    branch_id: Optional[str] = None  # Made optional for customer orders
    cashier_id: Optional[str] = None  # Made optional for orders without cashier
    status: str = "pending"  # Default status
    payment_method: str = "cash"  # Default payment method
    payment_status: str = "pending"  # Default for legacy orders
    delivery_app: Optional[str] = None
    delivery_app_name: Optional[str] = None  # اسم شركة التوصيل
    delivery_commission: float = 0.0
    driver_id: Optional[str] = None
    driver_name: Optional[str] = None  # اسم السائق
    driver_phone: Optional[str] = None  # هاتف السائق
    notes: Optional[str] = None
    created_at: Optional[str] = None  # Made optional for legacy orders
    updated_at: Optional[str] = None  # Made optional for legacy orders
    tenant_id: Optional[str] = None  # Added for tenant filtering

# Shift Models
class ShiftCreate(BaseModel):
    cashier_id: str
    branch_id: str
    opening_cash: float

class ShiftClose(BaseModel):
    closing_cash: float
    notes: Optional[str] = None

# نموذج إغلاق الصندوق المتقدم مع جرد الفئات
class CashRegisterClose(BaseModel):
    denominations: Dict[str, int] = {}  # {"250": 5, "500": 10, "1000": 20, ...}
    notes: Optional[str] = None

class ShiftResponse(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str
    cashier_id: str
    cashier_name: str
    branch_id: str
    opening_cash: float
    closing_cash: Optional[float] = None
    expected_cash: Optional[float] = None
    cash_difference: Optional[float] = None
    total_sales: float = 0.0
    total_cost: float = 0.0
    gross_profit: float = 0.0
    total_orders: int = 0
    card_sales: float = 0.0
    cash_sales: float = 0.0
    credit_sales: float = 0.0
    delivery_app_sales: Dict[str, float] = {}
    driver_sales: float = 0.0  # مبيعات السائقين
    total_expenses: float = 0.0
    net_profit: float = 0.0
    started_at: str
    ended_at: Optional[str] = None
    status: str
    denominations: Optional[Dict[str, int]] = None  # تفاصيل الجرد
    cancelled_orders: int = 0  # عدد الطلبات الملغاة
    cancelled_amount: float = 0.0  # إجمالي الإلغاءات
    discounts_total: float = 0.0  # إجمالي الخصومات
    cancelled_by: List[Dict] = []  # تفاصيل من قام بالإلغاء

# Delivery Driver Models
class DriverCreate(BaseModel):
    name: str
    phone: str
    branch_id: str
    pin: str = "1234"  # رمز PIN الافتراضي - يجب تغييره
    user_id: Optional[str] = None  # ربط بحساب مستخدم

class DriverResponse(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str
    name: str
    phone: str
    branch_id: str
    is_available: bool = True
    current_order_id: Optional[str] = None
    total_deliveries: int = 0
    user_id: Optional[str] = None
    # معلومات الموقع
    location_lat: Optional[float] = None
    location_lng: Optional[float] = None
    location_updated_at: Optional[str] = None
    # معلومات الطلب الحالي (اختياري)
    current_order: Optional[Dict[str, Any]] = None
    is_active: bool = True

class DriverLocationUpdate(BaseModel):
    latitude: float
    longitude: float

# Delivery App Settings - إعدادات شركات التوصيل
class DeliveryAppSettingCreate(BaseModel):
    app_id: str
    name: str
    name_en: Optional[str] = None
    commission_type: str = "percentage"  # percentage or fixed
    commission_rate: float = 0.0  # نسبة الاستقطاع
    is_active: bool = True
    payment_terms: str = "weekly"  # daily, weekly, monthly
    contact_info: Optional[str] = None

# Currency Models
class Currency(BaseModel):
    code: str
    name: str
    symbol: str
    exchange_rate: float

# ==================== HR MODELS - إدارة الموارد البشرية ====================

class EmployeeCreate(BaseModel):
    name: str
    phone: str
    email: Optional[str] = None
    national_id: Optional[str] = None  # رقم الهوية
    position: str  # المسمى الوظيفي
    department: Optional[str] = None  # القسم
    branch_id: str
    hire_date: str  # تاريخ التعيين
    salary: float  # الراتب الأساسي
    salary_type: str = "monthly"  # monthly, daily, hourly
    work_hours_per_day: float = 8.0  # ساعات العمل اليومية
    user_id: Optional[str] = None  # ربط بحساب مستخدم

class EmployeeResponse(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str
    name: str
    phone: str
    email: Optional[str] = None
    national_id: Optional[str] = None
    position: str
    department: Optional[str] = None
    branch_id: str
    hire_date: str
    salary: float
    salary_type: str
    work_hours_per_day: float
    user_id: Optional[str] = None
    is_active: bool = True
    created_at: str
    tenant_id: Optional[str] = None

class EmployeeUpdate(BaseModel):
    name: Optional[str] = None
    phone: Optional[str] = None
    email: Optional[str] = None
    national_id: Optional[str] = None
    position: Optional[str] = None
    department: Optional[str] = None
    branch_id: Optional[str] = None
    salary: Optional[float] = None
    salary_type: Optional[str] = None
    work_hours_per_day: Optional[float] = None
    is_active: Optional[bool] = None

# نموذج الحضور والانصراف
class AttendanceCreate(BaseModel):
    employee_id: str
    date: str  # YYYY-MM-DD
    check_in: Optional[str] = None  # وقت الحضور HH:MM
    check_out: Optional[str] = None  # وقت الانصراف HH:MM
    status: str = "present"  # present, absent, late, early_leave, holiday
    notes: Optional[str] = None
    source: str = "manual"  # manual, fingerprint, system

class AttendanceResponse(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str
    employee_id: str
    employee_name: Optional[str] = None
    date: str
    check_in: Optional[str] = None
    check_out: Optional[str] = None
    worked_hours: Optional[float] = None
    status: str
    notes: Optional[str] = None
    source: str
    created_at: str

# نموذج السلف
class AdvanceCreate(BaseModel):
    employee_id: str
    amount: float
    reason: Optional[str] = None
    deduction_months: int = 1  # عدد أشهر الاستقطاع
    date: Optional[str] = None

class AdvanceResponse(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str
    employee_id: str
    employee_name: Optional[str] = None
    amount: float
    remaining_amount: float
    deducted_amount: float = 0
    deduction_months: int
    monthly_deduction: float
    reason: Optional[str] = None
    status: str  # pending, approved, rejected, paid
    date: str
    created_by: str
    created_at: str

# نموذج الخصومات
class DeductionCreate(BaseModel):
    employee_id: str
    deduction_type: str  # absence, late, early_leave, violation, other
    amount: Optional[float] = None  # مبلغ ثابت
    hours: Optional[float] = None  # عدد الساعات
    days: Optional[float] = None  # عدد الأيام
    reason: str
    date: str

class DeductionResponse(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str
    employee_id: str
    employee_name: Optional[str] = None
    deduction_type: str
    amount: float
    hours: Optional[float] = None
    days: Optional[float] = None
    reason: str
    date: str
    created_by: str
    created_at: str

# نموذج المكافآت والوقت الإضافي
class BonusCreate(BaseModel):
    employee_id: str
    bonus_type: str  # performance, overtime, holiday, other
    amount: Optional[float] = None  # مبلغ ثابت
    hours: Optional[float] = None  # ساعات إضافية
    reason: str
    date: str

class BonusResponse(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str
    employee_id: str
    employee_name: Optional[str] = None
    bonus_type: str
    amount: float
    hours: Optional[float] = None
    reason: str
    date: str
    created_by: str
    created_at: str

# نموذج كشف الراتب
class PayrollCreate(BaseModel):
    employee_id: str
    month: str  # YYYY-MM
    basic_salary: float
    total_deductions: float = 0
    total_bonuses: float = 0
    advance_deduction: float = 0
    net_salary: float
    notes: Optional[str] = None

class PayrollResponse(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str
    employee_id: str
    employee_name: Optional[str] = None
    month: str
    basic_salary: float
    worked_days: int = 0
    absent_days: int = 0
    late_hours: float = 0
    overtime_hours: float = 0
    total_deductions: float
    total_bonuses: float
    advance_deduction: float
    net_salary: float
    status: str  # draft, approved, paid
    notes: Optional[str] = None
    created_by: str
    created_at: str
    paid_at: Optional[str] = None

# ==================== INVENTORY TRANSFER MODELS - تحويلات المخزون ====================

class InventoryTransferCreate(BaseModel):
    from_branch_id: str  # الفرع المرسل (أو المخزن الرئيسي)
    to_branch_id: str  # الفرع المستلم
    items: List[Dict[str, Any]]  # [{inventory_id, quantity, notes}]
    transfer_type: str = "warehouse_to_branch"  # warehouse_to_branch, branch_to_warehouse, branch_to_branch
    notes: Optional[str] = None

class InventoryTransferResponse(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str
    transfer_number: int
    from_branch_id: str
    from_branch_name: Optional[str] = None
    to_branch_id: str
    to_branch_name: Optional[str] = None
    items: List[Dict[str, Any]]
    transfer_type: str
    status: str  # pending, approved, shipped, received, cancelled
    notes: Optional[str] = None
    created_by: str
    created_at: str
    approved_by: Optional[str] = None
    approved_at: Optional[str] = None
    received_by: Optional[str] = None
    received_at: Optional[str] = None

# نموذج طلب شراء
class PurchaseRequestCreate(BaseModel):
    branch_id: str  # الفرع الطالب
    items: List[Dict[str, Any]]  # [{name, quantity, unit, notes}]
    priority: str = "normal"  # urgent, high, normal, low
    notes: Optional[str] = None

class PurchaseRequestResponse(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str
    request_number: int
    branch_id: str
    branch_name: Optional[str] = None
    items: List[Dict[str, Any]]
    priority: str
    status: str  # pending, approved, ordered, received, cancelled
    notes: Optional[str] = None
    created_by: str
    created_at: str
    approved_by: Optional[str] = None
    approved_at: Optional[str] = None

# ==================== AUTH HELPERS ====================

def hash_password(password: str) -> str:
    return bcrypt.hashpw(password.encode('utf-8'), bcrypt.gensalt()).decode('utf-8')

def verify_password(password: str, hashed: str) -> bool:
    return bcrypt.checkpw(password.encode('utf-8'), hashed.encode('utf-8'))

def create_token(user_id: str, role: str, branch_id: Optional[str] = None) -> str:
    payload = {
        "user_id": user_id,
        "role": role,
        "branch_id": branch_id,
        "exp": datetime.now(timezone.utc) + timedelta(hours=JWT_EXPIRATION_HOURS)
    }
    return jwt.encode(payload, JWT_SECRET, algorithm=JWT_ALGORITHM)

async def get_current_user(credentials: HTTPAuthorizationCredentials = Depends(security)):
    try:
        token = credentials.credentials
        payload = jwt.decode(token, JWT_SECRET, algorithms=[JWT_ALGORITHM])
        user = await db.users.find_one({"id": payload["user_id"]}, {"_id": 0})
        if not user:
            raise HTTPException(status_code=401, detail="المستخدم غير موجود")
        return user
    except jwt.ExpiredSignatureError:
        raise HTTPException(status_code=401, detail="انتهت صلاحية الجلسة")
    except jwt.InvalidTokenError:
        raise HTTPException(status_code=401, detail="رمز غير صالح")

# دالة مساعدة للحصول على tenant_id للمستخدم
def get_user_tenant_id(user: dict) -> Optional[str]:
    """الحصول على tenant_id للمستخدم - Super Admin يستخدم tenant النظام"""
    if user.get("role") == UserRole.SUPER_ADMIN:
        # Super Admin يستخدم tenant النظام الافتراضي
        return user.get("tenant_id") or "system"
    # إذا لم يكن للمستخدم tenant_id، نستخدم "default"
    return user.get("tenant_id") or "default"

# دالة مساعدة لبناء query مع tenant_id
def build_tenant_query(user: dict, base_query: dict = None) -> dict:
    """بناء query مع فلترة tenant_id"""
    query = base_query.copy() if base_query else {}
    
    # Super Admin يرى كل شيء
    if user.get("role") == UserRole.SUPER_ADMIN:
        return query
    
    tenant_id = get_user_tenant_id(user)
    if tenant_id:
        query["tenant_id"] = tenant_id
    
    return query

# دالة مساعدة لبناء query مع فلترة الفرع
def build_branch_query(user: dict, base_query: dict = None) -> dict:
    """بناء query مع فلترة الفرع للمستخدمين المقيدين بفرع معين"""
    query = build_tenant_query(user, base_query)
    
    # إذا كان المستخدم مرتبط بفرع معين (ليس admin أو manager)
    user_branch_id = user.get("branch_id")
    user_role = user.get("role")
    
    # المستخدمون العاديون (cashier, supervisor, delivery) يرون فقط فرعهم
    if user_branch_id and user_role not in [UserRole.SUPER_ADMIN, UserRole.ADMIN, UserRole.MANAGER]:
        query["branch_id"] = user_branch_id
    
    return query

def user_can_access_branch(user: dict, branch_id: str) -> bool:
    """التحقق من صلاحية المستخدم للوصول لفرع معين"""
    if user.get("role") in [UserRole.SUPER_ADMIN, UserRole.ADMIN, UserRole.MANAGER]:
        return True
    return user.get("branch_id") == branch_id

# ==================== EMAIL SERVICE ====================

async def send_shift_report_email(shift_data: dict, recipient_emails: List[str]):
    if not SENDGRID_API_KEY or not recipient_emails:
        logger.warning("SendGrid not configured or no recipients")
        return
    
    html_content = f"""
    <html dir="rtl" style="font-family: 'Cairo', Arial, sans-serif;">
    <body style="background: #f5f5f5; padding: 20px;">
        <div style="max-width: 600px; margin: 0 auto; background: white; border-radius: 10px; padding: 30px;">
            <h1 style="color: #D4AF37; text-align: center;">تقرير إغلاق الصندوق</h1>
            <h2 style="color: #333;">Maestro EGP</h2>
            <hr style="border: 1px solid #D4AF37;">
            
            <table style="width: 100%; border-collapse: collapse;">
                <tr><td style="padding: 10px; border-bottom: 1px solid #eee;"><strong>أمين الصندوق:</strong></td><td>{shift_data.get('cashier_name', 'N/A')}</td></tr>
                <tr><td style="padding: 10px; border-bottom: 1px solid #eee;"><strong>تاريخ البدء:</strong></td><td>{shift_data.get('started_at', 'N/A')}</td></tr>
                <tr><td style="padding: 10px; border-bottom: 1px solid #eee;"><strong>تاريخ الإنتهاء:</strong></td><td>{shift_data.get('ended_at', 'N/A')}</td></tr>
                <tr><td style="padding: 10px; border-bottom: 1px solid #eee;"><strong>إجمالي المبيعات:</strong></td><td style="color: #10B981; font-weight: bold;">{shift_data.get('total_sales', 0):,.0f} د.ع</td></tr>
                <tr><td style="padding: 10px; border-bottom: 1px solid #eee;"><strong>إجمالي التكاليف:</strong></td><td>{shift_data.get('total_cost', 0):,.0f} د.ع</td></tr>
                <tr><td style="padding: 10px; border-bottom: 1px solid #eee;"><strong>الربح الإجمالي:</strong></td><td style="color: #10B981;">{shift_data.get('gross_profit', 0):,.0f} د.ع</td></tr>
                <tr><td style="padding: 10px; border-bottom: 1px solid #eee;"><strong>المصاريف:</strong></td><td style="color: #EF4444;">{shift_data.get('total_expenses', 0):,.0f} د.ع</td></tr>
                <tr><td style="padding: 10px; border-bottom: 1px solid #eee;"><strong>صافي الربح:</strong></td><td style="color: {'#10B981' if shift_data.get('net_profit', 0) >= 0 else '#EF4444'}; font-weight: bold;">{shift_data.get('net_profit', 0):,.0f} د.ع</td></tr>
                <tr><td style="padding: 10px; border-bottom: 1px solid #eee;"><strong>عدد الطلبات:</strong></td><td>{shift_data.get('total_orders', 0)}</td></tr>
                <tr><td style="padding: 10px; border-bottom: 1px solid #eee;"><strong>النقد المتوقع:</strong></td><td>{shift_data.get('expected_cash', 0):,.0f} د.ع</td></tr>
                <tr><td style="padding: 10px; border-bottom: 1px solid #eee;"><strong>النقد الفعلي:</strong></td><td>{shift_data.get('closing_cash', 0):,.0f} د.ع</td></tr>
                <tr><td style="padding: 10px;"><strong>الفرق:</strong></td><td style="color: {'#EF4444' if shift_data.get('cash_difference', 0) < 0 else '#10B981'}; font-weight: bold;">{shift_data.get('cash_difference', 0):,.0f} د.ع</td></tr>
            </table>
            
            <p style="color: #666; font-size: 12px; margin-top: 30px; text-align: center;">
                تم إرسال هذا التقرير تلقائياً من نظام Maestro EGP
            </p>
        </div>
    </body>
    </html>
    """
    
    try:
        message = Mail(
            from_email=SENDER_EMAIL,
            to_emails=recipient_emails,
            subject=f"تقرير إغلاق الصندوق - {shift_data.get('cashier_name', '')} - {datetime.now().strftime('%Y-%m-%d')}",
            html_content=html_content
        )
        sg = SendGridAPIClient(SENDGRID_API_KEY)
        sg.send(message)
        logger.info(f"Shift report email sent to {recipient_emails}")
    except Exception as e:
        logger.error(f"Failed to send email: {e}")

async def send_welcome_email(recipient_email: str, tenant_name: str, owner_name: str, username: str, password: str):
    """إرسال بريد ترحيبي للعميل الجديد مع بيانات الدخول"""
    if not SENDGRID_API_KEY or not recipient_email:
        logger.warning("SendGrid not configured or no recipient")
        return
    
    frontend_url = os.environ.get('FRONTEND_URL', 'https://maestroegp.com')
    
    html_content = f"""
    <html dir="rtl" style="font-family: 'Cairo', Arial, sans-serif;">
    <body style="background: #1a1a2e; padding: 40px 20px;">
        <div style="max-width: 600px; margin: 0 auto; background: linear-gradient(145deg, #16213e, #1a1a2e); border-radius: 20px; padding: 40px; box-shadow: 0 20px 60px rgba(0,0,0,0.3); border: 1px solid rgba(212, 175, 55, 0.2);">
            
            <!-- Header -->
            <div style="text-align: center; margin-bottom: 30px;">
                <h1 style="color: #D4AF37; font-size: 32px; margin: 0;">👑 Maestro EGP</h1>
                <p style="color: #9ca3af; font-size: 14px; margin-top: 10px;">نظام إدارة المطاعم والكافيهات</p>
            </div>
            
            <!-- Welcome Message -->
            <div style="background: rgba(212, 175, 55, 0.1); border-radius: 15px; padding: 25px; margin-bottom: 25px; border: 1px solid rgba(212, 175, 55, 0.2);">
                <h2 style="color: #ffffff; margin: 0 0 15px 0; font-size: 22px;">مرحباً {owner_name}! 🎉</h2>
                <p style="color: #d1d5db; margin: 0; line-height: 1.8;">
                    تم إنشاء حسابك في <strong style="color: #D4AF37;">{tenant_name}</strong> بنجاح على منصة Maestro EGP.
                    يمكنك الآن البدء في إدارة مطعمك/الكافيه الخاص بك بكل سهولة.
                </p>
            </div>
            
            <!-- Credentials Box -->
            <div style="background: #0f172a; border-radius: 15px; padding: 25px; margin-bottom: 25px; border: 1px solid #334155;">
                <h3 style="color: #D4AF37; margin: 0 0 20px 0; font-size: 18px;">🔐 بيانات تسجيل الدخول</h3>
                
                <div style="margin-bottom: 15px;">
                    <p style="color: #9ca3af; margin: 0 0 5px 0; font-size: 12px;">البريد الإلكتروني / اسم المستخدم:</p>
                    <p style="color: #ffffff; background: #1e293b; padding: 12px 15px; border-radius: 8px; margin: 0; font-family: monospace; font-size: 14px;">{username}</p>
                </div>
                
                <div>
                    <p style="color: #9ca3af; margin: 0 0 5px 0; font-size: 12px;">كلمة المرور:</p>
                    <p style="color: #ffffff; background: #1e293b; padding: 12px 15px; border-radius: 8px; margin: 0; font-family: monospace; font-size: 14px;">{password}</p>
                </div>
            </div>
            
            <!-- Login Button -->
            <div style="text-align: center; margin-bottom: 30px;">
                <a href="{frontend_url}/login" style="display: inline-block; background: linear-gradient(145deg, #D4AF37, #B8860B); color: #000000; text-decoration: none; padding: 15px 40px; border-radius: 10px; font-weight: bold; font-size: 16px; box-shadow: 0 4px 15px rgba(212, 175, 55, 0.3);">
                    🚀 تسجيل الدخول الآن
                </a>
            </div>
            
            <!-- Features Section -->
            <div style="background: rgba(16, 185, 129, 0.1); border-radius: 15px; padding: 20px; margin-bottom: 25px; border: 1px solid rgba(16, 185, 129, 0.2);">
                <h3 style="color: #10B981; margin: 0 0 15px 0; font-size: 16px;">✨ ماذا يمكنك أن تفعل؟</h3>
                <ul style="color: #d1d5db; margin: 0; padding-right: 20px; line-height: 2;">
                    <li>إدارة الطلبات (محلي، سفري، توصيل)</li>
                    <li>تتبع السائقين على الخريطة</li>
                    <li>إدارة المخزون والمنتجات</li>
                    <li>تقارير المبيعات والأرباح</li>
                    <li>إدارة الموظفين والرواتب</li>
                    <li>نظام الكول سنتر</li>
                </ul>
            </div>
            
            <!-- Instructions -->
            <div style="background: rgba(59, 130, 246, 0.1); border-radius: 15px; padding: 20px; margin-bottom: 25px; border: 1px solid rgba(59, 130, 246, 0.2);">
                <h3 style="color: #3B82F6; margin: 0 0 15px 0; font-size: 16px;">📋 خطوات البدء</h3>
                <ol style="color: #d1d5db; margin: 0; padding-right: 20px; line-height: 2;">
                    <li>سجل دخولك باستخدام البيانات أعلاه</li>
                    <li>قم بتغيير كلمة المرور من الإعدادات</li>
                    <li>أضف الفروع والموظفين</li>
                    <li>أضف التصنيفات والمنتجات</li>
                    <li>ابدأ استقبال الطلبات! 🎯</li>
                </ol>
            </div>
            
            <!-- Warning -->
            <div style="background: rgba(239, 68, 68, 0.1); border-radius: 10px; padding: 15px; margin-bottom: 25px; border: 1px solid rgba(239, 68, 68, 0.2);">
                <p style="color: #EF4444; margin: 0; font-size: 13px;">
                    ⚠️ <strong>هام:</strong> يرجى تغيير كلمة المرور فور تسجيل الدخول للحفاظ على أمان حسابك.
                </p>
            </div>
            
            <!-- Footer -->
            <div style="text-align: center; border-top: 1px solid #334155; padding-top: 20px;">
                <p style="color: #6b7280; font-size: 12px; margin: 0;">
                    للدعم الفني: support@maestroegp.com
                </p>
                <p style="color: #6b7280; font-size: 11px; margin: 10px 0 0 0;">
                    © {datetime.now().year} Maestro EGP - جميع الحقوق محفوظة
                </p>
            </div>
        </div>
    </body>
    </html>
    """
    
    try:
        message = Mail(
            from_email=SENDER_EMAIL,
            to_emails=[recipient_email],
            subject=f"🎉 مرحباً في {tenant_name} - بيانات الدخول",
            html_content=html_content
        )
        sg = SendGridAPIClient(SENDGRID_API_KEY)
        sg.send(message)
        logger.info(f"Welcome email sent to {recipient_email}")
    except Exception as e:
        logger.error(f"Failed to send welcome email: {e}")

# ==================== HELPER FUNCTIONS ====================

async def get_delivery_app_commission(app_id: str) -> float:
    """Get commission rate for delivery app"""
    setting = await db.delivery_app_settings.find_one({"app_id": app_id}, {"_id": 0})
    if setting:
        return setting.get("commission_rate", 0)
    return 0

async def calculate_order_cost(items: List[Dict]) -> float:
    """Calculate total cost for order items"""
    total_cost = 0
    for item in items:
        product = await db.products.find_one({"id": item.get("product_id")}, {"_id": 0})
        if product:
            item_cost = (product.get("cost", 0) + product.get("operating_cost", 0)) * item.get("quantity", 1)
            total_cost += item_cost
    return total_cost

# ==================== AUTH ROUTES ====================

@api_router.post("/auth/register")
async def register(user: UserCreate):
    existing = await db.users.find_one({"$or": [{"email": user.email}, {"username": user.username}]})
    if existing:
        raise HTTPException(status_code=400, detail="المستخدم موجود بالفعل")
    
    user_doc = {
        "id": str(uuid.uuid4()),
        "username": user.username,
        "email": user.email,
        "password": hash_password(user.password),
        "full_name": user.full_name,
        "role": user.role,
        "branch_id": user.branch_id,
        "permissions": user.permissions,
        "is_active": True,
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    await db.users.insert_one(user_doc)
    del user_doc["password"]
    del user_doc["_id"]
    token = create_token(user_doc["id"], user_doc["role"], user_doc.get("branch_id"))
    return {"user": user_doc, "token": token}

@api_router.post("/auth/login")
async def login(credentials: UserLogin):
    user = await db.users.find_one({"email": credentials.email}, {"_id": 0})
    if not user or not verify_password(credentials.password, user.get("password_hash", user.get("password", ""))):
        raise HTTPException(status_code=401, detail="بيانات الدخول غير صحيحة")
    if not user.get("is_active", True):
        raise HTTPException(status_code=401, detail="الحساب معطل")
    
    # إزالة كلمة المرور من الاستجابة
    if "password" in user:
        del user["password"]
    if "password_hash" in user:
        del user["password_hash"]
    token = create_token(user["id"], user["role"], user.get("branch_id"))
    return {"user": user, "token": token}

@api_router.get("/auth/me")
async def get_me(current_user: dict = Depends(get_current_user)):
    user = dict(current_user)
    if "password" in user:
        del user["password"]
    return user

# ==================== USER ROUTES ====================

@api_router.get("/users", response_model=List[UserResponse])
async def get_users(current_user: dict = Depends(get_current_user)):
    if current_user["role"] not in [UserRole.ADMIN, UserRole.MANAGER, UserRole.SUPER_ADMIN]:
        raise HTTPException(status_code=403, detail="غير مصرح")
    
    # فلترة المستخدمين حسب tenant_id
    query = build_tenant_query(current_user)
    users = await db.users.find(query, {"_id": 0, "password": 0}).to_list(1000)
    return users

@api_router.put("/users/{user_id}")
async def update_user(user_id: str, update: UserUpdate, current_user: dict = Depends(get_current_user)):
    if current_user["role"] not in [UserRole.ADMIN, UserRole.MANAGER, UserRole.SUPER_ADMIN]:
        raise HTTPException(status_code=403, detail="غير مصرح")
    
    # التحقق من أن المستخدم ينتمي لنفس الـ tenant
    query = build_tenant_query(current_user, {"id": user_id})
    user = await db.users.find_one(query)
    if not user:
        raise HTTPException(status_code=404, detail="المستخدم غير موجود")
    
    update_data = {k: v for k, v in update.model_dump().items() if v is not None}
    
    # التحقق من عدم تكرار البريد الإلكتروني أو اسم المستخدم
    if update_data.get("email"):
        existing = await db.users.find_one({"email": update_data["email"], "id": {"$ne": user_id}})
        if existing:
            raise HTTPException(status_code=400, detail="البريد الإلكتروني مستخدم بالفعل")
    
    if update_data.get("username"):
        existing = await db.users.find_one({"username": update_data["username"], "id": {"$ne": user_id}})
        if existing:
            raise HTTPException(status_code=400, detail="اسم المستخدم مستخدم بالفعل")
    
    if update_data:
        await db.users.update_one({"id": user_id}, {"$set": update_data})
    
    user = await db.users.find_one({"id": user_id}, {"_id": 0, "password": 0})
    return user

@api_router.delete("/users/{user_id}")
async def delete_user(user_id: str, current_user: dict = Depends(get_current_user)):
    if current_user["role"] not in [UserRole.ADMIN, UserRole.SUPER_ADMIN]:
        raise HTTPException(status_code=403, detail="غير مصرح")
    
    # التحقق من أن المستخدم ينتمي لنفس الـ tenant
    query = build_tenant_query(current_user, {"id": user_id})
    user = await db.users.find_one(query)
    if not user:
        raise HTTPException(status_code=404, detail="المستخدم غير موجود")
    
    await db.users.delete_one({"id": user_id})
    return {"message": "تم حذف المستخدم"}

@api_router.post("/users", response_model=UserResponse)
async def create_user(user: UserCreate, current_user: dict = Depends(get_current_user)):
    """إنشاء مستخدم جديد مع tenant_id تلقائياً"""
    if current_user["role"] not in [UserRole.ADMIN, UserRole.MANAGER, UserRole.SUPER_ADMIN]:
        raise HTTPException(status_code=403, detail="غير مصرح")
    
    # التحقق من عدم وجود المستخدم
    existing = await db.users.find_one({"$or": [{"email": user.email}, {"username": user.username}]})
    if existing:
        raise HTTPException(status_code=400, detail="المستخدم موجود بالفعل")
    
    # الحصول على tenant_id من المستخدم الحالي
    tenant_id = get_user_tenant_id(current_user)
    
    user_doc = {
        "id": str(uuid.uuid4()),
        "username": user.username,
        "email": user.email,
        "password": hash_password(user.password),
        "full_name": user.full_name,
        "role": user.role,
        "branch_id": user.branch_id,
        "permissions": user.permissions,
        "tenant_id": tenant_id,  # إضافة tenant_id تلقائياً
        "is_active": True,
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    await db.users.insert_one(user_doc)
    del user_doc["password"]
    del user_doc["_id"]
    return user_doc

class PasswordReset(BaseModel):
    new_password: str

@api_router.put("/users/{user_id}/reset-password")
async def reset_user_password(user_id: str, data: PasswordReset, current_user: dict = Depends(get_current_user)):
    """إعادة تعيين كلمة مرور المستخدم"""
    if current_user["role"] not in [UserRole.ADMIN, UserRole.MANAGER, UserRole.SUPER_ADMIN]:
        raise HTTPException(status_code=403, detail="غير مصرح")
    
    # التحقق من أن المستخدم ينتمي لنفس الـ tenant
    query = build_tenant_query(current_user, {"id": user_id})
    user = await db.users.find_one(query)
    if not user:
        raise HTTPException(status_code=404, detail="المستخدم غير موجود")
    
    hashed = hash_password(data.new_password)
    await db.users.update_one({"id": user_id}, {"$set": {"password": hashed}})
    
    return {"message": "تم تغيير كلمة المرور بنجاح"}

# ==================== BRANCH ROUTES ====================

@api_router.post("/branches", response_model=BranchResponse)
async def create_branch(branch: BranchCreate, current_user: dict = Depends(get_current_user)):
    if current_user["role"] not in [UserRole.ADMIN, UserRole.SUPER_ADMIN]:
        raise HTTPException(status_code=403, detail="غير مصرح")
    
    branch_doc = {
        "id": str(uuid.uuid4()),
        **branch.model_dump(),
        "tenant_id": get_user_tenant_id(current_user),  # فصل البيانات
        "is_active": True,
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    await db.branches.insert_one(branch_doc)
    del branch_doc["_id"]
    return branch_doc

@api_router.get("/branches", response_model=List[BranchResponse])
async def get_branches(current_user: dict = Depends(get_current_user), include_inactive: bool = False):
    # Super Admin يرى الفروع الخاصة به (tenant_id الخاص به)
    if current_user.get("role") == UserRole.SUPER_ADMIN:
        owner_tenant_id = current_user.get("tenant_id") or "default"
        query = {"tenant_id": owner_tenant_id}
    else:
        query = build_tenant_query(current_user)  # فلترة حسب tenant_id
    
    # المستخدمون المرتبطون بفرع معين يرون فقط فرعهم
    user_branch_id = current_user.get("branch_id")
    user_role = current_user.get("role")
    
    if user_branch_id and user_role not in [UserRole.SUPER_ADMIN, UserRole.ADMIN, UserRole.MANAGER]:
        query["id"] = user_branch_id
    
    # إخفاء الفروع المعطّلة إلا إذا طُلب عرضها
    if not include_inactive:
        query["is_active"] = {"$ne": False}
    
    # إخفاء الفروع الافتراضية (الفرع الرئيسي، Main Branch، إلخ)
    default_branch_names = ["الفرع الرئيسي", "Main Branch", "الفرع الثاني", "فرع المالك الرئيسي"]
    query["name"] = {"$nin": default_branch_names}
    
    branches = await db.branches.find(query, {"_id": 0}).to_list(100)
    return branches

@api_router.get("/branches/{branch_id}", response_model=BranchResponse)
async def get_branch(branch_id: str, current_user: dict = Depends(get_current_user)):
    query = build_tenant_query(current_user, {"id": branch_id})
    branch = await db.branches.find_one(query, {"_id": 0})
    if not branch:
        raise HTTPException(status_code=404, detail="الفرع غير موجود")
    return branch

@api_router.put("/branches/{branch_id}")
async def update_branch(branch_id: str, branch: BranchCreate, current_user: dict = Depends(get_current_user)):
    if current_user["role"] not in [UserRole.ADMIN, UserRole.SUPER_ADMIN]:
        raise HTTPException(status_code=403, detail="غير مصرح")
    query = build_tenant_query(current_user, {"id": branch_id})
    await db.branches.update_one(query, {"$set": branch.model_dump()})
    return await db.branches.find_one({"id": branch_id}, {"_id": 0})

@api_router.delete("/branches/{branch_id}")
async def delete_branch(branch_id: str, current_user: dict = Depends(get_current_user)):
    if current_user["role"] not in [UserRole.ADMIN, UserRole.SUPER_ADMIN]:
        raise HTTPException(status_code=403, detail="غير مصرح")
    # Check if branch has users or orders
    users_count = await db.users.count_documents({"branch_id": branch_id})
    if users_count > 0:
        raise HTTPException(status_code=400, detail="لا يمكن حذف الفرع - يوجد مستخدمين مرتبطين به")
    await db.branches.update_one({"id": branch_id}, {"$set": {"is_active": False}})
    return {"message": "تم تعطيل الفرع"}

# ==================== KITCHEN SECTIONS ROUTES ====================

@api_router.post("/kitchen-sections")
async def create_kitchen_section(section: dict, current_user: dict = Depends(get_current_user)):
    if current_user["role"] not in [UserRole.ADMIN, UserRole.MANAGER]:
        raise HTTPException(status_code=403, detail="غير مصرح")
    
    tenant_id = get_user_tenant_id(current_user)
    section_doc = {
        "id": str(uuid.uuid4()),
        "name": section.get("name"),
        "name_en": section.get("name_en"),
        "color": section.get("color", "#D4AF37"),
        "icon": section.get("icon", "🍳"),
        "printer_id": section.get("printer_id"),
        "branch_id": section.get("branch_id"),
        "tenant_id": tenant_id,  # فصل البيانات لكل عميل
        "sort_order": section.get("sort_order", 0),
        "is_active": True,
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    await db.kitchen_sections.insert_one(section_doc)
    del section_doc["_id"]
    return section_doc

@api_router.get("/kitchen-sections")
async def get_kitchen_sections(branch_id: Optional[str] = None, current_user: dict = Depends(get_current_user)):
    query = build_tenant_query(current_user)  # فلترة حسب tenant_id
    if branch_id:
        query["branch_id"] = branch_id
    sections = await db.kitchen_sections.find(query, {"_id": 0}).sort("sort_order", 1).to_list(100)
    return sections

@api_router.put("/kitchen-sections/{section_id}")
async def update_kitchen_section(section_id: str, section: dict, current_user: dict = Depends(get_current_user)):
    if current_user["role"] not in [UserRole.ADMIN, UserRole.MANAGER]:
        raise HTTPException(status_code=403, detail="غير مصرح")
    
    query = build_tenant_query(current_user, {"id": section_id})
    update_data = {k: v for k, v in section.items() if k != "id"}
    await db.kitchen_sections.update_one(query, {"$set": update_data})
    return await db.kitchen_sections.find_one({"id": section_id}, {"_id": 0})

@api_router.delete("/kitchen-sections/{section_id}")
async def delete_kitchen_section(section_id: str, current_user: dict = Depends(get_current_user)):
    if current_user["role"] not in [UserRole.ADMIN, UserRole.MANAGER]:
        raise HTTPException(status_code=403, detail="غير مصرح")
    query = build_tenant_query(current_user, {"id": section_id})
    await db.kitchen_sections.delete_one(query)
    return {"message": "تم الحذف"}

@api_router.put("/categories/{category_id}/kitchen-section")
async def assign_category_to_kitchen_section(category_id: str, data: dict, current_user: dict = Depends(get_current_user)):
    """Assign a category to a kitchen section"""
    if current_user["role"] not in [UserRole.ADMIN, UserRole.MANAGER]:
        raise HTTPException(status_code=403, detail="غير مصرح")
    
    kitchen_section_id = data.get("kitchen_section_id")
    await db.categories.update_one(
        {"id": category_id}, 
        {"$set": {"kitchen_section_id": kitchen_section_id}}
    )
    return await db.categories.find_one({"id": category_id}, {"_id": 0})

# ==================== CATEGORY ROUTES ====================

@api_router.post("/categories", response_model=CategoryResponse)
async def create_category(category: CategoryCreate, current_user: dict = Depends(get_current_user)):
    if current_user["role"] not in [UserRole.ADMIN, UserRole.MANAGER, UserRole.SUPER_ADMIN]:
        raise HTTPException(status_code=403, detail="غير مصرح")
    
    cat_doc = {
        "id": str(uuid.uuid4()),
        **category.model_dump(),
        "tenant_id": get_user_tenant_id(current_user),  # فصل البيانات
        "is_active": True
    }
    await db.categories.insert_one(cat_doc)
    del cat_doc["_id"]
    return cat_doc

@api_router.get("/categories", response_model=List[CategoryResponse])
async def get_categories(current_user: dict = Depends(get_current_user)):
    # Super Admin يرى الفئات الخاصة به (tenant_id الخاص به)
    if current_user.get("role") == UserRole.SUPER_ADMIN:
        owner_tenant_id = current_user.get("tenant_id") or "default"
        query = {"tenant_id": owner_tenant_id}
    else:
        query = build_tenant_query(current_user)  # فلترة حسب tenant_id
    categories = await db.categories.find(query, {"_id": 0}).sort("sort_order", 1).to_list(100)
    return categories

@api_router.put("/categories/{category_id}")
async def update_category(category_id: str, category: CategoryCreate, current_user: dict = Depends(get_current_user)):
    if current_user["role"] not in [UserRole.ADMIN, UserRole.MANAGER, UserRole.SUPER_ADMIN]:
        raise HTTPException(status_code=403, detail="غير مصرح")
    query = build_tenant_query(current_user, {"id": category_id})
    await db.categories.update_one(query, {"$set": category.model_dump()})
    return await db.categories.find_one({"id": category_id}, {"_id": 0})

@api_router.delete("/categories/{category_id}")
async def delete_category(category_id: str, current_user: dict = Depends(get_current_user)):
    if current_user["role"] not in [UserRole.ADMIN, UserRole.MANAGER, UserRole.SUPER_ADMIN]:
        raise HTTPException(status_code=403, detail="غير مصرح")
    query = build_tenant_query(current_user, {"id": category_id})
    await db.categories.delete_one(query)
    return {"message": "تم الحذف"}

# ==================== PRODUCT ROUTES ====================

@api_router.post("/products", response_model=ProductResponse)
async def create_product(product: ProductCreate, current_user: dict = Depends(get_current_user)):
    if current_user["role"] not in [UserRole.ADMIN, UserRole.MANAGER, UserRole.SUPER_ADMIN]:
        raise HTTPException(status_code=403, detail="غير مصرح")
    
    # Calculate profit
    profit = product.price - product.cost - product.operating_cost
    
    prod_doc = {
        "id": str(uuid.uuid4()),
        **product.model_dump(),
        "tenant_id": get_user_tenant_id(current_user),  # فصل البيانات
        "profit": profit
    }
    await db.products.insert_one(prod_doc)
    del prod_doc["_id"]
    return prod_doc

@api_router.get("/products", response_model=List[ProductResponse])
async def get_products(category_id: Optional[str] = None, current_user: dict = Depends(get_current_user)):
    # Super Admin يرى المنتجات الخاصة به (tenant_id الخاص به)
    if current_user.get("role") == UserRole.SUPER_ADMIN:
        owner_tenant_id = current_user.get("tenant_id") or "default"
        query = {"tenant_id": owner_tenant_id}
    else:
        query = build_tenant_query(current_user)  # فلترة حسب tenant_id
    if category_id:
        query["category_id"] = category_id
    products = await db.products.find(query, {"_id": 0}).to_list(1000)
    # Calculate profit for each product
    for p in products:
        p["profit"] = p.get("price", 0) - p.get("cost", 0) - p.get("operating_cost", 0)
    return products

@api_router.get("/products/{product_id}")
async def get_product(product_id: str, current_user: dict = Depends(get_current_user)):
    query = build_tenant_query(current_user, {"id": product_id})
    product = await db.products.find_one(query, {"_id": 0})
    if not product:
        raise HTTPException(status_code=404, detail="المنتج غير موجود")
    product["profit"] = product.get("price", 0) - product.get("cost", 0) - product.get("operating_cost", 0)
    return product

@api_router.put("/products/{product_id}")
async def update_product(product_id: str, product: ProductCreate, current_user: dict = Depends(get_current_user)):
    if current_user["role"] not in [UserRole.ADMIN, UserRole.MANAGER, UserRole.SUPER_ADMIN]:
        raise HTTPException(status_code=403, detail="غير مصرح")
    
    profit = product.price - product.cost - product.operating_cost
    update_data = {**product.model_dump(), "profit": profit}
    
    query = build_tenant_query(current_user, {"id": product_id})
    await db.products.update_one(query, {"$set": update_data})
    return await db.products.find_one({"id": product_id}, {"_id": 0})

@api_router.delete("/products/{product_id}")
async def delete_product(product_id: str, current_user: dict = Depends(get_current_user)):
    if current_user["role"] not in [UserRole.ADMIN, UserRole.MANAGER]:
        raise HTTPException(status_code=403, detail="غير مصرح")
    await db.products.delete_one({"id": product_id})
    return {"message": "تم الحذف"}

# ==================== INVENTORY ROUTES ====================

@api_router.post("/inventory", response_model=InventoryResponse)
async def create_inventory_item(item: InventoryItemCreate, current_user: dict = Depends(get_current_user)):
    if current_user["role"] not in [UserRole.ADMIN, UserRole.MANAGER, UserRole.SUPERVISOR]:
        raise HTTPException(status_code=403, detail="غير مصرح")
    
    inv_doc = {
        "id": str(uuid.uuid4()),
        **item.model_dump(),
        "last_updated": datetime.now(timezone.utc).isoformat()
    }
    await db.inventory.insert_one(inv_doc)
    del inv_doc["_id"]
    return inv_doc

@api_router.get("/inventory", response_model=List[InventoryResponse])
async def get_inventory(branch_id: Optional[str] = None, item_type: Optional[str] = None, current_user: dict = Depends(get_current_user)):
    query = build_tenant_query(current_user)  # فلترة حسب tenant_id
    if branch_id:
        query["branch_id"] = branch_id
    if item_type:
        query["item_type"] = item_type
    items = await db.inventory.find(query, {"_id": 0}).to_list(1000)
    return items

@api_router.post("/inventory/transaction")
async def inventory_transaction(transaction: InventoryTransaction, current_user: dict = Depends(get_current_user)):
    query = build_tenant_query(current_user, {"id": transaction.inventory_id})
    item = await db.inventory.find_one(query)
    if not item:
        raise HTTPException(status_code=404, detail="الصنف غير موجود")
    
    new_qty = item["quantity"]
    if transaction.transaction_type == "in":
        new_qty += transaction.quantity
    else:
        new_qty -= transaction.quantity
        if new_qty < 0:
            raise HTTPException(status_code=400, detail="الكمية غير كافية")
    
    await db.inventory.update_one(
        {"id": transaction.inventory_id},
        {"$set": {"quantity": new_qty, "last_updated": datetime.now(timezone.utc).isoformat()}}
    )
    
    # Log transaction
    trans_doc = {
        "id": str(uuid.uuid4()),
        "inventory_id": transaction.inventory_id,
        "transaction_type": transaction.transaction_type,
        "quantity": transaction.quantity,
        "notes": transaction.notes,
        "user_id": current_user["id"],
        "tenant_id": get_user_tenant_id(current_user),  # فصل البيانات
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    await db.inventory_transactions.insert_one(trans_doc)
    
    return {"message": "تمت العملية بنجاح", "new_quantity": new_qty}

# ==================== PURCHASE ROUTES - المشتريات ====================

@api_router.post("/purchases")
async def create_purchase(purchase: PurchaseCreate, current_user: dict = Depends(get_current_user)):
    if current_user["role"] not in [UserRole.ADMIN, UserRole.MANAGER, UserRole.SUPERVISOR]:
        raise HTTPException(status_code=403, detail="غير مصرح")
    
    purchase_doc = {
        "id": str(uuid.uuid4()),
        **purchase.model_dump(),
        "created_by": current_user["id"],
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    await db.purchases.insert_one(purchase_doc)
    
    # Update inventory quantities
    for item in purchase.items:
        await db.inventory.update_one(
            {"id": item.get("inventory_id")},
            {
                "$inc": {"quantity": item.get("quantity", 0)},
                "$set": {
                    "cost_per_unit": item.get("cost_per_unit", 0),
                    "last_updated": datetime.now(timezone.utc).isoformat()
                }
            }
        )
    
    del purchase_doc["_id"]
    return purchase_doc

@api_router.get("/purchases")
async def get_purchases(
    branch_id: Optional[str] = None,
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    query = {}
    if branch_id:
        query["branch_id"] = branch_id
    if start_date:
        query["created_at"] = {"$gte": start_date}
    if end_date:
        query.setdefault("created_at", {})["$lte"] = end_date + "T23:59:59"
    
    purchases = await db.purchases.find(query, {"_id": 0}).sort("created_at", -1).to_list(500)
    return purchases

# ==================== EXPENSE ROUTES - المصاريف ====================

@api_router.post("/expenses")
async def create_expense(expense: ExpenseCreate, current_user: dict = Depends(get_current_user)):
    if current_user["role"] not in [UserRole.ADMIN, UserRole.MANAGER, UserRole.SUPERVISOR, UserRole.SUPER_ADMIN]:
        raise HTTPException(status_code=403, detail="غير مصرح")
    
    expense_doc = {
        "id": str(uuid.uuid4()),
        **expense.model_dump(),
        "tenant_id": get_user_tenant_id(current_user),  # فصل البيانات
        "date": expense.date or datetime.now(timezone.utc).strftime("%Y-%m-%d"),
        "created_by": current_user["id"],
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    await db.expenses.insert_one(expense_doc)
    del expense_doc["_id"]
    return expense_doc

@api_router.get("/expenses")
async def get_expenses(
    branch_id: Optional[str] = None,
    category: Optional[str] = None,
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    query = build_tenant_query(current_user)  # فلترة حسب tenant_id
    if branch_id:
        query["branch_id"] = branch_id
    if category:
        query["category"] = category
    if start_date:
        query["date"] = {"$gte": start_date}
    if end_date:
        query.setdefault("date", {})["$lte"] = end_date
    
    expenses = await db.expenses.find(query, {"_id": 0}).sort("date", -1).to_list(500)
    return expenses

@api_router.get("/expenses/categories")
async def get_expense_categories():
    """جلب التصنيفات الافتراضية"""
    return [
        {"id": "rent", "name": "إيجار"},
        {"id": "utilities", "name": "كهرباء وماء"},
        {"id": "gas", "name": "غاز"},
        {"id": "salaries", "name": "رواتب"},
        {"id": "advance", "name": "سلف"},
        {"id": "maintenance", "name": "صيانة"},
        {"id": "supplies", "name": "مستلزمات"},
        {"id": "marketing", "name": "تسويق"},
        {"id": "transport", "name": "نقل"},
        {"id": "other", "name": "أخرى"}
    ]

@api_router.get("/expense-categories")
async def get_custom_expense_categories(current_user: dict = Depends(get_current_user)):
    """جلب التصنيفات المخصصة"""
    tenant_id = current_user.get("tenant_id")
    query = {"tenant_id": tenant_id} if tenant_id else {}
    
    categories = await db.expense_categories.find(
        query,
        {"_id": 0}
    ).to_list(100)
    
    return categories

@api_router.post("/expense-categories")
async def create_expense_category(category: Dict[str, Any], current_user: dict = Depends(get_current_user)):
    """إنشاء تصنيف مصاريف جديد"""
    tenant_id = current_user.get("tenant_id")
    
    # التحقق من عدم وجود التصنيف مسبقاً
    existing = await db.expense_categories.find_one({
        "id": category.get("id"),
        "tenant_id": tenant_id
    })
    
    if existing:
        return {"message": "التصنيف موجود بالفعل", "category": existing}
    
    category_doc = {
        "id": category.get("id"),
        "name": category.get("name"),
        "icon": category.get("icon", "🏷️"),
        "tenant_id": tenant_id,
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    
    await db.expense_categories.insert_one(category_doc)
    category_doc.pop("_id", None)
    
    return category_doc

# ==================== OPERATING COST ROUTES - التكاليف التشغيلية ====================

@api_router.post("/operating-costs")
async def create_operating_cost(cost: OperatingCostCreate, current_user: dict = Depends(get_current_user)):
    if current_user["role"] not in [UserRole.ADMIN, UserRole.MANAGER]:
        raise HTTPException(status_code=403, detail="غير مصرح")
    
    cost_doc = {
        "id": str(uuid.uuid4()),
        **cost.model_dump(),
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    await db.operating_costs.insert_one(cost_doc)
    del cost_doc["_id"]
    return cost_doc

@api_router.get("/operating-costs")
async def get_operating_costs(branch_id: Optional[str] = None, current_user: dict = Depends(get_current_user)):
    query = {"branch_id": branch_id} if branch_id else {}
    costs = await db.operating_costs.find(query, {"_id": 0}).to_list(100)
    return costs

# ==================== HR ROUTES - إدارة الموارد البشرية ====================

# --- الموظفين ---

@api_router.post("/employees", response_model=EmployeeResponse)
async def create_employee(employee: EmployeeCreate, current_user: dict = Depends(get_current_user)):
    """إنشاء موظف جديد"""
    if current_user["role"] not in [UserRole.ADMIN, UserRole.MANAGER, UserRole.SUPER_ADMIN]:
        raise HTTPException(status_code=403, detail="غير مصرح")
    
    employee_doc = {
        "id": str(uuid.uuid4()),
        **employee.model_dump(),
        "tenant_id": get_user_tenant_id(current_user),
        "is_active": True,
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    await db.employees.insert_one(employee_doc)
    del employee_doc["_id"]
    return employee_doc

@api_router.get("/employees", response_model=List[EmployeeResponse])
async def get_employees(
    branch_id: Optional[str] = None,
    department: Optional[str] = None,
    is_active: Optional[bool] = None,
    current_user: dict = Depends(get_current_user)
):
    """جلب قائمة الموظفين"""
    query = build_tenant_query(current_user)
    if branch_id:
        query["branch_id"] = branch_id
    if department:
        query["department"] = department
    if is_active is not None:
        query["is_active"] = is_active
    
    employees = await db.employees.find(query, {"_id": 0}).to_list(1000)
    return employees

@api_router.get("/employees/{employee_id}", response_model=EmployeeResponse)
async def get_employee(employee_id: str, current_user: dict = Depends(get_current_user)):
    """جلب موظف محدد"""
    query = build_tenant_query(current_user, {"id": employee_id})
    employee = await db.employees.find_one(query, {"_id": 0})
    if not employee:
        raise HTTPException(status_code=404, detail="الموظف غير موجود")
    return employee

@api_router.put("/employees/{employee_id}")
async def update_employee(employee_id: str, update: EmployeeUpdate, current_user: dict = Depends(get_current_user)):
    """تحديث بيانات موظف"""
    if current_user["role"] not in [UserRole.ADMIN, UserRole.MANAGER, UserRole.SUPER_ADMIN]:
        raise HTTPException(status_code=403, detail="غير مصرح")
    
    query = build_tenant_query(current_user, {"id": employee_id})
    employee = await db.employees.find_one(query)
    if not employee:
        raise HTTPException(status_code=404, detail="الموظف غير موجود")
    
    update_data = {k: v for k, v in update.model_dump().items() if v is not None}
    if update_data:
        await db.employees.update_one({"id": employee_id}, {"$set": update_data})
    
    return await db.employees.find_one({"id": employee_id}, {"_id": 0})

@api_router.delete("/employees/{employee_id}")
async def delete_employee(employee_id: str, current_user: dict = Depends(get_current_user)):
    """حذف موظف"""
    if current_user["role"] not in [UserRole.ADMIN, UserRole.SUPER_ADMIN]:
        raise HTTPException(status_code=403, detail="غير مصرح")
    
    query = build_tenant_query(current_user, {"id": employee_id})
    employee = await db.employees.find_one(query)
    if not employee:
        raise HTTPException(status_code=404, detail="الموظف غير موجود")
    
    # تعطيل بدلاً من الحذف للحفاظ على السجلات
    await db.employees.update_one({"id": employee_id}, {"$set": {"is_active": False}})
    return {"message": "تم تعطيل الموظف"}

# --- الحضور والانصراف ---

@api_router.post("/attendance", response_model=AttendanceResponse)
async def create_attendance(attendance: AttendanceCreate, current_user: dict = Depends(get_current_user)):
    """تسجيل حضور/انصراف"""
    if current_user["role"] not in [UserRole.ADMIN, UserRole.MANAGER, UserRole.SUPERVISOR, UserRole.SUPER_ADMIN]:
        raise HTTPException(status_code=403, detail="غير مصرح")
    
    # التحقق من وجود الموظف
    employee = await db.employees.find_one({"id": attendance.employee_id}, {"_id": 0})
    if not employee:
        raise HTTPException(status_code=404, detail="الموظف غير موجود")
    
    # حساب ساعات العمل إذا توفر وقت الحضور والانصراف
    worked_hours = None
    if attendance.check_in and attendance.check_out:
        try:
            check_in = datetime.strptime(attendance.check_in, "%H:%M")
            check_out = datetime.strptime(attendance.check_out, "%H:%M")
            worked_hours = (check_out - check_in).seconds / 3600
        except:
            pass
    
    attendance_doc = {
        "id": str(uuid.uuid4()),
        **attendance.model_dump(),
        "employee_name": employee.get("name"),
        "worked_hours": worked_hours,
        "tenant_id": get_user_tenant_id(current_user),
        "created_by": current_user["id"],
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    await db.attendance.insert_one(attendance_doc)
    del attendance_doc["_id"]
    return attendance_doc

@api_router.get("/attendance")
async def get_attendance(
    employee_id: Optional[str] = None,
    branch_id: Optional[str] = None,
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    """جلب سجلات الحضور"""
    query = build_tenant_query(current_user)
    if employee_id:
        query["employee_id"] = employee_id
    if branch_id:
        # جلب الموظفين في الفرع
        employees = await db.employees.find({"branch_id": branch_id}, {"id": 1}).to_list(1000)
        emp_ids = [e["id"] for e in employees]
        query["employee_id"] = {"$in": emp_ids}
    if start_date:
        query["date"] = {"$gte": start_date}
    if end_date:
        query.setdefault("date", {})["$lte"] = end_date
    
    records = await db.attendance.find(query, {"_id": 0}).sort("date", -1).to_list(1000)
    return records

@api_router.put("/attendance/{attendance_id}")
async def update_attendance(attendance_id: str, update: dict, current_user: dict = Depends(get_current_user)):
    """تحديث سجل حضور"""
    if current_user["role"] not in [UserRole.ADMIN, UserRole.MANAGER, UserRole.SUPERVISOR, UserRole.SUPER_ADMIN]:
        raise HTTPException(status_code=403, detail="غير مصرح")
    
    query = build_tenant_query(current_user, {"id": attendance_id})
    record = await db.attendance.find_one(query)
    if not record:
        raise HTTPException(status_code=404, detail="السجل غير موجود")
    
    # حساب ساعات العمل الجديدة
    check_in = update.get("check_in", record.get("check_in"))
    check_out = update.get("check_out", record.get("check_out"))
    worked_hours = record.get("worked_hours")
    
    if check_in and check_out:
        try:
            ci = datetime.strptime(check_in, "%H:%M")
            co = datetime.strptime(check_out, "%H:%M")
            worked_hours = (co - ci).seconds / 3600
        except:
            pass
    
    update["worked_hours"] = worked_hours
    await db.attendance.update_one({"id": attendance_id}, {"$set": update})
    return await db.attendance.find_one({"id": attendance_id}, {"_id": 0})

# --- السلف ---

@api_router.post("/advances", response_model=AdvanceResponse)
async def create_advance(advance: AdvanceCreate, current_user: dict = Depends(get_current_user)):
    """طلب سلفة"""
    if current_user["role"] not in [UserRole.ADMIN, UserRole.MANAGER, UserRole.SUPER_ADMIN]:
        raise HTTPException(status_code=403, detail="غير مصرح")
    
    employee = await db.employees.find_one({"id": advance.employee_id}, {"_id": 0})
    if not employee:
        raise HTTPException(status_code=404, detail="الموظف غير موجود")
    
    monthly_deduction = advance.amount / advance.deduction_months
    
    advance_doc = {
        "id": str(uuid.uuid4()),
        **advance.model_dump(),
        "employee_name": employee.get("name"),
        "remaining_amount": advance.amount,
        "deducted_amount": 0,
        "monthly_deduction": monthly_deduction,
        "status": "approved",  # يمكن إضافة workflow للموافقة
        "date": advance.date or datetime.now(timezone.utc).strftime("%Y-%m-%d"),
        "tenant_id": get_user_tenant_id(current_user),
        "created_by": current_user["id"],
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    await db.advances.insert_one(advance_doc)
    
    # إضافة السلفة كمصروف
    expense_doc = {
        "id": str(uuid.uuid4()),
        "category": "advance",
        "description": f"سلفة للموظف {employee.get('name')}",
        "amount": advance.amount,
        "payment_method": "cash",
        "branch_id": employee.get("branch_id"),
        "employee_id": advance.employee_id,
        "advance_id": advance_doc["id"],
        "tenant_id": get_user_tenant_id(current_user),
        "date": advance_doc["date"],
        "created_by": current_user["id"],
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    await db.expenses.insert_one(expense_doc)
    
    del advance_doc["_id"]
    return advance_doc

@api_router.get("/advances")
async def get_advances(
    employee_id: Optional[str] = None,
    status: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    """جلب قائمة السلف"""
    query = build_tenant_query(current_user)
    if employee_id:
        query["employee_id"] = employee_id
    if status:
        query["status"] = status
    
    advances = await db.advances.find(query, {"_id": 0}).sort("created_at", -1).to_list(500)
    return advances

# ==================== EMPLOYEE RATINGS - تقييم الموظفين التلقائي ====================

@api_router.get("/employee-ratings")
async def get_employee_ratings(
    month: str = None,  # YYYY-MM
    branch_id: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    """حساب تقييم الموظفين بناءً على الحضور والانصراف والخصومات"""
    tenant_id = get_user_tenant_id(current_user)
    
    if not month:
        month = datetime.now().strftime("%Y-%m")
    
    # جلب الموظفين
    emp_query = {"tenant_id": tenant_id, "is_active": True}
    if branch_id:
        emp_query["branch_id"] = branch_id
    
    employees = await db.employees.find(emp_query, {"_id": 0}).to_list(500)
    
    ratings = []
    
    for emp in employees:
        emp_id = emp["id"]
        
        # حساب أيام العمل في الشهر (26 يوم افتراضياً)
        work_days_expected = 26
        
        # جلب سجلات الحضور للشهر
        attendance_records = await db.attendance.find({
            "tenant_id": tenant_id,
            "employee_id": emp_id,
            "date": {"$regex": f"^{month}"}
        }, {"_id": 0}).to_list(100)
        
        # حساب أيام الحضور
        attendance_days = len(attendance_records)
        
        # حساب التأخير (إذا الحضور بعد الساعة 9 صباحاً)
        late_count = 0
        early_leave_count = 0
        total_work_hours = 0
        
        for record in attendance_records:
            check_in = record.get("check_in")
            check_out = record.get("check_out")
            
            if check_in:
                try:
                    check_in_time = datetime.strptime(check_in, "%H:%M")
                    expected_time = datetime.strptime(emp.get("work_start", "09:00"), "%H:%M")
                    if check_in_time > expected_time:
                        late_count += 1
                except:
                    pass
            
            if check_out:
                try:
                    check_out_time = datetime.strptime(check_out, "%H:%M")
                    expected_end = datetime.strptime(emp.get("work_end", "17:00"), "%H:%M")
                    if check_out_time < expected_end:
                        early_leave_count += 1
                except:
                    pass
            
            # حساب ساعات العمل
            if check_in and check_out:
                try:
                    ci = datetime.strptime(check_in, "%H:%M")
                    co = datetime.strptime(check_out, "%H:%M")
                    hours = (co - ci).seconds / 3600
                    total_work_hours += hours
                except:
                    pass
        
        # جلب الخصومات للشهر
        deductions = await db.deductions.find({
            "tenant_id": tenant_id,
            "employee_id": emp_id,
            "month": month
        }, {"_id": 0}).to_list(100)
        
        total_deductions = sum(d.get("amount", 0) for d in deductions)
        deduction_count = len(deductions)
        
        # جلب المكافآت للشهر
        bonuses = await db.bonuses.find({
            "tenant_id": tenant_id,
            "employee_id": emp_id,
            "month": month
        }, {"_id": 0}).to_list(100)
        
        total_bonuses = sum(b.get("amount", 0) for b in bonuses)
        bonus_count = len(bonuses)
        
        # ========== حساب التقييم ==========
        # التقييم من 100 نقطة
        
        # 1. تقييم الحضور (40 نقطة)
        attendance_percentage = (attendance_days / work_days_expected) * 100 if work_days_expected > 0 else 0
        attendance_score = min(40, (attendance_percentage / 100) * 40)
        
        # 2. تقييم الالتزام بالمواعيد (30 نقطة)
        punctuality_issues = late_count + early_leave_count
        punctuality_deduction = min(30, punctuality_issues * 3)  # خصم 3 نقاط لكل تأخير/خروج مبكر
        punctuality_score = max(0, 30 - punctuality_deduction)
        
        # 3. تقييم عدم وجود خصومات (20 نقطة)
        deduction_penalty = min(20, deduction_count * 5)  # خصم 5 نقاط لكل خصم
        discipline_score = max(0, 20 - deduction_penalty)
        
        # 4. نقاط المكافآت (10 نقاط إضافية)
        bonus_score = min(10, bonus_count * 2)  # 2 نقطة لكل مكافأة
        
        # المجموع
        total_score = attendance_score + punctuality_score + discipline_score + bonus_score
        
        # تحديد المستوى
        if total_score >= 90:
            level = "ممتاز"
            level_color = "green"
        elif total_score >= 75:
            level = "جيد جداً"
            level_color = "blue"
        elif total_score >= 60:
            level = "جيد"
            level_color = "yellow"
        elif total_score >= 50:
            level = "مقبول"
            level_color = "orange"
        else:
            level = "ضعيف"
            level_color = "red"
        
        ratings.append({
            "employee_id": emp_id,
            "employee_name": emp.get("name", ""),
            "branch_id": emp.get("branch_id"),
            "position": emp.get("position", ""),
            "month": month,
            
            # إحصائيات الحضور
            "attendance_days": attendance_days,
            "work_days_expected": work_days_expected,
            "attendance_percentage": round(attendance_percentage, 1),
            "late_count": late_count,
            "early_leave_count": early_leave_count,
            "total_work_hours": round(total_work_hours, 1),
            
            # إحصائيات الخصومات والمكافآت
            "deduction_count": deduction_count,
            "total_deductions": total_deductions,
            "bonus_count": bonus_count,
            "total_bonuses": total_bonuses,
            
            # التقييم
            "scores": {
                "attendance": round(attendance_score, 1),
                "punctuality": round(punctuality_score, 1),
                "discipline": round(discipline_score, 1),
                "bonus": round(bonus_score, 1)
            },
            "total_score": round(total_score, 1),
            "level": level,
            "level_color": level_color
        })
    
    # ترتيب حسب التقييم
    ratings.sort(key=lambda x: x["total_score"], reverse=True)
    
    # إحصائيات عامة
    if ratings:
        avg_score = sum(r["total_score"] for r in ratings) / len(ratings)
        excellent_count = len([r for r in ratings if r["total_score"] >= 90])
        good_count = len([r for r in ratings if 75 <= r["total_score"] < 90])
        average_count = len([r for r in ratings if 60 <= r["total_score"] < 75])
        poor_count = len([r for r in ratings if r["total_score"] < 60])
    else:
        avg_score = 0
        excellent_count = good_count = average_count = poor_count = 0
    
    return {
        "month": month,
        "ratings": ratings,
        "summary": {
            "total_employees": len(ratings),
            "average_score": round(avg_score, 1),
            "excellent_count": excellent_count,
            "good_count": good_count,
            "average_count": average_count,
            "poor_count": poor_count
        }
    }

# ==================== PAYROLL REPORTS & EXPORT - تقارير الرواتب والتصدير ====================

@api_router.get("/reports/payroll-summary")
async def get_payroll_summary_report(
    month: str,  # YYYY-MM
    branch_id: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    """تقرير شامل للرواتب - إجمالي الرواتب، الخصومات، المكافآت، السلف، المستحقات"""
    tenant_id = get_user_tenant_id(current_user)
    
    # بناء استعلام الموظفين
    emp_query = {"is_active": True}
    if tenant_id:
        emp_query["tenant_id"] = tenant_id
    
    # فلترة حسب الفرع
    user_branch_id = current_user.get("branch_id")
    user_role = current_user.get("role")
    
    if user_branch_id and user_role not in [UserRole.SUPER_ADMIN, UserRole.ADMIN, UserRole.MANAGER]:
        emp_query["branch_id"] = user_branch_id
    elif branch_id:
        emp_query["branch_id"] = branch_id
    
    employees = await db.employees.find(emp_query, {"_id": 0}).to_list(500)
    
    start_date = f"{month}-01"
    end_date = f"{month}-31"
    
    # بناء بيانات التقرير لكل موظف
    employee_data = []
    totals = {
        "basic_salary": 0,
        "total_deductions": 0,
        "total_bonuses": 0,
        "total_advances": 0,
        "net_payable": 0
    }
    
    for emp in employees:
        # الخصومات
        deductions = await db.deductions.find({
            "employee_id": emp["id"],
            "date": {"$gte": start_date, "$lte": end_date}
        }, {"_id": 0}).to_list(100)
        emp_deductions = sum(d.get("amount", 0) for d in deductions)
        
        # المكافآت
        bonuses = await db.bonuses.find({
            "employee_id": emp["id"],
            "date": {"$gte": start_date, "$lte": end_date}
        }, {"_id": 0}).to_list(100)
        emp_bonuses = sum(b.get("amount", 0) for b in bonuses)
        
        # السلف المعلقة
        advances = await db.advances.find({
            "employee_id": emp["id"],
            "status": "approved",
            "remaining_amount": {"$gt": 0}
        }, {"_id": 0}).to_list(100)
        emp_advances = sum(a.get("monthly_deduction", 0) for a in advances)
        pending_advances = sum(a.get("remaining_amount", 0) for a in advances)
        
        basic_salary = emp.get("salary", 0)
        net_payable = basic_salary + emp_bonuses - emp_deductions - emp_advances
        
        # جلب اسم الفرع
        branch = await db.branches.find_one({"id": emp.get("branch_id")}, {"_id": 0, "name": 1})
        
        employee_data.append({
            "id": emp["id"],
            "name": emp.get("name"),
            "position": emp.get("position"),
            "branch_id": emp.get("branch_id"),
            "branch_name": branch.get("name") if branch else "-",
            "basic_salary": basic_salary,
            "deductions": emp_deductions,
            "deductions_details": deductions,
            "bonuses": emp_bonuses,
            "bonuses_details": bonuses,
            "advances_deduction": emp_advances,
            "pending_advances": pending_advances,
            "net_payable": net_payable
        })
        
        totals["basic_salary"] += basic_salary
        totals["total_deductions"] += emp_deductions
        totals["total_bonuses"] += emp_bonuses
        totals["total_advances"] += emp_advances
        totals["net_payable"] += net_payable
    
    return {
        "month": month,
        "employee_count": len(employees),
        "employees": employee_data,
        "totals": totals
    }

@api_router.get("/reports/employee-salary-slip/{employee_id}")
async def get_employee_salary_slip(
    employee_id: str,
    month: str,  # YYYY-MM
    current_user: dict = Depends(get_current_user)
):
    """مفردات مرتب موظف واحد"""
    tenant_id = get_user_tenant_id(current_user)
    
    # جلب الموظف
    emp_query = {"id": employee_id}
    if tenant_id:
        emp_query["tenant_id"] = tenant_id
    
    employee = await db.employees.find_one(emp_query, {"_id": 0})
    if not employee:
        raise HTTPException(status_code=404, detail="الموظف غير موجود")
    
    # التحقق من صلاحية الفرع
    user_branch_id = current_user.get("branch_id")
    user_role = current_user.get("role")
    if user_branch_id and user_role not in [UserRole.SUPER_ADMIN, UserRole.ADMIN, UserRole.MANAGER]:
        if employee.get("branch_id") != user_branch_id:
            raise HTTPException(status_code=403, detail="غير مصرح")
    
    start_date = f"{month}-01"
    end_date = f"{month}-31"
    
    # جلب الفرع
    branch = await db.branches.find_one({"id": employee.get("branch_id")}, {"_id": 0})
    
    # جلب معلومات العميل
    tenant_info = await db.tenants.find_one({"id": tenant_id}, {"_id": 0}) if tenant_id else None
    
    # الخصومات التفصيلية
    deductions = await db.deductions.find({
        "employee_id": employee_id,
        "date": {"$gte": start_date, "$lte": end_date}
    }, {"_id": 0}).to_list(100)
    
    # تصنيف الخصومات
    deductions_by_type = {}
    for d in deductions:
        dtype = d.get("deduction_type", "other")
        if dtype not in deductions_by_type:
            deductions_by_type[dtype] = {"items": [], "total": 0}
        deductions_by_type[dtype]["items"].append(d)
        deductions_by_type[dtype]["total"] += d.get("amount", 0)
    
    # المكافآت التفصيلية
    bonuses = await db.bonuses.find({
        "employee_id": employee_id,
        "date": {"$gte": start_date, "$lte": end_date}
    }, {"_id": 0}).to_list(100)
    
    # تصنيف المكافآت
    bonuses_by_type = {}
    for b in bonuses:
        btype = b.get("bonus_type", "other")
        if btype not in bonuses_by_type:
            bonuses_by_type[btype] = {"items": [], "total": 0}
        bonuses_by_type[btype]["items"].append(b)
        bonuses_by_type[btype]["total"] += b.get("amount", 0)
    
    # السلف
    advances = await db.advances.find({
        "employee_id": employee_id,
        "status": {"$in": ["approved", "paid"]}
    }, {"_id": 0}).to_list(100)
    
    # الحضور
    attendance = await db.attendance.find({
        "employee_id": employee_id,
        "date": {"$gte": start_date, "$lte": end_date}
    }, {"_id": 0}).to_list(31)
    
    # إحصائيات الحضور
    attendance_stats = {
        "present": len([a for a in attendance if a.get("status") == "present"]),
        "absent": len([a for a in attendance if a.get("status") == "absent"]),
        "late": len([a for a in attendance if a.get("status") == "late"]),
        "early_leave": len([a for a in attendance if a.get("status") == "early_leave"]),
        "holiday": len([a for a in attendance if a.get("status") == "holiday"])
    }
    
    # حساب الإجماليات
    total_deductions = sum(d.get("amount", 0) for d in deductions)
    total_bonuses = sum(b.get("amount", 0) for b in bonuses)
    advance_deduction = sum(a.get("monthly_deduction", 0) for a in advances if a.get("status") == "approved")
    pending_advances = sum(a.get("remaining_amount", 0) for a in advances if a.get("status") == "approved")
    
    basic_salary = employee.get("salary", 0)
    net_salary = basic_salary + total_bonuses - total_deductions - advance_deduction
    
    return {
        "employee": employee,
        "branch": branch,
        "tenant": tenant_info,
        "month": month,
        "salary_details": {
            "basic_salary": basic_salary,
            "salary_type": employee.get("salary_type", "monthly"),
            "work_hours_per_day": employee.get("work_hours_per_day", 8)
        },
        "deductions": {
            "items": deductions,
            "by_type": deductions_by_type,
            "total": total_deductions
        },
        "bonuses": {
            "items": bonuses,
            "by_type": bonuses_by_type,
            "total": total_bonuses
        },
        "advances": {
            "items": advances,
            "deduction_this_month": advance_deduction,
            "pending_total": pending_advances
        },
        "attendance": {
            "records": attendance,
            "stats": attendance_stats
        },
        "summary": {
            "basic_salary": basic_salary,
            "total_additions": total_bonuses,
            "total_deductions": total_deductions + advance_deduction,
            "net_salary": net_salary
        },
        "generated_at": datetime.now(timezone.utc).isoformat()
    }

@api_router.get("/reports/payroll/export/excel")
async def export_payroll_excel(
    month: str,
    branch_id: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    """تصدير تقرير الرواتب إلى Excel"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from io import BytesIO
    
    # جلب بيانات التقرير
    tenant_id = get_user_tenant_id(current_user)
    
    emp_query = {"is_active": True}
    if tenant_id:
        emp_query["tenant_id"] = tenant_id
    
    user_branch_id = current_user.get("branch_id")
    user_role = current_user.get("role")
    
    if user_branch_id and user_role not in [UserRole.SUPER_ADMIN, UserRole.ADMIN, UserRole.MANAGER]:
        emp_query["branch_id"] = user_branch_id
    elif branch_id:
        emp_query["branch_id"] = branch_id
    
    employees = await db.employees.find(emp_query, {"_id": 0}).to_list(500)
    
    start_date = f"{month}-01"
    end_date = f"{month}-31"
    
    # إنشاء ملف Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "تقرير الرواتب"
    
    # التنسيق
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # العنوان
    ws.merge_cells('A1:H1')
    ws['A1'] = f"تقرير الرواتب - {month}"
    ws['A1'].font = Font(bold=True, size=14)
    ws['A1'].alignment = Alignment(horizontal='center')
    
    # الرؤوس
    headers = ['#', 'الموظف', 'الفرع', 'الراتب الأساسي', 'المكافآت', 'الخصومات', 'السلف', 'صافي الراتب']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border
    
    # البيانات
    row_num = 4
    totals = [0, 0, 0, 0, 0]
    
    for idx, emp in enumerate(employees, 1):
        # الخصومات
        deductions = await db.deductions.find({
            "employee_id": emp["id"],
            "date": {"$gte": start_date, "$lte": end_date}
        }, {"_id": 0}).to_list(100)
        emp_deductions = sum(d.get("amount", 0) for d in deductions)
        
        # المكافآت
        bonuses = await db.bonuses.find({
            "employee_id": emp["id"],
            "date": {"$gte": start_date, "$lte": end_date}
        }, {"_id": 0}).to_list(100)
        emp_bonuses = sum(b.get("amount", 0) for b in bonuses)
        
        # السلف
        advances = await db.advances.find({
            "employee_id": emp["id"],
            "status": "approved",
            "remaining_amount": {"$gt": 0}
        }, {"_id": 0}).to_list(100)
        emp_advances = sum(a.get("monthly_deduction", 0) for a in advances)
        
        basic_salary = emp.get("salary", 0)
        net_salary = basic_salary + emp_bonuses - emp_deductions - emp_advances
        
        # جلب اسم الفرع
        branch = await db.branches.find_one({"id": emp.get("branch_id")}, {"_id": 0, "name": 1})
        
        data = [
            idx,
            emp.get("name", ""),
            branch.get("name", "-") if branch else "-",
            basic_salary,
            emp_bonuses,
            emp_deductions,
            emp_advances,
            net_salary
        ]
        
        for col, value in enumerate(data, 1):
            cell = ws.cell(row=row_num, column=col, value=value)
            cell.border = thin_border
            if col >= 4:
                cell.number_format = '#,##0'
        
        totals[0] += basic_salary
        totals[1] += emp_bonuses
        totals[2] += emp_deductions
        totals[3] += emp_advances
        totals[4] += net_salary
        
        row_num += 1
    
    # الإجماليات
    ws.cell(row=row_num, column=2, value="الإجمالي").font = Font(bold=True)
    ws.cell(row=row_num, column=2).border = thin_border
    for col, total in enumerate(totals, 4):
        cell = ws.cell(row=row_num, column=col, value=total)
        cell.font = Font(bold=True)
        cell.border = thin_border
        cell.number_format = '#,##0'
    
    # عرض الأعمدة
    ws.column_dimensions['A'].width = 5
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 12
    ws.column_dimensions['G'].width = 12
    ws.column_dimensions['H'].width = 15
    
    # حفظ الملف
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=payroll_report_{month}.xlsx"}
    )

@api_router.get("/reports/employee-salary-slip/{employee_id}/export/excel")
async def export_employee_salary_slip_excel(
    employee_id: str,
    month: str,
    current_user: dict = Depends(get_current_user)
):
    """تصدير مفردات مرتب موظف إلى Excel"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from io import BytesIO
    
    # جلب بيانات الموظف
    slip_data = await get_employee_salary_slip(employee_id, month, current_user)
    employee = slip_data["employee"]
    
    wb = Workbook()
    ws = wb.active
    ws.title = "مفردات المرتب"
    
    # التنسيق
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # العنوان
    ws.merge_cells('A1:D1')
    ws['A1'] = "مفردات المرتب"
    ws['A1'].font = Font(bold=True, size=16)
    ws['A1'].alignment = Alignment(horizontal='center')
    
    # معلومات الموظف
    ws['A3'] = "اسم الموظف:"
    ws['B3'] = employee.get("name", "")
    ws['C3'] = "الشهر:"
    ws['D3'] = month
    
    ws['A4'] = "الوظيفة:"
    ws['B4'] = employee.get("position", "")
    ws['C4'] = "الفرع:"
    ws['D4'] = slip_data["branch"].get("name", "-") if slip_data["branch"] else "-"
    
    # الراتب الأساسي
    ws['A6'] = "الراتب الأساسي"
    ws['A6'].font = Font(bold=True)
    ws['B6'] = slip_data["salary_details"]["basic_salary"]
    ws['B6'].number_format = '#,##0'
    
    # المكافآت
    row = 8
    ws[f'A{row}'] = "المكافآت"
    ws[f'A{row}'].font = Font(bold=True)
    ws[f'A{row}'].fill = PatternFill(start_color="92D050", end_color="92D050", fill_type="solid")
    row += 1
    
    for bonus in slip_data["bonuses"]["items"]:
        ws[f'A{row}'] = bonus.get("reason", bonus.get("bonus_type", ""))
        ws[f'B{row}'] = bonus.get("amount", 0)
        ws[f'B{row}'].number_format = '#,##0'
        row += 1
    
    ws[f'A{row}'] = "إجمالي المكافآت"
    ws[f'A{row}'].font = Font(bold=True)
    ws[f'B{row}'] = slip_data["bonuses"]["total"]
    ws[f'B{row}'].number_format = '#,##0'
    row += 2
    
    # الخصومات
    ws[f'A{row}'] = "الخصومات"
    ws[f'A{row}'].font = Font(bold=True)
    ws[f'A{row}'].fill = PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid")
    row += 1
    
    for deduction in slip_data["deductions"]["items"]:
        ws[f'A{row}'] = deduction.get("reason", deduction.get("deduction_type", ""))
        ws[f'B{row}'] = deduction.get("amount", 0)
        ws[f'B{row}'].number_format = '#,##0'
        row += 1
    
    ws[f'A{row}'] = "إجمالي الخصومات"
    ws[f'A{row}'].font = Font(bold=True)
    ws[f'B{row}'] = slip_data["deductions"]["total"]
    ws[f'B{row}'].number_format = '#,##0'
    row += 2
    
    # السلف
    ws[f'A{row}'] = "خصم السلف"
    ws[f'A{row}'].font = Font(bold=True)
    ws[f'B{row}'] = slip_data["advances"]["deduction_this_month"]
    ws[f'B{row}'].number_format = '#,##0'
    row += 2
    
    # صافي الراتب
    ws[f'A{row}'] = "صافي الراتب"
    ws[f'A{row}'].font = Font(bold=True, size=14)
    ws[f'A{row}'].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    ws[f'A{row}'].font = Font(bold=True, size=14, color="FFFFFF")
    ws[f'B{row}'] = slip_data["summary"]["net_salary"]
    ws[f'B{row}'].number_format = '#,##0'
    ws[f'B{row}'].font = Font(bold=True, size=14)
    
    # عرض الأعمدة
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 15
    
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=salary_slip_{employee.get('name', '')}_{month}.xlsx"}
    )

# ==================== COUPONS & PROMOTIONS ROUTES - الكوبونات والعروض ====================

class CouponCreate(BaseModel):
    code: str
    name: str
    description: Optional[str] = None
    discount_type: str = "percentage"  # percentage, fixed
    discount_value: float
    min_order_amount: float = 0
    max_discount: Optional[float] = None  # للنسبة المئوية فقط
    usage_limit: Optional[int] = None  # عدد مرات الاستخدام الكلي
    usage_per_customer: int = 1  # عدد مرات الاستخدام لكل عميل
    valid_from: str
    valid_until: str
    is_active: bool = True
    applicable_to: str = "all"  # all, category, product
    applicable_ids: List[str] = []  # category_ids أو product_ids
    loyalty_tier_required: Optional[str] = None  # bronze, silver, gold, platinum
    first_order_only: bool = False

class PromotionCreate(BaseModel):
    name: str
    description: Optional[str] = None
    promotion_type: str = "buy_x_get_y"  # buy_x_get_y, bundle, happy_hour, flash_sale
    buy_quantity: int = 1
    get_quantity: int = 1
    discount_percent: float = 100  # للعنصر المجاني
    bundle_price: Optional[float] = None
    start_time: Optional[str] = None  # HH:MM للـ happy_hour
    end_time: Optional[str] = None
    valid_from: str
    valid_until: str
    applicable_products: List[str] = []
    applicable_categories: List[str] = []
    is_active: bool = True
    loyalty_tier_required: Optional[str] = None

@api_router.get("/coupons")
async def get_coupons(current_user: dict = Depends(get_current_user)):
    """قائمة الكوبونات"""
    query = build_tenant_query(current_user)
    coupons = await db.coupons.find(query, {"_id": 0}).sort("created_at", -1).to_list(200)
    return coupons

@api_router.post("/coupons")
async def create_coupon(coupon: CouponCreate, current_user: dict = Depends(get_current_user)):
    """إنشاء كوبون"""
    if current_user["role"] not in [UserRole.ADMIN, UserRole.MANAGER, UserRole.SUPER_ADMIN]:
        raise HTTPException(status_code=403, detail="غير مصرح")
    
    # التحقق من عدم تكرار الكود
    existing = await db.coupons.find_one({"code": coupon.code.upper()})
    if existing:
        raise HTTPException(status_code=400, detail="كود الكوبون موجود مسبقاً")
    
    coupon_doc = {
        "id": str(uuid.uuid4()),
        **coupon.model_dump(),
        "code": coupon.code.upper(),
        "used_count": 0,
        "total_discount_given": 0,
        "tenant_id": get_user_tenant_id(current_user),
        "created_by": current_user["id"],
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    
    await db.coupons.insert_one(coupon_doc)
    if "_id" in coupon_doc:
        del coupon_doc["_id"]
    
    return {"message": "تم إنشاء الكوبون", "coupon": coupon_doc}

@api_router.put("/coupons/{coupon_id}")
async def update_coupon(coupon_id: str, coupon: CouponCreate, current_user: dict = Depends(get_current_user)):
    """تحديث كوبون"""
    if current_user["role"] not in [UserRole.ADMIN, UserRole.MANAGER, UserRole.SUPER_ADMIN]:
        raise HTTPException(status_code=403, detail="غير مصرح")
    
    query = build_tenant_query(current_user, {"id": coupon_id})
    await db.coupons.update_one(
        query,
        {"$set": {**coupon.model_dump(), "code": coupon.code.upper(), "updated_at": datetime.now(timezone.utc).isoformat()}}
    )
    return {"message": "تم التحديث"}

@api_router.delete("/coupons/{coupon_id}")
async def delete_coupon(coupon_id: str, current_user: dict = Depends(get_current_user)):
    """حذف كوبون"""
    if current_user["role"] not in [UserRole.ADMIN, UserRole.SUPER_ADMIN]:
        raise HTTPException(status_code=403, detail="غير مصرح")
    
    query = build_tenant_query(current_user, {"id": coupon_id})
    await db.coupons.delete_one(query)
    return {"message": "تم الحذف"}

@api_router.post("/coupons/validate")
async def validate_coupon(
    code: str,
    order_total: float,
    customer_id: Optional[str] = None,
    customer_phone: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    """التحقق من صلاحية الكوبون"""
    coupon = await db.coupons.find_one({"code": code.upper(), "is_active": True}, {"_id": 0})
    
    if not coupon:
        raise HTTPException(status_code=404, detail="الكوبون غير صالح")
    
    now = datetime.now(timezone.utc).isoformat()
    
    # التحقق من التاريخ
    if coupon.get("valid_from") > now:
        raise HTTPException(status_code=400, detail="الكوبون لم يبدأ بعد")
    
    if coupon.get("valid_until") < now:
        raise HTTPException(status_code=400, detail="الكوبون منتهي الصلاحية")
    
    # التحقق من الحد الأدنى للطلب
    if order_total < coupon.get("min_order_amount", 0):
        raise HTTPException(
            status_code=400, 
            detail=f"الحد الأدنى للطلب {coupon.get('min_order_amount')} د.ع"
        )
    
    # التحقق من عدد الاستخدامات الكلي
    if coupon.get("usage_limit") and coupon.get("used_count", 0) >= coupon.get("usage_limit"):
        raise HTTPException(status_code=400, detail="الكوبون استُنفد")
    
    # التحقق من استخدام العميل
    if customer_phone:
        customer_uses = await db.coupon_usage.count_documents({
            "coupon_id": coupon["id"],
            "customer_phone": customer_phone
        })
        if customer_uses >= coupon.get("usage_per_customer", 1):
            raise HTTPException(status_code=400, detail="لقد استخدمت هذا الكوبون مسبقاً")
    
    # التحقق من مستوى الولاء المطلوب
    if coupon.get("loyalty_tier_required"):
        if customer_phone:
            member = await db.loyalty_members.find_one({"phone": customer_phone}, {"_id": 0})
            if not member:
                raise HTTPException(status_code=400, detail="يجب أن تكون عضواً في برنامج الولاء")
            
            tier_order = {"bronze": 1, "silver": 2, "gold": 3, "platinum": 4}
            required_tier = tier_order.get(coupon.get("loyalty_tier_required").lower(), 0)
            member_tier = tier_order.get(member.get("current_tier", "bronze").lower(), 1)
            
            if member_tier < required_tier:
                raise HTTPException(
                    status_code=400, 
                    detail=f"هذا الكوبون متاح لأعضاء {coupon.get('loyalty_tier_required')} فأعلى"
                )
    
    # التحقق من الطلب الأول
    if coupon.get("first_order_only") and customer_phone:
        previous_orders = await db.orders.count_documents({"customer_phone": customer_phone})
        if previous_orders > 0:
            raise HTTPException(status_code=400, detail="هذا الكوبون للطلب الأول فقط")
    
    # حساب الخصم
    if coupon.get("discount_type") == "percentage":
        discount = order_total * (coupon.get("discount_value", 0) / 100)
        if coupon.get("max_discount"):
            discount = min(discount, coupon.get("max_discount"))
    else:
        discount = coupon.get("discount_value", 0)
    
    return {
        "valid": True,
        "coupon": coupon,
        "discount": round(discount, 2),
        "final_total": round(order_total - discount, 2)
    }

@api_router.post("/coupons/{coupon_id}/use")
async def use_coupon(
    coupon_id: str,
    order_id: str,
    discount_amount: float,
    customer_phone: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    """تسجيل استخدام كوبون"""
    # تحديث عداد الاستخدام
    await db.coupons.update_one(
        {"id": coupon_id},
        {
            "$inc": {"used_count": 1, "total_discount_given": discount_amount}
        }
    )
    
    # تسجيل الاستخدام
    usage_doc = {
        "id": str(uuid.uuid4()),
        "coupon_id": coupon_id,
        "order_id": order_id,
        "customer_phone": customer_phone,
        "discount_amount": discount_amount,
        "used_at": datetime.now(timezone.utc).isoformat()
    }
    await db.coupon_usage.insert_one(usage_doc)
    
    return {"message": "تم تسجيل الاستخدام"}

# العروض الترويجية
@api_router.get("/promotions")
async def get_promotions(current_user: dict = Depends(get_current_user)):
    """قائمة العروض"""
    query = build_tenant_query(current_user)
    promotions = await db.promotions.find(query, {"_id": 0}).sort("created_at", -1).to_list(100)
    return promotions

@api_router.post("/promotions")
async def create_promotion(promotion: PromotionCreate, current_user: dict = Depends(get_current_user)):
    """إنشاء عرض"""
    if current_user["role"] not in [UserRole.ADMIN, UserRole.MANAGER, UserRole.SUPER_ADMIN]:
        raise HTTPException(status_code=403, detail="غير مصرح")
    
    promotion_doc = {
        "id": str(uuid.uuid4()),
        **promotion.model_dump(),
        "used_count": 0,
        "tenant_id": get_user_tenant_id(current_user),
        "created_by": current_user["id"],
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    
    await db.promotions.insert_one(promotion_doc)
    if "_id" in promotion_doc:
        del promotion_doc["_id"]
    
    return {"message": "تم إنشاء العرض", "promotion": promotion_doc}

@api_router.put("/promotions/{promotion_id}")
async def update_promotion(promotion_id: str, promotion: PromotionCreate, current_user: dict = Depends(get_current_user)):
    """تحديث عرض"""
    await db.promotions.update_one(
        {"id": promotion_id},
        {"$set": {**promotion.model_dump(), "updated_at": datetime.now(timezone.utc).isoformat()}}
    )
    return {"message": "تم التحديث"}

@api_router.delete("/promotions/{promotion_id}")
async def delete_promotion(promotion_id: str, current_user: dict = Depends(get_current_user)):
    """حذف عرض"""
    await db.promotions.delete_one({"id": promotion_id})
    return {"message": "تم الحذف"}

@api_router.get("/promotions/active")
async def get_active_promotions(current_user: dict = Depends(get_current_user)):
    """العروض النشطة حالياً"""
    now = datetime.now(timezone.utc).isoformat()
    current_time = datetime.now(timezone.utc).strftime("%H:%M")
    
    query = build_tenant_query(current_user, {
        "is_active": True,
        "valid_from": {"$lte": now},
        "valid_until": {"$gte": now}
    })
    
    promotions = await db.promotions.find(query, {"_id": 0}).to_list(50)
    
    # تصفية Happy Hour
    active_promotions = []
    for promo in promotions:
        if promo.get("promotion_type") == "happy_hour":
            start = promo.get("start_time", "00:00")
            end = promo.get("end_time", "23:59")
            if start <= current_time <= end:
                active_promotions.append(promo)
        else:
            active_promotions.append(promo)
    
    return active_promotions

# ==================== INVENTORY TRANSFER ROUTES - تحويلات المخزون ====================

async def get_next_transfer_number() -> int:
    """الحصول على رقم التحويل التالي"""
    counter = await db.counters.find_one_and_update(
        {"type": "transfer"},
        {"$inc": {"counter": 1}},
        upsert=True,
        return_document=True
    )
    return counter["counter"]

@api_router.post("/inventory-transfers", response_model=InventoryTransferResponse)
async def create_inventory_transfer(transfer: InventoryTransferCreate, current_user: dict = Depends(get_current_user)):
    """إنشاء طلب تحويل مخزون"""
    if current_user["role"] not in [UserRole.ADMIN, UserRole.MANAGER, UserRole.SUPERVISOR, UserRole.SUPER_ADMIN]:
        raise HTTPException(status_code=403, detail="غير مصرح")
    
    transfer_number = await get_next_transfer_number()
    
    # جلب أسماء الفروع
    from_branch = await db.branches.find_one({"id": transfer.from_branch_id}, {"name": 1})
    to_branch = await db.branches.find_one({"id": transfer.to_branch_id}, {"name": 1})
    
    transfer_doc = {
        "id": str(uuid.uuid4()),
        "transfer_number": transfer_number,
        **transfer.model_dump(),
        "from_branch_name": from_branch.get("name") if from_branch else None,
        "to_branch_name": to_branch.get("name") if to_branch else None,
        "status": "pending",
        "tenant_id": get_user_tenant_id(current_user),
        "created_by": current_user["id"],
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    await db.inventory_transfers.insert_one(transfer_doc)
    del transfer_doc["_id"]
    return transfer_doc

@api_router.get("/inventory-transfers")
async def get_inventory_transfers(
    from_branch_id: Optional[str] = None,
    to_branch_id: Optional[str] = None,
    status: Optional[str] = None,
    transfer_type: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    """جلب تحويلات المخزون"""
    query = build_tenant_query(current_user)
    if from_branch_id:
        query["from_branch_id"] = from_branch_id
    if to_branch_id:
        query["to_branch_id"] = to_branch_id
    if status:
        query["status"] = status
    if transfer_type:
        query["transfer_type"] = transfer_type
    
    transfers = await db.inventory_transfers.find(query, {"_id": 0}).sort("created_at", -1).to_list(500)
    return transfers

@api_router.put("/inventory-transfers/{transfer_id}/approve")
async def approve_inventory_transfer(transfer_id: str, current_user: dict = Depends(get_current_user)):
    """الموافقة على التحويل"""
    if current_user["role"] not in [UserRole.ADMIN, UserRole.MANAGER, UserRole.SUPER_ADMIN]:
        raise HTTPException(status_code=403, detail="غير مصرح")
    
    query = build_tenant_query(current_user, {"id": transfer_id})
    transfer = await db.inventory_transfers.find_one(query)
    if not transfer:
        raise HTTPException(status_code=404, detail="التحويل غير موجود")
    
    if transfer.get("status") != "pending":
        raise HTTPException(status_code=400, detail="لا يمكن الموافقة على هذا التحويل")
    
    await db.inventory_transfers.update_one(
        {"id": transfer_id},
        {"$set": {
            "status": "approved",
            "approved_by": current_user["id"],
            "approved_at": datetime.now(timezone.utc).isoformat()
        }}
    )
    return {"message": "تمت الموافقة على التحويل"}

@api_router.put("/inventory-transfers/{transfer_id}/ship")
async def ship_inventory_transfer(transfer_id: str, current_user: dict = Depends(get_current_user)):
    """شحن التحويل (خصم من المخزن المرسل)"""
    if current_user["role"] not in [UserRole.ADMIN, UserRole.MANAGER, UserRole.SUPERVISOR, UserRole.SUPER_ADMIN]:
        raise HTTPException(status_code=403, detail="غير مصرح")
    
    query = build_tenant_query(current_user, {"id": transfer_id})
    transfer = await db.inventory_transfers.find_one(query)
    if not transfer:
        raise HTTPException(status_code=404, detail="التحويل غير موجود")
    
    if transfer.get("status") != "approved":
        raise HTTPException(status_code=400, detail="يجب الموافقة على التحويل أولاً")
    
    # خصم الكميات من المخزن المرسل
    for item in transfer.get("items", []):
        await db.inventory.update_one(
            {"id": item.get("inventory_id"), "branch_id": transfer["from_branch_id"]},
            {"$inc": {"quantity": -item.get("quantity", 0)}}
        )
    
    await db.inventory_transfers.update_one(
        {"id": transfer_id},
        {"$set": {
            "status": "shipped",
            "shipped_by": current_user["id"],
            "shipped_at": datetime.now(timezone.utc).isoformat()
        }}
    )
    return {"message": "تم شحن التحويل"}

@api_router.put("/inventory-transfers/{transfer_id}/receive")
async def receive_inventory_transfer(transfer_id: str, current_user: dict = Depends(get_current_user)):
    """استلام التحويل (إضافة للمخزن المستلم)"""
    if current_user["role"] not in [UserRole.ADMIN, UserRole.MANAGER, UserRole.SUPERVISOR, UserRole.SUPER_ADMIN]:
        raise HTTPException(status_code=403, detail="غير مصرح")
    
    query = build_tenant_query(current_user, {"id": transfer_id})
    transfer = await db.inventory_transfers.find_one(query)
    if not transfer:
        raise HTTPException(status_code=404, detail="التحويل غير موجود")
    
    if transfer.get("status") != "shipped":
        raise HTTPException(status_code=400, detail="يجب شحن التحويل أولاً")
    
    # إضافة الكميات للمخزن المستلم
    for item in transfer.get("items", []):
        # التحقق من وجود الصنف في المخزن المستلم
        existing = await db.inventory.find_one({
            "id": item.get("inventory_id"),
            "branch_id": transfer["to_branch_id"]
        })
        
        if existing:
            await db.inventory.update_one(
                {"id": item.get("inventory_id"), "branch_id": transfer["to_branch_id"]},
                {"$inc": {"quantity": item.get("quantity", 0)}}
            )
        else:
            # إنشاء صنف جديد في المخزن المستلم
            source_item = await db.inventory.find_one({"id": item.get("inventory_id")}, {"_id": 0})
            if source_item:
                new_item = {
                    **source_item,
                    "id": str(uuid.uuid4()),
                    "branch_id": transfer["to_branch_id"],
                    "quantity": item.get("quantity", 0),
                    "last_updated": datetime.now(timezone.utc).isoformat()
                }
                await db.inventory.insert_one(new_item)
    
    await db.inventory_transfers.update_one(
        {"id": transfer_id},
        {"$set": {
            "status": "received",
            "received_by": current_user["id"],
            "received_at": datetime.now(timezone.utc).isoformat()
        }}
    )
    return {"message": "تم استلام التحويل"}

# ==================== PURCHASE REQUEST ROUTES - طلبات الشراء ====================

async def get_next_request_number() -> int:
    """الحصول على رقم الطلب التالي"""
    counter = await db.counters.find_one_and_update(
        {"type": "purchase_request"},
        {"$inc": {"counter": 1}},
        upsert=True,
        return_document=True
    )
    return counter["counter"]

@api_router.post("/purchase-requests", response_model=PurchaseRequestResponse)
async def create_purchase_request(request: PurchaseRequestCreate, current_user: dict = Depends(get_current_user)):
    """إنشاء طلب شراء"""
    if current_user["role"] not in [UserRole.ADMIN, UserRole.MANAGER, UserRole.SUPERVISOR, UserRole.SUPER_ADMIN]:
        raise HTTPException(status_code=403, detail="غير مصرح")
    
    request_number = await get_next_request_number()
    
    branch = await db.branches.find_one({"id": request.branch_id}, {"name": 1})
    
    request_doc = {
        "id": str(uuid.uuid4()),
        "request_number": request_number,
        **request.model_dump(),
        "branch_name": branch.get("name") if branch else None,
        "status": "pending",
        "tenant_id": get_user_tenant_id(current_user),
        "created_by": current_user["id"],
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    await db.purchase_requests.insert_one(request_doc)
    del request_doc["_id"]
    return request_doc

@api_router.get("/purchase-requests")
async def get_purchase_requests(
    branch_id: Optional[str] = None,
    status: Optional[str] = None,
    priority: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    """جلب طلبات الشراء"""
    query = build_tenant_query(current_user)
    if branch_id:
        query["branch_id"] = branch_id
    if status:
        query["status"] = status
    if priority:
        query["priority"] = priority
    
    requests = await db.purchase_requests.find(query, {"_id": 0}).sort("created_at", -1).to_list(500)
    return requests

@api_router.put("/purchase-requests/{request_id}/approve")
async def approve_purchase_request(request_id: str, current_user: dict = Depends(get_current_user)):
    """الموافقة على طلب الشراء"""
    if current_user["role"] not in [UserRole.ADMIN, UserRole.MANAGER, UserRole.SUPER_ADMIN]:
        raise HTTPException(status_code=403, detail="غير مصرح")
    
    query = build_tenant_query(current_user, {"id": request_id})
    request = await db.purchase_requests.find_one(query)
    if not request:
        raise HTTPException(status_code=404, detail="الطلب غير موجود")
    
    await db.purchase_requests.update_one(
        {"id": request_id},
        {"$set": {
            "status": "approved",
            "approved_by": current_user["id"],
            "approved_at": datetime.now(timezone.utc).isoformat()
        }}
    )
    return {"message": "تمت الموافقة على طلب الشراء"}

@api_router.put("/purchase-requests/{request_id}/status")
async def update_purchase_request_status(request_id: str, status: str, current_user: dict = Depends(get_current_user)):
    """تحديث حالة طلب الشراء"""
    if current_user["role"] not in [UserRole.ADMIN, UserRole.MANAGER, UserRole.SUPERVISOR, UserRole.SUPER_ADMIN]:
        raise HTTPException(status_code=403, detail="غير مصرح")
    
    query = build_tenant_query(current_user, {"id": request_id})
    request = await db.purchase_requests.find_one(query)
    if not request:
        raise HTTPException(status_code=404, detail="الطلب غير موجود")
    
    await db.purchase_requests.update_one(
        {"id": request_id},
        {"$set": {"status": status}}
    )
    return {"message": "تم تحديث الحالة"}

# ==================== TABLE ROUTES ====================

@api_router.post("/tables", response_model=TableResponse)
async def create_table(table: TableCreate, current_user: dict = Depends(get_current_user)):
    # السماح للمدير والأدمن أو من لديه صلاحية tables
    user_permissions = current_user.get("permissions", [])
    if current_user["role"] not in [UserRole.SUPER_ADMIN, UserRole.ADMIN, UserRole.MANAGER] and "tables" not in user_permissions:
        raise HTTPException(status_code=403, detail="غير مصرح")
    
    tenant_id = get_user_tenant_id(current_user)
    table_doc = {
        "id": str(uuid.uuid4()),
        **table.model_dump(),
        "tenant_id": tenant_id,  # فصل البيانات لكل عميل
        "status": "available",
        "current_order_id": None
    }
    await db.tables.insert_one(table_doc)
    del table_doc["_id"]
    return table_doc

@api_router.get("/tables", response_model=List[TableResponse])
async def get_tables(branch_id: Optional[str] = None, current_user: dict = Depends(get_current_user)):
    # Super Admin لا يرى طاولات (ليس لديه مطعم)
    if current_user.get("role") == UserRole.SUPER_ADMIN:
        return []
    
    query = build_tenant_query(current_user)  # فلترة حسب tenant_id
    if branch_id:
        query["branch_id"] = branch_id
    tables = await db.tables.find(query, {"_id": 0}).sort("number", 1).to_list(100)
    return tables

@api_router.put("/tables/{table_id}/status")
async def update_table_status(table_id: str, status: str, current_user: dict = Depends(get_current_user)):
    query = build_tenant_query(current_user, {"id": table_id})
    await db.tables.update_one(query, {"$set": {"status": status}})
    return {"message": "تم التحديث"}

@api_router.post("/tables/transfer")
async def transfer_table_order(
    from_table_id: str = Body(...),
    to_table_id: str = Body(...),
    order_id: str = Body(None),
    current_user: dict = Depends(get_current_user)
):
    """تحويل الطلب من طاولة إلى أخرى"""
    
    # التحقق من الطاولة المصدر مع فلترة tenant_id
    query = build_tenant_query(current_user, {"id": from_table_id})
    from_table = await db.tables.find_one(query)
    if not from_table:
        raise HTTPException(status_code=404, detail="الطاولة المصدر غير موجودة")
    
    if from_table.get("status") != "occupied":
        raise HTTPException(status_code=400, detail="الطاولة المصدر ليست مشغولة")
    
    # التحقق من الطاولة المستهدفة مع فلترة tenant_id
    query = build_tenant_query(current_user, {"id": to_table_id})
    to_table = await db.tables.find_one(query)
    if not to_table:
        raise HTTPException(status_code=404, detail="الطاولة المستهدفة غير موجودة")
    
    if to_table.get("status") != "available":
        raise HTTPException(status_code=400, detail="الطاولة المستهدفة غير متاحة")
    
    # الحصول على الطلب الحالي
    actual_order_id = order_id or from_table.get("current_order_id")
    if actual_order_id:
        # تحديث الطلب
        await db.orders.update_one(
            {"id": actual_order_id},
            {"$set": {
                "table_id": to_table_id,
                "table_number": to_table.get("number"),
                "updated_at": datetime.now(timezone.utc).isoformat(),
                "transfer_history": {
                    "from_table": from_table.get("number"),
                    "to_table": to_table.get("number"),
                    "transferred_at": datetime.now(timezone.utc).isoformat(),
                    "transferred_by": current_user["id"]
                }
            }}
        )
    
    # تحديث حالة الطاولات
    await db.tables.update_one(
        {"id": from_table_id},
        {"$set": {"status": "available", "current_order_id": None}}
    )
    
    await db.tables.update_one(
        {"id": to_table_id},
        {"$set": {"status": "occupied", "current_order_id": actual_order_id}}
    )
    
    logger.info(f"Order transferred from table {from_table.get('number')} to table {to_table.get('number')}")
    
    return {
        "message": f"تم تحويل الطلب من طاولة {from_table.get('number')} إلى طاولة {to_table.get('number')}",
        "from_table": from_table.get("number"),
        "to_table": to_table.get("number")
    }

@api_router.delete("/tables/{table_id}")
async def delete_table(table_id: str, current_user: dict = Depends(get_current_user)):
    """حذف طاولة - فقط للمالك أو المدير"""
    if current_user.get("role") not in ["super_admin", "admin", "manager"]:
        raise HTTPException(status_code=403, detail="ليس لديك صلاحية حذف الطاولات")
    
    query = build_tenant_query(current_user, {"id": table_id})
    table = await db.tables.find_one(query)
    if not table:
        raise HTTPException(status_code=404, detail="الطاولة غير موجودة")
    
    # التحقق من أن الطاولة ليست مشغولة
    if table.get("status") == "occupied":
        raise HTTPException(status_code=400, detail="لا يمكن حذف طاولة مشغولة")
    
    await db.tables.delete_one(query)
    return {"message": "تم حذف الطاولة"}

# ==================== CUSTOMER ROUTES - إدارة العملاء ====================

@api_router.post("/customers", response_model=CustomerResponse)
async def create_customer(customer: CustomerCreate, current_user: dict = Depends(get_current_user)):
    # التحقق من عدم وجود العميل بنفس الرقم في نفس الـ tenant
    tenant_id = get_user_tenant_id(current_user)
    query = {"phone": customer.phone}
    if tenant_id:
        query["tenant_id"] = tenant_id
    existing = await db.customers.find_one(query)
    if existing:
        raise HTTPException(status_code=400, detail="رقم الهاتف موجود مسبقاً")
    
    customer_doc = {
        "id": str(uuid.uuid4()),
        **customer.model_dump(),
        "tenant_id": tenant_id,  # فصل البيانات
        "total_orders": 0,
        "total_spent": 0.0,
        "last_order_date": None,
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    await db.customers.insert_one(customer_doc)
    del customer_doc["_id"]
    return customer_doc

@api_router.get("/customers", response_model=List[CustomerResponse])
async def get_customers(search: Optional[str] = None, phone: Optional[str] = None, current_user: dict = Depends(get_current_user)):
    query = build_tenant_query(current_user)  # فلترة حسب tenant_id
    if phone:
        query["$or"] = [{"phone": phone}, {"phone2": phone}]
    elif search:
        query["$or"] = [
            {"name": {"$regex": search, "$options": "i"}},
            {"phone": {"$regex": search}},
            {"phone2": {"$regex": search}},
            {"area": {"$regex": search, "$options": "i"}}
        ]
    customers = await db.customers.find(query, {"_id": 0}).sort("name", 1).to_list(500)
    return customers

@api_router.get("/customers/{customer_id}", response_model=CustomerResponse)
async def get_customer(customer_id: str, current_user: dict = Depends(get_current_user)):
    query = build_tenant_query(current_user, {"id": customer_id})
    customer = await db.customers.find_one(query, {"_id": 0})
    if not customer:
        raise HTTPException(status_code=404, detail="العميل غير موجود")
    return customer

@api_router.get("/customers/by-phone/{phone}")
async def get_customer_by_phone(phone: str, current_user: dict = Depends(get_current_user)):
    """البحث عن عميل بالهاتف مع سجل الطلبات"""
    tenant_id = get_user_tenant_id(current_user)
    
    # بناء query للبحث عن العميل بالهاتف مع مراعاة tenant_id
    phone_conditions = [{"phone": phone}, {"phone2": phone}]
    
    if tenant_id:
        # المستخدم العميل يرى فقط بياناته
        query = {"$and": [{"tenant_id": tenant_id}, {"$or": phone_conditions}]}
    elif current_user.get("role") == UserRole.SUPER_ADMIN:
        # Super Admin يرى الكل
        query = {"$or": phone_conditions}
    else:
        # المستخدم الرئيسي (بدون tenant_id) يرى البيانات الرئيسية فقط
        query = {"$and": [
            {"$or": [{"tenant_id": {"$exists": False}}, {"tenant_id": None}]},
            {"$or": phone_conditions}
        ]}
    
    customer = await db.customers.find_one(query, {"_id": 0})
    
    if not customer:
        return {"found": False, "customer": None, "orders": []}
    
    # جلب آخر 10 طلبات للعميل
    orders_query = build_tenant_query(current_user, {"customer_phone": phone})
    orders = await db.orders.find(orders_query, {"_id": 0}).sort("created_at", -1).limit(10).to_list(10)
    
    return {
        "found": True,
        "customer": customer,
        "orders": orders
    }

@api_router.put("/customers/{customer_id}")
async def update_customer(customer_id: str, customer: CustomerCreate, current_user: dict = Depends(get_current_user)):
    query = build_tenant_query(current_user, {"id": customer_id})
    await db.customers.update_one(query, {"$set": customer.model_dump()})
    return await db.customers.find_one({"id": customer_id}, {"_id": 0})

@api_router.delete("/customers/{customer_id}")
async def delete_customer(customer_id: str, current_user: dict = Depends(get_current_user)):
    await db.customers.delete_one({"id": customer_id})
    return {"message": "تم الحذف"}

# ==================== ORDER ROUTES ====================

async def get_next_order_number(branch_id: str) -> int:
    today = datetime.now(timezone.utc).strftime("%Y-%m-%d")
    counter = await db.order_counters.find_one_and_update(
        {"branch_id": branch_id, "date": today},
        {"$inc": {"counter": 1}},
        upsert=True,
        return_document=True
    )
    return counter["counter"]

@api_router.post("/orders", response_model=OrderResponse)
async def create_order(order: OrderCreate, current_user: dict = Depends(get_current_user)):
    order_number = await get_next_order_number(order.branch_id)
    
    tenant_id = get_user_tenant_id(current_user)
    
    # البحث عن الوردية المفتوحة للكاشير
    shift_query = {
        "cashier_id": current_user["id"],
        "status": "open"
    }
    if tenant_id:
        shift_query["tenant_id"] = tenant_id
    
    current_shift = await db.shifts.find_one(shift_query, {"_id": 0, "id": 1})
    shift_id = current_shift["id"] if current_shift else None
    
    subtotal = sum(item.price * item.quantity for item in order.items)
    tax = subtotal * 0.0  # No tax for Iraq
    total = subtotal - order.discount + tax
    
    # Calculate total cost including packaging for delivery/takeaway
    total_cost = 0
    total_packaging_cost = 0
    items_with_cost = []
    is_delivery_or_takeaway = order.order_type in ["delivery", "takeaway"]
    
    for item in order.items:
        product = await db.products.find_one({"id": item.product_id}, {"_id": 0})
        item_cost = 0
        packaging_cost = 0
        
        if product:
            # حساب تكلفة المنتج الأساسية
            base_cost = product.get("cost", 0) + product.get("operating_cost", 0)
            
            # إضافة تكلفة التغليف للتوصيل والسفري فقط
            if is_delivery_or_takeaway:
                packaging_cost = product.get("packaging_cost", 0) * item.quantity
                total_packaging_cost += packaging_cost
            
            # إذا كان المنتج مرتبط بمنتج مصنع، احصل على تكلفة المواد الخام
            manufactured_product_id = product.get("manufactured_product_id")
            if manufactured_product_id:
                mfg_product = await db.manufactured_products.find_one(
                    {"id": manufactured_product_id},
                    {"_id": 0, "raw_material_cost": 1}
                )
                if mfg_product:
                    base_cost = mfg_product.get("raw_material_cost", 0) + product.get("operating_cost", 0)
            
            item_cost = base_cost * item.quantity + packaging_cost
        
        total_cost += item_cost
        item_dict = item.model_dump()
        item_dict["cost"] = item_cost
        item_dict["packaging_cost"] = packaging_cost
        items_with_cost.append(item_dict)
    
    # Calculate delivery commission if applicable
    delivery_commission = 0
    if order.delivery_app:
        commission_rate = await get_delivery_app_commission(order.delivery_app)
        delivery_commission = total * (commission_rate / 100)
    
    # Calculate profit
    profit = total - total_cost - delivery_commission
    
    # Determine payment status
    if order.payment_method == PaymentMethod.PENDING:
        payment_status = "pending"
        # الطلب جاهز تلقائياً إذا تم تحديده
        order_status = OrderStatus.READY if order.auto_ready else OrderStatus.PENDING
    elif order.payment_method == PaymentMethod.CREDIT:
        payment_status = "credit"
        order_status = OrderStatus.READY if order.auto_ready else OrderStatus.PREPARING
    else:
        payment_status = "paid"
        order_status = OrderStatus.READY if order.auto_ready else OrderStatus.PREPARING
    
    # الحصول على اسم شركة التوصيل
    delivery_app_name = None
    if order.delivery_app:
        delivery_app_doc = await db.delivery_apps.find_one({"id": order.delivery_app})
        if delivery_app_doc:
            delivery_app_name = delivery_app_doc.get("name")
    
    order_doc = {
        "id": str(uuid.uuid4()),
        "order_number": order_number,
        "order_type": order.order_type,
        "table_id": order.table_id,
        "customer_name": order.customer_name,
        "customer_phone": order.customer_phone,
        "delivery_address": order.delivery_address,
        "buzzer_number": order.buzzer_number,  # رقم جهاز التنبيه
        "items": items_with_cost,
        "subtotal": subtotal,
        "discount": order.discount,
        "tax": tax,
        "packaging_cost": total_packaging_cost,  # تكلفة التغليف الإجمالية
        "total": total,
        "total_cost": total_cost,
        "profit": profit,
        "branch_id": order.branch_id,
        "cashier_id": current_user["id"],
        "shift_id": shift_id,  # ربط الطلب بالوردية الحالية
        "tenant_id": tenant_id,  # فصل البيانات لكل عميل
        "status": order_status,
        "payment_method": order.payment_method,
        "payment_status": payment_status,
        "delivery_app": order.delivery_app,
        "delivery_app_name": delivery_app_name,  # اسم شركة التوصيل
        "delivery_commission": delivery_commission,
        "driver_id": order.driver_id,
        "notes": order.notes,
        "credit_transferred": False,
        "created_at": datetime.now(timezone.utc).isoformat(),
        "updated_at": datetime.now(timezone.utc).isoformat()
    }
    
    await db.orders.insert_one(order_doc)
    del order_doc["_id"]
    
    # تحديث معلومات العميل إذا كان موجوداً
    if order.customer_phone:
        customer = await db.customers.find_one({"$or": [{"phone": order.customer_phone}, {"phone2": order.customer_phone}]})
        if customer:
            await db.customers.update_one(
                {"id": customer["id"]},
                {
                    "$inc": {"total_orders": 1, "total_spent": total},
                    "$set": {"last_order_date": datetime.now(timezone.utc).isoformat()}
                }
            )
        elif order.customer_name:
            # إنشاء عميل جديد تلقائياً
            new_customer = {
                "id": str(uuid.uuid4()),
                "name": order.customer_name,
                "phone": order.customer_phone,
                "phone2": None,
                "address": order.delivery_address,
                "area": None,
                "notes": None,
                "is_blocked": False,
                "total_orders": 1,
                "total_spent": total,
                "last_order_date": datetime.now(timezone.utc).isoformat(),
                "created_at": datetime.now(timezone.utc).isoformat()
            }
            await db.customers.insert_one(new_customer)
    
    # Update table status if dine-in
    if order.table_id:
        await db.tables.update_one(
            {"id": order.table_id},
            {"$set": {"status": "occupied", "current_order_id": order_doc["id"]}}
        )
    
    # Assign driver if specified
    if order.driver_id:
        driver = await db.drivers.find_one({"id": order.driver_id}, {"_id": 0})
        if driver:
            # Update order with driver info
            await db.orders.update_one(
                {"id": order_doc["id"]},
                {"$set": {
                    "driver_name": driver.get("name"),
                    "driver_phone": driver.get("phone")
                }}
            )
            order_doc["driver_name"] = driver.get("name")
            order_doc["driver_phone"] = driver.get("phone")
            
            # Update driver availability
            await db.drivers.update_one(
                {"id": order.driver_id},
                {"$set": {"is_available": False, "current_order_id": order_doc["id"]}}
            )
    
    # Deduct inventory - خصم المواد الخام من مخزون الفرع بناءً على الوصفات
    # جلب إعدادات المخزون
    inventory_settings = await db.settings.find_one({"type": "inventory_settings"}, {"_id": 0})
    inventory_mode = inventory_settings.get("inventory_mode", "centralized") if inventory_settings else "centralized"
    
    for item in order.items:
        product = await db.products.find_one({"id": item.product_id})
        if product:
            # النظام الجديد: المنتجات المصنعة
            manufactured_product_id = product.get("manufactured_product_id")
            if manufactured_product_id:
                # خصم من مخزون الفرع (branch_inventory)
                branch_item = await db.branch_inventory.find_one({
                    "branch_id": order.branch_id,
                    "product_id": manufactured_product_id
                })
                
                if branch_item:
                    await db.branch_inventory.update_one(
                        {"id": branch_item["id"]},
                        {
                            "$inc": {"quantity": -item.quantity},
                            "$set": {"last_updated": datetime.now(timezone.utc).isoformat()}
                        }
                    )
                elif inventory_mode == "centralized":
                    # في حالة المخزون المركزي، خصم من المنتجات المصنعة مباشرة
                    await db.manufactured_products.update_one(
                        {"id": manufactured_product_id},
                        {
                            "$inc": {"quantity": -item.quantity},
                            "$set": {"last_updated": datetime.now(timezone.utc).isoformat()}
                        }
                    )
                continue
            
            # النظام القديم: المنتجات النهائية
            finished_product_id = product.get("finished_product_id")
            if finished_product_id:
                finished_product = await db.inventory.find_one(
                    {"id": finished_product_id, "item_type": "finished"},
                    {"_id": 0}
                )
                if finished_product and finished_product.get("recipe"):
                    # خصم من المواد الخام بناءً على الوصفة
                    for ingredient in finished_product["recipe"]:
                        raw_material_id = ingredient.get("raw_material_id")
                        qty_per_unit = ingredient.get("quantity", 0)
                        total_to_deduct = qty_per_unit * item.quantity
                        
                        # خصم من مخزون الفرع
                        await db.inventory.update_one(
                            {"id": raw_material_id, "branch_id": order.branch_id},
                            {"$inc": {"quantity": -total_to_deduct}}
                        )
                        # أو من المخزون الرئيسي إذا لم يكن موجوداً في الفرع
                        await db.inventory.update_one(
                            {"id": raw_material_id, "branch_id": "main"},
                            {"$inc": {"quantity": -total_to_deduct}}
                        )
            
            # ثالثاً: استخدم المكونات القديمة إذا كانت موجودة (للتوافق مع البيانات القديمة)
            elif product.get("ingredients"):
                for ing in product["ingredients"]:
                    await db.inventory.update_one(
                        {"id": ing["inventory_id"]},
                        {"$inc": {"quantity": -ing["quantity"] * item.quantity}}
                    )
    
    return order_doc

@api_router.get("/orders", response_model=List[OrderResponse])
async def get_orders(
    branch_id: Optional[str] = None,
    status: Optional[str] = None,
    date: Optional[str] = None,
    payment_status: Optional[str] = None,
    order_type: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    # فلترة حسب tenant_id و branch_id للمستخدم
    query = build_branch_query(current_user)
    
    # إذا تم تحديد فرع في الطلب، تحقق من صلاحية الوصول
    if branch_id:
        if not user_can_access_branch(current_user, branch_id):
            raise HTTPException(status_code=403, detail="لا يمكنك الوصول لهذا الفرع")
        query["branch_id"] = branch_id
    
    if status:
        # دعم حالات متعددة مفصولة بفاصلة
        statuses = [s.strip() for s in status.split(',')]
        if len(statuses) > 1:
            query["status"] = {"$in": statuses}
        else:
            query["status"] = status
    if date:
        query["created_at"] = {"$regex": f"^{date}"}
    if payment_status:
        query["payment_status"] = payment_status
    if order_type:
        query["order_type"] = order_type
    
    orders = await db.orders.find(query, {"_id": 0}).sort("created_at", -1).to_list(500)
    return orders

@api_router.get("/orders/{order_id}", response_model=OrderResponse)
async def get_order(order_id: str, current_user: dict = Depends(get_current_user)):
    query = build_tenant_query(current_user, {"id": order_id})
    order = await db.orders.find_one(query, {"_id": 0})
    if not order:
        raise HTTPException(status_code=404, detail="الطلب غير موجود")
    return order

@api_router.put("/orders/{order_id}/add-items")
async def add_items_to_order(order_id: str, items: List[OrderItemCreate], current_user: dict = Depends(get_current_user)):
    """إضافة عناصر جديدة لطلب موجود"""
    query = build_tenant_query(current_user, {"id": order_id})
    order = await db.orders.find_one(query)
    if not order:
        raise HTTPException(status_code=404, detail="الطلب غير موجود")
    
    # إضافة العناصر الجديدة
    new_items = []
    product_query = build_tenant_query(current_user)
    for item in items:
        product_query["id"] = item.product_id
        product = await db.products.find_one(product_query)
        new_items.append({
            "product_id": item.product_id,
            "product_name": item.product_name,
            "quantity": item.quantity,
            "price": item.price,
            "cost": product.get("cost", 0) if product else 0,
            "notes": item.notes
        })
    
    # دمج العناصر الجديدة مع القديمة
    existing_items = order.get("items", [])
    all_items = existing_items + new_items
    
    # إعادة حساب المجاميع
    subtotal = sum(i["price"] * i["quantity"] for i in all_items)
    total_cost = sum(i.get("cost", 0) * i["quantity"] for i in all_items)
    discount = order.get("discount", 0)
    tax = 0
    total = subtotal - discount + tax
    profit = total - total_cost - order.get("delivery_commission", 0)
    
    await db.orders.update_one(
        {"id": order_id},
        {"$set": {
            "items": all_items,
            "subtotal": subtotal,
            "total_cost": total_cost,
            "total": total,
            "profit": profit,
            "updated_at": datetime.now(timezone.utc).isoformat()
        }}
    )
    
    query = build_tenant_query(current_user, {"id": order_id})
    return await db.orders.find_one(query, {"_id": 0})

@api_router.put("/orders/{order_id}/status")
async def update_order_status(order_id: str, status: str, current_user: dict = Depends(get_current_user)):
    query = build_tenant_query(current_user, {"id": order_id})
    order = await db.orders.find_one(query)
    if not order:
        raise HTTPException(status_code=404, detail="الطلب غير موجود")
    
    # التحقق من صلاحية الإلغاء
    if status == OrderStatus.CANCELLED:
        # فقط المالك أو المدير يمكنهم الإلغاء
        if current_user.get("role") not in ["admin", "manager"]:
            raise HTTPException(status_code=403, detail="ليس لديك صلاحية إلغاء الطلبات")
    
    await db.orders.update_one(
        query,
        {"$set": {"status": status, "updated_at": datetime.now(timezone.utc).isoformat()}}
    )
    
    # Free table if completed
    if status in [OrderStatus.DELIVERED, OrderStatus.CANCELLED] and order.get("table_id"):
        await db.tables.update_one(
            {"id": order["table_id"]},
            {"$set": {"status": "available", "current_order_id": None}}
        )
    
    return {"message": "تم التحديث"}

@api_router.put("/orders/{order_id}/kitchen-status")
async def update_order_kitchen_status(order_id: str, kitchen_status: str, current_user: dict = Depends(get_current_user)):
    """
    تحديث حالة الطلب في المطبخ بشكل مستقل عن حالة الطلب الرئيسية
    kitchen_status: pending_kitchen, preparing_kitchen, ready_kitchen, completed_kitchen
    """
    query = build_tenant_query(current_user, {"id": order_id})
    order = await db.orders.find_one(query)
    if not order:
        raise HTTPException(status_code=404, detail="الطلب غير موجود")
    
    valid_statuses = ["pending_kitchen", "preparing_kitchen", "ready_kitchen", "completed_kitchen"]
    if kitchen_status not in valid_statuses:
        raise HTTPException(status_code=400, detail="حالة المطبخ غير صالحة")
    
    await db.orders.update_one(
        query,
        {"$set": {
            "kitchen_status": kitchen_status, 
            "kitchen_updated_at": datetime.now(timezone.utc).isoformat()
        }}
    )
    
    return {"message": "تم تحديث حالة المطبخ"}

@api_router.get("/kitchen/orders")
async def get_kitchen_orders(current_user: dict = Depends(get_current_user)):
    """
    جلب الطلبات لشاشة المطبخ
    تظهر الطلبات التي لم يتم تحديدها كـ completed_kitchen بغض النظر عن حالتها الرئيسية
    """
    tenant_id = get_user_tenant_id(current_user)
    
    # جلب الطلبات التي ليست completed_kitchen وليست ملغاة
    query = {
        "tenant_id": tenant_id,
        "kitchen_status": {"$nin": ["completed_kitchen", None]},
        "status": {"$ne": "cancelled"}
    }
    
    # أيضاً جلب الطلبات الجديدة التي ليس لها kitchen_status بعد
    query_new = {
        "tenant_id": tenant_id,
        "kitchen_status": {"$exists": False},
        "status": {"$in": ["pending", "preparing", "ready"]}
    }
    
    # دمج النتائج
    orders_with_kitchen_status = await db.orders.find(query, {"_id": 0}).sort("created_at", 1).to_list(100)
    orders_new = await db.orders.find(query_new, {"_id": 0}).sort("created_at", 1).to_list(100)
    
    # إضافة kitchen_status الافتراضي للطلبات الجديدة
    for order in orders_new:
        if "kitchen_status" not in order:
            order["kitchen_status"] = "pending_kitchen"
            # تحديث في قاعدة البيانات
            await db.orders.update_one(
                {"id": order["id"]},
                {"$set": {"kitchen_status": "pending_kitchen"}}
            )
    
    all_orders = orders_with_kitchen_status + orders_new
    
    # إزالة التكرارات وترتيب حسب التاريخ
    seen_ids = set()
    unique_orders = []
    for order in all_orders:
        if order["id"] not in seen_ids:
            seen_ids.add(order["id"])
            unique_orders.append(order)
    
    unique_orders.sort(key=lambda x: x.get("created_at", ""))
    
    # إضافة اسم الفرع لكل طلب
    branch_ids = list(set(o.get("branch_id") for o in unique_orders if o.get("branch_id")))
    branches = {}
    if branch_ids:
        branches_list = await db.branches.find({"id": {"$in": branch_ids}}, {"_id": 0, "id": 1, "name": 1}).to_list(100)
        branches = {b["id"]: b["name"] for b in branches_list}
    
    for order in unique_orders:
        order["branch_name"] = branches.get(order.get("branch_id"), "")
    
    return unique_orders

@api_router.put("/orders/{order_id}/payment")
async def update_order_payment(order_id: str, payment_method: str, current_user: dict = Depends(get_current_user)):
    query = build_tenant_query(current_user, {"id": order_id})
    order = await db.orders.find_one(query)
    if not order:
        raise HTTPException(status_code=404, detail="الطلب غير موجود")
    
    update_data = {
        "payment_method": payment_method,
        "updated_at": datetime.now(timezone.utc).isoformat()
    }
    
    if payment_method == PaymentMethod.CREDIT:
        update_data["payment_status"] = "credit"
        if order.get("delivery_app"):
            update_data["credit_transferred"] = True
    elif payment_method in [PaymentMethod.CASH, PaymentMethod.CARD]:
        update_data["payment_status"] = "paid"
    
    await db.orders.update_one({"id": order_id}, {"$set": update_data})
    
    return {"message": "تم تحديث طريقة الدفع"}

@api_router.put("/orders/{order_id}/cancel")
async def cancel_order(order_id: str, current_user: dict = Depends(get_current_user)):
    """
    إلغاء الطلب:
    - أقل من دقيقة: يُحذف نهائياً (لا يظهر في التقارير)
    - أكثر من دقيقتين: يُسجل كطلب ملغي (يظهر في التقارير)
    - بين دقيقة ودقيقتين: يُسجل كطلب ملغي فقط للمدير/المالك
    """
    query = build_tenant_query(current_user, {"id": order_id})
    order = await db.orders.find_one(query)
    if not order:
        raise HTTPException(status_code=404, detail="الطلب غير موجود")
    
    # حساب الفرق في الوقت بالثواني
    created_at = datetime.fromisoformat(order["created_at"].replace("Z", "+00:00"))
    time_diff = (datetime.now(timezone.utc) - created_at).total_seconds()
    
    is_within_minute = time_diff < 60  # أقل من دقيقة
    is_within_two_minutes = time_diff < 120  # أقل من دقيقتين
    is_admin_or_manager = current_user.get("role") in ["admin", "manager"]
    
    # التحقق من الصلاحيات
    if not is_within_minute and not is_admin_or_manager:
        raise HTTPException(status_code=403, detail="فقط المالك أو المدير يمكنهم إلغاء الطلبات بعد دقيقة من إنشائها")
    
    # تحرير الطاولة إذا كان الطلب على طاولة
    if order.get("table_id"):
        await db.tables.update_one(
            {"id": order["table_id"]},
            {"$set": {"status": "available", "current_order_id": None}}
        )
    
    # أقل من دقيقة: حذف نهائي
    if is_within_minute:
        await db.orders.delete_one({"id": order_id})
        return {
            "message": "تم حذف الطلب نهائياً",
            "was_quick_delete": True,
            "in_reports": False
        }
    
    # أكثر من دقيقة: تسجيل كطلب ملغي
    await db.orders.update_one(
        {"id": order_id},
        {"$set": {
            "status": OrderStatus.CANCELLED,
            "cancelled_at": datetime.now(timezone.utc).isoformat(),
            "cancelled_by": current_user["id"],
            "cancellation_reason": "إلغاء بواسطة المدير" if is_admin_or_manager else "إلغاء",
            "time_to_cancel_seconds": int(time_diff),
            "updated_at": datetime.now(timezone.utc).isoformat()
        }}
    )
    
    return {
        "message": "تم إلغاء الطلب وتسجيله في التقارير",
        "was_quick_delete": False,
        "in_reports": True
    }

# ==================== REFUND/RETURN ROUTES - إرجاع الطلبات ====================

class RefundCreate(BaseModel):
    """نموذج إنشاء إرجاع"""
    order_id: str  # رقم الطلب (يمكن أن يكون order_number أو id)
    reason: str  # سبب الإرجاع (مطلوب)
    refund_type: str = "full"  # full أو partial
    refund_amount: Optional[float] = None  # المبلغ المسترد (للإرجاع الجزئي)
    items_to_refund: Optional[List[Dict[str, Any]]] = None  # العناصر المسترجعة (للإرجاع الجزئي)
    notes: Optional[str] = None

class RefundResponse(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str
    refund_number: int
    order_id: str
    order_number: int
    order_type: str
    original_total: float
    refund_amount: float
    refund_type: str
    reason: str
    items_refunded: List[Dict[str, Any]] = []
    refunded_by: str
    refunded_by_name: str
    branch_id: str
    status: str  # pending, approved, completed, rejected
    notes: Optional[str] = None
    created_at: str
    completed_at: Optional[str] = None

async def get_next_refund_number(branch_id: str) -> int:
    """الحصول على رقم الإرجاع التالي"""
    today = datetime.now(timezone.utc).strftime("%Y-%m-%d")
    counter = await db.refund_counters.find_one_and_update(
        {"branch_id": branch_id, "date": today},
        {"$inc": {"counter": 1}},
        upsert=True,
        return_document=True
    )
    return counter["counter"]

@api_router.post("/refunds")
async def create_refund(refund: RefundCreate, current_user: dict = Depends(get_current_user)):
    """إنشاء طلب إرجاع - يتطلب صلاحية can_refund"""
    
    # التحقق من الصلاحية
    user_permissions = current_user.get("permissions", [])
    user_role = current_user.get("role", "")
    
    # المدير والمالك لديهم صلاحية كاملة، أو المستخدم لديه صلاحية can_refund
    if user_role not in [UserRole.ADMIN, UserRole.SUPER_ADMIN, UserRole.MANAGER] and "can_refund" not in user_permissions:
        raise HTTPException(status_code=403, detail="ليس لديك صلاحية إرجاع الطلبات")
    
    # التحقق من وجود سبب الإرجاع (شرط إلزامي)
    if not refund.reason or len(refund.reason.strip()) < 3:
        raise HTTPException(status_code=400, detail="يجب كتابة سبب الإرجاع (3 أحرف على الأقل)")
    
    tenant_id = get_user_tenant_id(current_user)
    
    # البحث عن الطلب برقم الفاتورة أو الـ ID
    order_query = {"$or": [{"id": refund.order_id}]}
    
    # محاولة تحويل order_id إلى رقم للبحث برقم الفاتورة
    try:
        order_number = int(refund.order_id)
        order_query["$or"].append({"order_number": order_number})
    except ValueError:
        pass
    
    if tenant_id:
        order_query["tenant_id"] = tenant_id
    
    # جلب آخر طلب بهذا الرقم
    orders = await db.orders.find(order_query, {"_id": 0}).sort("created_at", -1).to_list(1)
    
    if not orders:
        raise HTTPException(status_code=404, detail="الطلب غير موجود. تأكد من رقم الفاتورة")
    
    order = orders[0]
    
    # التحقق من أن الطلب من نفس اليوم
    order_date = order.get("created_at", "")
    if order_date:
        if isinstance(order_date, str):
            order_datetime = datetime.fromisoformat(order_date.replace("Z", "+00:00"))
        else:
            order_datetime = order_date
        
        today = datetime.now(timezone.utc).date()
        order_day = order_datetime.date()
        
        if order_day != today:
            raise HTTPException(
                status_code=400, 
                detail=f"لا يمكن إرجاع هذا الطلب. الإرجاع متاح فقط لطلبات اليوم. تاريخ الطلب: {order_day.strftime('%Y-%m-%d')}"
            )
    
    # التحقق من أن الطلب مدفوع
    if order.get("payment_status") not in ["paid", "credit"]:
        raise HTTPException(status_code=400, detail="لا يمكن إرجاع طلب غير مدفوع")
    
    # التحقق من عدم وجود إرجاع سابق لنفس الطلب (إرجاع كامل)
    existing_refund = await db.refunds.find_one({
        "order_id": order["id"],
        "refund_type": "full",
        "status": {"$in": ["pending", "approved", "completed"]}
    })
    if existing_refund:
        raise HTTPException(status_code=400, detail="تم إرجاع هذا الطلب مسبقاً")
    
    # حساب مبلغ الإرجاع
    original_total = order.get("total", 0)
    
    if refund.refund_type == "full":
        refund_amount = original_total
        items_refunded = order.get("items", [])
    else:
        # إرجاع جزئي
        if not refund.refund_amount and not refund.items_to_refund:
            raise HTTPException(status_code=400, detail="يجب تحديد المبلغ أو العناصر للإرجاع الجزئي")
        refund_amount = refund.refund_amount or 0
        items_refunded = refund.items_to_refund or []
    
    # إنشاء سجل الإرجاع
    refund_number = await get_next_refund_number(order.get("branch_id", ""))
    
    refund_doc = {
        "id": str(uuid.uuid4()),
        "refund_number": refund_number,
        "order_id": order["id"],
        "order_number": order.get("order_number", 0),
        "order_type": order.get("order_type", ""),
        "original_total": original_total,
        "refund_amount": refund_amount,
        "refund_type": refund.refund_type,
        "reason": refund.reason.strip(),
        "items_refunded": items_refunded,
        "refunded_by": current_user["id"],
        "refunded_by_name": current_user.get("full_name", current_user.get("username", "")),
        "branch_id": order.get("branch_id", ""),
        "tenant_id": tenant_id,
        "status": "completed",  # يتم الإرجاع مباشرة
        "notes": refund.notes,
        "created_at": datetime.now(timezone.utc).isoformat(),
        "completed_at": datetime.now(timezone.utc).isoformat(),
        # معلومات إضافية للتقارير
        "customer_name": order.get("customer_name"),
        "customer_phone": order.get("customer_phone"),
        "original_payment_method": order.get("payment_method"),
        "shift_id": order.get("shift_id")
    }
    
    await db.refunds.insert_one(refund_doc)
    
    # تحديث حالة الطلب الأصلي
    update_data = {
        "is_refunded": True,
        "refund_id": refund_doc["id"],
        "refund_amount": refund_amount,
        "refund_reason": refund.reason,
        "refunded_at": datetime.now(timezone.utc).isoformat(),
        "refunded_by": current_user["id"]
    }
    
    if refund.refund_type == "full":
        update_data["status"] = "refunded"
    
    await db.orders.update_one({"id": order["id"]}, {"$set": update_data})
    
    # إضافة سجل في المصاريف (اختياري - للمحاسبة)
    expense_doc = {
        "id": str(uuid.uuid4()),
        "category": "refund",
        "description": f"إرجاع طلب #{order.get('order_number', 0)} - {refund.reason}",
        "amount": refund_amount,
        "payment_method": order.get("payment_method", "cash"),
        "reference_number": f"REF-{refund_number}",
        "branch_id": order.get("branch_id", ""),
        "tenant_id": tenant_id,
        "created_by": current_user["id"],
        "date": datetime.now(timezone.utc).strftime("%Y-%m-%d"),
        "created_at": datetime.now(timezone.utc).isoformat(),
        "refund_id": refund_doc["id"],
        "order_id": order["id"]
    }
    await db.expenses.insert_one(expense_doc)
    
    if "_id" in refund_doc:
        del refund_doc["_id"]
    return refund_doc

@api_router.get("/refunds")
async def get_refunds(
    branch_id: Optional[str] = None,
    date_from: Optional[str] = None,
    date_to: Optional[str] = None,
    status: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    """جلب قائمة الإرجاعات"""
    query = build_tenant_query(current_user)
    
    if branch_id:
        query["branch_id"] = branch_id
    
    if status:
        query["status"] = status
    
    if date_from:
        query["created_at"] = {"$gte": date_from}
    
    if date_to:
        if "created_at" in query:
            query["created_at"]["$lte"] = date_to + "T23:59:59"
        else:
            query["created_at"] = {"$lte": date_to + "T23:59:59"}
    
    refunds = await db.refunds.find(query, {"_id": 0}).sort("created_at", -1).to_list(500)
    return refunds

@api_router.get("/refunds/{refund_id}")
async def get_refund(refund_id: str, current_user: dict = Depends(get_current_user)):
    """جلب تفاصيل إرجاع محدد"""
    query = build_tenant_query(current_user, {"id": refund_id})
    refund = await db.refunds.find_one(query, {"_id": 0})
    
    if not refund:
        raise HTTPException(status_code=404, detail="الإرجاع غير موجود")
    
    return refund

@api_router.get("/orders/{order_id}/refund-status")
async def check_order_refund_status(order_id: str, current_user: dict = Depends(get_current_user)):
    """التحقق من حالة إرجاع طلب معين"""
    tenant_id = get_user_tenant_id(current_user)
    
    # البحث عن الطلب برقم الفاتورة أو الـ ID
    or_conditions = [{"id": order_id}]
    try:
        order_number = int(order_id)
        or_conditions.append({"order_number": order_number})
    except ValueError:
        pass
    
    # بناء الاستعلام مع tenant_id
    if tenant_id:
        order_query = {"$and": [{"$or": or_conditions}, {"tenant_id": tenant_id}]}
    else:
        order_query = {"$or": or_conditions}
    
    # جلب آخر طلب بهذا الرقم (الأحدث)
    orders = await db.orders.find(order_query, {"_id": 0}).sort("created_at", -1).to_list(1)
    
    if not orders:
        raise HTTPException(status_code=404, detail="الطلب غير موجود. تأكد من رقم الفاتورة")
    
    order = orders[0]
    
    # التحقق من تاريخ الطلب
    order_date = order.get("created_at", "")
    is_today = False
    order_date_str = ""
    
    if order_date:
        if isinstance(order_date, str):
            order_datetime = datetime.fromisoformat(order_date.replace("Z", "+00:00"))
        else:
            order_datetime = order_date
        
        today = datetime.now(timezone.utc).date()
        order_day = order_datetime.date()
        is_today = (order_day == today)
        order_date_str = order_day.strftime('%Y-%m-%d')
    
    # البحث عن إرجاعات لهذا الطلب
    refunds = await db.refunds.find({"order_id": order["id"]}, {"_id": 0}).to_list(10)
    
    # تحديد إمكانية الإرجاع
    can_refund = (
        order.get("payment_status") in ["paid", "credit"] 
        and not order.get("is_refunded") 
        and is_today
    )
    
    # رسالة توضيحية إذا لم يكن قابل للإرجاع
    refund_message = None
    if not can_refund:
        if order.get("is_refunded"):
            refund_message = "تم إرجاع هذا الطلب مسبقاً"
        elif order.get("payment_status") not in ["paid", "credit"]:
            refund_message = "الطلب غير مدفوع"
        elif not is_today:
            refund_message = f"لا يمكن إرجاع طلبات الأيام السابقة. تاريخ الطلب: {order_date_str}"
    
    return {
        "order_id": order["id"],
        "order_number": order.get("order_number"),
        "order_type": order.get("order_type"),
        "total": order.get("total", 0),
        "payment_status": order.get("payment_status"),
        "customer_name": order.get("customer_name"),
        "created_at": order.get("created_at"),
        "order_date": order_date_str,
        "is_today": is_today,
        "is_refunded": order.get("is_refunded", False),
        "can_refund": can_refund,
        "refund_message": refund_message,
        "refunds": refunds
    }

# ==================== DELIVERY APP SETTINGS ====================

@api_router.post("/delivery-app-settings")
async def create_delivery_app_setting(setting: DeliveryAppSettingCreate, current_user: dict = Depends(get_current_user)):
    if current_user["role"] not in [UserRole.ADMIN, UserRole.MANAGER]:
        raise HTTPException(status_code=403, detail="غير مصرح")
    
    tenant_id = get_user_tenant_id(current_user)
    
    # Check if exists for this tenant
    query = {"app_id": setting.app_id}
    if tenant_id:
        query["tenant_id"] = tenant_id
    
    existing = await db.delivery_app_settings.find_one(query)
    
    setting_data = setting.model_dump()
    setting_data["tenant_id"] = tenant_id
    
    if existing:
        await db.delivery_app_settings.update_one(query, {"$set": setting_data})
    else:
        await db.delivery_app_settings.insert_one(setting_data)
    
    return {"message": "تم الحفظ"}

@api_router.get("/delivery-app-settings")
async def get_delivery_app_settings(current_user: dict = Depends(get_current_user)):
    query = build_tenant_query(current_user)
    settings = await db.delivery_app_settings.find(query, {"_id": 0}).to_list(20)
    return settings

@api_router.delete("/delivery-app-settings/{app_id}")
async def delete_delivery_app_setting(app_id: str, current_user: dict = Depends(get_current_user)):
    """حذف شركة توصيل"""
    query = build_tenant_query(current_user, {"app_id": app_id})
    result = await db.delivery_app_settings.delete_one(query)
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="شركة التوصيل غير موجودة")
    return {"success": True, "message": "تم حذف شركة التوصيل بنجاح"}

@api_router.get("/delivery-apps")
async def get_delivery_apps(current_user: dict = Depends(get_current_user)):
    # Get default apps
    default_apps = [
        {"id": "toters", "name": "توترز", "name_en": "Toters", "icon": "Truck", "is_default": True},
        {"id": "talabat", "name": "طلبات", "name_en": "Talabat", "icon": "ShoppingBag", "is_default": True},
        {"id": "baly", "name": "بالي", "name_en": "Baly", "icon": "Package", "is_default": True},
        {"id": "alsaree3", "name": "عالسريع", "name_en": "Al-Sari3", "icon": "Zap", "is_default": True},
        {"id": "talabati", "name": "طلباتي", "name_en": "Talabati", "icon": "Box", "is_default": True},
    ]
    
    # Get all settings from database for this tenant
    query = build_tenant_query(current_user)
    all_settings = await db.delivery_app_settings.find(query, {"_id": 0}).to_list(50)
    
    # Create a map of app_id to settings
    settings_map = {s["app_id"]: s for s in all_settings}
    
    # Update default apps with their settings
    result_apps = []
    for app in default_apps:
        setting = settings_map.get(app["id"])
        if setting:
            app["commission_rate"] = setting.get("commission_rate", 0)
            app["is_active"] = setting.get("is_active", True)
        else:
            app["commission_rate"] = 0
            app["is_active"] = True
        result_apps.append(app)
    
    # Add custom apps (apps that are not in default list)
    default_ids = {a["id"] for a in default_apps}
    for setting in all_settings:
        if setting["app_id"] not in default_ids:
            result_apps.append({
                "id": setting["app_id"],
                "name": setting.get("name", setting["app_id"]),
                "name_en": setting.get("name_en", setting["app_id"]),
                "icon": "Truck",
                "is_default": False,
                "commission_rate": setting.get("commission_rate", 0),
                "is_active": setting.get("is_active", True)
            })
    
    return result_apps

# ==================== EXPORT TO EXCEL ====================

from io import BytesIO
from fastapi.responses import StreamingResponse

@api_router.get("/reports/export/excel")
async def export_sales_to_excel(
    report_type: str = "sales",  # sales, products, delivery, expenses
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    branch_id: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    """تصدير التقارير إلى Excel"""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        raise HTTPException(status_code=500, detail="مكتبة Excel غير متوفرة")
    
    # Create workbook
    wb = Workbook()
    ws = wb.active
    
    # Styling
    header_font = Font(bold=True, color="FFFFFF", size=12)
    header_fill = PatternFill(start_color="1F2937", end_color="1F2937", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Query dates
    if not start_date:
        start_date = datetime.now(timezone.utc).strftime("%Y-%m-%d")
    if not end_date:
        end_date = start_date
    
    query = build_tenant_query(current_user, {
        "created_at": {"$gte": start_date, "$lte": end_date + "T23:59:59"}
    })
    if branch_id:
        query["branch_id"] = branch_id
    
    if report_type == "sales":
        ws.title = "تقرير المبيعات"
        
        # Get orders
        orders = await db.orders.find(query, {"_id": 0}).to_list(10000)
        
        # Headers
        headers = ["رقم الطلب", "التاريخ", "الوقت", "النوع", "العميل", "الفرع", "طريقة الدفع", "المبلغ", "الحالة"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border
        
        # Data
        order_types = {"dine_in": "محلي", "takeaway": "سفري", "delivery": "توصيل"}
        payment_methods = {"cash": "نقدي", "card": "بطاقة", "credit": "آجل"}
        statuses = {"pending": "معلق", "preparing": "قيد التحضير", "ready": "جاهز", "completed": "مكتمل", "delivered": "تم التوصيل", "cancelled": "ملغي"}
        
        for row, order in enumerate(orders, 2):
            created_at = datetime.fromisoformat(order["created_at"].replace("Z", "+00:00")) if order.get("created_at") else datetime.now()
            
            data = [
                order.get("order_number", ""),
                created_at.strftime("%Y-%m-%d"),
                created_at.strftime("%H:%M"),
                order_types.get(order.get("order_type", ""), order.get("order_type", "")),
                order.get("customer_name", "بدون اسم"),
                order.get("branch_name", ""),
                payment_methods.get(order.get("payment_method", ""), order.get("payment_method", "")),
                order.get("total", 0),
                statuses.get(order.get("status", ""), order.get("status", ""))
            ]
            
            for col, value in enumerate(data, 1):
                cell = ws.cell(row=row, column=col, value=value)
                cell.border = thin_border
                if col == 8:  # Amount column
                    cell.number_format = '#,##0'
        
        # Summary row
        summary_row = len(orders) + 3
        ws.cell(row=summary_row, column=7, value="الإجمالي:").font = Font(bold=True)
        total_cell = ws.cell(row=summary_row, column=8, value=sum(o.get("total", 0) for o in orders))
        total_cell.font = Font(bold=True)
        total_cell.number_format = '#,##0'
        
    elif report_type == "products":
        ws.title = "تقرير المنتجات"
        
        # Get orders with items
        orders = await db.orders.find(query, {"_id": 0, "items": 1, "total": 1, "status": 1}).to_list(10000)
        
        # Aggregate products
        products = {}
        for order in orders:
            if order.get("status") == "cancelled":
                continue
            for item in order.get("items", []):
                name = item.get("name", "Unknown")
                if name not in products:
                    products[name] = {"quantity": 0, "revenue": 0}
                products[name]["quantity"] += item.get("quantity", 1)
                products[name]["revenue"] += item.get("price", 0) * item.get("quantity", 1)
        
        # Headers
        headers = ["المنتج", "الكمية المباعة", "الإيرادات"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border
        
        # Data sorted by quantity
        sorted_products = sorted(products.items(), key=lambda x: x[1]["quantity"], reverse=True)
        for row, (name, data) in enumerate(sorted_products, 2):
            ws.cell(row=row, column=1, value=name).border = thin_border
            ws.cell(row=row, column=2, value=data["quantity"]).border = thin_border
            revenue_cell = ws.cell(row=row, column=3, value=data["revenue"])
            revenue_cell.border = thin_border
            revenue_cell.number_format = '#,##0'
        
    elif report_type == "expenses":
        ws.title = "تقرير المصاريف"
        
        # Get expenses
        expenses = await db.expenses.find(query, {"_id": 0}).to_list(10000)
        
        # Headers
        headers = ["التاريخ", "الفئة", "الوصف", "المبلغ"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border
        
        # Data
        for row, expense in enumerate(expenses, 2):
            created_at = datetime.fromisoformat(expense["created_at"].replace("Z", "+00:00")) if expense.get("created_at") else datetime.now()
            
            data = [
                created_at.strftime("%Y-%m-%d"),
                expense.get("category", ""),
                expense.get("description", ""),
                expense.get("amount", 0)
            ]
            
            for col, value in enumerate(data, 1):
                cell = ws.cell(row=row, column=col, value=value)
                cell.border = thin_border
                if col == 4:
                    cell.number_format = '#,##0'
        
        # Summary
        summary_row = len(expenses) + 3
        ws.cell(row=summary_row, column=3, value="الإجمالي:").font = Font(bold=True)
        total_cell = ws.cell(row=summary_row, column=4, value=sum(e.get("amount", 0) for e in expenses))
        total_cell.font = Font(bold=True)
        total_cell.number_format = '#,##0'
    
    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2) * 1.2
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Save to bytes
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    filename = f"report_{report_type}_{start_date}_to_{end_date}.xlsx"
    
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )

# ==================== EXPORT TO PDF ====================

@api_router.get("/reports/export/pdf")
async def export_report_to_pdf(
    report_type: str = "sales",  # sales, payroll, expenses, inventory
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    branch_id: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    """تصدير التقارير إلى PDF"""
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.lib.units import cm
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    
    tenant_id = get_user_tenant_id(current_user)
    
    if not start_date:
        start_date = datetime.now(timezone.utc).strftime("%Y-%m-%d")
    if not end_date:
        end_date = start_date
    
    # Build query with branch filtering
    user_branch_id = current_user.get("branch_id")
    user_role = current_user.get("role")
    
    effective_branch_id = None
    if user_branch_id and user_role not in [UserRole.SUPER_ADMIN, UserRole.ADMIN, UserRole.MANAGER]:
        effective_branch_id = user_branch_id
    elif branch_id:
        effective_branch_id = branch_id
    
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(A4), rightMargin=1*cm, leftMargin=1*cm, topMargin=1*cm, bottomMargin=1*cm)
    
    styles = getSampleStyleSheet()
    elements = []
    
    # Title
    title_text = ""
    headers = []
    data_rows = []
    totals_row = None
    
    if report_type == "sales":
        title_text = f"تقرير المبيعات - من {start_date} إلى {end_date}"
        
        query = {"status": {"$ne": "cancelled"}, "created_at": {"$gte": start_date, "$lte": end_date + "T23:59:59"}}
        if tenant_id:
            query["tenant_id"] = tenant_id
        if effective_branch_id:
            query["branch_id"] = effective_branch_id
        
        orders = await db.orders.find(query, {"_id": 0}).to_list(10000)
        
        headers = ["#", "رقم الطلب", "التاريخ", "النوع", "طريقة الدفع", "المبلغ"]
        order_types = {"dine_in": "محلي", "takeaway": "سفري", "delivery": "توصيل"}
        payment_methods = {"cash": "نقدي", "card": "بطاقة", "credit": "آجل"}
        
        total_amount = 0
        for idx, order in enumerate(orders, 1):
            created_at = order.get("created_at", "")[:10]
            amount = order.get("total", 0)
            total_amount += amount
            
            data_rows.append([
                str(idx),
                order.get("order_number", ""),
                created_at,
                order_types.get(order.get("order_type"), order.get("order_type", "")),
                payment_methods.get(order.get("payment_method"), order.get("payment_method", "")),
                f"{amount:,.0f}"
            ])
        
        totals_row = ["", "", "", "", "الإجمالي:", f"{total_amount:,.0f}"]
        
    elif report_type == "expenses":
        title_text = f"تقرير المصاريف - من {start_date} إلى {end_date}"
        
        query = {"date": {"$gte": start_date, "$lte": end_date}}
        if tenant_id:
            query["tenant_id"] = tenant_id
        if effective_branch_id:
            query["branch_id"] = effective_branch_id
        
        expenses = await db.expenses.find(query, {"_id": 0}).to_list(1000)
        
        headers = ["#", "التاريخ", "الفئة", "الوصف", "المبلغ"]
        category_names = {
            "supplies": "مستلزمات", "utilities": "خدمات", "salaries": "رواتب",
            "rent": "إيجار", "maintenance": "صيانة", "marketing": "تسويق", "other": "أخرى"
        }
        
        total_amount = 0
        for idx, expense in enumerate(expenses, 1):
            amount = expense.get("amount", 0)
            total_amount += amount
            
            data_rows.append([
                str(idx),
                expense.get("date", ""),
                category_names.get(expense.get("category"), expense.get("category", "")),
                expense.get("description", "")[:30],
                f"{amount:,.0f}"
            ])
        
        totals_row = ["", "", "", "الإجمالي:", f"{total_amount:,.0f}"]
        
    elif report_type == "inventory":
        title_text = "تقرير المخزون"
        
        query = {}
        if tenant_id:
            query["tenant_id"] = tenant_id
        if effective_branch_id:
            query["branch_id"] = effective_branch_id
        
        items = await db.inventory.find(query, {"_id": 0}).to_list(1000)
        
        headers = ["#", "الصنف", "النوع", "الكمية", "الحد الأدنى", "سعر الوحدة", "القيمة"]
        type_names = {"raw": "خام", "finished": "منتج نهائي"}
        
        total_value = 0
        for idx, item in enumerate(items, 1):
            qty = item.get("quantity", 0)
            cost = item.get("cost_per_unit", 0)
            value = qty * cost
            total_value += value
            
            data_rows.append([
                str(idx),
                item.get("name", ""),
                type_names.get(item.get("item_type"), item.get("item_type", "")),
                str(qty),
                str(item.get("min_quantity", 0)),
                f"{cost:,.0f}",
                f"{value:,.0f}"
            ])
        
        totals_row = ["", "", "", "", "", "الإجمالي:", f"{total_value:,.0f}"]
    
    elif report_type == "payroll":
        title_text = f"تقرير الرواتب - {start_date[:7]}"
        
        query = {"is_active": True}
        if tenant_id:
            query["tenant_id"] = tenant_id
        if effective_branch_id:
            query["branch_id"] = effective_branch_id
        
        employees = await db.employees.find(query, {"_id": 0}).to_list(500)
        month = start_date[:7]
        month_start = f"{month}-01"
        month_end = f"{month}-31"
        
        headers = ["#", "الموظف", "الوظيفة", "الراتب", "المكافآت", "الخصومات", "السلف", "الصافي"]
        
        totals = [0, 0, 0, 0, 0]
        for idx, emp in enumerate(employees, 1):
            # Get deductions, bonuses, advances
            deductions = await db.deductions.find({
                "employee_id": emp["id"],
                "date": {"$gte": month_start, "$lte": month_end}
            }, {"_id": 0}).to_list(100)
            emp_deductions = sum(d.get("amount", 0) for d in deductions)
            
            bonuses = await db.bonuses.find({
                "employee_id": emp["id"],
                "date": {"$gte": month_start, "$lte": month_end}
            }, {"_id": 0}).to_list(100)
            emp_bonuses = sum(b.get("amount", 0) for b in bonuses)
            
            advances = await db.advances.find({
                "employee_id": emp["id"],
                "status": "approved",
                "remaining_amount": {"$gt": 0}
            }, {"_id": 0}).to_list(100)
            emp_advances = sum(a.get("monthly_deduction", 0) for a in advances)
            
            basic = emp.get("salary", 0)
            net = basic + emp_bonuses - emp_deductions - emp_advances
            
            totals[0] += basic
            totals[1] += emp_bonuses
            totals[2] += emp_deductions
            totals[3] += emp_advances
            totals[4] += net
            
            data_rows.append([
                str(idx),
                emp.get("name", ""),
                emp.get("position", ""),
                f"{basic:,.0f}",
                f"{emp_bonuses:,.0f}",
                f"{emp_deductions:,.0f}",
                f"{emp_advances:,.0f}",
                f"{net:,.0f}"
            ])
        
        totals_row = ["", "", "الإجمالي:", f"{totals[0]:,.0f}", f"{totals[1]:,.0f}", f"{totals[2]:,.0f}", f"{totals[3]:,.0f}", f"{totals[4]:,.0f}"]
    
    # Build PDF
    title_style = styles['Heading1']
    title_style.alignment = 1
    elements.append(Paragraph(title_text, title_style))
    elements.append(Spacer(1, 20))
    
    if headers and data_rows:
        table_data = [headers] + data_rows
        if totals_row:
            table_data.append(totals_row)
        
        col_widths = [doc.width / len(headers)] * len(headers)
        table = Table(table_data, colWidths=col_widths)
        
        style = TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4472C4')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('FONTSIZE', (0, 1), (-1, -1), 9),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.white),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('ROWBACKGROUNDS', (0, 1), (-1, -2), [colors.white, colors.HexColor('#F5F5F5')]),
        ])
        
        if totals_row:
            style.add('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#E6E6E6'))
            style.add('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold')
        
        table.setStyle(style)
        elements.append(table)
    
    doc.build(elements)
    buffer.seek(0)
    
    filename = f"report_{report_type}_{start_date}_to_{end_date}.pdf"
    
    return StreamingResponse(
        buffer,
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )

@api_router.get("/reports/payroll/export/pdf")
async def export_payroll_pdf(
    month: str,
    branch_id: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    """تصدير تقرير الرواتب إلى PDF"""
    return await export_report_to_pdf(
        report_type="payroll",
        start_date=f"{month}-01",
        end_date=f"{month}-31",
        branch_id=branch_id,
        current_user=current_user
    )

@api_router.get("/reports/employee-salary-slip/{employee_id}/export/pdf")
async def export_employee_salary_slip_pdf(
    employee_id: str,
    month: str,
    current_user: dict = Depends(get_current_user)
):
    """تصدير مفردات مرتب موظف إلى PDF"""
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.lib.units import cm
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    
    # Get salary slip data
    slip_data = await get_employee_salary_slip(employee_id, month, current_user)
    employee = slip_data["employee"]
    
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, rightMargin=2*cm, leftMargin=2*cm, topMargin=2*cm, bottomMargin=2*cm)
    
    styles = getSampleStyleSheet()
    elements = []
    
    # Title
    title_style = styles['Heading1']
    title_style.alignment = 1
    elements.append(Paragraph("مفردات المرتب", title_style))
    elements.append(Spacer(1, 20))
    
    # Employee info
    info_data = [
        ["الموظف:", employee.get("name", ""), "الشهر:", month],
        ["الوظيفة:", employee.get("position", ""), "الفرع:", slip_data.get("branch", {}).get("name", "-") if slip_data.get("branch") else "-"]
    ]
    info_table = Table(info_data, colWidths=[doc.width/4]*4)
    info_table.setStyle(TableStyle([
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
    ]))
    elements.append(info_table)
    elements.append(Spacer(1, 20))
    
    # Salary details
    salary_data = [
        ["البند", "المبلغ"],
        ["الراتب الأساسي", f"{slip_data['salary_details']['basic_salary']:,.0f}"],
        ["المكافآت", f"{slip_data['bonuses']['total']:,.0f}"],
        ["الخصومات", f"-{slip_data['deductions']['total']:,.0f}"],
        ["خصم السلف", f"-{slip_data['advances']['deduction_this_month']:,.0f}"],
        ["صافي الراتب", f"{slip_data['summary']['net_salary']:,.0f}"],
    ]
    
    salary_table = Table(salary_data, colWidths=[doc.width/2]*2)
    salary_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4472C4')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTSIZE', (0, 0), (-1, -1), 11),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#E6E6E6')),
        ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
    ]))
    elements.append(salary_table)
    
    doc.build(elements)
    buffer.seek(0)
    
    filename = f"salary_slip_{employee.get('name', 'employee')}_{month}.pdf"
    
    return StreamingResponse(
        buffer,
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )

# ==================== CASH REGISTER ROUTES ====================

# Note: The main cash register close endpoint is defined above at /api/cash-register/close 
# with the complete functionality including shift management

@api_router.get("/cash-register/today")
async def get_today_cash_register(current_user: dict = Depends(get_current_user)):
    """جلب بيانات صندوق اليوم"""
    today = datetime.now(timezone.utc).strftime("%Y-%m-%d")
    branch_id = current_user.get("branch_id")
    
    # مبيعات اليوم النقدية
    sales_query = {
        "created_at": {"$regex": f"^{today}"},
        "payment_method": "cash",
        "status": {"$ne": "cancelled"}
    }
    if branch_id:
        sales_query["branch_id"] = branch_id
    
    cash_orders = await db.orders.find(sales_query, {"_id": 0, "total": 1, "order_number": 1}).to_list(500)
    total_cash_sales = sum(o.get("total", 0) for o in cash_orders)
    
    # مبيعات البطاقة
    card_query = sales_query.copy()
    card_query["payment_method"] = "card"
    card_orders = await db.orders.find(card_query, {"total": 1}).to_list(500)
    total_card_sales = sum(o.get("total", 0) for o in card_orders)
    
    # الآجل
    credit_query = sales_query.copy()
    credit_query["payment_method"] = "credit"
    credit_orders = await db.orders.find(credit_query, {"total": 1}).to_list(500)
    total_credit = sum(o.get("total", 0) for o in credit_orders)
    
    # المصاريف
    expenses_query = {"date": today}
    if branch_id:
        expenses_query["branch_id"] = branch_id
    expenses = await db.expenses.find(expenses_query, {"_id": 0}).to_list(100)
    total_expenses = sum(e.get("amount", 0) for e in expenses)
    
    # آخر إغلاق
    last_close_query = {"branch_id": branch_id} if branch_id else {}
    last_close = await db.cash_register_closes.find_one(
        last_close_query, 
        {"_id": 0},
        sort=[("closed_at", -1)]
    )
    
    return {
        "date": today,
        "cash_sales": total_cash_sales,
        "card_sales": total_card_sales,
        "credit_sales": total_credit,
        "total_sales": total_cash_sales + total_card_sales + total_credit,
        "orders_count": len(cash_orders) + len(card_orders) + len(credit_orders),
        "expenses": expenses,
        "total_expenses": total_expenses,
        "expected_cash": total_cash_sales - total_expenses,
        "last_close": last_close
    }

# ==================== SETTINGS ROUTES ====================

@api_router.get("/settings")
async def get_settings(current_user: dict = Depends(get_current_user)):
    settings = await db.settings.find({}, {"_id": 0}).to_list(100)
    return {s["type"]: s.get("value") or s for s in settings}

@api_router.post("/settings/email-recipients")
async def set_email_recipients(emails: List[str], current_user: dict = Depends(get_current_user)):
    if current_user["role"] != UserRole.ADMIN:
        raise HTTPException(status_code=403, detail="غير مصرح")
    
    await db.settings.update_one(
        {"type": "email_recipients"},
        {"$set": {"type": "email_recipients", "emails": emails}},
        upsert=True
    )
    return {"message": "تم الحفظ"}

@api_router.post("/settings/currencies")
async def set_currencies(currencies: List[Currency], current_user: dict = Depends(get_current_user)):
    if current_user["role"] != UserRole.ADMIN:
        raise HTTPException(status_code=403, detail="غير مصرح")
    
    await db.settings.update_one(
        {"type": "currencies"},
        {"$set": {"type": "currencies", "value": [c.model_dump() for c in currencies]}},
        upsert=True
    )
    return {"message": "تم الحفظ"}

@api_router.post("/settings/general")
async def set_general_settings(settings: Dict[str, Any], current_user: dict = Depends(get_current_user)):
    if current_user["role"] != UserRole.ADMIN:
        raise HTTPException(status_code=403, detail="غير مصرح")
    
    await db.settings.update_one(
        {"type": "general"},
        {"$set": {"type": "general", "value": settings}},
        upsert=True
    )
    return {"message": "تم الحفظ"}

@api_router.get("/settings/general")
async def get_general_settings():
    settings = await db.settings.find_one({"type": "general"}, {"_id": 0})
    return settings.get("value", {}) if settings else {}

@api_router.put("/settings/restaurant")
async def update_restaurant_settings(settings: Dict[str, Any], current_user: dict = Depends(get_current_user)):
    """حفظ إعدادات المطعم (الاسم والشعار)"""
    # السماح للمدير (admin) والمالك (super_admin)
    if current_user["role"] not in [UserRole.ADMIN, UserRole.SUPER_ADMIN]:
        raise HTTPException(status_code=403, detail="غير مصرح")
    
    # جلب tenant_id من المستخدم
    tenant_id = current_user.get("tenant_id")
    
    # تحديث tenant إذا كان موجود
    if tenant_id:
        await db.tenants.update_one(
            {"id": tenant_id},
            {"$set": {
                "name": settings.get("name"),
                "name_ar": settings.get("name_ar"),
                "logo_url": settings.get("logo_url"),
                "updated_at": datetime.now(timezone.utc).isoformat()
            }}
        )
    
    # تحديث default tenant أيضاً إذا لم يكن هناك tenant محدد
    if not tenant_id or tenant_id == "default":
        await db.tenants.update_one(
            {"id": "default"},
            {"$set": {
                "name": settings.get("name"),
                "name_ar": settings.get("name_ar"),
                "logo_url": settings.get("logo_url"),
                "updated_at": datetime.now(timezone.utc).isoformat()
            }},
            upsert=True
        )
    
    # حفظ في settings أيضاً
    await db.settings.update_one(
        {"type": "restaurant"},
        {"$set": {
            "type": "restaurant",
            "name": settings.get("name"),
            "name_ar": settings.get("name_ar"),
            "logo_url": settings.get("logo_url"),
            "updated_at": datetime.now(timezone.utc).isoformat()
        }},
        upsert=True
    )
    
    return {"message": "تم حفظ إعدادات المطعم بنجاح"}

@api_router.get("/settings/restaurant")
async def get_restaurant_settings(current_user: dict = Depends(get_current_user)):
    """جلب إعدادات المطعم"""
    tenant_id = current_user.get("tenant_id", "default")
    
    # محاولة جلب من tenant
    tenant = await db.tenants.find_one({"id": tenant_id}, {"_id": 0})
    if tenant and tenant.get("name"):
        return {
            "name": tenant.get("name"),
            "name_ar": tenant.get("name_ar"),
            "logo_url": tenant.get("logo_url")
        }
    
    # محاولة جلب من settings
    settings = await db.settings.find_one({"type": "restaurant"}, {"_id": 0})
    if settings:
        return {
            "name": settings.get("name", ""),
            "name_ar": settings.get("name_ar", ""),
            "logo_url": settings.get("logo_url", "")
        }
    
    return {"name": "", "name_ar": "", "logo_url": ""}

@api_router.put("/settings/dashboard")
async def set_dashboard_settings(settings: Dict[str, Any], current_user: dict = Depends(get_current_user)):
    """حفظ إعدادات الصفحة الرئيسية - التحكم في الصفحات الظاهرة"""
    if current_user["role"] not in [UserRole.ADMIN, UserRole.MANAGER]:
        raise HTTPException(status_code=403, detail="غير مصرح")
    
    await db.settings.update_one(
        {"type": "dashboard_settings"},
        {"$set": {"type": "dashboard_settings", "value": settings}},
        upsert=True
    )
    return {"message": "تم حفظ إعدادات الصفحة الرئيسية"}

@api_router.get("/settings/dashboard")
async def get_dashboard_settings(current_user: dict = Depends(get_current_user)):
    """جلب إعدادات الصفحة الرئيسية مع مراعاة ميزات العميل"""
    
    # الإعدادات الافتراضية الكاملة
    default_settings = {
        "showPOS": True,
        "showTables": True,
        "showOrders": True,
        "showExpenses": True,
        "showInventory": True,
        "showDelivery": True,
        "showReports": True,
        "showSettings": True,
        "showHR": True,
        "showWarehouse": True,
        "showCallLogs": True,
        "showCallCenter": True,
        "showKitchen": True,
        "showLoyalty": True,
        "showCoupons": True,
        "showRecipes": True,
        "showReservations": True,
        "showReviews": True,
        "showRatings": True,
        "showSmartReports": True,
        "showPurchasing": True,
        "showBranchOrders": True,
        "showCustomerMenu": True,
        # خيارات الإعدادات
        "settingsUsers": True,
        "settingsCustomers": True,
        "settingsBranches": True,
        "settingsCategories": True,
        "settingsProducts": True,
        "settingsPrinters": True,
        "settingsDeliveryCompanies": True,
        "settingsCallCenter": True,
        "settingsNotifications": True
    }
    
    # جلب إعدادات لوحة القيادة المحفوظة
    settings = await db.settings.find_one({"type": "dashboard_settings"}, {"_id": 0})
    if settings and settings.get("value"):
        default_settings = {**default_settings, **settings.get("value", {})}
    
    # إذا كان المستخدم Super Admin، أرجع كل الميزات
    if current_user.get("role") == UserRole.SUPER_ADMIN:
        return default_settings
    
    # إذا كان المستخدم بدون tenant_id (النظام الرئيسي)، أرجع كل الميزات
    tenant_id = get_user_tenant_id(current_user)
    if not tenant_id:
        return default_settings
    
    # جلب ميزات العميل
    tenant = await db.tenants.find_one({"id": tenant_id}, {"_id": 0, "enabled_features": 1})
    if tenant and tenant.get("enabled_features"):
        tenant_features = tenant["enabled_features"]
        # دمج الميزات - العميل يرى فقط الميزات المفعّلة له
        for key in default_settings:
            if key in tenant_features:
                default_settings[key] = tenant_features[key] and default_settings[key]
    
    return default_settings

@api_router.get("/tenant/info")
async def get_tenant_info(current_user: dict = Depends(get_current_user)):
    """جلب معلومات العميل (الشعار والاسم) للعرض في Dashboard"""
    tenant_id = get_user_tenant_id(current_user)
    
    # إذا كان النظام الرئيسي (بدون tenant)
    if not tenant_id:
        # جلب إعدادات النظام العامة
        settings = await db.settings.find_one({"type": "system_branding"}, {"_id": 0})
        if settings and settings.get("value"):
            return settings["value"]
        return {
            "name": "Maestro",
            "name_ar": "Maestro",
            "name_en": "Maestro",
            "logo_url": None
        }
    
    # جلب معلومات العميل
    tenant = await db.tenants.find_one(
        {"id": tenant_id}, 
        {"_id": 0, "name": 1, "name_ar": 1, "name_en": 1, "logo_url": 1}
    )
    
    if not tenant:
        return {"name": "Maestro", "logo_url": None}
    
    return tenant

# ==================== LOGIN BACKGROUNDS API ====================

# كلمة سر خاصة للـ Super Admin - من متغيرات البيئة
SUPER_ADMIN_SECRET = os.environ.get("SUPER_ADMIN_SECRET", "271018")

# التحقق من صلاحية Super Admin
async def verify_super_admin(current_user: dict = Depends(get_current_user)):
    if current_user.get("role") != UserRole.SUPER_ADMIN:
        raise HTTPException(status_code=403, detail="صلاحيات Super Admin مطلوبة")
    return current_user

# ==================== FIX DATA API ====================

@api_router.post("/fix-data")
async def fix_data_endpoint(current_user: dict = Depends(verify_super_admin)):
    """
    Endpoint لإصلاح البيانات القديمة
    يمكن استدعاؤه فقط من قبل المالك
    """
    try:
        results = {
            "tables_fixed": 0,
            "tables_deleted_duplicates": 0,
            "categories_fixed": 0,
            "products_fixed": 0,
            "tenant_tables_created": 0,
        }
        
        # 1. تحديث الطاولات القديمة التي ليس لها tenant_id لتصبح "default"
        tables_result = await db.tables.update_many(
            {"$or": [{"tenant_id": {"$exists": False}}, {"tenant_id": None}, {"tenant_id": ""}]},
            {"$set": {"tenant_id": "default"}}
        )
        results["tables_fixed"] = tables_result.modified_count
        
        # 2. حذف الطاولات المكررة (نفس رقم الطاولة ونفس tenant_id)
        pipeline = [
            {"$group": {
                "_id": {"number": "$number", "tenant_id": "$tenant_id"},
                "count": {"$sum": 1},
                "ids": {"$push": "$id"}
            }},
            {"$match": {"count": {"$gt": 1}}}
        ]
        duplicates = await db.tables.aggregate(pipeline).to_list(100)
        for dup in duplicates:
            # الاحتفاظ بأول طاولة وحذف الباقي
            ids_to_delete = dup["ids"][1:]
            delete_result = await db.tables.delete_many({"id": {"$in": ids_to_delete}})
            results["tables_deleted_duplicates"] += delete_result.deleted_count
        
        # 3. إنشاء طاولات لكل عميل ليس لديه طاولات
        tenants = await db.tenants.find({}, {"_id": 0, "id": 1, "name": 1}).to_list(None)
        for tenant in tenants:
            tenant_tables = await db.tables.count_documents({"tenant_id": tenant["id"]})
            if tenant_tables == 0:
                tenant_branch = await db.branches.find_one({"tenant_id": tenant["id"]})
                branch_id = tenant_branch["id"] if tenant_branch else None
                
                default_tables = []
                for i in range(1, 6):
                    default_tables.append({
                        "id": str(uuid.uuid4()),
                        "number": i,
                        "capacity": 4,
                        "section": "القاعة الرئيسية",
                        "status": "available",
                        "current_order_id": None,
                        "branch_id": branch_id,
                        "tenant_id": tenant["id"]
                    })
                await db.tables.insert_many(default_tables)
                results["tenant_tables_created"] += 5
        
        # 4. تحديث صور الفئات للنظام الرئيسي
        category_images = {
            "برغر": "https://images.unsplash.com/photo-1635275650933-7b0911815a2e?w=400",
            "بيتزا": "https://images.unsplash.com/photo-1703073186021-021fb5a0bde1?w=400",
            "مشروبات": "https://images.unsplash.com/photo-1657958977261-d75e81b4713f?w=400",
            "حلويات": "https://images.unsplash.com/photo-1546902189-eaaf09f8e38f?w=400",
            "سلطات": "https://images.unsplash.com/photo-1677653805080-59c57727c84e?w=400",
        }
        for cat_name, cat_image in category_images.items():
            cat_result = await db.categories.update_many(
                {"name": cat_name, "tenant_id": "default", "$or": [{"image": {"$exists": False}}, {"image": None}, {"image": ""}]},
                {"$set": {"image": cat_image}}
            )
            results["categories_fixed"] += cat_result.modified_count
        
        # 5. تحديث صور المنتجات للنظام الرئيسي
        product_images = {
            "برغر كلاسيك": "https://images.unsplash.com/photo-1656439659132-24c68e36b553?w=400",
            "برغر دبل": "https://images.unsplash.com/photo-1635275650933-7b0911815a2e?w=400",
            "بيتزا مارغريتا": "https://images.unsplash.com/photo-1681567604770-0dc826c870ae?w=400",
            "بيتزا خضار": "https://images.unsplash.com/photo-1602104980741-b87a33837f9f?w=400",
            "كولا": "https://images.unsplash.com/photo-1657958977261-d75e81b4713f?w=400",
            "عصير برتقال": "https://images.unsplash.com/photo-1716925539259-ce0115263d37?w=400",
        }
        for prod_name, prod_image in product_images.items():
            prod_result = await db.products.update_many(
                {"name": prod_name, "tenant_id": "default", "$or": [{"image": {"$exists": False}}, {"image": None}, {"image": ""}]},
                {"$set": {"image": prod_image}}
            )
            results["products_fixed"] += prod_result.modified_count
        
        return {
            "status": "success",
            "message": "تم إصلاح البيانات بنجاح",
            "results": results
        }
        
    except Exception as e:
        return {
            "status": "error",
            "message": f"حدث خطأ: {str(e)}"
        }

# ==================== SYSTEM BRANDING API ====================

@api_router.get("/system/branding")
async def get_system_branding(current_user: dict = Depends(verify_super_admin)):
    """جلب إعدادات هوية النظام الرئيسي"""
    settings = await db.settings.find_one({"type": "system_branding"}, {"_id": 0})
    if settings and settings.get("value"):
        return settings["value"]
    return {
        "name": "Maestro",
        "name_ar": "Maestro",
        "name_en": "Maestro",
        "logo_url": None
    }

@api_router.put("/system/branding")
async def update_system_branding(branding: dict, current_user: dict = Depends(verify_super_admin)):
    """تحديث إعدادات هوية النظام الرئيسي (الاسم والشعار)"""
    allowed_fields = ["name", "name_ar", "name_en", "logo_url"]
    update_data = {k: v for k, v in branding.items() if k in allowed_fields}
    
    await db.settings.update_one(
        {"type": "system_branding"},
        {"$set": {"type": "system_branding", "value": update_data}},
        upsert=True
    )
    
    return {"message": "تم تحديث هوية النظام بنجاح", "branding": update_data}

class LoginBackgroundCreate(BaseModel):
    image_url: str
    title: Optional[str] = None
    animation_type: str = "fade"  # fade, slide, zoom, kenburns, parallax
    animation_duration: int = 8  # بالثواني
    overlay_opacity: float = 0.5
    is_active: bool = True
    sort_order: int = 0

class LoginBackgroundSettings(BaseModel):
    backgrounds: List[Dict[str, Any]] = []
    animation_enabled: bool = True
    transition_type: str = "fade"  # fade, slide, crossfade
    transition_duration: float = 1.5  # بالثواني
    auto_play: bool = True
    show_logo: bool = True
    logo_url: Optional[str] = None
    logo_animation: str = "pulse"  # pulse, bounce, glow, none
    overlay_color: str = "rgba(0,0,0,0.5)"
    text_color: str = "#ffffff"

@api_router.get("/login-backgrounds")
async def get_login_backgrounds():
    """جلب إعدادات خلفيات صفحة الدخول (عام - بدون مصادقة)"""
    settings = await db.settings.find_one({"type": "login_backgrounds"}, {"_id": 0})
    
    default_settings = {
        "backgrounds": [],
        "animation_enabled": True,
        "transition_type": "fade",
        "transition_duration": 1.5,
        "auto_play": True,
        "show_logo": True,
        "logo_url": None,
        "logo_animation": "pulse",
        "overlay_color": "rgba(0,0,0,0.5)",
        "text_color": "#ffffff"
    }
    
    if settings and settings.get("value"):
        return {**default_settings, **settings.get("value", {})}
    return default_settings

@api_router.put("/login-backgrounds")
async def update_login_backgrounds(settings: LoginBackgroundSettings, current_user: dict = Depends(verify_super_admin)):
    """تحديث إعدادات خلفيات صفحة الدخول (Super Admin فقط)"""
    
    await db.settings.update_one(
        {"type": "login_backgrounds"},
        {"$set": {"type": "login_backgrounds", "value": settings.model_dump()}},
        upsert=True
    )
    return {"message": "تم حفظ إعدادات الخلفيات"}

@api_router.post("/login-backgrounds/upload")
async def upload_login_background(
    file_url: str,
    title: Optional[str] = None,
    animation_type: str = "fade",
    current_user: dict = Depends(verify_super_admin)
):
    """إضافة خلفية جديدة"""
    
    # جلب الإعدادات الحالية
    settings = await db.settings.find_one({"type": "login_backgrounds"}, {"_id": 0})
    current_backgrounds = []
    current_value = {}
    
    if settings and settings.get("value"):
        current_value = settings["value"]
        current_backgrounds = current_value.get("backgrounds", [])
    
    # إضافة الخلفية الجديدة
    new_background = {
        "id": str(uuid.uuid4()),
        "image_url": file_url,
        "title": title,
        "animation_type": animation_type,
        "animation_duration": 8,
        "overlay_opacity": 0.5,
        "is_active": True,
        "sort_order": len(current_backgrounds),
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    
    current_backgrounds.append(new_background)
    
    # تحديث مع الحفاظ على الإعدادات الأخرى
    default_settings = {
        "backgrounds": current_backgrounds,
        "animation_enabled": current_value.get("animation_enabled", True),
        "transition_type": current_value.get("transition_type", "fade"),
        "transition_duration": current_value.get("transition_duration", 1.5),
        "auto_play": current_value.get("auto_play", True),
        "show_logo": current_value.get("show_logo", True),
        "logo_url": current_value.get("logo_url", None),
        "logo_animation": current_value.get("logo_animation", "pulse"),
        "overlay_color": current_value.get("overlay_color", "rgba(0,0,0,0.5)"),
        "text_color": current_value.get("text_color", "#ffffff")
    }
    
    # حفظ التحديث
    await db.settings.update_one(
        {"type": "login_backgrounds"},
        {"$set": {"type": "login_backgrounds", "value": default_settings}},
        upsert=True
    )
    
    return {"message": "تم إضافة الخلفية", "background": new_background}

@api_router.delete("/login-backgrounds/{background_id}")
async def delete_login_background(background_id: str, current_user: dict = Depends(verify_super_admin)):
    """حذف خلفية"""
    
    settings = await db.settings.find_one({"type": "login_backgrounds"}, {"_id": 0})
    if not settings or not settings.get("value"):
        raise HTTPException(status_code=404, detail="لا توجد خلفيات")
    
    backgrounds = settings["value"].get("backgrounds", [])
    backgrounds = [b for b in backgrounds if b.get("id") != background_id]
    
    await db.settings.update_one(
        {"type": "login_backgrounds"},
        {"$set": {"value.backgrounds": backgrounds}}
    )
    
    return {"message": "تم حذف الخلفية"}

@api_router.post("/login-backgrounds/upload-logo")
async def upload_login_page_logo(
    file: UploadFile = File(...),
    current_user: dict = Depends(verify_super_admin)
):
    """رفع شعار صفحة تسجيل الدخول - للمالك فقط"""
    
    # التحقق من نوع الملف
    allowed_types = ['image/jpeg', 'image/png', 'image/gif', 'image/webp', 'image/svg+xml']
    if file.content_type not in allowed_types:
        raise HTTPException(status_code=400, detail="نوع الملف غير مدعوم. يرجى استخدام JPG, PNG, GIF, WebP أو SVG")
    
    # معالجة وحفظ الصورة
    filename = await process_and_save_image(file, LOGOS_DIR, max_size=(512, 512), quality=90)
    
    # إنشاء URL نسبي للشعار
    logo_url = f"/api/uploads/logos/{filename}"
    
    # تحديث login_backgrounds مع الشعار الجديد
    settings = await db.settings.find_one({"type": "login_backgrounds"}, {"_id": 0})
    current_value = settings.get("value", {}) if settings else {}
    
    # تحديث logo_url فقط
    await db.settings.update_one(
        {"type": "login_backgrounds"},
        {"$set": {"value.logo_url": logo_url}},
        upsert=True
    )
    
    return {"message": "تم رفع شعار صفحة تسجيل الدخول بنجاح", "logo_url": logo_url}

@api_router.put("/login-backgrounds/logo-url")
async def update_login_page_logo_url(
    logo_url: str = Body(..., embed=True),
    current_user: dict = Depends(verify_super_admin)
):
    """تحديث شعار صفحة تسجيل الدخول برابط خارجي - للمالك فقط"""
    
    # تحديث login_backgrounds مع الشعار الجديد
    await db.settings.update_one(
        {"type": "login_backgrounds"},
        {"$set": {"value.logo_url": logo_url}},
        upsert=True
    )
    
    return {"message": "تم تحديث شعار صفحة تسجيل الدخول", "logo_url": logo_url}

@api_router.delete("/login-backgrounds/logo")
async def delete_login_page_logo(current_user: dict = Depends(verify_super_admin)):
    """حذف شعار صفحة تسجيل الدخول - للمالك فقط"""
    
    # تحديث login_backgrounds بإزالة الشعار
    await db.settings.update_one(
        {"type": "login_backgrounds"},
        {"$set": {"value.logo_url": None}},
        upsert=True
    )
    
    return {"message": "تم حذف شعار صفحة تسجيل الدخول"}

# ==================== INVOICE/RECEIPT SETTINGS - إعدادات الفاتورة ====================

class SystemInvoiceSettings(BaseModel):
    """إعدادات الفاتورة للنظام (يتحكم فيها المالك)"""
    system_name: Optional[str] = None  # اسم النظام
    system_logo_url: Optional[str] = None  # شعار النظام
    thank_you_message: str = "شكراً لزيارتكم"  # رسالة الشكر
    system_phone: Optional[str] = None  # رقم هاتف النظام
    system_phone2: Optional[str] = None  # رقم هاتف ثاني
    system_email: Optional[str] = None  # بريد النظام
    system_website: Optional[str] = None  # موقع النظام
    footer_text: Optional[str] = None  # نص إضافي في التذييل
    show_system_branding: bool = True  # عرض شعار وبيانات النظام
    promo_text: Optional[str] = "نظام إدارة متكامل للمطاعم والكافيهات"  # نص الدعاية
    cta_text: Optional[str] = "للحصول على نسختك تواصل معنا"  # نص التحفيز للتواصل

class TenantInvoiceSettings(BaseModel):
    """إعدادات الفاتورة للعميل (المطعم)"""
    show_logo: bool = True  # عرض الشعار
    invoice_logo: Optional[str] = None  # شعار الفاتورة المخصص
    phone: Optional[str] = None  # رقم الهاتف
    phone2: Optional[str] = None  # رقم هاتف ثاني
    address: Optional[str] = None  # العنوان
    tax_number: Optional[str] = None  # الرقم الضريبي
    show_tax: bool = True  # إظهار الرقم الضريبي
    custom_header: Optional[str] = None  # نص إضافي في الترويسة
    custom_footer: Optional[str] = None  # نص إضافي في التذييل

@api_router.get("/system/invoice-settings")
async def get_system_invoice_settings():
    """جلب إعدادات الفاتورة للنظام (عام - للطباعة)"""
    settings = await db.settings.find_one({"type": "system_invoice_settings"}, {"_id": 0})
    
    default_settings = {
        "system_name": None,
        "system_logo_url": None,
        "thank_you_message": "شكراً لزيارتكم",
        "system_phone": None,
        "system_phone2": None,
        "system_email": None,
        "system_website": None,
        "footer_text": None,
        "show_system_branding": True
    }
    
    if settings and settings.get("value"):
        return {**default_settings, **settings.get("value", {})}
    return default_settings

@api_router.put("/system/invoice-settings")
async def update_system_invoice_settings(settings: SystemInvoiceSettings, current_user: dict = Depends(verify_super_admin)):
    """تحديث إعدادات الفاتورة للنظام (المالك فقط)"""
    
    await db.settings.update_one(
        {"type": "system_invoice_settings"},
        {"$set": {"value": settings.model_dump(), "updated_at": datetime.now(timezone.utc).isoformat()}},
        upsert=True
    )
    
    return {"message": "تم تحديث إعدادات الفاتورة", "settings": settings.model_dump()}

@api_router.get("/tenant/invoice-settings")
async def get_tenant_invoice_settings(current_user: dict = Depends(get_current_user)):
    """جلب إعدادات الفاتورة للعميل"""
    tenant_id = get_user_tenant_id(current_user)
    
    default_settings = {
        "show_logo": True,
        "phone": None,
        "phone2": None,
        "address": None,
        "tax_number": None,
        "custom_header": None,
        "custom_footer": None
    }
    
    if tenant_id:
        settings = await db.tenant_invoice_settings.find_one({"tenant_id": tenant_id}, {"_id": 0})
        if settings:
            return {**default_settings, **settings}
    
    return default_settings

@api_router.put("/tenant/invoice-settings")
async def update_tenant_invoice_settings(settings: TenantInvoiceSettings, current_user: dict = Depends(get_current_user)):
    """تحديث إعدادات الفاتورة للعميل"""
    if current_user["role"] not in [UserRole.ADMIN, UserRole.MANAGER]:
        raise HTTPException(status_code=403, detail="غير مصرح")
    
    tenant_id = get_user_tenant_id(current_user)
    
    await db.tenant_invoice_settings.update_one(
        {"tenant_id": tenant_id},
        {"$set": {**settings.model_dump(), "tenant_id": tenant_id, "updated_at": datetime.now(timezone.utc).isoformat()}},
        upsert=True
    )
    
    return {"message": "تم تحديث إعدادات الفاتورة", "settings": settings.model_dump()}

# ==================== Currency & Language Settings ====================

# العملات المدعومة مع أسعار الصرف التقريبية
SUPPORTED_CURRENCIES = {
    "IQD": {"name": "دينار عراقي", "name_en": "Iraqi Dinar", "symbol": "د.ع", "rate_to_usd": 0.00076, "decimal_places": 0},
    "USD": {"name": "دولار أمريكي", "name_en": "US Dollar", "symbol": "$", "rate_to_usd": 1, "decimal_places": 2},
    "SAR": {"name": "ريال سعودي", "name_en": "Saudi Riyal", "symbol": "ر.س", "rate_to_usd": 0.27, "decimal_places": 2},
    "AED": {"name": "درهم إماراتي", "name_en": "UAE Dirham", "symbol": "د.إ", "rate_to_usd": 0.27, "decimal_places": 2},
    "KWD": {"name": "دينار كويتي", "name_en": "Kuwaiti Dinar", "symbol": "د.ك", "rate_to_usd": 3.25, "decimal_places": 3},
    "EGP": {"name": "جنيه مصري", "name_en": "Egyptian Pound", "symbol": "ج.م", "rate_to_usd": 0.032, "decimal_places": 2},
    "JOD": {"name": "دينار أردني", "name_en": "Jordanian Dinar", "symbol": "د.أ", "rate_to_usd": 1.41, "decimal_places": 3},
    "EUR": {"name": "يورو", "name_en": "Euro", "symbol": "€", "rate_to_usd": 1.08, "decimal_places": 2},
    "GBP": {"name": "جنيه استرليني", "name_en": "British Pound", "symbol": "£", "rate_to_usd": 1.27, "decimal_places": 2},
    "TRY": {"name": "ليرة تركية", "name_en": "Turkish Lira", "symbol": "₺", "rate_to_usd": 0.031, "decimal_places": 2},
}

# اللغات المدعومة
SUPPORTED_LANGUAGES = {
    "ar": {"name": "العربية", "name_en": "Arabic", "dir": "rtl"},
    "en": {"name": "English", "name_en": "English", "dir": "ltr"},
    "ku": {"name": "کوردی", "name_en": "Kurdish", "dir": "rtl"},
    "fa": {"name": "فارسی", "name_en": "Persian", "dir": "rtl"},
    "tr": {"name": "Türkçe", "name_en": "Turkish", "dir": "ltr"},
}

# البلدان مع العملات الافتراضية
COUNTRIES = {
    "IQ": {"name": "العراق", "name_en": "Iraq", "currency": "IQD", "language": "ar"},
    "SA": {"name": "السعودية", "name_en": "Saudi Arabia", "currency": "SAR", "language": "ar"},
    "AE": {"name": "الإمارات", "name_en": "UAE", "currency": "AED", "language": "ar"},
    "KW": {"name": "الكويت", "name_en": "Kuwait", "currency": "KWD", "language": "ar"},
    "EG": {"name": "مصر", "name_en": "Egypt", "currency": "EGP", "language": "ar"},
    "JO": {"name": "الأردن", "name_en": "Jordan", "currency": "JOD", "language": "ar"},
    "US": {"name": "أمريكا", "name_en": "United States", "currency": "USD", "language": "en"},
    "GB": {"name": "بريطانيا", "name_en": "United Kingdom", "currency": "GBP", "language": "en"},
    "TR": {"name": "تركيا", "name_en": "Turkey", "currency": "TRY", "language": "tr"},
}

class TenantRegionalSettings(BaseModel):
    """إعدادات المنطقة والعملة للعميل"""
    country: str = "IQ"
    currency: str = "IQD"
    language: str = "ar"
    secondary_currency: Optional[str] = "USD"  # عملة ثانوية للعرض
    show_secondary_currency: bool = False  # عرض السعر بالعملة الثانوية
    custom_exchange_rate: Optional[float] = None  # سعر صرف مخصص

@api_router.get("/system/currencies")
async def get_supported_currencies():
    """جلب قائمة العملات المدعومة"""
    return {"currencies": SUPPORTED_CURRENCIES}

@api_router.get("/system/languages")
async def get_supported_languages():
    """جلب قائمة اللغات المدعومة"""
    return {"languages": SUPPORTED_LANGUAGES}

@api_router.get("/system/countries")
async def get_supported_countries():
    """جلب قائمة البلدان المدعومة"""
    return {"countries": COUNTRIES}

@api_router.get("/tenant/regional-settings")
async def get_tenant_regional_settings(current_user: dict = Depends(get_current_user)):
    """جلب إعدادات المنطقة والعملة للعميل"""
    tenant_id = get_user_tenant_id(current_user)
    
    settings = await db.tenant_regional_settings.find_one({"tenant_id": tenant_id}, {"_id": 0})
    
    default_settings = {
        "country": "IQ",
        "currency": "IQD",
        "language": "ar",
        "secondary_currency": "USD",
        "show_secondary_currency": False,
        "custom_exchange_rate": None
    }
    
    if settings:
        return {**default_settings, **settings}
    return default_settings

@api_router.put("/tenant/regional-settings")
async def update_tenant_regional_settings(settings: TenantRegionalSettings, current_user: dict = Depends(get_current_user)):
    """تحديث إعدادات المنطقة والعملة للعميل"""
    if current_user["role"] not in [UserRole.ADMIN, UserRole.MANAGER]:
        raise HTTPException(status_code=403, detail="غير مصرح")
    
    tenant_id = get_user_tenant_id(current_user)
    
    await db.tenant_regional_settings.update_one(
        {"tenant_id": tenant_id},
        {"$set": {**settings.model_dump(), "tenant_id": tenant_id, "updated_at": datetime.now(timezone.utc).isoformat()}},
        upsert=True
    )
    
    return {"message": "تم تحديث إعدادات المنطقة", "settings": settings.model_dump()}

@api_router.get("/customer/regional-settings/{tenant_id}")
async def get_customer_regional_settings(tenant_id: str):
    """جلب إعدادات المنطقة للزبون (بدون تسجيل دخول)"""
    # البحث بـ menu_slug أو tenant_id
    tenant = await db.tenants.find_one({"menu_slug": tenant_id})
    if not tenant:
        tenant = await db.tenants.find_one({"id": tenant_id})
    
    if not tenant:
        raise HTTPException(status_code=404, detail="المطعم غير موجود")
    
    actual_tenant_id = tenant.get("id")
    settings = await db.tenant_regional_settings.find_one({"tenant_id": actual_tenant_id}, {"_id": 0})
    
    default_settings = {
        "country": "IQ",
        "currency": "IQD",
        "language": "ar",
        "secondary_currency": "USD",
        "show_secondary_currency": False
    }
    
    result = {**default_settings}
    if settings:
        result.update(settings)
    
    # إضافة معلومات العملة
    currency_code = result.get("currency", "IQD")
    if currency_code in SUPPORTED_CURRENCIES:
        result["currency_info"] = SUPPORTED_CURRENCIES[currency_code]
    
    return result

@api_router.post("/convert-currency")
async def convert_currency(
    amount: float,
    from_currency: str = "IQD",
    to_currency: str = "USD"
):
    """تحويل المبلغ بين العملات"""
    if from_currency not in SUPPORTED_CURRENCIES or to_currency not in SUPPORTED_CURRENCIES:
        raise HTTPException(status_code=400, detail="عملة غير مدعومة")
    
    # التحويل عبر الدولار كوسيط
    from_rate = SUPPORTED_CURRENCIES[from_currency]["rate_to_usd"]
    to_rate = SUPPORTED_CURRENCIES[to_currency]["rate_to_usd"]
    
    # المبلغ بالدولار
    usd_amount = amount * from_rate
    # المبلغ بالعملة المستهدفة
    converted_amount = usd_amount / to_rate
    
    decimal_places = SUPPORTED_CURRENCIES[to_currency]["decimal_places"]
    converted_amount = round(converted_amount, decimal_places)
    
    return {
        "original_amount": amount,
        "original_currency": from_currency,
        "converted_amount": converted_amount,
        "target_currency": to_currency,
        "exchange_rate": from_rate / to_rate
    }

# ==================== Owner Currency Settings ====================

class OwnerCurrencySettings(BaseModel):
    """إعدادات تحويل العملات للمالك"""
    preferred_currency: str = "USD"
    use_live_rates: bool = False
    custom_rates: Optional[dict] = None  # {"IQD_USD": 0.00076, "SAR_USD": 0.27}

@api_router.get("/super-admin/currency-settings")
async def get_owner_currency_settings(current_user: dict = Depends(verify_super_admin)):
    """جلب إعدادات العملة للمالك"""
    settings = await db.settings.find_one({"type": "owner_currency_settings"}, {"_id": 0})
    
    if not settings:
        return {
            "preferred_currency": "USD",
            "use_live_rates": False,
            "custom_rates": {},
            "supported_currencies": list(SUPPORTED_CURRENCIES.keys())
        }
    
    return {
        **settings,
        "supported_currencies": list(SUPPORTED_CURRENCIES.keys())
    }

@api_router.put("/super-admin/currency-settings")
async def update_owner_currency_settings(
    settings: OwnerCurrencySettings,
    current_user: dict = Depends(verify_super_admin)
):
    """تحديث إعدادات العملة للمالك"""
    await db.settings.update_one(
        {"type": "owner_currency_settings"},
        {"$set": {
            "type": "owner_currency_settings",
            "preferred_currency": settings.preferred_currency,
            "use_live_rates": settings.use_live_rates,
            "custom_rates": settings.custom_rates or {},
            "updated_at": datetime.utcnow()
        }},
        upsert=True
    )
    return {"message": "تم حفظ إعدادات العملة"}

@api_router.put("/super-admin/custom-exchange-rate")
async def update_custom_exchange_rate(
    from_currency: str,
    to_currency: str,
    rate: float,
    current_user: dict = Depends(verify_super_admin)
):
    """تحديث سعر صرف مخصص"""
    rate_key = f"{from_currency}_{to_currency}"
    
    await db.settings.update_one(
        {"type": "owner_currency_settings"},
        {"$set": {f"custom_rates.{rate_key}": rate, "updated_at": datetime.utcnow()}},
        upsert=True
    )
    return {"message": f"تم تحديث سعر صرف {from_currency} إلى {to_currency}"}

@api_router.get("/super-admin/live-exchange-rates")
async def get_live_exchange_rates(current_user: dict = Depends(verify_super_admin)):
    """جلب أسعار الصرف الحية (من الإنترنت)"""
    import httpx
    
    try:
        # استخدام API مجاني لأسعار الصرف
        async with httpx.AsyncClient() as client:
            response = await client.get(
                "https://api.exchangerate-api.com/v4/latest/USD",
                timeout=10.0
            )
            
            if response.status_code == 200:
                data = response.json()
                rates = data.get("rates", {})
                
                # تحويل الأسعار للعملات المدعومة
                live_rates = {}
                for code in SUPPORTED_CURRENCIES.keys():
                    if code in rates:
                        live_rates[code] = {
                            "rate_to_usd": 1 / rates[code] if rates[code] > 0 else 0,
                            "rate_from_usd": rates[code]
                        }
                
                return {
                    "success": True,
                    "base": "USD",
                    "rates": live_rates,
                    "fetched_at": datetime.utcnow().isoformat()
                }
    except Exception as e:
        logger.error(f"Error fetching live rates: {str(e)}")
    
    # إرجاع الأسعار الثابتة في حالة الفشل
    return {
        "success": False,
        "message": "تعذر جلب الأسعار الحية، يتم استخدام الأسعار الثابتة",
        "rates": {code: {"rate_to_usd": info["rate_to_usd"], "rate_from_usd": 1/info["rate_to_usd"] if info["rate_to_usd"] > 0 else 0} 
                 for code, info in SUPPORTED_CURRENCIES.items()}
    }

# ==================== Super Admin Currency Dashboard ====================

@api_router.get("/super-admin/sales-summary")
async def get_super_admin_sales_summary(
    display_currency: str = "USD",
    current_user: dict = Depends(verify_super_admin)
):
    """جلب ملخص المبيعات لجميع العملاء مع تحويل العملات"""
    
    # جلب جميع العملاء النشطين (ليس تجريبي)
    tenants = await db.tenants.find(
        {"is_demo": {"$ne": True}, "subscription_type": {"$ne": "demo"}}, 
        {"_id": 0, "id": 1, "name": 1}
    ).to_list(100)
    
    total_sales_usd = 0
    total_orders = 0
    tenant_sales = []
    active_tenants = 0
    
    for tenant in tenants:
        tenant_id = tenant.get("id")
        
        # جلب إعدادات العملة للعميل
        regional = await db.tenant_regional_settings.find_one({"tenant_id": tenant_id}, {"_id": 0})
        tenant_currency = regional.get("currency", "IQD") if regional else "IQD"
        
        # جلب إجمالي المبيعات
        pipeline = [
            {"$match": {"tenant_id": tenant_id, "status": {"$in": ["completed", "delivered"]}}},
            {"$group": {"_id": None, "total": {"$sum": "$total"}, "count": {"$sum": 1}}}
        ]
        result = await db.orders.aggregate(pipeline).to_list(1)
        
        if result:
            sales_in_tenant_currency = result[0].get("total", 0)
            orders_count = result[0].get("count", 0)
            
            if sales_in_tenant_currency > 0:
                active_tenants += 1
            
            # تحويل للدولار
            if tenant_currency in SUPPORTED_CURRENCIES:
                rate = SUPPORTED_CURRENCIES[tenant_currency]["rate_to_usd"]
                sales_in_usd = sales_in_tenant_currency * rate
            else:
                sales_in_usd = sales_in_tenant_currency
            
            total_sales_usd += sales_in_usd
            total_orders += orders_count
            
            # حساب المبيعات بالعملة المطلوبة
            if display_currency in SUPPORTED_CURRENCIES:
                display_rate = SUPPORTED_CURRENCIES[display_currency]["rate_to_usd"]
                converted_sales = sales_in_usd / display_rate if display_rate > 0 else 0
            else:
                converted_sales = sales_in_usd
            
            tenant_sales.append({
                "name": tenant.get("name"),
                "original_sales": sales_in_tenant_currency,
                "original_currency": tenant_currency,
                "converted_sales": round(converted_sales, 2),
                "orders_count": orders_count
            })
    
    # ترتيب حسب المبيعات المحولة
    tenant_sales.sort(key=lambda x: x["converted_sales"], reverse=True)
    
    # تحويل الإجمالي للعملة المطلوبة
    if display_currency in SUPPORTED_CURRENCIES:
        rate = SUPPORTED_CURRENCIES[display_currency]["rate_to_usd"]
        decimal_places = SUPPORTED_CURRENCIES[display_currency].get("decimal_places", 2)
        total_in_display = total_sales_usd / rate if rate > 0 else 0
        total_in_display = round(total_in_display, decimal_places)
    else:
        total_in_display = total_sales_usd
    
    return {
        "total_sales_converted": total_in_display,
        "total_sales_usd": round(total_sales_usd, 2),
        "total_orders": total_orders,
        "active_tenants": active_tenants,
        "display_currency": display_currency,
        "display_currency_symbol": SUPPORTED_CURRENCIES.get(display_currency, {}).get("symbol", "$"),
        "tenant_sales": tenant_sales
    }

# ==================== Login Page Settings ====================

@api_router.get("/system/login-page-settings")
async def get_login_page_settings():
    """جلب إعدادات صفحة الدخول"""
    settings = await db.settings.find_one({"type": "login_page_settings"}, {"_id": 0})
    
    default_settings = {
        "enable_animation": True,
        "transition_type": "fade",
        "transition_duration": 1.5,
        "auto_change": True,
        "logo_animation": "pulse",
        "backgrounds": [],
        "login_logo_enabled": True,
        "login_logo_url": "",
        "accent_color": "rgba(147, 51, 234, 0.5)"
    }
    
    if settings and settings.get("value"):
        return {**default_settings, **settings.get("value")}
    
    return default_settings

@api_router.put("/system/login-page-settings")
async def update_login_page_settings(settings: dict, current_user: dict = Depends(verify_super_admin)):
    """تحديث إعدادات صفحة الدخول (المالك فقط)"""
    
    await db.settings.update_one(
        {"type": "login_page_settings"},
        {"$set": {"value": settings, "updated_at": datetime.now(timezone.utc).isoformat()}},
        upsert=True
    )
    
    return {"message": "تم تحديث إعدادات صفحة الدخول", "settings": settings}

@api_router.get("/invoice-data/{order_id}")
async def get_invoice_data(order_id: str, current_user: dict = Depends(get_current_user)):
    """جلب بيانات الفاتورة الكاملة للطباعة"""
    
    # جلب الطلب
    query = build_tenant_query(current_user, {"id": order_id})
    order = await db.orders.find_one(query, {"_id": 0})
    if not order:
        raise HTTPException(status_code=404, detail="الطلب غير موجود")
    
    tenant_id = get_user_tenant_id(current_user)
    
    # جلب بيانات العميل (المطعم)
    tenant = await db.tenants.find_one({"id": tenant_id}, {"_id": 0}) if tenant_id else None
    
    # جلب إعدادات فاتورة العميل
    tenant_invoice = await db.tenant_invoice_settings.find_one({"tenant_id": tenant_id}, {"_id": 0}) if tenant_id else {}
    
    # جلب إعدادات النظام
    system_settings = await db.settings.find_one({"type": "system_invoice_settings"}, {"_id": 0})
    system_invoice = system_settings.get("value", {}) if system_settings else {}
    
    # جلب الفرع
    branch = await db.branches.find_one({"id": order.get("branch_id")}, {"_id": 0, "name": 1, "address": 1, "phone": 1})
    
    return {
        "order": order,
        "tenant": {
            "name": tenant.get("name") if tenant else "المطعم",
            "logo_url": tenant.get("logo_url") if tenant else None,
            "phone": tenant_invoice.get("phone") or (tenant.get("owner_phone") if tenant else None),
            "phone2": tenant_invoice.get("phone2"),
            "address": tenant_invoice.get("address") or (branch.get("address") if branch else None),
            "tax_number": tenant_invoice.get("tax_number"),
            "custom_header": tenant_invoice.get("custom_header"),
            "custom_footer": tenant_invoice.get("custom_footer")
        },
        "system": {
            "logo_url": system_invoice.get("system_logo_url"),
            "thank_you_message": system_invoice.get("thank_you_message", "شكراً لزيارتكم"),
            "phone": system_invoice.get("system_phone"),
            "phone2": system_invoice.get("system_phone2"),
            "email": system_invoice.get("system_email"),
            "website": system_invoice.get("system_website"),
            "footer_text": system_invoice.get("footer_text"),
            "show_branding": system_invoice.get("show_system_branding", True)
        },
        "branch": branch
    }

# ==================== ROLES & STAFF MANAGEMENT - إدارة الأدوار والموظفين ====================
# نظام إدارة الموظفين والصلاحيات للعملاء

class StaffCreate(BaseModel):
    """نموذج إنشاء موظف جديد"""
    full_name: str
    email: str
    phone: Optional[str] = None
    password: str
    role: str = "cashier"  # cashier, supervisor, delivery, branch_manager
    branch_id: str
    job_title: Optional[str] = None  # المسمى الوظيفي المخصص
    permissions: Optional[List[str]] = None  # صلاحيات الموظف

class StaffUpdate(BaseModel):
    """نموذج تحديث بيانات موظف"""
    full_name: Optional[str] = None
    phone: Optional[str] = None
    role: Optional[str] = None
    branch_id: Optional[str] = None
    job_title: Optional[str] = None
    is_active: Optional[bool] = None
    permissions: Optional[List[str]] = None  # صلاحيات الموظف

class StaffResponse(BaseModel):
    """نموذج الاستجابة لبيانات الموظف"""
    model_config = ConfigDict(extra="ignore")
    id: str
    full_name: str
    email: str
    phone: Optional[str] = None
    role: str
    branch_id: Optional[str] = None
    branch_name: Optional[str] = None
    job_title: Optional[str] = None
    is_active: bool = True
    last_login: Optional[str] = None
    created_at: Optional[str] = None
    permissions: Optional[List[str]] = None  # صلاحيات الموظف

# الأدوار المتاحة للموظفين (غير Admin و SuperAdmin)
STAFF_ROLES = {
    "branch_manager": "مدير فرع",
    "supervisor": "مشرف",
    "cashier": "كاشير",
    "waiter": "جرسون",
    "kitchen": "مطبخ"
}

# ملاحظة: تم إزالة دور "delivery" (سائق توصيل) من هنا
# السائقين يتم إنشاؤهم وإدارتهم فقط من قسم التوصيل (Delivery)

@api_router.get("/staff/roles")
async def get_staff_roles(current_user: dict = Depends(get_current_user)):
    """جلب قائمة الأدوار المتاحة للموظفين"""
    if current_user.get("role") not in [UserRole.ADMIN, UserRole.MANAGER, UserRole.SUPER_ADMIN]:
        raise HTTPException(status_code=403, detail="غير مصرح")
    return STAFF_ROLES

@api_router.get("/staff", response_model=List[StaffResponse])
async def get_staff_members(
    branch_id: Optional[str] = None,
    role: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    """جلب قائمة الموظفين - للعميل فقط"""
    if current_user.get("role") not in [UserRole.ADMIN, UserRole.MANAGER, UserRole.SUPER_ADMIN]:
        raise HTTPException(status_code=403, detail="غير مصرح")
    
    query = build_tenant_query(current_user)
    
    # استثناء المستخدمين الرئيسيين (Admin و SuperAdmin)
    query["role"] = {"$nin": [UserRole.ADMIN, UserRole.SUPER_ADMIN]}
    
    if branch_id:
        query["branch_id"] = branch_id
    if role:
        query["role"] = role
    
    staff = await db.users.find(query, {"_id": 0, "password": 0}).to_list(500)
    
    # إضافة اسم الفرع لكل موظف
    branch_ids = list(set([s.get("branch_id") for s in staff if s.get("branch_id")]))
    branches = await db.branches.find({"id": {"$in": branch_ids}}, {"_id": 0, "id": 1, "name": 1}).to_list(100)
    branch_map = {b["id"]: b["name"] for b in branches}
    
    for s in staff:
        s["branch_name"] = branch_map.get(s.get("branch_id"), "غير محدد")
    
    return staff

@api_router.post("/staff", response_model=StaffResponse)
async def create_staff_member(staff: StaffCreate, current_user: dict = Depends(get_current_user)):
    """إنشاء موظف جديد"""
    if current_user.get("role") not in [UserRole.ADMIN, UserRole.MANAGER, UserRole.SUPER_ADMIN]:
        raise HTTPException(status_code=403, detail="غير مصرح")
    
    # التحقق من أن الدور صحيح
    if staff.role not in STAFF_ROLES:
        raise HTTPException(status_code=400, detail=f"الدور غير صحيح. الأدوار المتاحة: {', '.join(STAFF_ROLES.keys())}")
    
    # التحقق من عدم تكرار البريد
    existing = await db.users.find_one({"email": staff.email})
    if existing:
        raise HTTPException(status_code=400, detail="البريد الإلكتروني مستخدم")
    
    # التحقق من الفرع
    tenant_id = get_user_tenant_id(current_user)
    branch_query = {"id": staff.branch_id}
    if tenant_id:
        branch_query["tenant_id"] = tenant_id
    
    branch = await db.branches.find_one(branch_query, {"_id": 0})
    if not branch:
        raise HTTPException(status_code=404, detail="الفرع غير موجود")
    
    # إنشاء الموظف
    now = datetime.now(timezone.utc).isoformat()
    staff_doc = {
        "id": str(uuid.uuid4()),
        "full_name": staff.full_name,
        "username": staff.email.split("@")[0],
        "email": staff.email,
        "phone": staff.phone,
        "password": hash_password(staff.password),
        "role": staff.role,
        "branch_id": staff.branch_id,
        "job_title": staff.job_title or STAFF_ROLES.get(staff.role, staff.role),
        "permissions": staff.permissions or [],
        "tenant_id": tenant_id,
        "is_active": True,
        "last_login": None,
        "created_at": now,
        "updated_at": now
    }
    
    await db.users.insert_one(staff_doc)
    
    # إزالة كلمة المرور من الاستجابة
    del staff_doc["password"]
    del staff_doc["_id"]
    staff_doc["branch_name"] = branch.get("name", "")
    
    return staff_doc

@api_router.get("/staff/{staff_id}", response_model=StaffResponse)
async def get_staff_member(staff_id: str, current_user: dict = Depends(get_current_user)):
    """جلب بيانات موظف محدد"""
    if current_user.get("role") not in [UserRole.ADMIN, UserRole.MANAGER, UserRole.SUPER_ADMIN]:
        raise HTTPException(status_code=403, detail="غير مصرح")
    
    query = build_tenant_query(current_user, {"id": staff_id})
    staff = await db.users.find_one(query, {"_id": 0, "password": 0})
    
    if not staff:
        raise HTTPException(status_code=404, detail="الموظف غير موجود")
    
    # جلب اسم الفرع
    if staff.get("branch_id"):
        branch = await db.branches.find_one({"id": staff["branch_id"]}, {"_id": 0, "name": 1})
        staff["branch_name"] = branch.get("name", "") if branch else ""
    
    return staff

@api_router.put("/staff/{staff_id}", response_model=StaffResponse)
async def update_staff_member(staff_id: str, update: StaffUpdate, current_user: dict = Depends(get_current_user)):
    """تحديث بيانات موظف"""
    if current_user.get("role") not in [UserRole.ADMIN, UserRole.MANAGER, UserRole.SUPER_ADMIN]:
        raise HTTPException(status_code=403, detail="غير مصرح")
    
    query = build_tenant_query(current_user, {"id": staff_id})
    staff = await db.users.find_one(query)
    
    if not staff:
        raise HTTPException(status_code=404, detail="الموظف غير موجود")
    
    # التحقق من أن الموظف ليس Admin أو SuperAdmin
    if staff.get("role") in [UserRole.ADMIN, UserRole.SUPER_ADMIN]:
        raise HTTPException(status_code=403, detail="لا يمكن تعديل هذا المستخدم")
    
    # بناء التحديث
    update_data = {}
    if update.full_name is not None:
        update_data["full_name"] = update.full_name
    if update.phone is not None:
        update_data["phone"] = update.phone
    if update.role is not None:
        if update.role not in STAFF_ROLES:
            raise HTTPException(status_code=400, detail="الدور غير صحيح")
        update_data["role"] = update.role
    if update.branch_id is not None:
        # التحقق من الفرع
        tenant_id = get_user_tenant_id(current_user)
        branch_query = {"id": update.branch_id}
        if tenant_id:
            branch_query["tenant_id"] = tenant_id
        branch = await db.branches.find_one(branch_query)
        if not branch:
            raise HTTPException(status_code=404, detail="الفرع غير موجود")
        update_data["branch_id"] = update.branch_id
    if update.job_title is not None:
        update_data["job_title"] = update.job_title
    if update.is_active is not None:
        update_data["is_active"] = update.is_active
    if update.permissions is not None:
        update_data["permissions"] = update.permissions
    
    update_data["updated_at"] = datetime.now(timezone.utc).isoformat()
    
    await db.users.update_one(query, {"$set": update_data})
    
    # جلب البيانات المحدثة
    updated_staff = await db.users.find_one({"id": staff_id}, {"_id": 0, "password": 0})
    
    # جلب اسم الفرع
    if updated_staff.get("branch_id"):
        branch = await db.branches.find_one({"id": updated_staff["branch_id"]}, {"_id": 0, "name": 1})
        updated_staff["branch_name"] = branch.get("name", "") if branch else ""
    
    return updated_staff

@api_router.delete("/staff/{staff_id}")
async def delete_staff_member(staff_id: str, current_user: dict = Depends(get_current_user)):
    """حذف (تعطيل) موظف"""
    if current_user.get("role") not in [UserRole.ADMIN, UserRole.MANAGER, UserRole.SUPER_ADMIN]:
        raise HTTPException(status_code=403, detail="غير مصرح")
    
    query = build_tenant_query(current_user, {"id": staff_id})
    staff = await db.users.find_one(query)
    
    if not staff:
        raise HTTPException(status_code=404, detail="الموظف غير موجود")
    
    # التحقق من أن الموظف ليس Admin أو SuperAdmin
    if staff.get("role") in [UserRole.ADMIN, UserRole.SUPER_ADMIN]:
        raise HTTPException(status_code=403, detail="لا يمكن حذف هذا المستخدم")
    
    # تعطيل بدلاً من الحذف
    await db.users.update_one(query, {"$set": {"is_active": False, "updated_at": datetime.now(timezone.utc).isoformat()}})
    
    return {"message": "تم تعطيل الموظف"}

@api_router.post("/staff/{staff_id}/reset-password")
async def reset_staff_password(staff_id: str, new_password: str = Body(..., embed=True), current_user: dict = Depends(get_current_user)):
    """إعادة تعيين كلمة مرور موظف"""
    if current_user.get("role") not in [UserRole.ADMIN, UserRole.MANAGER, UserRole.SUPER_ADMIN]:
        raise HTTPException(status_code=403, detail="غير مصرح")
    
    query = build_tenant_query(current_user, {"id": staff_id})
    staff = await db.users.find_one(query)
    
    if not staff:
        raise HTTPException(status_code=404, detail="الموظف غير موجود")
    
    # التحقق من أن الموظف ليس Admin أو SuperAdmin
    if staff.get("role") in [UserRole.ADMIN, UserRole.SUPER_ADMIN]:
        raise HTTPException(status_code=403, detail="لا يمكن تعديل هذا المستخدم")
    
    await db.users.update_one(query, {"$set": {"password": hash_password(new_password), "updated_at": datetime.now(timezone.utc).isoformat()}})
    
    return {"message": "تم تغيير كلمة المرور"}

# ==================== FILE UPLOAD ROUTES ====================

async def process_and_save_image(file: UploadFile, target_dir: Path, max_size: tuple = (1920, 1080), quality: int = 85) -> str:
    """معالجة وحفظ الصورة بالحجم والصيغة المناسبة"""
    try:
        # قراءة محتوى الملف
        content = await file.read()
        
        # فتح الصورة
        image = Image.open(io.BytesIO(content))
        
        # تحويل RGBA إلى RGB إذا لزم الأمر
        if image.mode in ('RGBA', 'LA', 'P'):
            background = Image.new('RGB', image.size, (255, 255, 255))
            if image.mode == 'P':
                image = image.convert('RGBA')
            background.paste(image, mask=image.split()[-1] if image.mode == 'RGBA' else None)
            image = background
        elif image.mode != 'RGB':
            image = image.convert('RGB')
        
        # تغيير الحجم مع الحفاظ على النسبة
        image.thumbnail(max_size, Image.Resampling.LANCZOS)
        
        # إنشاء اسم ملف فريد
        file_id = str(uuid.uuid4())
        filename = f"{file_id}.jpg"
        filepath = target_dir / filename
        
        # حفظ الصورة بصيغة JPEG مضغوطة
        image.save(filepath, "JPEG", quality=quality, optimize=True)
        
        return filename
    except Exception as e:
        logger.error(f"Error processing image: {e}")
        raise HTTPException(status_code=400, detail=f"فشل في معالجة الصورة: {str(e)}")

@api_router.post("/upload/background")
async def upload_background_file(
    file: UploadFile = File(...),
    title: str = Form(None),
    animation_type: str = Form("fade"),
    current_user: dict = Depends(verify_super_admin)
):
    """رفع خلفية من الجهاز مع تحويل تلقائي"""
    
    # التحقق من نوع الملف
    allowed_types = ['image/jpeg', 'image/png', 'image/gif', 'image/webp', 'image/heic', 'image/heif', 'image/bmp', 'image/tiff']
    if file.content_type not in allowed_types:
        raise HTTPException(status_code=400, detail="نوع الملف غير مدعوم. الأنواع المدعومة: JPEG, PNG, GIF, WEBP, HEIC, BMP, TIFF")
    
    # معالجة وحفظ الصورة
    filename = await process_and_save_image(file, BACKGROUNDS_DIR, max_size=(1920, 1080), quality=85)
    
    # إنشاء URL للصورة
    base_url = os.environ.get('REACT_APP_BACKEND_URL', '')
    image_url = f"{base_url}/api/uploads/backgrounds/{filename}"
    
    # جلب الإعدادات الحالية
    settings = await db.settings.find_one({"type": "login_backgrounds"}, {"_id": 0})
    current_backgrounds = []
    current_value = {}
    
    if settings and settings.get("value"):
        current_value = settings["value"]
        current_backgrounds = current_value.get("backgrounds", [])
    
    # إضافة الخلفية الجديدة
    new_background = {
        "id": str(uuid.uuid4()),
        "image_url": image_url,
        "title": title,
        "animation_type": animation_type,
        "animation_duration": 8,
        "overlay_opacity": 0.5,
        "is_active": True,
        "sort_order": len(current_backgrounds),
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    
    current_backgrounds.append(new_background)
    
    # تحديث مع الحفاظ على الإعدادات الأخرى
    default_settings = {
        "backgrounds": current_backgrounds,
        "animation_enabled": current_value.get("animation_enabled", True),
        "transition_type": current_value.get("transition_type", "fade"),
        "transition_duration": current_value.get("transition_duration", 1.5),
        "auto_play": current_value.get("auto_play", True),
        "show_logo": current_value.get("show_logo", True),
        "logo_url": current_value.get("logo_url", None),
        "logo_animation": current_value.get("logo_animation", "pulse"),
        "overlay_color": current_value.get("overlay_color", "rgba(0,0,0,0.5)"),
        "text_color": current_value.get("text_color", "#ffffff")
    }
    
    # حفظ التحديث
    await db.settings.update_one(
        {"type": "login_backgrounds"},
        {"$set": {"type": "login_backgrounds", "value": default_settings}},
        upsert=True
    )
    
    return {"message": "تم رفع الخلفية بنجاح", "background": new_background}

@api_router.post("/upload/logo")
async def upload_logo_file(
    file: UploadFile = File(...),
    tenant_id: str = Form(None),
    current_user: dict = Depends(verify_super_admin)
):
    """رفع شعار للعميل - للمالك فقط"""
    
    # التحقق من نوع الملف
    allowed_types = ['image/jpeg', 'image/png', 'image/gif', 'image/webp', 'image/svg+xml']
    if file.content_type not in allowed_types:
        raise HTTPException(status_code=400, detail="نوع الملف غير مدعوم")
    
    # معالجة وحفظ الصورة
    filename = await process_and_save_image(file, LOGOS_DIR, max_size=(512, 512), quality=90)
    
    # إنشاء URL نسبي للشعار
    logo_url = f"/api/uploads/logos/{filename}"
    
    # تحديث الشعار للعميل إذا تم تحديد tenant_id
    if tenant_id:
        await db.tenants.update_one(
            {"id": tenant_id},
            {"$set": {"logo_url": logo_url}}
        )
    
    return {"message": "تم رفع الشعار بنجاح", "logo_url": logo_url, "url": logo_url}

@api_router.post("/upload/restaurant-logo")
async def upload_restaurant_logo(
    file: UploadFile = File(...),
    current_user: dict = Depends(get_current_user)
):
    """رفع شعار المطعم - للمدير أو المالك"""
    
    # السماح للمدير (admin) والمالك (super_admin)
    if current_user["role"] not in [UserRole.ADMIN, UserRole.SUPER_ADMIN]:
        raise HTTPException(status_code=403, detail="غير مصرح")
    
    # التحقق من نوع الملف
    allowed_types = ['image/jpeg', 'image/png', 'image/gif', 'image/webp', 'image/svg+xml']
    if file.content_type not in allowed_types:
        raise HTTPException(status_code=400, detail="نوع الملف غير مدعوم. يرجى استخدام JPG, PNG, GIF, WebP أو SVG")
    
    # معالجة وحفظ الصورة
    filename = await process_and_save_image(file, LOGOS_DIR, max_size=(512, 512), quality=90)
    
    # إنشاء URL نسبي للشعار
    logo_url = f"/api/uploads/logos/{filename}"
    
    return {"message": "تم رفع الشعار بنجاح", "url": logo_url, "logo_url": logo_url}

@api_router.post("/upload/image")
async def upload_general_image(
    file: UploadFile = File(...),
    type: str = Form("product"),  # product, category, general
    current_user: dict = Depends(get_current_user)
):
    """رفع صورة عامة للمنتجات أو الفئات"""
    
    # التحقق من نوع الملف
    allowed_types = [
        'image/jpeg', 'image/png', 'image/gif', 'image/webp', 
        'image/heic', 'image/heif', 'image/bmp', 'image/tiff',
        'image/svg+xml', 'image/avif'
    ]
    
    # السماح بأي نوع صورة
    content_type = file.content_type or ''
    if not content_type.startswith('image/'):
        raise HTTPException(status_code=400, detail="يجب أن يكون الملف صورة")
    
    # تحديد المجلد حسب النوع
    if type == "product":
        target_dir = PRODUCTS_DIR
        subfolder = "products"
        max_size = (800, 800)
    elif type == "category":
        target_dir = CATEGORIES_DIR
        subfolder = "categories"
        max_size = (400, 400)
    else:
        target_dir = IMAGES_DIR
        subfolder = "images"
        max_size = (1024, 1024)
    
    # معالجة وحفظ الصورة
    filename = await process_and_save_image(file, target_dir, max_size=max_size, quality=85)
    
    # إنشاء URL للصورة
    image_url = f"/api/uploads/images/{subfolder}/{filename}"
    
    return {
        "message": "تم رفع الصورة بنجاح",
        "image_url": image_url,
        "filename": filename
    }

@api_router.post("/upload/product-image")
async def upload_product_image(
    file: UploadFile = File(...),
    current_user: dict = Depends(get_current_user)
):
    """رفع صورة منتج"""
    return await upload_general_image(file, "product", current_user)

@api_router.post("/upload/category-image")
async def upload_category_image(
    file: UploadFile = File(...),
    current_user: dict = Depends(get_current_user)
):
    """رفع صورة فئة"""
    return await upload_general_image(file, "category", current_user)

# ==================== PRINTER ROUTES ====================

class PrinterCreate(BaseModel):
    name: str
    ip_address: str
    port: int = 9100
    branch_id: str
    printer_type: str = "receipt"  # النوع: receipt, kitchen, bar, packaging, label, custom
    custom_type_name: Optional[str] = None  # اسم مخصص للنوع إذا كان custom
    # صلاحيات الطباعة
    print_mode: str = "full_receipt"  # full_receipt, orders_only, selected_products
    show_prices: bool = True  # عرض الأسعار في الطباعة
    print_individual_items: bool = False  # طباعة كل صنف على حدة
    auto_print_on_order: bool = True  # طباعة تلقائية عند الطلب

@api_router.post("/printers")
async def create_printer(printer: PrinterCreate, current_user: dict = Depends(get_current_user)):
    if current_user["role"] not in [UserRole.ADMIN, UserRole.MANAGER, UserRole.SUPER_ADMIN]:
        raise HTTPException(status_code=403, detail="غير مصرح")
    
    tenant_id = get_user_tenant_id(current_user)
    printer_doc = {
        "id": str(uuid.uuid4()),
        **printer.model_dump(),
        "tenant_id": tenant_id,  # ربط الطابعة بالعميل
        "is_active": True,
        "is_online": False,
        "last_check": None
    }
    await db.printers.insert_one(printer_doc)
    del printer_doc["_id"]
    return printer_doc

@api_router.get("/printer-types")
async def get_printer_types(current_user: dict = Depends(get_current_user)):
    """جلب أنواع الطابعات المتاحة"""
    default_types = [
        {"id": "receipt", "name": "طابعة إيصالات", "name_en": "Receipt Printer", "icon": "Receipt"},
        {"id": "kitchen", "name": "طابعة مطبخ", "name_en": "Kitchen Printer", "icon": "ChefHat"},
        {"id": "bar", "name": "طابعة بار/مشروبات", "name_en": "Bar Printer", "icon": "Wine"},
        {"id": "packaging", "name": "طابعة تغليف", "name_en": "Packaging Printer", "icon": "Package"},
        {"id": "label", "name": "طابعة ملصقات", "name_en": "Label Printer", "icon": "Tag"},
    ]
    
    # جلب الأنواع المخصصة للعميل
    tenant_id = get_user_tenant_id(current_user)
    query = {"tenant_id": tenant_id} if tenant_id else {}
    custom_types = await db.printer_types.find(query, {"_id": 0}).to_list(50)
    
    return {"default": default_types, "custom": custom_types}

@api_router.post("/printer-types")
async def create_printer_type(type_data: dict, current_user: dict = Depends(get_current_user)):
    """إضافة نوع طابعة مخصص"""
    if current_user["role"] not in [UserRole.ADMIN, UserRole.MANAGER]:
        raise HTTPException(status_code=403, detail="غير مصرح")
    
    tenant_id = get_user_tenant_id(current_user)
    type_doc = {
        "id": str(uuid.uuid4()),
        "name": type_data.get("name"),
        "name_en": type_data.get("name_en", type_data.get("name")),
        "icon": type_data.get("icon", "Printer"),
        "tenant_id": tenant_id,
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    await db.printer_types.insert_one(type_doc)
    del type_doc["_id"]
    return type_doc

@api_router.delete("/printer-types/{type_id}")
async def delete_printer_type(type_id: str, current_user: dict = Depends(get_current_user)):
    """حذف نوع طابعة مخصص"""
    query = build_tenant_query(current_user, {"id": type_id})
    await db.printer_types.delete_one(query)
    return {"message": "تم الحذف"}

@api_router.get("/printers")
async def get_printers(branch_id: Optional[str] = None, current_user: dict = Depends(get_current_user)):
    # فلترة حسب tenant_id لفصل بيانات كل عميل
    query = build_tenant_query(current_user)
    if branch_id:
        query["branch_id"] = branch_id
    printers = await db.printers.find(query, {"_id": 0}).to_list(50)
    return printers

@api_router.put("/printers/{printer_id}")
async def update_printer(printer_id: str, printer: PrinterCreate, current_user: dict = Depends(get_current_user)):
    if current_user["role"] not in [UserRole.ADMIN, UserRole.MANAGER, UserRole.SUPER_ADMIN]:
        raise HTTPException(status_code=403, detail="غير مصرح")
    
    # التحقق من أن الطابعة تنتمي لنفس العميل
    query = build_tenant_query(current_user, {"id": printer_id})
    existing = await db.printers.find_one(query)
    if not existing:
        raise HTTPException(status_code=404, detail="الطابعة غير موجودة")
    
    update_data = printer.model_dump()
    await db.printers.update_one({"id": printer_id}, {"$set": update_data})
    updated = await db.printers.find_one({"id": printer_id}, {"_id": 0})
    return updated

@api_router.delete("/printers/{printer_id}")
async def delete_printer(printer_id: str, current_user: dict = Depends(get_current_user)):
    if current_user["role"] not in [UserRole.ADMIN, UserRole.MANAGER, UserRole.SUPER_ADMIN]:
        raise HTTPException(status_code=403, detail="غير مصرح")
    
    # التحقق من أن الطابعة تنتمي لنفس العميل
    query = build_tenant_query(current_user, {"id": printer_id})
    result = await db.printers.delete_one(query)
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="الطابعة غير موجودة")
    return {"message": "تم حذف الطابعة بنجاح"}

@api_router.post("/printers/{printer_id}/test")
async def test_printer_connection(printer_id: str, current_user: dict = Depends(get_current_user)):
    """اختبار اتصال الطابعة"""
    import socket
    
    # التحقق من أن الطابعة تنتمي لنفس العميل
    query = build_tenant_query(current_user, {"id": printer_id})
    printer = await db.printers.find_one(query, {"_id": 0})
    if not printer:
        raise HTTPException(status_code=404, detail="الطابعة غير موجودة")
    
    ip = printer.get("ip_address", "")
    port = printer.get("port", 9100)
    
    try:
        sock = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
        sock.settimeout(3)  # timeout 3 ثواني
        result = sock.connect_ex((ip, port))
        sock.close()
        
        if result == 0:
            # تحديث حالة الطابعة
            await db.printers.update_one(
                query,
                {"$set": {"is_online": True, "last_check": datetime.now(timezone.utc).isoformat()}}
            )
            return {"status": "online", "message": "الطابعة متصلة"}
        else:
            await db.printers.update_one(
                query,
                {"$set": {"is_online": False, "last_check": datetime.now(timezone.utc).isoformat()}}
            )
            return {"status": "offline", "message": "الطابعة غير متصلة"}
    except socket.error as e:
        await db.printers.update_one(
            query,
            {"$set": {"is_online": False, "last_check": datetime.now(timezone.utc).isoformat()}}
        )
        return {"status": "error", "message": f"خطأ في الاتصال: {str(e)}"}

# ==================== SUPER ADMIN & TENANT MANAGEMENT ====================
# نظام إدارة المستأجرين - لوحة تحكم المالك الرئيسي

class SuperAdminLoginRequest(BaseModel):
    email: str
    password: str
    secret_key: str

@api_router.post("/super-admin/login")
async def super_admin_login(request: SuperAdminLoginRequest):
    """تسجيل دخول Super Admin"""
    user = await db.users.find_one({"email": request.email, "role": UserRole.SUPER_ADMIN})
    if not user:
        raise HTTPException(status_code=401, detail="المستخدم غير موجود")
    
    # التحقق من المفتاح السري (من قاعدة البيانات أو القيمة الافتراضية)
    stored_secret = user.get("secret_key") or SUPER_ADMIN_SECRET
    if request.secret_key != stored_secret:
        raise HTTPException(status_code=403, detail="المفتاح السري غير صحيح")
    
    # التحقق من كلمة المرور (الحقل قد يكون password أو password_hash)
    password_field = user.get("password") or user.get("password_hash")
    if not password_field or not verify_password(request.password, password_field):
        raise HTTPException(status_code=401, detail="كلمة المرور غير صحيحة")
    
    token = create_token(user["id"], user["role"], user.get("branch_id"))
    
    return {
        "token": token,
        "user": {
            "id": user["id"],
            "email": user["email"],
            "full_name": user["full_name"],
            "role": user["role"]
        }
    }

class SuperAdminRegisterRequest(BaseModel):
    email: str
    password: str
    full_name: str
    secret_key: str

@api_router.post("/super-admin/register")
async def register_super_admin(request: SuperAdminRegisterRequest):
    """إنشاء حساب Super Admin (مرة واحدة فقط)"""
    if request.secret_key != SUPER_ADMIN_SECRET:
        raise HTTPException(status_code=403, detail="مفتاح السر غير صحيح")
    
    # التحقق من عدم وجود Super Admin
    existing = await db.users.find_one({"role": UserRole.SUPER_ADMIN})
    if existing:
        raise HTTPException(status_code=400, detail="يوجد Super Admin بالفعل")
    
    # التحقق من عدم وجود البريد
    email_exists = await db.users.find_one({"email": request.email})
    if email_exists:
        raise HTTPException(status_code=400, detail="البريد الإلكتروني مستخدم")
    
    user_doc = {
        "id": str(uuid.uuid4()),
        "username": "super_admin",
        "email": request.email,
        "password": hash_password(request.password),
        "full_name": request.full_name,
        "role": UserRole.SUPER_ADMIN,
        "branch_id": None,
        "tenant_id": None,
        "permissions": ["all", "super_admin"],
        "is_active": True,
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    
    await db.users.insert_one(user_doc)
    
    token = create_token(user_doc["id"], user_doc["role"], None)
    
    return {
        "message": "تم إنشاء حساب Super Admin بنجاح",
        "token": token,
        "user": {
            "id": user_doc["id"],
            "email": user_doc["email"],
            "full_name": user_doc["full_name"],
            "role": user_doc["role"]
        }
    }

@api_router.get("/super-admin/tenants")
async def get_all_tenants(current_user: dict = Depends(verify_super_admin)):
    """جلب جميع العملاء (المستأجرين)"""
    tenants = await db.tenants.find({}, {"_id": 0}).sort("created_at", -1).to_list(1000)
    
    # إضافة إحصائيات لكل مستأجر
    for tenant in tenants:
        tenant["users_count"] = await db.users.count_documents({"tenant_id": tenant["id"]})
        tenant["branches_count"] = await db.branches.count_documents({"tenant_id": tenant["id"]})
        tenant["orders_count"] = await db.orders.count_documents({"tenant_id": tenant["id"]})
    
    # إرجاع قائمة العملاء فقط (النظام الرئيسي هو المالك ولا يظهر كعميل)
    return tenants

@api_router.post("/super-admin/tenants")
async def create_tenant(tenant: TenantCreate, background_tasks: BackgroundTasks, current_user: dict = Depends(verify_super_admin)):
    """إنشاء مستأجر جديد (عميل جديد) مع إرسال بريد ترحيبي وإشعار"""
    
    # التحقق من عدم وجود slug مكرر
    existing = await db.tenants.find_one({"slug": tenant.slug})
    if existing:
        raise HTTPException(status_code=400, detail="الرابط المختصر مستخدم")
    
    # التحقق من عدم وجود البريد
    email_exists = await db.users.find_one({"email": tenant.owner_email})
    if email_exists:
        raise HTTPException(status_code=400, detail="البريد الإلكتروني مستخدم")
    
    tenant_id = str(uuid.uuid4())
    
    # تحديد تاريخ انتهاء الاشتراك بناءً على المدة المحددة
    subscription_duration = getattr(tenant, 'subscription_duration', 1)  # افتراضي شهر واحد
    
    if tenant.subscription_type == "trial":
        expires_at = (datetime.now(timezone.utc) + timedelta(days=14)).isoformat()
    elif tenant.subscription_type == "demo":
        expires_at = (datetime.now(timezone.utc) + timedelta(days=30)).isoformat()
    else:
        # استخدام مدة الاشتراك المحددة بالأشهر
        expires_at = (datetime.now(timezone.utc) + timedelta(days=30 * subscription_duration)).isoformat()
    
    tenant_doc = {
        "id": tenant_id,
        "name": tenant.name,
        "slug": tenant.slug,
        "owner_name": tenant.owner_name,
        "owner_email": tenant.owner_email,
        "owner_phone": tenant.owner_phone,
        "subscription_type": tenant.subscription_type,
        "subscription_duration": subscription_duration,
        "max_branches": tenant.max_branches,
        "max_users": tenant.max_users,
        "is_active": True,
        "is_demo": getattr(tenant, 'is_demo', False),
        "created_at": datetime.now(timezone.utc).isoformat(),
        "expires_at": expires_at,
        "created_by": current_user["id"]
    }
    
    await db.tenants.insert_one(tenant_doc)
    
    # إنشاء مستخدم admin للمستأجر
    admin_password = f"{tenant.slug}123"  # كلمة مرور افتراضية
    admin_doc = {
        "id": str(uuid.uuid4()),
        "username": f"{tenant.slug}_admin",
        "email": tenant.owner_email,
        "password": hash_password(admin_password),
        "full_name": tenant.owner_name,
        "role": UserRole.ADMIN,
        "branch_id": None,
        "tenant_id": tenant_id,
        "permissions": ["all"],
        "is_active": True,
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    
    await db.users.insert_one(admin_doc)
    
    # إنشاء فرع افتراضي للمستأجر
    branch_doc = {
        "id": str(uuid.uuid4()),
        "name": "الفرع الرئيسي",
        "address": "",
        "phone": tenant.owner_phone,
        "email": tenant.owner_email,
        "tenant_id": tenant_id,
        "is_active": True,
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    
    await db.branches.insert_one(branch_doc)
    
    del tenant_doc["_id"]
    
    # إرسال بريد ترحيبي تلقائياً
    background_tasks.add_task(
        send_welcome_email,
        recipient_email=tenant.owner_email,
        tenant_name=tenant.name,
        owner_name=tenant.owner_name,
        username=tenant.owner_email,
        password=admin_password
    )
    
    # إنشاء إشعار للمالك عن العميل الجديد
    notification_doc = {
        "id": str(uuid.uuid4()),
        "type": "new_tenant",
        "title": "عميل جديد 🎉",
        "message": f"تم إنشاء عميل جديد: {tenant.name} ({tenant.owner_name})",
        "tenant_id": tenant_id,
        "data": {
            "tenant_name": tenant.name,
            "owner_name": tenant.owner_name,
            "owner_email": tenant.owner_email,
            "subscription_type": tenant.subscription_type,
            "subscription_duration": subscription_duration,
            "expires_at": expires_at
        },
        "is_read": False,
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    await db.notifications.insert_one(notification_doc)
    
    return {
        "tenant": tenant_doc,
        "admin_credentials": {
            "email": tenant.owner_email,
            "password": admin_password,
            "message": "يرجى تغيير كلمة المرور فور تسجيل الدخول"
        },
        "access_url": f"/tenant/{tenant.slug}",
        "email_sent": True
    }

@api_router.get("/super-admin/tenants/{tenant_id}")
async def get_tenant_details(tenant_id: str, current_user: dict = Depends(verify_super_admin)):
    """تفاصيل مستأجر معين"""
    
    # التحقق إذا كان النظام الرئيسي
    if tenant_id == "main-system":
        # استعلام النظام الرئيسي يشمل: بدون tenant_id أو tenant_id = null أو tenant_id = "default"
        main_system_query = {
            "$or": [
                {"tenant_id": {"$exists": False}}, 
                {"tenant_id": None},
                {"tenant_id": "default"}
            ]
        }
        
        # جلب مستخدمي النظام الرئيسي
        users = await db.users.find({
            **main_system_query,
            "role": {"$ne": UserRole.SUPER_ADMIN}
        }, {"_id": 0, "password": 0}).to_list(100)
        
        branches = await db.branches.find(main_system_query, {"_id": 0}).to_list(50)
        
        # إحصائيات المبيعات للنظام الرئيسي
        today = datetime.now(timezone.utc).replace(hour=0, minute=0, second=0, microsecond=0).isoformat()
        orders_today = await db.orders.count_documents({
            **main_system_query,
            "created_at": {"$gte": today}
        })
        
        total_sales_cursor = db.orders.aggregate([
            {"$match": {**main_system_query, "status": {"$ne": "cancelled"}}},
            {"$group": {"_id": None, "total": {"$sum": "$total"}}}
        ])
        total_sales_result = await total_sales_cursor.to_list(1)
        total_sales = total_sales_result[0]["total"] if total_sales_result else 0
        
        # إجمالي الطلبات
        total_orders = await db.orders.count_documents(main_system_query)
        
        return {
            "tenant": {
                "id": "main-system",
                "name": "🏠 النظام الرئيسي",
                "slug": "main",
                "owner_name": "المالك",
                "owner_email": "admin@maestroegp.com",
                "owner_phone": "",
                "subscription_type": "premium",
                "is_active": True,
                "is_main_system": True,
                "created_at": "2024-01-01T00:00:00"
            },
            "users": users,
            "branches": branches,
            "stats": {
                "users_count": len(users),
                "branches_count": len(branches),
                "orders_today": orders_today,
                "total_sales": total_sales,
                "total_orders": total_orders
            }
        }
    
    # للعملاء العاديين
    tenant = await db.tenants.find_one({"id": tenant_id}, {"_id": 0})
    if not tenant:
        raise HTTPException(status_code=404, detail="المستأجر غير موجود")
    
    # إحصائيات تفصيلية
    users = await db.users.find({"tenant_id": tenant_id}, {"_id": 0, "password": 0}).to_list(100)
    branches = await db.branches.find({"tenant_id": tenant_id}, {"_id": 0}).to_list(50)
    
    # إحصائيات المبيعات
    today = datetime.now(timezone.utc).replace(hour=0, minute=0, second=0, microsecond=0).isoformat()
    orders_today = await db.orders.count_documents({
        "tenant_id": tenant_id,
        "created_at": {"$gte": today}
    })
    
    total_sales_cursor = db.orders.aggregate([
        {"$match": {"tenant_id": tenant_id, "status": {"$ne": "cancelled"}}},
        {"$group": {"_id": None, "total": {"$sum": "$total"}}}
    ])
    total_sales_result = await total_sales_cursor.to_list(1)
    total_sales = total_sales_result[0]["total"] if total_sales_result else 0
    
    return {
        "tenant": tenant,
        "users": users,
        "branches": branches,
        "stats": {
            "users_count": len(users),
            "branches_count": len(branches),
            "orders_today": orders_today,
            "total_sales": total_sales
        }
    }

@api_router.put("/super-admin/tenants/{tenant_id}/features")
async def update_tenant_features(tenant_id: str, features: dict, current_user: dict = Depends(verify_super_admin)):
    """تحديث ميزات العميل المتاحة"""
    tenant = await db.tenants.find_one({"id": tenant_id})
    if not tenant:
        raise HTTPException(status_code=404, detail="المستأجر غير موجود")
    
    # قائمة الميزات المسموح بها
    allowed_features = [
        "showPOS", "showTables", "showOrders", "showExpenses",
        "showInventory", "showDelivery", "showReports", "showSettings",
        "showHR", "showWarehouse", "showCallLogs", "showCallCenter", "showKitchen",
        "showLoyalty", "showCoupons", "showRecipes", "showReservations",
        "showReviews", "showSmartReports", "showPurchasing", "showBranchOrders",
        "showCustomerMenu",
        # خيارات الإعدادات
        "settingsUsers", "settingsCustomers", "settingsBranches", 
        "settingsCategories", "settingsProducts", "settingsPrinters",
        "settingsDeliveryCompanies", "settingsCallCenter", "settingsNotifications"
    ]
    
    # فلترة الميزات المرسلة
    enabled_features = {k: v for k, v in features.items() if k in allowed_features}
    
    # تحديث العميل
    await db.tenants.update_one(
        {"id": tenant_id},
        {"$set": {"enabled_features": enabled_features}}
    )
    
    return {"message": "تم تحديث ميزات العميل", "features": enabled_features}

@api_router.get("/super-admin/tenants/{tenant_id}/features")
async def get_tenant_features(tenant_id: str, current_user: dict = Depends(verify_super_admin)):
    """جلب ميزات العميل المتاحة"""
    tenant = await db.tenants.find_one({"id": tenant_id}, {"_id": 0, "enabled_features": 1, "name": 1})
    if not tenant:
        raise HTTPException(status_code=404, detail="المستأجر غير موجود")
    
    # الميزات الافتراضية للعميل الجديد
    default_features = {
        "showPOS": True,
        "showTables": True,
        "showOrders": True,
        "showExpenses": True,
        "showInventory": True,
        "showDelivery": True,
        "showReports": True,
        "showSettings": True,
        "showHR": False,
        "showWarehouse": False,
        "showCallLogs": False,
        "showCallCenter": False,
        "showKitchen": False,
        "showLoyalty": True,
        "showCoupons": True,
        "showRecipes": False,
        "showReservations": True,
        "showReviews": True,
        "showRatings": True,
        "showSmartReports": True,
        "showPurchasing": False,
        "showBranchOrders": False,
        "showCustomerMenu": True,
        # خيارات الإعدادات
        "settingsUsers": True,
        "settingsCustomers": True,
        "settingsBranches": True,
        "settingsCategories": True,
        "settingsProducts": True,
        "settingsPrinters": True,
        "settingsDeliveryCompanies": True,
        "settingsCallCenter": True,
        "settingsNotifications": True
    }
    
    # دمج الميزات المحفوظة مع الافتراضية
    saved_features = tenant.get("enabled_features", {})
    features = {**default_features, **saved_features}
    
    return {"tenant_name": tenant.get("name"), "features": features}

@api_router.put("/super-admin/tenants/{tenant_id}")
async def update_tenant(tenant_id: str, updates: dict, background_tasks: BackgroundTasks, current_user: dict = Depends(verify_super_admin)):
    """تحديث بيانات مستأجر مع إرسال بريد إلكتروني تلقائي"""
    tenant = await db.tenants.find_one({"id": tenant_id})
    if not tenant:
        raise HTTPException(status_code=404, detail="المستأجر غير موجود")
    
    # قائمة الحقول المسموح بتحديثها
    allowed_updates = [
        "name", "name_en", "name_ar", "owner_name", "owner_email", "owner_phone", 
        "subscription_type", "max_branches", "max_users", 
        "is_active", "expires_at", "logo_url"
    ]
    update_data = {k: v for k, v in updates.items() if k in allowed_updates}
    
    # التحقق من تغيير البريد الإلكتروني
    email_changed = False
    new_email = updates.get("owner_email")
    if new_email and new_email != tenant.get("owner_email"):
        # التحقق من عدم استخدام البريد من قبل
        existing = await db.users.find_one({"email": new_email})
        if existing:
            raise HTTPException(status_code=400, detail="البريد الإلكتروني مستخدم من قبل")
        email_changed = True
    
    if update_data:
        await db.tenants.update_one({"id": tenant_id}, {"$set": update_data})
    
    # تحديث حالة المستخدمين عند تغيير is_active
    if "is_active" in updates:
        await db.users.update_many(
            {"tenant_id": tenant_id},
            {"$set": {"is_active": updates["is_active"]}}
        )
    
    # تحديث بيانات المستخدم الأدمن إذا تم تغيير البريد أو الاسم
    admin_update = {}
    if new_email and email_changed:
        admin_update["email"] = new_email
    if updates.get("owner_name"):
        admin_update["full_name"] = updates.get("owner_name")
    
    if admin_update:
        await db.users.update_one(
            {"tenant_id": tenant_id, "role": UserRole.ADMIN},
            {"$set": admin_update}
        )
    
    # إرسال بريد إلكتروني إذا طُلب ذلك
    if updates.get("send_welcome_email"):
        admin = await db.users.find_one({"tenant_id": tenant_id, "role": UserRole.ADMIN}, {"_id": 0})
        if admin:
            # إعادة تعيين كلمة مرور مؤقتة للإرسال
            temp_password = updates.get("temp_password") or f"{tenant.get('slug')}123"
            await db.users.update_one(
                {"id": admin["id"]},
                {"$set": {"password": hash_password(temp_password)}}
            )
            
            # إرسال البريد في الخلفية
            background_tasks.add_task(
                send_welcome_email,
                recipient_email=admin.get("email"),
                tenant_name=update_data.get("name", tenant.get("name")),
                owner_name=update_data.get("owner_name", tenant.get("owner_name")),
                username=admin.get("email"),
                password=temp_password
            )
    
    updated = await db.tenants.find_one({"id": tenant_id}, {"_id": 0})
    return updated

@api_router.delete("/super-admin/tenants/{tenant_id}")
async def delete_tenant(tenant_id: str, permanent: bool = False, current_user: dict = Depends(verify_super_admin)):
    """حذف مستأجر (نهائي أو تعطيل)"""
    tenant = await db.tenants.find_one({"id": tenant_id})
    if not tenant:
        raise HTTPException(status_code=404, detail="المستأجر غير موجود")
    
    if permanent:
        # حذف نهائي - حذف جميع بيانات العميل
        await db.users.delete_many({"tenant_id": tenant_id})
        await db.branches.delete_many({"tenant_id": tenant_id})
        await db.categories.delete_many({"tenant_id": tenant_id})
        await db.products.delete_many({"tenant_id": tenant_id})
        await db.orders.delete_many({"tenant_id": tenant_id})
        await db.tables.delete_many({"tenant_id": tenant_id})
        await db.shifts.delete_many({"tenant_id": tenant_id})
        await db.inventory.delete_many({"tenant_id": tenant_id})
        await db.customers.delete_many({"tenant_id": tenant_id})
        await db.drivers.delete_many({"tenant_id": tenant_id})
        await db.suppliers.delete_many({"tenant_id": tenant_id})
        await db.recipes.delete_many({"tenant_id": tenant_id})
        await db.printers.delete_many({"tenant_id": tenant_id})
        await db.tenants.delete_one({"id": tenant_id})
        
        return {"message": "تم حذف المستأجر نهائياً مع جميع بياناته"}
    else:
        # تعطيل بدلاً من الحذف
        await db.tenants.update_one({"id": tenant_id}, {"$set": {"is_active": False}})
        await db.users.update_many({"tenant_id": tenant_id}, {"$set": {"is_active": False}})
        
        return {"message": "تم تعطيل المستأجر وجميع مستخدميه"}

@api_router.put("/super-admin/tenants/{tenant_id}/reactivate")
async def reactivate_tenant(tenant_id: str, current_user: dict = Depends(verify_super_admin)):
    """إعادة تفعيل مستأجر معطل"""
    tenant = await db.tenants.find_one({"id": tenant_id})
    if not tenant:
        raise HTTPException(status_code=404, detail="المستأجر غير موجود")
    
    # إعادة التفعيل
    await db.tenants.update_one({"id": tenant_id}, {"$set": {"is_active": True}})
    await db.users.update_many({"tenant_id": tenant_id}, {"$set": {"is_active": True}})
    
    # إنشاء إشعار عن التفعيل
    notification_doc = {
        "id": str(uuid.uuid4()),
        "type": "tenant_activated",
        "title": "تم تفعيل عميل ✅",
        "message": f"تم إعادة تفعيل العميل: {tenant.get('name', 'Unknown')}",
        "tenant_id": tenant_id,
        "data": {
            "tenant_name": tenant.get("name"),
            "owner_name": tenant.get("owner_name"),
            "action": "activated"
        },
        "is_read": False,
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    await db.notifications.insert_one(notification_doc)
    
    return {"message": "تم إعادة تفعيل المستأجر وجميع مستخدميه"}

@api_router.put("/super-admin/tenants/{tenant_id}/deactivate")
async def deactivate_tenant(tenant_id: str, current_user: dict = Depends(verify_super_admin)):
    """تعطيل مستأجر"""
    tenant = await db.tenants.find_one({"id": tenant_id})
    if not tenant:
        raise HTTPException(status_code=404, detail="المستأجر غير موجود")
    
    # التعطيل
    await db.tenants.update_one({"id": tenant_id}, {"$set": {"is_active": False}})
    await db.users.update_many({"tenant_id": tenant_id}, {"$set": {"is_active": False}})
    
    # إنشاء إشعار عن التعطيل
    notification_doc = {
        "id": str(uuid.uuid4()),
        "type": "tenant_deactivated",
        "title": "تم تعطيل عميل ⚠️",
        "message": f"تم تعطيل العميل: {tenant.get('name', 'Unknown')}",
        "tenant_id": tenant_id,
        "data": {
            "tenant_name": tenant.get("name"),
            "owner_name": tenant.get("owner_name"),
            "action": "deactivated"
        },
        "is_read": False,
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    await db.notifications.insert_one(notification_doc)
    
    return {"message": "تم تعطيل المستأجر وجميع مستخدميه"}


@api_router.post("/super-admin/tenants/{tenant_id}/reset-password")
async def reset_tenant_admin_password(tenant_id: str, new_password: str, current_user: dict = Depends(verify_super_admin)):
    """إعادة تعيين كلمة مرور مدير المستأجر"""
    
    # التحقق إذا كان النظام الرئيسي
    if tenant_id == "main-system":
        # البحث عن admin النظام الرئيسي
        admin = await db.users.find_one({
            "$or": [{"tenant_id": {"$exists": False}}, {"tenant_id": None}],
            "role": UserRole.ADMIN
        })
        if not admin:
            raise HTTPException(status_code=404, detail="مدير النظام الرئيسي غير موجود")
        
        await db.users.update_one(
            {"id": admin["id"]},
            {"$set": {"password": hash_password(new_password)}}
        )
        
        return {"message": "تم إعادة تعيين كلمة مرور مدير النظام الرئيسي", "email": admin["email"]}
    
    # للعملاء العاديين
    tenant = await db.tenants.find_one({"id": tenant_id})
    if not tenant:
        raise HTTPException(status_code=404, detail="المستأجر غير موجود")
    
    # البحث عن admin المستأجر
    admin = await db.users.find_one({"tenant_id": tenant_id, "role": UserRole.ADMIN})
    if not admin:
        raise HTTPException(status_code=404, detail="مدير المستأجر غير موجود")
    
    await db.users.update_one(
        {"id": admin["id"]},
        {"$set": {"password": hash_password(new_password)}}
    )
    
    return {"message": "تم إعادة تعيين كلمة المرور", "email": admin["email"]}

# إعدادات المالك
@api_router.get("/super-admin/owner-settings")
async def get_owner_settings(current_user: dict = Depends(verify_super_admin)):
    """جلب إعدادات المالك"""
    owner = await db.users.find_one({"role": "super_admin"}, {"_id": 0, "email": 1, "username": 1})
    return owner or {}

@api_router.put("/super-admin/owner-settings")
async def update_owner_settings(
    settings: dict,
    current_user: dict = Depends(verify_super_admin)
):
    """تحديث إعدادات المالك (كلمة المرور والمفتاح السري)"""
    try:
        update_data = {}
        
        if settings.get("password"):
            hashed_password = bcrypt.hashpw(settings["password"].encode('utf-8'), bcrypt.gensalt()).decode('utf-8')
            update_data["password"] = hashed_password
        
        if settings.get("secret_key"):
            update_data["secret_key"] = settings["secret_key"]
        
        if not update_data:
            raise HTTPException(status_code=400, detail="لم يتم تقديم أي بيانات للتحديث")
        
        result = await db.users.update_one(
            {"role": "super_admin"},
            {"$set": update_data}
        )
        
        if result.matched_count == 0:
            raise HTTPException(status_code=404, detail="لم يتم العثور على حساب المالك")
        
        return {"message": "تم تحديث إعدادات المالك بنجاح"}
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error updating owner settings: {str(e)}")
        raise HTTPException(status_code=500, detail=f"خطأ في التحديث: {str(e)}")

@api_router.get("/super-admin/stats")
async def get_super_admin_stats(current_user: dict = Depends(verify_super_admin)):
    """إحصائيات شاملة للـ Super Admin"""
    
    # إجمالي العملاء (بدون الحسابات التجريبية)
    total_tenants = await db.tenants.count_documents({"is_demo": {"$ne": True}})
    active_tenants = await db.tenants.count_documents({"is_active": True, "is_demo": {"$ne": True}})
    
    # عدد الحسابات التجريبية
    demo_tenants = await db.tenants.count_documents({"$or": [{"is_demo": True}, {"subscription_type": "demo"}]})
    
    # استبعاد مستخدمي النظام الرئيسي (super_admin و admin و default) والحسابات التجريبية
    # جلب IDs الحسابات التجريبية
    demo_tenant_ids = await db.tenants.find(
        {"$or": [{"is_demo": True}, {"subscription_type": "demo"}]},
        {"id": 1}
    ).to_list(100)
    demo_ids = [t["id"] for t in demo_tenant_ids]
    
    total_users = await db.users.count_documents({
        "role": {"$nin": [UserRole.SUPER_ADMIN, UserRole.ADMIN]},
        "tenant_id": {"$exists": True, "$ne": None, "$ne": "default", "$nin": demo_ids}
    })
    
    # حساب الطلبات فقط للعملاء الفعليين (استبعاد النظام الرئيسي والتجريبي)
    total_orders = await db.orders.count_documents({
        "tenant_id": {"$exists": True, "$ne": None, "$ne": "default", "$nin": demo_ids}
    })
    
    # المبيعات الإجمالية - فقط للعملاء الفعليين (استبعاد النظام الرئيسي والتجريبي)
    sales_cursor = db.orders.aggregate([
        {"$match": {
            "status": {"$ne": "cancelled"}, 
            "tenant_id": {"$exists": True, "$ne": None, "$ne": "default", "$nin": demo_ids}
        }},
        {"$group": {"_id": None, "total": {"$sum": "$total"}}}
    ])
    sales_result = await sales_cursor.to_list(1)
    total_sales = sales_result[0]["total"] if sales_result else 0
    
    # المستأجرين حسب نوع الاشتراك (بدون التجريبي)
    subscription_stats = await db.tenants.aggregate([
        {"$match": {"is_demo": {"$ne": True}}},
        {"$group": {"_id": "$subscription_type", "count": {"$sum": 1}}}
    ]).to_list(10)
    
    # أحدث المستأجرين
    recent_tenants = await db.tenants.find({}, {"_id": 0}).sort("created_at", -1).limit(5).to_list(5)
    
    return {
        "total_tenants": total_tenants,
        "active_tenants": active_tenants,
        "inactive_tenants": total_tenants - active_tenants,
        "demo_tenants": demo_tenants,
        "total_users": total_users,
        "total_orders": total_orders,
        "total_sales": total_sales,
        "subscription_stats": {item["_id"]: item["count"] for item in subscription_stats},
        "recent_tenants": recent_tenants
    }

# ==================== نقاط نهاية الإشعارات ====================

@api_router.get("/super-admin/notifications")
async def get_notifications(
    current_user: dict = Depends(verify_super_admin),
    unread_only: bool = False,
    limit: int = 50
):
    """جلب إشعارات المالك"""
    query = {}
    if unread_only:
        query["is_read"] = False
    
    notifications = await db.notifications.find(
        query, {"_id": 0}
    ).sort("created_at", -1).limit(limit).to_list(limit)
    
    # عدد الإشعارات غير المقروءة
    unread_count = await db.notifications.count_documents({"is_read": False})
    
    return {
        "notifications": notifications,
        "unread_count": unread_count
    }

@api_router.put("/super-admin/notifications/{notification_id}/read")
async def mark_notification_read(notification_id: str, current_user: dict = Depends(verify_super_admin)):
    """تعليم إشعار كمقروء"""
    result = await db.notifications.update_one(
        {"id": notification_id},
        {"$set": {"is_read": True}}
    )
    if result.modified_count == 0:
        raise HTTPException(status_code=404, detail="الإشعار غير موجود")
    return {"message": "تم تعليم الإشعار كمقروء"}

@api_router.put("/super-admin/notifications/read-all")
async def mark_all_notifications_read(current_user: dict = Depends(verify_super_admin)):
    """تعليم جميع الإشعارات كمقروءة"""
    await db.notifications.update_many(
        {"is_read": False},
        {"$set": {"is_read": True}}
    )
    return {"message": "تم تعليم جميع الإشعارات كمقروءة"}

@api_router.delete("/super-admin/notifications/{notification_id}")
async def delete_notification(notification_id: str, current_user: dict = Depends(verify_super_admin)):
    """حذف إشعار"""
    result = await db.notifications.delete_one({"id": notification_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="الإشعار غير موجود")
    return {"message": "تم حذف الإشعار"}

@api_router.delete("/super-admin/notifications")
async def clear_all_notifications(current_user: dict = Depends(verify_super_admin)):
    """حذف جميع الإشعارات"""
    await db.notifications.delete_many({})
    return {"message": "تم حذف جميع الإشعارات"}

# ==================== إعدادات الإشعارات ====================

@api_router.get("/super-admin/notification-settings")
async def get_notification_settings(current_user: dict = Depends(verify_super_admin)):
    """جلب إعدادات الإشعارات"""
    settings = await db.settings.find_one({"type": "notification_settings"}, {"_id": 0})
    
    if not settings:
        # إعدادات افتراضية
        default_settings = {
            "type": "notification_settings",
            "value": {
                "days_before_expiry": 15,
                "email_notifications": False,
                "push_notifications": True,
                "notify_new_tenant": True,
                "notify_tenant_status": True
            }
        }
        await db.settings.insert_one(default_settings)
        return default_settings["value"]
    
    return settings.get("value", {})

@api_router.put("/super-admin/notification-settings")
async def update_notification_settings(settings: NotificationSettings, current_user: dict = Depends(verify_super_admin)):
    """تحديث إعدادات الإشعارات"""
    await db.settings.update_one(
        {"type": "notification_settings"},
        {"$set": {"value": settings.model_dump()}},
        upsert=True
    )
    return {"message": "تم حفظ إعدادات الإشعارات", "settings": settings.model_dump()}

# ==================== فحص الاشتراكات المنتهية ====================

@api_router.get("/super-admin/expiring-subscriptions")
async def get_expiring_subscriptions(current_user: dict = Depends(verify_super_admin)):
    """جلب الاشتراكات القريبة من الانتهاء"""
    
    # جلب إعدادات الإشعارات
    settings = await db.settings.find_one({"type": "notification_settings"}, {"_id": 0})
    days_before = 15
    if settings and settings.get("value"):
        days_before = settings["value"].get("days_before_expiry", 15)
    
    # حساب التاريخ المستهدف
    target_date = (datetime.now(timezone.utc) + timedelta(days=days_before)).isoformat()
    now = datetime.now(timezone.utc).isoformat()
    
    # جلب الاشتراكات القريبة من الانتهاء
    expiring = await db.tenants.find({
        "is_active": True,
        "expires_at": {"$lte": target_date, "$gte": now}
    }, {"_id": 0}).to_list(100)
    
    # جلب الاشتراكات المنتهية بالفعل
    expired = await db.tenants.find({
        "is_active": True,
        "expires_at": {"$lt": now}
    }, {"_id": 0}).to_list(100)
    
    return {
        "expiring_soon": expiring,
        "already_expired": expired,
        "days_before_alert": days_before
    }

# ==================== لوحة معلومات الاشتراكات ====================

@api_router.get("/super-admin/subscriptions-dashboard")
async def get_subscriptions_dashboard(current_user: dict = Depends(verify_super_admin)):
    """لوحة معلومات شاملة للاشتراكات"""
    
    now = datetime.now(timezone.utc)
    
    # جلب إعدادات التنبيه
    settings = await db.settings.find_one({"type": "notification_settings"}, {"_id": 0})
    days_before = 15
    if settings and settings.get("value"):
        days_before = settings["value"].get("days_before_expiry", 15)
    
    # حساب التواريخ
    target_date = (now + timedelta(days=days_before)).isoformat()
    now_iso = now.isoformat()
    
    # جلب جميع العملاء (غير التجريبية)
    all_tenants = await db.tenants.find(
        {"is_demo": {"$ne": True}},
        {"_id": 0}
    ).to_list(1000)
    
    # تصنيف الاشتراكات
    active_subscriptions = []
    expiring_soon = []
    already_expired = []
    
    for tenant in all_tenants:
        expires_at = tenant.get("expires_at")
        if expires_at:
            if expires_at < now_iso:
                already_expired.append(tenant)
            elif expires_at <= target_date:
                expiring_soon.append(tenant)
                if tenant.get("is_active"):
                    active_subscriptions.append(tenant)
            else:
                if tenant.get("is_active"):
                    active_subscriptions.append(tenant)
        else:
            if tenant.get("is_active"):
                active_subscriptions.append(tenant)
    
    # إحصائيات حسب نوع الاشتراك
    subscription_types = {}
    for tenant in all_tenants:
        sub_type = tenant.get("subscription_type", "unknown")
        if sub_type not in subscription_types:
            subscription_types[sub_type] = {"count": 0, "active": 0, "expired": 0}
        subscription_types[sub_type]["count"] += 1
        if tenant.get("is_active") and tenant not in already_expired:
            subscription_types[sub_type]["active"] += 1
        elif tenant in already_expired:
            subscription_types[sub_type]["expired"] += 1
    
    # جلب أسعار الاشتراكات من قاعدة البيانات
    prices_doc = await db.settings.find_one({"type": "subscription_prices"}, {"_id": 0})
    
    # الأسعار الافتراضية بالدولار
    default_prices = {
        "bronze": {"monthly": 15, "name": "برونزية"},
        "silver": {"monthly": 30, "name": "فضية"},
        "gold": {"monthly": 50, "name": "ذهبية"},
        "basic": {"monthly": 25, "name": "أساسي"},
        "premium": {"monthly": 50, "name": "مميز"},
        "trial": {"monthly": 0, "name": "تجريبي"},
        "demo": {"monthly": 0, "name": "عرض"}
    }
    
    if prices_doc and prices_doc.get("value"):
        subscription_prices = prices_doc["value"]
    else:
        subscription_prices = default_prices
    
    expected_revenue = {
        "from_expiring": 0,
        "from_active": 0,
        "total_monthly": 0,
        "currency": "USD",
        "details": []
    }
    
    for tenant in expiring_soon:
        sub_type = tenant.get("subscription_type", "basic")
        duration = tenant.get("subscription_duration", 1)
        price_per_month = subscription_prices.get(sub_type, {}).get("monthly", 0)
        expected_revenue["from_expiring"] += price_per_month * duration
        expected_revenue["details"].append({
            "tenant_name": tenant.get("name"),
            "subscription_type": sub_type,
            "duration_months": duration,
            "expected_amount": price_per_month * duration,
            "expires_at": tenant.get("expires_at")
        })
    
    for tenant in active_subscriptions:
        sub_type = tenant.get("subscription_type", "basic")
        price_per_month = subscription_prices.get(sub_type, {}).get("monthly", 0)
        expected_revenue["from_active"] += price_per_month
    
    expected_revenue["total_monthly"] = expected_revenue["from_active"]
    
    # ترتيب الاشتراكات القريبة من الانتهاء حسب التاريخ
    expiring_soon.sort(key=lambda x: x.get("expires_at", ""))
    already_expired.sort(key=lambda x: x.get("expires_at", ""), reverse=True)
    
    # حساب عدد الأيام المتبقية لكل اشتراك
    for tenant in expiring_soon:
        if tenant.get("expires_at"):
            try:
                exp_date = datetime.fromisoformat(tenant["expires_at"].replace("Z", "+00:00"))
                days_left = (exp_date - now).days
                tenant["days_left"] = max(0, days_left)
            except:
                tenant["days_left"] = None
    
    for tenant in already_expired:
        if tenant.get("expires_at"):
            try:
                exp_date = datetime.fromisoformat(tenant["expires_at"].replace("Z", "+00:00"))
                days_ago = (now - exp_date).days
                tenant["days_expired"] = days_ago
            except:
                tenant["days_expired"] = None
    
    return {
        "summary": {
            "total_tenants": len(all_tenants),
            "active_subscriptions": len(active_subscriptions),
            "expiring_soon": len(expiring_soon),
            "already_expired": len(already_expired),
            "days_before_alert": days_before
        },
        "subscription_types": subscription_types,
        "expected_revenue": expected_revenue,
        "expiring_soon_list": expiring_soon[:10],  # أول 10
        "expired_list": already_expired[:10],  # أول 10
        "subscription_prices": subscription_prices
    }

# ==================== أسعار الاشتراكات ====================

class SubscriptionPriceUpdate(BaseModel):
    """تحديث سعر اشتراك واحد"""
    subscription_type: str  # basic, premium
    monthly_price: float  # السعر الشهري بالدولار

class SubscriptionPricesUpdate(BaseModel):
    """تحديث جميع أسعار الاشتراكات"""
    bronze: float = 15  # السعر الشهري للبرونزية بالدولار
    silver: float = 30  # السعر الشهري للفضية بالدولار
    gold: float = 50  # السعر الشهري للذهبية بالدولار
    basic: float = 25  # السعر الشهري للأساسي بالدولار
    premium: float = 50  # السعر الشهري للمميز بالدولار

@api_router.get("/super-admin/subscription-prices")
async def get_subscription_prices(current_user: dict = Depends(verify_super_admin)):
    """جلب أسعار الاشتراكات"""
    prices_doc = await db.settings.find_one({"type": "subscription_prices"}, {"_id": 0})
    
    # الأسعار الافتراضية بالدولار
    default_prices = {
        "bronze": {"monthly": 15, "name": "برونزية"},
        "silver": {"monthly": 30, "name": "فضية"},
        "gold": {"monthly": 50, "name": "ذهبية"},
        "basic": {"monthly": 25, "name": "أساسي"},
        "premium": {"monthly": 50, "name": "مميز"},
        "trial": {"monthly": 0, "name": "تجريبي"},
        "demo": {"monthly": 0, "name": "عرض"}
    }
    
    if prices_doc and prices_doc.get("value"):
        return {
            "prices": prices_doc["value"],
            "currency": "USD"
        }
    
    return {
        "prices": default_prices,
        "currency": "USD"
    }

@api_router.put("/super-admin/subscription-prices")
async def update_subscription_prices(prices: SubscriptionPricesUpdate, current_user: dict = Depends(verify_super_admin)):
    """تحديث أسعار الاشتراكات بالدولار"""
    
    new_prices = {
        "bronze": {"monthly": prices.bronze, "name": "برونزية"},
        "silver": {"monthly": prices.silver, "name": "فضية"},
        "gold": {"monthly": prices.gold, "name": "ذهبية"},
        "basic": {"monthly": prices.basic, "name": "أساسي"},
        "premium": {"monthly": prices.premium, "name": "مميز"},
        "trial": {"monthly": 0, "name": "تجريبي"},
        "demo": {"monthly": 0, "name": "عرض"}
    }
    
    await db.settings.update_one(
        {"type": "subscription_prices"},
        {"$set": {"value": new_prices, "updated_at": datetime.now(timezone.utc).isoformat()}},
        upsert=True
    )
    
    return {
        "message": "تم تحديث أسعار الاشتراكات",
        "prices": new_prices,
        "currency": "USD"
    }

@api_router.post("/super-admin/impersonate/{tenant_id}")
async def impersonate_tenant(tenant_id: str, current_user: dict = Depends(verify_super_admin)):
    """الدخول كعميل - للمشاهدة والتحكم المباشر"""
    
    # التحقق إذا كان النظام الرئيسي
    if tenant_id == "main-system":
        # البحث عن admin النظام الرئيسي
        admin = await db.users.find_one({
            "$or": [{"tenant_id": {"$exists": False}}, {"tenant_id": None}],
            "role": UserRole.ADMIN
        })
        if not admin:
            raise HTTPException(status_code=404, detail="مدير النظام الرئيسي غير موجود")
        
        # إنشاء token للدخول
        token = create_token(admin["id"], admin["role"], admin.get("branch_id"))
        
        return {
            "token": token,
            "user": {
                "id": admin["id"],
                "email": admin["email"],
                "full_name": admin.get("full_name") or admin.get("name", ""),
                "role": admin["role"],
                "tenant_id": None
            },
            "tenant": {
                "id": "main-system",
                "name": "🏠 النظام الرئيسي",
                "is_main_system": True
            },
            "impersonated": True,
            "original_super_admin": current_user["id"]
        }
    
    # للعملاء العاديين
    tenant = await db.tenants.find_one({"id": tenant_id}, {"_id": 0})
    if not tenant:
        raise HTTPException(status_code=404, detail="العميل غير موجود")
    
    # البحث عن admin العميل
    admin = await db.users.find_one({"tenant_id": tenant_id, "role": UserRole.ADMIN})
    if not admin:
        raise HTTPException(status_code=404, detail="مدير العميل غير موجود")
    
    # إنشاء token للدخول كالعميل مع علامة impersonation
    token = create_token(admin["id"], admin["role"], admin.get("branch_id"))
    
    return {
        "token": token,
        "user": {
            "id": admin["id"],
            "email": admin["email"],
            "full_name": admin.get("full_name") or admin.get("name", ""),
            "role": admin["role"],
            "tenant_id": tenant_id
        },
        "tenant": tenant,
        "impersonated": True,
        "original_super_admin": current_user["id"]
    }

@api_router.get("/super-admin/tenants/{tenant_id}/live-stats")
async def get_tenant_live_stats(tenant_id: str, current_user: dict = Depends(verify_super_admin)):
    """إحصائيات حية للعميل"""
    
    # التحقق إذا كان النظام الرئيسي
    if tenant_id == "main-system":
        tenant = {
            "id": "main-system",
            "name": "🏠 النظام الرئيسي",
            "is_main_system": True
        }
        tenant_query = {"$or": [{"tenant_id": {"$exists": False}}, {"tenant_id": None}]}
    else:
        tenant = await db.tenants.find_one({"id": tenant_id}, {"_id": 0})
        if not tenant:
            raise HTTPException(status_code=404, detail="العميل غير موجود")
        tenant_query = {"tenant_id": tenant_id}
    
    # إحصائيات اليوم
    today = datetime.now(timezone.utc).replace(hour=0, minute=0, second=0, microsecond=0).isoformat()
    
    today_query = {**tenant_query, "created_at": {"$gte": today}}
    today_orders = await db.orders.find(today_query, {"_id": 0}).to_list(500)
    
    # حساب الإحصائيات
    total_today = sum(o["total"] for o in today_orders if o.get("status") != "cancelled")
    pending_orders = len([o for o in today_orders if o.get("status") == "pending"])
    preparing_orders = len([o for o in today_orders if o.get("status") == "preparing"])
    delivered_orders = len([o for o in today_orders if o.get("status") == "delivered"])
    cancelled_orders = len([o for o in today_orders if o.get("status") == "cancelled"])
    
    # المنتجات الأكثر مبيعاً اليوم
    product_sales = {}
    for order in today_orders:
        if order.get("status") != "cancelled":
            for item in order.get("items", []):
                name = item.get("name", "Unknown")
                if name not in product_sales:
                    product_sales[name] = {"quantity": 0, "total": 0}
                product_sales[name]["quantity"] += item.get("quantity", 0)
                product_sales[name]["total"] += item.get("subtotal", 0)
    
    top_products = sorted(product_sales.items(), key=lambda x: x[1]["total"], reverse=True)[:5]
    
    # المستخدمين النشطين (لديهم وردية مفتوحة)
    active_shifts = await db.shifts.count_documents({
        "status": "open"
    })
    
    return {
        "tenant": tenant,
        "today": {
            "total_sales": total_today,
            "total_orders": len(today_orders),
            "pending_orders": pending_orders,
            "preparing_orders": preparing_orders,
            "delivered_orders": delivered_orders,
            "cancelled_orders": cancelled_orders
        },
        "top_products": [{"name": p[0], **p[1]} for p in top_products],
        "active_shifts": active_shifts,
        "recent_orders": today_orders[:10]  # آخر 10 طلبات
    }

@api_router.get("/super-admin/tenants/{tenant_id}/orders")
async def get_tenant_orders(tenant_id: str, date: Optional[str] = None, status: Optional[str] = None, current_user: dict = Depends(verify_super_admin)):
    """جلب طلبات عميل معين"""
    
    query = {"tenant_id": tenant_id}
    if date:
        query["created_at"] = {"$regex": f"^{date}"}
    if status:
        query["status"] = status
    
    orders = await db.orders.find(query, {"_id": 0}).sort("created_at", -1).limit(100).to_list(100)
    return orders

@api_router.get("/super-admin/tenants/{tenant_id}/products")
async def get_tenant_products(tenant_id: str, current_user: dict = Depends(verify_super_admin)):
    """جلب منتجات عميل معين"""
    
    products = await db.products.find({"tenant_id": tenant_id}, {"_id": 0}).to_list(500)
    return products

@api_router.delete("/super-admin/tenants/{tenant_id}/permanent")
async def permanently_delete_tenant(tenant_id: str, confirm: bool = False, current_user: dict = Depends(verify_super_admin)):
    """حذف عميل نهائياً مع جميع بياناته"""
    
    if not confirm:
        raise HTTPException(status_code=400, detail="يجب تأكيد الحذف بإرسال confirm=true")
    
    tenant = await db.tenants.find_one({"id": tenant_id})
    if not tenant:
        raise HTTPException(status_code=404, detail="العميل غير موجود")
    
    # حذف جميع البيانات المرتبطة
    await db.users.delete_many({"tenant_id": tenant_id})
    await db.branches.delete_many({"tenant_id": tenant_id})
    await db.orders.delete_many({"tenant_id": tenant_id})
    await db.products.delete_many({"tenant_id": tenant_id})
    await db.categories.delete_many({"tenant_id": tenant_id})
    await db.inventory.delete_many({"tenant_id": tenant_id})
    await db.customers.delete_many({"tenant_id": tenant_id})
    await db.shifts.delete_many({"tenant_id": tenant_id})
    await db.expenses.delete_many({"tenant_id": tenant_id})
    await db.drivers.delete_many({"tenant_id": tenant_id})
    await db.tenants.delete_one({"id": tenant_id})
    
    return {"message": f"تم حذف العميل '{tenant['name']}' وجميع بياناته نهائياً"}

@api_router.post("/super-admin/reset-sales")
async def reset_all_sales(confirm: bool = False, current_user: dict = Depends(verify_super_admin)):
    """تصفير جميع المبيعات والطلبات - للتجربة"""
    
    if not confirm:
        raise HTTPException(status_code=400, detail="يجب تأكيد التصفير بإرسال confirm=true")
    
    # حذف جميع الطلبات
    orders_result = await db.orders.delete_many({})
    
    # حذف جميع الورديات
    shifts_result = await db.shifts.delete_many({})
    
    # إعادة تعيين إحصائيات العملاء
    await db.customers.update_many({}, {"$set": {
        "total_orders": 0,
        "total_spent": 0.0,
        "last_order_date": None
    }})
    
    # إعادة تعيين إحصائيات السائقين
    await db.drivers.update_many({}, {"$set": {
        "total_deliveries": 0,
        "is_available": True,
        "current_order_id": None
    }})
    
    return {
        "message": "تم تصفير جميع المبيعات بنجاح",
        "deleted_orders": orders_result.deleted_count,
        "deleted_shifts": shifts_result.deleted_count
    }

@api_router.post("/super-admin/tenants/{tenant_id}/reset-sales")
async def reset_tenant_sales(tenant_id: str, confirm: bool = False, current_user: dict = Depends(verify_super_admin)):
    """تصفير مبيعات عميل معين - للتجربة"""
    
    if not confirm:
        raise HTTPException(status_code=400, detail="يجب تأكيد التصفير بإرسال confirm=true")
    
    # التحقق إذا كان النظام الرئيسي
    if tenant_id == "main-system":
        # حذف طلبات النظام الرئيسي (بدون tenant_id)
        orders_result = await db.orders.delete_many({
            "$or": [{"tenant_id": {"$exists": False}}, {"tenant_id": None}]
        })
        
        # حذف ورديات النظام الرئيسي
        shifts_result = await db.shifts.delete_many({
            "$or": [{"tenant_id": {"$exists": False}}, {"tenant_id": None}]
        })
        
        # إعادة تعيين إحصائيات عملاء النظام الرئيسي
        await db.customers.update_many({
            "$or": [{"tenant_id": {"$exists": False}}, {"tenant_id": None}]
        }, {"$set": {
            "total_orders": 0,
            "total_spent": 0.0,
            "last_order_date": None
        }})
        
        return {
            "message": "تم تصفير مبيعات النظام الرئيسي بنجاح",
            "deleted_orders": orders_result.deleted_count,
            "deleted_shifts": shifts_result.deleted_count
        }
    
    # للعملاء العاديين
    tenant = await db.tenants.find_one({"id": tenant_id})
    if not tenant:
        raise HTTPException(status_code=404, detail="العميل غير موجود")
    
    # حذف طلبات العميل
    orders_result = await db.orders.delete_many({"tenant_id": tenant_id})
    
    # حذف ورديات العميل
    shifts_result = await db.shifts.delete_many({"tenant_id": tenant_id})
    
    # إعادة تعيين إحصائيات عملاء هذا العميل
    await db.customers.update_many({"tenant_id": tenant_id}, {"$set": {
        "total_orders": 0,
        "total_spent": 0.0,
        "last_order_date": None
    }})
    
    return {
        "message": f"تم تصفير مبيعات '{tenant['name']}' بنجاح",
        "deleted_orders": orders_result.deleted_count,
        "deleted_shifts": shifts_result.deleted_count
    }

@api_router.post("/super-admin/tenants/{tenant_id}/reset-inventory")
async def reset_tenant_inventory(tenant_id: str, confirm: bool = False, current_user: dict = Depends(verify_super_admin)):
    """تصفير بيانات المخزون والمشتريات لعميل معين - للمالك فقط"""
    
    if not confirm:
        raise HTTPException(status_code=400, detail="يجب تأكيد التصفير بإرسال confirm=true")
    
    results = {
        "reset_counts": {},
        "tenant_name": ""
    }
    
    # التحقق إذا كان النظام الرئيسي
    if tenant_id == "main-system":
        query = {"$or": [{"tenant_id": {"$exists": False}}, {"tenant_id": None}]}
        results["tenant_name"] = "النظام الرئيسي"
    else:
        tenant = await db.tenants.find_one({"id": tenant_id})
        if not tenant:
            raise HTTPException(status_code=404, detail="العميل غير موجود")
        query = {"tenant_id": tenant_id}
        results["tenant_name"] = tenant.get("name", tenant_id)
    
    # حذف طلبات الفروع
    deleted = await db.branch_orders_new.delete_many(query)
    results["reset_counts"]["branch_orders"] = deleted.deleted_count
    
    # حذف فواتير الشراء
    deleted_purchases = await db.purchases_new.delete_many(query)
    results["reset_counts"]["purchases"] = deleted_purchases.deleted_count
    
    # حذف طلبات الشراء
    deleted_requests = await db.purchase_requests.delete_many(query)
    results["reset_counts"]["purchase_requests"] = deleted_requests.deleted_count
    
    # حذف سجلات التصنيع
    deleted_mfg = await db.manufacturing_records.delete_many(query)
    results["reset_counts"]["manufacturing_records"] = deleted_mfg.deleted_count
    
    # حذف حركات المخزون
    deleted_movements = await db.inventory_movements.delete_many(query)
    results["reset_counts"]["inventory_movements"] = deleted_movements.deleted_count
    
    # تصفير كميات المواد الخام (دون حذف المواد نفسها)
    updated_raw = await db.raw_materials.update_many(
        query,
        {"$set": {"quantity": 0}}
    )
    results["reset_counts"]["raw_materials_qty_reset"] = updated_raw.modified_count
    
    # تصفير مخزون التصنيع
    updated_mfg_inv = await db.manufacturing_inventory.update_many(
        query,
        {"$set": {"quantity": 0}}
    )
    results["reset_counts"]["manufacturing_inventory_reset"] = updated_mfg_inv.modified_count
    
    # تصفير كميات المنتجات المصنعة
    updated_products = await db.manufactured_products.update_many(
        query,
        {"$set": {"quantity": 0}}
    )
    results["reset_counts"]["manufactured_products_qty_reset"] = updated_products.modified_count
    
    # حذف مخزون الفروع
    deleted_branch_inv = await db.branch_inventory.delete_many(query)
    results["reset_counts"]["branch_inventory"] = deleted_branch_inv.deleted_count
    
    return {
        "message": f"تم تصفير بيانات المخزون لـ '{results['tenant_name']}' بنجاح",
        "success": True,
        **results
    }

# ==================== SUPPLIERS & PURCHASING ====================

class SupplierCreate(BaseModel):
    name: str
    phone: str
    email: Optional[str] = None
    address: Optional[str] = None
    notes: Optional[str] = None
    contact_person: Optional[str] = None
    payment_terms: str = "cash"

class SupplierUpdate(BaseModel):
    name: Optional[str] = None
    phone: Optional[str] = None
    email: Optional[str] = None
    address: Optional[str] = None
    notes: Optional[str] = None
    contact_person: Optional[str] = None
    payment_terms: Optional[str] = None
    is_active: Optional[bool] = None

class PurchaseOrderCreate(BaseModel):
    supplier_id: str
    items: List[Dict[str, Any]]
    expected_delivery: Optional[str] = None
    notes: Optional[str] = None
    branch_id: Optional[str] = None

class PurchaseOrderStatusUpdate(BaseModel):
    status: str
    notes: Optional[str] = None

class RawMaterialCreate(BaseModel):
    name: str
    name_en: Optional[str] = None
    unit: str
    category: Optional[str] = None
    min_stock: float = 0
    current_stock: float = 0
    price: float = 0
    supplier_id: Optional[str] = None
    branch_id: Optional[str] = None

@api_router.get("/suppliers")
async def get_suppliers(current_user: dict = Depends(get_current_user)):
    """جلب قائمة الموردين"""
    query = build_tenant_query(current_user)
    suppliers = await db.suppliers.find(query, {"_id": 0}).sort("created_at", -1).to_list(500)
    return suppliers

@api_router.post("/suppliers")
async def create_supplier(supplier: SupplierCreate, current_user: dict = Depends(get_current_user)):
    """إضافة مورد جديد"""
    tenant_id = get_user_tenant_id(current_user)
    
    supplier_doc = {
        "id": str(uuid.uuid4()),
        **supplier.model_dump(),
        "tenant_id": tenant_id,
        "is_active": True,
        "total_orders": 0,
        "total_amount": 0,
        "created_at": datetime.now(timezone.utc).isoformat(),
        "created_by": current_user["id"]
    }
    
    await db.suppliers.insert_one(supplier_doc)
    del supplier_doc["_id"]
    return supplier_doc

@api_router.put("/suppliers/{supplier_id}")
async def update_supplier(supplier_id: str, update: SupplierUpdate, current_user: dict = Depends(get_current_user)):
    """تحديث بيانات مورد"""
    query = build_tenant_query(current_user, {"id": supplier_id})
    
    supplier = await db.suppliers.find_one(query)
    if not supplier:
        raise HTTPException(status_code=404, detail="المورد غير موجود")
    
    update_data = {k: v for k, v in update.model_dump().items() if v is not None}
    if update_data:
        await db.suppliers.update_one({"id": supplier_id}, {"$set": update_data})
    
    return await db.suppliers.find_one({"id": supplier_id}, {"_id": 0})

@api_router.delete("/suppliers/{supplier_id}")
async def delete_supplier(supplier_id: str, current_user: dict = Depends(get_current_user)):
    """حذف مورد"""
    query = build_tenant_query(current_user, {"id": supplier_id})
    
    supplier = await db.suppliers.find_one(query)
    if not supplier:
        raise HTTPException(status_code=404, detail="المورد غير موجود")
    
    await db.suppliers.update_one({"id": supplier_id}, {"$set": {"is_active": False}})
    return {"message": "تم تعطيل المورد"}

@api_router.get("/purchase-orders")
async def get_purchase_orders(
    status: Optional[str] = None,
    supplier_id: Optional[str] = None,
    branch_id: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    """جلب أوامر الشراء"""
    query = build_tenant_query(current_user)
    
    if status:
        query["status"] = status
    if supplier_id:
        query["supplier_id"] = supplier_id
    if branch_id:
        query["branch_id"] = branch_id
    
    orders = await db.purchase_orders.find(query, {"_id": 0}).sort("created_at", -1).to_list(500)
    
    for order in orders:
        supplier = await db.suppliers.find_one({"id": order.get("supplier_id")}, {"_id": 0, "name": 1})
        order["supplier"] = supplier
    
    return orders

@api_router.post("/purchase-orders")
async def create_purchase_order(order: PurchaseOrderCreate, current_user: dict = Depends(get_current_user)):
    """إنشاء أمر شراء جديد"""
    tenant_id = get_user_tenant_id(current_user)
    
    supplier_query = build_tenant_query(current_user, {"id": order.supplier_id})
    supplier = await db.suppliers.find_one(supplier_query)
    if not supplier:
        raise HTTPException(status_code=404, detail="المورد غير موجود")
    
    total_amount = sum(item.get("total", item.get("quantity", 0) * item.get("unit_price", 0)) for item in order.items)
    
    last_order = await db.purchase_orders.find_one(
        {"tenant_id": tenant_id} if tenant_id else {},
        {"_id": 0, "order_number": 1},
        sort=[("created_at", -1)]
    )
    order_num = 1
    if last_order and last_order.get("order_number"):
        try:
            order_num = int(last_order["order_number"].replace("PO-", "")) + 1
        except:
            order_num = 1
    
    order_doc = {
        "id": str(uuid.uuid4()),
        "order_number": f"PO-{str(order_num).zfill(4)}",
        "supplier_id": order.supplier_id,
        "items": order.items,
        "total_amount": total_amount,
        "status": "pending",
        "expected_delivery": order.expected_delivery,
        "notes": order.notes,
        "branch_id": order.branch_id or current_user.get("branch_id"),
        "tenant_id": tenant_id,
        "created_by": current_user["id"],
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    
    await db.purchase_orders.insert_one(order_doc)
    del order_doc["_id"]
    
    order_doc["supplier"] = {"id": supplier["id"], "name": supplier["name"]}
    return order_doc

@api_router.put("/purchase-orders/{order_id}/status")
async def update_purchase_order_status(order_id: str, update: PurchaseOrderStatusUpdate, current_user: dict = Depends(get_current_user)):
    """تحديث حالة أمر الشراء"""
    query = build_tenant_query(current_user, {"id": order_id})
    
    order = await db.purchase_orders.find_one(query)
    if not order:
        raise HTTPException(status_code=404, detail="أمر الشراء غير موجود")
    
    update_data = {"status": update.status}
    
    if update.status == "approved":
        update_data["approved_by"] = current_user["id"]
        update_data["approved_at"] = datetime.now(timezone.utc).isoformat()
    elif update.status == "delivered":
        update_data["delivered_at"] = datetime.now(timezone.utc).isoformat()
        update_data["received_by"] = current_user["id"]
        
        for item in order.get("items", []):
            material_id = item.get("material_id")
            quantity = item.get("quantity", 0)
            if material_id and quantity > 0:
                await db.raw_materials.update_one(
                    {"id": material_id},
                    {"$inc": {"current_stock": quantity}}
                )
        
        await db.suppliers.update_one(
            {"id": order["supplier_id"]},
            {"$inc": {"total_orders": 1, "total_amount": order.get("total_amount", 0)}}
        )
    
    if update.notes:
        update_data["status_notes"] = update.notes
    
    await db.purchase_orders.update_one({"id": order_id}, {"$set": update_data})
    return await db.purchase_orders.find_one({"id": order_id}, {"_id": 0})

@api_router.delete("/purchase-orders/{order_id}")
async def delete_purchase_order(order_id: str, current_user: dict = Depends(get_current_user)):
    """حذف أمر شراء"""
    query = build_tenant_query(current_user, {"id": order_id})
    
    order = await db.purchase_orders.find_one(query)
    if not order:
        raise HTTPException(status_code=404, detail="أمر الشراء غير موجود")
    
    if order["status"] not in ["pending", "cancelled"]:
        raise HTTPException(status_code=400, detail="لا يمكن حذف أمر شراء تمت معالجته")
    
    await db.purchase_orders.delete_one({"id": order_id})
    return {"message": "تم حذف أمر الشراء"}

@api_router.get("/raw-materials")
async def get_raw_materials(
    branch_id: Optional[str] = None,
    category: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    """جلب المواد الخام من جدول inventory"""
    query = build_tenant_query(current_user, {"item_type": "raw"})
    
    if branch_id:
        query["branch_id"] = branch_id
    if category:
        query["category"] = category
    
    materials = await db.inventory.find(query, {"_id": 0}).sort("name", 1).to_list(500)
    return materials

@api_router.post("/raw-materials")
async def create_raw_material(material: RawMaterialCreate, current_user: dict = Depends(get_current_user)):
    """إضافة مادة خام جديدة"""
    tenant_id = get_user_tenant_id(current_user)
    
    material_doc = {
        "id": str(uuid.uuid4()),
        **material.model_dump(),
        "tenant_id": tenant_id,
        "is_active": True,
        "created_at": datetime.now(timezone.utc).isoformat(),
        "created_by": current_user["id"]
    }
    
    await db.raw_materials.insert_one(material_doc)
    del material_doc["_id"]
    return material_doc

@api_router.put("/raw-materials/{material_id}")
async def update_raw_material(material_id: str, update: dict, current_user: dict = Depends(get_current_user)):
    """تحديث مادة خام"""
    query = build_tenant_query(current_user, {"id": material_id})
    
    material = await db.raw_materials.find_one(query)
    if not material:
        raise HTTPException(status_code=404, detail="المادة غير موجودة")
    
    allowed_fields = ["name", "name_en", "unit", "category", "min_stock", "current_stock", "price", "supplier_id", "is_active"]
    update_data = {k: v for k, v in update.items() if k in allowed_fields and v is not None}
    
    if update_data:
        await db.raw_materials.update_one({"id": material_id}, {"$set": update_data})
    
    return await db.raw_materials.find_one({"id": material_id}, {"_id": 0})

@api_router.get("/inventory/low-stock-alerts")
async def get_low_stock_alerts(current_user: dict = Depends(get_current_user)):
    """جلب تنبيهات انخفاض المخزون"""
    query = build_tenant_query(current_user)
    
    materials = await db.raw_materials.find(query, {"_id": 0}).to_list(500)
    
    alerts = []
    for material in materials:
        current_stock = material.get("current_stock", 0)
        min_stock = material.get("min_stock", 0)
        
        if current_stock < min_stock:
            alerts.append({
                "id": material["id"],
                "material_name": material["name"],
                "current_stock": current_stock,
                "min_stock": min_stock,
                "unit": material.get("unit", ""),
                "shortage": min_stock - current_stock,
                "price": material.get("price", 0)
            })
    
    alerts.sort(key=lambda x: x["shortage"], reverse=True)
    return alerts

# ==================== FINISHED PRODUCTS - المنتجات النهائية ====================

class FinishedProductCreate(BaseModel):
    """إنشاء منتج نهائي مع وصفته"""
    name: str
    name_en: Optional[str] = None
    unit: str = "قطعة"
    quantity: float = 0.0  # الكمية المتوفرة
    min_quantity: float = 0.0  # الحد الأدنى للتنبيه
    cost_per_unit: float = 0.0  # سيتم حسابها تلقائياً من الوصفة
    selling_price: float = 0.0  # سعر البيع
    recipe: List[Dict[str, Any]] = []  # [{raw_material_id, quantity}]
    description: Optional[str] = None
    category: str = "general"

class FinishedProductUpdate(BaseModel):
    """تحديث منتج نهائي"""
    name: Optional[str] = None
    name_en: Optional[str] = None
    unit: Optional[str] = None
    quantity: Optional[float] = None
    min_quantity: Optional[float] = None
    selling_price: Optional[float] = None
    recipe: Optional[List[Dict[str, Any]]] = None
    description: Optional[str] = None
    category: Optional[str] = None
    is_active: Optional[bool] = None

@api_router.post("/finished-products")
async def create_finished_product(product: FinishedProductCreate, current_user: dict = Depends(get_current_user)):
    """إنشاء منتج نهائي جديد مع وصفته (المواد الخام المكونة له)"""
    if current_user["role"] not in [UserRole.ADMIN, UserRole.MANAGER, UserRole.SUPER_ADMIN]:
        raise HTTPException(status_code=403, detail="غير مصرح")
    
    tenant_id = get_user_tenant_id(current_user)
    
    # حساب تكلفة الوحدة من الوصفة
    recipe_cost = 0.0
    recipe_details = []
    
    for ingredient in product.recipe:
        raw_material_id = ingredient.get("raw_material_id")
        qty = ingredient.get("quantity", 0)
        
        # جلب المادة الخام من المخزون
        raw_material = await db.inventory.find_one(
            {"id": raw_material_id, "item_type": "raw"},
            {"_id": 0}
        )
        
        if raw_material:
            ingredient_cost = qty * raw_material.get("cost_per_unit", 0)
            recipe_cost += ingredient_cost
            recipe_details.append({
                "raw_material_id": raw_material_id,
                "raw_material_name": raw_material.get("name", ""),
                "quantity": qty,
                "unit": raw_material.get("unit", ""),
                "cost_per_unit": raw_material.get("cost_per_unit", 0),
                "total_cost": ingredient_cost
            })
    
    product_doc = {
        "id": str(uuid.uuid4()),
        "name": product.name,
        "name_en": product.name_en,
        "unit": product.unit,
        "quantity": product.quantity,
        "min_quantity": product.min_quantity,
        "cost_per_unit": recipe_cost,  # التكلفة المحسوبة من الوصفة
        "selling_price": product.selling_price,
        "recipe": recipe_details,
        "description": product.description,
        "category": product.category,
        "item_type": "finished",
        "tenant_id": tenant_id,
        "branch_id": "main",  # المنتجات النهائية في المخزن الرئيسي
        "is_active": True,
        "created_by": current_user["id"],
        "created_at": datetime.now(timezone.utc).isoformat(),
        "last_updated": datetime.now(timezone.utc).isoformat()
    }
    
    await db.inventory.insert_one(product_doc)
    del product_doc["_id"]
    return product_doc

@api_router.get("/finished-products")
async def get_finished_products(
    category: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    """جلب جميع المنتجات النهائية"""
    query = build_tenant_query(current_user, {"item_type": "finished"})
    
    if category:
        query["category"] = category
    
    products = await db.inventory.find(query, {"_id": 0}).to_list(500)
    return products

@api_router.get("/finished-products/{product_id}")
async def get_finished_product(product_id: str, current_user: dict = Depends(get_current_user)):
    """جلب منتج نهائي محدد مع وصفته"""
    query = build_tenant_query(current_user, {"id": product_id, "item_type": "finished"})
    product = await db.inventory.find_one(query, {"_id": 0})
    
    if not product:
        raise HTTPException(status_code=404, detail="المنتج غير موجود")
    
    return product

@api_router.put("/finished-products/{product_id}")
async def update_finished_product(
    product_id: str, 
    update: FinishedProductUpdate, 
    current_user: dict = Depends(get_current_user)
):
    """تحديث منتج نهائي ووصفته"""
    if current_user["role"] not in [UserRole.ADMIN, UserRole.MANAGER, UserRole.SUPER_ADMIN]:
        raise HTTPException(status_code=403, detail="غير مصرح")
    
    query = build_tenant_query(current_user, {"id": product_id, "item_type": "finished"})
    product = await db.inventory.find_one(query)
    
    if not product:
        raise HTTPException(status_code=404, detail="المنتج غير موجود")
    
    update_data = {k: v for k, v in update.model_dump().items() if v is not None}
    
    # إذا تم تحديث الوصفة، أعد حساب التكلفة
    if "recipe" in update_data and update_data["recipe"]:
        recipe_cost = 0.0
        recipe_details = []
        
        for ingredient in update_data["recipe"]:
            raw_material_id = ingredient.get("raw_material_id")
            qty = ingredient.get("quantity", 0)
            
            raw_material = await db.inventory.find_one(
                {"id": raw_material_id, "item_type": "raw"},
                {"_id": 0}
            )
            
            if raw_material:
                ingredient_cost = qty * raw_material.get("cost_per_unit", 0)
                recipe_cost += ingredient_cost
                recipe_details.append({
                    "raw_material_id": raw_material_id,
                    "raw_material_name": raw_material.get("name", ""),
                    "quantity": qty,
                    "unit": raw_material.get("unit", ""),
                    "cost_per_unit": raw_material.get("cost_per_unit", 0),
                    "total_cost": ingredient_cost
                })
        
        update_data["recipe"] = recipe_details
        update_data["cost_per_unit"] = recipe_cost
    
    update_data["last_updated"] = datetime.now(timezone.utc).isoformat()
    
    await db.inventory.update_one({"id": product_id}, {"$set": update_data})
    return await db.inventory.find_one({"id": product_id}, {"_id": 0})

@api_router.delete("/finished-products/{product_id}")
async def delete_finished_product(product_id: str, current_user: dict = Depends(get_current_user)):
    """حذف منتج نهائي"""
    if current_user["role"] not in [UserRole.ADMIN, UserRole.SUPER_ADMIN]:
        raise HTTPException(status_code=403, detail="غير مصرح")
    
    query = build_tenant_query(current_user, {"id": product_id, "item_type": "finished"})
    product = await db.inventory.find_one(query)
    
    if not product:
        raise HTTPException(status_code=404, detail="المنتج غير موجود")
    
    await db.inventory.delete_one({"id": product_id})
    return {"message": "تم حذف المنتج"}

@api_router.post("/finished-products/{product_id}/manufacture")
async def manufacture_finished_product(
    product_id: str,
    data: dict = Body(...),
    current_user: dict = Depends(get_current_user)
):
    """تصنيع منتج نهائي (خصم المواد الخام وزيادة كمية المنتج النهائي)"""
    if current_user["role"] not in [UserRole.ADMIN, UserRole.MANAGER, UserRole.SUPER_ADMIN]:
        raise HTTPException(status_code=403, detail="غير مصرح")
    
    quantity_to_manufacture = data.get("quantity", 1)
    
    query = build_tenant_query(current_user, {"id": product_id, "item_type": "finished"})
    product = await db.inventory.find_one(query, {"_id": 0})
    
    if not product:
        raise HTTPException(status_code=404, detail="المنتج غير موجود")
    
    if not product.get("recipe"):
        raise HTTPException(status_code=400, detail="المنتج ليس له وصفة محددة")
    
    # التحقق من توفر المواد الخام
    insufficient_materials = []
    for ingredient in product["recipe"]:
        raw_material = await db.inventory.find_one(
            {"id": ingredient["raw_material_id"], "item_type": "raw"},
            {"_id": 0}
        )
        
        if not raw_material:
            insufficient_materials.append({
                "name": ingredient.get("raw_material_name", "مادة Unknownة"),
                "required": ingredient["quantity"] * quantity_to_manufacture,
                "available": 0
            })
        else:
            required_qty = ingredient["quantity"] * quantity_to_manufacture
            if raw_material.get("quantity", 0) < required_qty:
                insufficient_materials.append({
                    "name": raw_material["name"],
                    "required": required_qty,
                    "available": raw_material.get("quantity", 0)
                })
    
    if insufficient_materials:
        raise HTTPException(
            status_code=400, 
            detail={
                "message": "المواد الخام غير كافية للتصنيع",
                "insufficient_materials": insufficient_materials
            }
        )
    
    # خصم المواد الخام
    for ingredient in product["recipe"]:
        required_qty = ingredient["quantity"] * quantity_to_manufacture
        await db.inventory.update_one(
            {"id": ingredient["raw_material_id"]},
            {"$inc": {"quantity": -required_qty}}
        )
        
        # تسجيل حركة المخزون
        movement_doc = {
            "id": str(uuid.uuid4()),
            "inventory_id": ingredient["raw_material_id"],
            "transaction_type": "out",
            "quantity": required_qty,
            "notes": f"تصنيع {quantity_to_manufacture} {product['unit']} من {product['name']}",
            "reference_type": "manufacturing",
            "reference_id": product_id,
            "created_by": current_user["id"],
            "created_at": datetime.now(timezone.utc).isoformat()
        }
        await db.inventory_transactions.insert_one(movement_doc)
    
    # زيادة كمية المنتج النهائي
    await db.inventory.update_one(
        {"id": product_id},
        {
            "$inc": {"quantity": quantity_to_manufacture},
            "$set": {"last_updated": datetime.now(timezone.utc).isoformat()}
        }
    )
    
    # تسجيل عملية التصنيع
    manufacturing_doc = {
        "id": str(uuid.uuid4()),
        "product_id": product_id,
        "product_name": product["name"],
        "quantity_manufactured": quantity_to_manufacture,
        "recipe_used": product["recipe"],
        "total_cost": product.get("cost_per_unit", 0) * quantity_to_manufacture,
        "tenant_id": get_user_tenant_id(current_user),
        "created_by": current_user["id"],
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    await db.manufacturing_logs.insert_one(manufacturing_doc)
    
    updated_product = await db.inventory.find_one({"id": product_id}, {"_id": 0})
    return {
        "message": f"تم تصنيع {quantity_to_manufacture} {product['unit']} من {product['name']}",
        "product": updated_product
    }

# ==================== BRANCH ORDERS - طلبات الفروع ====================

class BranchOrderCreate(BaseModel):
    to_branch_id: str
    items: List[Dict[str, Any]]  # [{product_id, quantity}] - منتجات نهائية فقط
    priority: str = "normal"  # low, normal, high
    notes: Optional[str] = None

class BranchOrderStatusUpdate(BaseModel):
    status: str  # pending, approved, in_transit, delivered, rejected
    notes: Optional[str] = None

@api_router.get("/branch-orders")
async def get_branch_orders(
    type: Optional[str] = None,  # outgoing, incoming
    status: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    """جلب طلبات الفروع"""
    query = build_tenant_query(current_user)
    
    user_branch_id = current_user.get("branch_id")
    
    if type == "outgoing" and user_branch_id:
        query["from_branch_id"] = user_branch_id
    elif type == "incoming" and user_branch_id:
        query["to_branch_id"] = user_branch_id
    
    if status:
        query["status"] = status
    
    orders = await db.branch_orders.find(query, {"_id": 0}).sort("created_at", -1).to_list(500)
    
    # جلب أسماء الفروع
    for order in orders:
        from_branch = await db.branches.find_one({"id": order.get("from_branch_id")}, {"_id": 0, "name": 1})
        to_branch = await db.branches.find_one({"id": order.get("to_branch_id")}, {"_id": 0, "name": 1})
        order["from_branch"] = {"id": order.get("from_branch_id"), "name": from_branch.get("name") if from_branch else "المخزن الرئيسي"}
        order["to_branch"] = {"id": order.get("to_branch_id"), "name": to_branch.get("name") if to_branch else "المخزن الرئيسي"}
    
    return orders

@api_router.post("/branch-orders")
async def create_branch_order(order: BranchOrderCreate, current_user: dict = Depends(get_current_user)):
    """
    إنشاء طلب فرع جديد - يخصم المواد الخام تلقائياً من المخزون المركزي
    
    النظام:
    1. الفرع يطلب منتجات نهائية (مثل: برغر لحم)
    2. المنتج النهائي له وصفة (مكونات من المواد الخام)
    3. عند إرسال الطلب، يتم خصم المواد الخام مباشرة من المخزون المركزي
    4. لا يُشترط وجود كمية مسبقة من المنتج النهائي (الخصم من المواد الخام مباشرة)
    """
    tenant_id = get_user_tenant_id(current_user)
    
    # تجهيز تفاصيل الطلب
    order_items_details = []
    raw_materials_to_deduct = {}  # {raw_material_id: total_quantity_needed}
    products_without_recipe = []
    insufficient_materials = []
    
    for item in order.items:
        product_id = item.get("product_id")
        requested_qty = item.get("quantity", 0)
        
        if requested_qty <= 0:
            continue
        
        # جلب المنتج النهائي
        product = await db.inventory.find_one(
            {"id": product_id, "item_type": "finished"},
            {"_id": 0}
        )
        
        if not product:
            continue
        
        # التحقق من وجود وصفة للمنتج
        recipe = product.get("recipe", [])
        if not recipe:
            products_without_recipe.append(product["name"])
            continue
        
        # تجميع المواد الخام المطلوبة من الوصفة
        for ingredient in recipe:
            raw_material_id = ingredient.get("raw_material_id")
            qty_per_unit = ingredient.get("quantity", 0)
            total_needed = qty_per_unit * requested_qty
            
            if raw_material_id in raw_materials_to_deduct:
                raw_materials_to_deduct[raw_material_id]["quantity"] += total_needed
            else:
                raw_materials_to_deduct[raw_material_id] = {
                    "quantity": total_needed,
                    "name": ingredient.get("raw_material_name", ""),
                    "unit": ingredient.get("unit", "")
                }
        
        order_items_details.append({
            "product_id": product_id,
            "product_name": product["name"],
            "quantity": requested_qty,
            "unit": product.get("unit", "قطعة"),
            "cost_per_unit": product.get("cost_per_unit", 0),
            "recipe": recipe
        })
    
    # إذا كانت هناك منتجات بدون وصفة
    if products_without_recipe:
        raise HTTPException(
            status_code=400,
            detail={
                "message": "بعض المنتجات ليس لها وصفة محددة",
                "products_without_recipe": products_without_recipe
            }
        )
    
    # التحقق من توفر المواد الخام
    for raw_material_id, details in raw_materials_to_deduct.items():
        raw_material = await db.inventory.find_one(
            {"id": raw_material_id, "item_type": "raw"},
            {"_id": 0}
        )
        
        if not raw_material:
            insufficient_materials.append({
                "name": details["name"],
                "required": details["quantity"],
                "available": 0
            })
        elif raw_material.get("quantity", 0) < details["quantity"]:
            insufficient_materials.append({
                "name": raw_material["name"],
                "required": details["quantity"],
                "available": raw_material.get("quantity", 0),
                "unit": raw_material.get("unit", "")
            })
    
    # إذا كانت هناك مواد خام غير كافية
    if insufficient_materials:
        raise HTTPException(
            status_code=400,
            detail={
                "message": "المواد الخام غير كافية في المخزون الرئيسي",
                "insufficient_materials": insufficient_materials
            }
        )
    
    if not order_items_details:
        raise HTTPException(status_code=400, detail="لا توجد منتجات صالحة في الطلب")
    
    # إنشاء رقم الطلب
    last_order = await db.branch_orders.find_one(
        {"tenant_id": tenant_id} if tenant_id else {},
        {"_id": 0, "order_number": 1},
        sort=[("created_at", -1)]
    )
    order_num = 1
    if last_order and last_order.get("order_number"):
        try:
            order_num = int(last_order["order_number"].replace("BO-", "")) + 1
        except:
            order_num = 1
    
    order_id = str(uuid.uuid4())
    
    # خصم المواد الخام من المخزون الرئيسي
    deducted_materials = []
    for raw_material_id, details in raw_materials_to_deduct.items():
        await db.inventory.update_one(
            {"id": raw_material_id},
            {"$inc": {"quantity": -details["quantity"]}}
        )
        
        # تسجيل حركة المخزون
        movement_doc = {
            "id": str(uuid.uuid4()),
            "inventory_id": raw_material_id,
            "transaction_type": "out",
            "quantity": details["quantity"],
            "notes": f"طلب فرع - {order_num}",
            "reference_type": "branch_order",
            "reference_id": order_id,
            "created_by": current_user["id"],
            "created_at": datetime.now(timezone.utc).isoformat()
        }
        await db.inventory_transactions.insert_one(movement_doc)
        
        deducted_materials.append({
            "raw_material_id": raw_material_id,
            "raw_material_name": details["name"],
            "quantity_deducted": details["quantity"],
            "unit": details["unit"]
        })
    
    # حساب إجمالي التكلفة
    total_cost = sum(item["quantity"] * item["cost_per_unit"] for item in order_items_details)
    
    order_doc = {
        "id": order_id,
        "order_number": f"BO-{str(order_num).zfill(4)}",
        "from_branch_id": "warehouse",  # من المخزن الرئيسي دائماً
        "to_branch_id": order.to_branch_id,
        "items": order_items_details,
        "raw_materials_deducted": deducted_materials,
        "total_cost": total_cost,
        "status": "pending",
        "priority": order.priority,
        "notes": order.notes,
        "tenant_id": tenant_id,
        "created_by": current_user["id"],
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    
    await db.branch_orders.insert_one(order_doc)
    del order_doc["_id"]
    
    # إضافة معلومات الفروع
    to_branch = await db.branches.find_one({"id": order_doc["to_branch_id"]}, {"_id": 0, "name": 1})
    order_doc["from_branch"] = {"id": "warehouse", "name": "المخزن الرئيسي"}
    order_doc["to_branch"] = {"id": order_doc["to_branch_id"], "name": to_branch.get("name") if to_branch else "الفرع"}
    
    return order_doc

@api_router.put("/branch-orders/{order_id}/status")
async def update_branch_order_status(order_id: str, update: BranchOrderStatusUpdate, current_user: dict = Depends(get_current_user)):
    """تحديث حالة طلب الفرع"""
    query = build_tenant_query(current_user, {"id": order_id})
    
    order = await db.branch_orders.find_one(query)
    if not order:
        raise HTTPException(status_code=404, detail="الطلب غير موجود")
    
    update_data = {"status": update.status}
    
    if update.status == "approved":
        update_data["approved_by"] = current_user["id"]
        update_data["approved_at"] = datetime.now(timezone.utc).isoformat()
    elif update.status == "in_transit":
        update_data["shipped_at"] = datetime.now(timezone.utc).isoformat()
        update_data["shipped_by"] = current_user["id"]
    elif update.status == "delivered":
        update_data["delivered_at"] = datetime.now(timezone.utc).isoformat()
        update_data["received_by"] = current_user["id"]
    elif update.status == "rejected":
        update_data["rejected_at"] = datetime.now(timezone.utc).isoformat()
        update_data["rejected_by"] = current_user["id"]
    
    if update.notes:
        update_data["status_notes"] = update.notes
    
    await db.branch_orders.update_one({"id": order_id}, {"$set": update_data})
    return await db.branch_orders.find_one({"id": order_id}, {"_id": 0})

@api_router.delete("/branch-orders/{order_id}")
async def delete_branch_order(order_id: str, current_user: dict = Depends(get_current_user)):
    """حذف طلب فرع"""
    query = build_tenant_query(current_user, {"id": order_id})
    
    order = await db.branch_orders.find_one(query)
    if not order:
        raise HTTPException(status_code=404, detail="الطلب غير موجود")
    
    if order["status"] not in ["pending", "rejected"]:
        raise HTTPException(status_code=400, detail="لا يمكن حذف طلب تمت معالجته")
    
    await db.branch_orders.delete_one({"id": order_id})
    return {"message": "تم حذف الطلب"}

# ==================== DASHBOARD BACKGROUNDS ====================

@api_router.get("/dashboard-backgrounds")
async def get_dashboard_backgrounds(current_user: dict = Depends(get_current_user)):
    """جلب خلفيات Dashboard المتاحة للعميل"""
    tenant_id = get_user_tenant_id(current_user)
    
    # جلب الخلفيات الافتراضية (متاحة للجميع)
    default_backgrounds = [
        "https://images.unsplash.com/photo-1517248135467-4c7edcad34c4?w=1920",
        "https://images.unsplash.com/photo-1554679665-f5537f187268?w=1920",
        "https://images.unsplash.com/photo-1466978913421-dad2ebd01d17?w=1920",
        "https://images.unsplash.com/photo-1552566626-52f8b828add9?w=1920",
        "https://images.unsplash.com/photo-1559329007-40df8a9345d8?w=1920",
        "https://images.unsplash.com/photo-1514933651103-005eec06c04b?w=1920"
    ]
    
    # جلب الخلفيات المرفوعة من قبل العميل
    tenant_backgrounds = await db.dashboard_backgrounds.find(
        {"tenant_id": tenant_id} if tenant_id else {},
        {"_id": 0}
    ).to_list(50)
    
    # جلب الخلفية المحددة حالياً
    settings = await db.tenant_settings.find_one(
        {"tenant_id": tenant_id} if tenant_id else {"tenant_id": None},
        {"_id": 0, "dashboard_background": 1}
    )
    
    return {
        "backgrounds": default_backgrounds + [bg["url"] for bg in tenant_backgrounds],
        "selected": settings.get("dashboard_background") if settings else None
    }

@api_router.put("/dashboard-backgrounds/select")
async def select_dashboard_background(data: dict, current_user: dict = Depends(get_current_user)):
    """اختيار خلفية Dashboard للعميل"""
    tenant_id = get_user_tenant_id(current_user)
    background_url = data.get("background_url")
    
    # تحديث أو إنشاء إعدادات العميل
    await db.tenant_settings.update_one(
        {"tenant_id": tenant_id} if tenant_id else {"tenant_id": None},
        {
            "$set": {
                "dashboard_background": background_url,
                "updated_at": datetime.now(timezone.utc).isoformat(),
                "updated_by": current_user["id"]
            },
            "$setOnInsert": {
                "tenant_id": tenant_id,
                "created_at": datetime.now(timezone.utc).isoformat()
            }
        },
        upsert=True
    )
    
    return {"message": "تم تحديث الخلفية", "background_url": background_url}

# ==================== RESERVATIONS - الحجوزات ====================

class ReservationCreate(BaseModel):
    customer_name: str
    customer_phone: str
    customer_email: Optional[str] = None
    date: str  # YYYY-MM-DD
    time: str  # HH:MM
    guests: int
    table_id: Optional[str] = None
    notes: Optional[str] = None
    branch_id: Optional[str] = None

class ReservationUpdate(BaseModel):
    customer_name: Optional[str] = None
    customer_phone: Optional[str] = None
    customer_email: Optional[str] = None
    date: Optional[str] = None
    time: Optional[str] = None
    guests: Optional[int] = None
    table_id: Optional[str] = None
    status: Optional[str] = None  # pending, confirmed, cancelled, completed, no_show
    notes: Optional[str] = None

@api_router.get("/reservations")
async def get_reservations(
    date: Optional[str] = None,
    status: Optional[str] = None,
    branch_id: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    """جلب قائمة الحجوزات"""
    query = build_tenant_query(current_user)
    
    if date:
        query["date"] = date
    if status:
        query["status"] = status
    if branch_id:
        query["branch_id"] = branch_id
    
    reservations = await db.reservations.find(query, {"_id": 0}).sort("date", -1).to_list(500)
    return reservations

@api_router.post("/reservations")
async def create_reservation(reservation: ReservationCreate, current_user: dict = Depends(get_current_user)):
    """إنشاء حجز جديد"""
    tenant_id = get_user_tenant_id(current_user)
    
    # إنشاء رقم الحجز
    last_reservation = await db.reservations.find_one(
        {"tenant_id": tenant_id} if tenant_id else {},
        {"_id": 0, "reservation_number": 1},
        sort=[("created_at", -1)]
    )
    res_num = 1
    if last_reservation and last_reservation.get("reservation_number"):
        try:
            res_num = int(last_reservation["reservation_number"].replace("RES-", "")) + 1
        except:
            res_num = 1
    
    reservation_doc = {
        "id": str(uuid.uuid4()),
        "reservation_number": f"RES-{str(res_num).zfill(4)}",
        **reservation.model_dump(),
        "status": "pending",
        "tenant_id": tenant_id,
        "created_by": current_user["id"],
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    
    await db.reservations.insert_one(reservation_doc)
    del reservation_doc["_id"]
    return reservation_doc

@api_router.put("/reservations/{reservation_id}")
async def update_reservation(reservation_id: str, update: ReservationUpdate, current_user: dict = Depends(get_current_user)):
    """تحديث حجز"""
    query = build_tenant_query(current_user, {"id": reservation_id})
    
    reservation = await db.reservations.find_one(query)
    if not reservation:
        raise HTTPException(status_code=404, detail="الحجز غير موجود")
    
    update_data = {k: v for k, v in update.model_dump().items() if v is not None}
    if update_data:
        update_data["updated_at"] = datetime.now(timezone.utc).isoformat()
        await db.reservations.update_one({"id": reservation_id}, {"$set": update_data})
    
    return await db.reservations.find_one({"id": reservation_id}, {"_id": 0})

@api_router.delete("/reservations/{reservation_id}")
async def delete_reservation(reservation_id: str, current_user: dict = Depends(get_current_user)):
    """حذف حجز"""
    query = build_tenant_query(current_user, {"id": reservation_id})
    
    reservation = await db.reservations.find_one(query)
    if not reservation:
        raise HTTPException(status_code=404, detail="الحجز غير موجود")
    
    await db.reservations.delete_one({"id": reservation_id})
    return {"message": "تم حذف الحجز"}

@api_router.get("/reservations/stats")
async def get_reservations_stats(current_user: dict = Depends(get_current_user)):
    """إحصائيات الحجوزات"""
    query = build_tenant_query(current_user)
    today = datetime.now(timezone.utc).strftime("%Y-%m-%d")
    
    total = await db.reservations.count_documents(query)
    today_count = await db.reservations.count_documents({**query, "date": today})
    pending = await db.reservations.count_documents({**query, "status": "pending"})
    confirmed = await db.reservations.count_documents({**query, "status": "confirmed"})
    
    return {
        "total": total,
        "today": today_count,
        "pending": pending,
        "confirmed": confirmed
    }

# ==================== REVIEWS - التقييمات ====================

class ReviewCreate(BaseModel):
    customer_name: str
    customer_phone: Optional[str] = None
    rating: int  # 1-5
    food_rating: Optional[int] = None
    service_rating: Optional[int] = None
    cleanliness_rating: Optional[int] = None
    comment: Optional[str] = None
    order_id: Optional[str] = None
    branch_id: Optional[str] = None

class ReviewResponse(BaseModel):
    response: str

@api_router.get("/reviews")
async def get_reviews(
    rating: Optional[int] = None,
    branch_id: Optional[str] = None,
    responded: Optional[bool] = None,
    current_user: dict = Depends(get_current_user)
):
    """جلب قائمة التقييمات"""
    query = build_tenant_query(current_user)
    
    if rating:
        query["rating"] = rating
    if branch_id:
        query["branch_id"] = branch_id
    if responded is not None:
        if responded:
            query["response"] = {"$ne": None}
        else:
            query["response"] = None
    
    reviews = await db.reviews.find(query, {"_id": 0}).sort("created_at", -1).to_list(500)
    return reviews

@api_router.post("/reviews")
async def create_review(review: ReviewCreate, current_user: dict = Depends(get_current_user)):
    """إضافة تقييم جديد"""
    tenant_id = get_user_tenant_id(current_user)
    
    review_doc = {
        "id": str(uuid.uuid4()),
        **review.model_dump(),
        "response": None,
        "responded_at": None,
        "responded_by": None,
        "tenant_id": tenant_id,
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    
    await db.reviews.insert_one(review_doc)
    del review_doc["_id"]
    return review_doc

@api_router.put("/reviews/{review_id}/respond")
async def respond_to_review(review_id: str, response: ReviewResponse, current_user: dict = Depends(get_current_user)):
    """الرد على تقييم"""
    query = build_tenant_query(current_user, {"id": review_id})
    
    review = await db.reviews.find_one(query)
    if not review:
        raise HTTPException(status_code=404, detail="التقييم غير موجود")
    
    await db.reviews.update_one(
        {"id": review_id},
        {"$set": {
            "response": response.response,
            "responded_at": datetime.now(timezone.utc).isoformat(),
            "responded_by": current_user["id"]
        }}
    )
    
    return await db.reviews.find_one({"id": review_id}, {"_id": 0})

@api_router.delete("/reviews/{review_id}")
async def delete_review(review_id: str, current_user: dict = Depends(get_current_user)):
    """حذف تقييم"""
    query = build_tenant_query(current_user, {"id": review_id})
    
    review = await db.reviews.find_one(query)
    if not review:
        raise HTTPException(status_code=404, detail="التقييم غير موجود")
    
    await db.reviews.delete_one({"id": review_id})
    return {"message": "تم حذف التقييم"}

@api_router.get("/reviews/stats")
async def get_reviews_stats(current_user: dict = Depends(get_current_user)):
    """إحصائيات التقييمات"""
    query = build_tenant_query(current_user)
    
    reviews = await db.reviews.find(query, {"_id": 0, "rating": 1, "response": 1}).to_list(1000)
    
    total = len(reviews)
    if total == 0:
        return {
            "total": 0,
            "average_rating": 0,
            "five_star": 0,
            "four_star": 0,
            "three_star": 0,
            "two_star": 0,
            "one_star": 0,
            "responded": 0,
            "pending_response": 0
        }
    
    ratings = [r["rating"] for r in reviews]
    avg_rating = sum(ratings) / total
    
    return {
        "total": total,
        "average_rating": round(avg_rating, 1),
        "five_star": ratings.count(5),
        "four_star": ratings.count(4),
        "three_star": ratings.count(3),
        "two_star": ratings.count(2),
        "one_star": ratings.count(1),
        "responded": len([r for r in reviews if r.get("response")]),
        "pending_response": len([r for r in reviews if not r.get("response")])
    }

# ==================== SMART REPORTS - التقارير الذكية ====================

@api_router.get("/smart-reports/sales")
async def get_sales_report(
    period: str = "today",  # today, week, month, year, custom
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    branch_id: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    """تقرير المبيعات"""
    query = build_tenant_query(current_user)
    
    if branch_id:
        query["branch_id"] = branch_id
    
    # تحديد الفترة
    now = datetime.now(timezone.utc)
    if period == "today":
        start = now.replace(hour=0, minute=0, second=0, microsecond=0)
    elif period == "week":
        start = now - timedelta(days=7)
    elif period == "month":
        start = now - timedelta(days=30)
    elif period == "year":
        start = now - timedelta(days=365)
    elif period == "custom" and start_date:
        start = datetime.fromisoformat(start_date)
    else:
        start = now.replace(hour=0, minute=0, second=0, microsecond=0)
    
    query["created_at"] = {"$gte": start.isoformat()}
    if end_date:
        query["created_at"]["$lte"] = end_date
    
    orders = await db.orders.find(query, {"_id": 0}).to_list(10000)
    
    # حساب الإحصائيات
    total_sales = sum(o.get("total", 0) for o in orders)
    total_orders = len(orders)
    avg_order_value = total_sales / total_orders if total_orders > 0 else 0
    
    # تقسيم حسب نوع الطلب
    dine_in = len([o for o in orders if o.get("order_type") == "dine_in"])
    takeaway = len([o for o in orders if o.get("order_type") == "takeaway"])
    delivery = len([o for o in orders if o.get("order_type") == "delivery"])
    
    # تقسيم حسب طريقة الدفع
    cash_orders = len([o for o in orders if o.get("payment_method") == "cash"])
    card_orders = len([o for o in orders if o.get("payment_method") == "card"])
    
    return {
        "period": period,
        "total_sales": total_sales,
        "total_orders": total_orders,
        "average_order_value": round(avg_order_value, 2),
        "by_type": {
            "dine_in": dine_in,
            "takeaway": takeaway,
            "delivery": delivery
        },
        "by_payment": {
            "cash": cash_orders,
            "card": card_orders
        }
    }

@api_router.get("/smart-reports/products")
async def get_products_report(
    period: str = "month",
    branch_id: Optional[str] = None,
    limit: int = 10,
    current_user: dict = Depends(get_current_user)
):
    """تقرير المنتجات الأكثر مبيعاً"""
    query = build_tenant_query(current_user)
    
    if branch_id:
        query["branch_id"] = branch_id
    
    # تحديد الفترة
    now = datetime.now(timezone.utc)
    if period == "week":
        start = now - timedelta(days=7)
    elif period == "month":
        start = now - timedelta(days=30)
    else:
        start = now - timedelta(days=30)
    
    query["created_at"] = {"$gte": start.isoformat()}
    
    orders = await db.orders.find(query, {"_id": 0, "items": 1}).to_list(10000)
    
    # حساب المنتجات الأكثر مبيعاً
    product_sales = {}
    for order in orders:
        for item in order.get("items", []):
            product_id = item.get("product_id", item.get("name", "unknown"))
            product_name = item.get("name", "Unknown")
            quantity = item.get("quantity", 1)
            total = item.get("total", 0)
            
            if product_id not in product_sales:
                product_sales[product_id] = {
                    "name": product_name,
                    "quantity": 0,
                    "revenue": 0
                }
            
            product_sales[product_id]["quantity"] += quantity
            product_sales[product_id]["revenue"] += total
    
    # ترتيب حسب الكمية
    sorted_products = sorted(product_sales.items(), key=lambda x: x[1]["quantity"], reverse=True)[:limit]
    
    return {
        "period": period,
        "top_products": [
            {
                "product_id": p[0],
                "name": p[1]["name"],
                "quantity": p[1]["quantity"],
                "revenue": p[1]["revenue"]
            }
            for p in sorted_products
        ]
    }

@api_router.get("/smart-reports/hourly")
async def get_hourly_report(
    date: Optional[str] = None,
    branch_id: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    """تقرير المبيعات حسب الساعة"""
    query = build_tenant_query(current_user)
    
    if branch_id:
        query["branch_id"] = branch_id
    
    if date:
        query["created_at"] = {"$regex": f"^{date}"}
    else:
        today = datetime.now(timezone.utc).strftime("%Y-%m-%d")
        query["created_at"] = {"$regex": f"^{today}"}
    
    orders = await db.orders.find(query, {"_id": 0, "created_at": 1, "total": 1}).to_list(10000)
    
    # تقسيم حسب الساعة
    hourly_data = {str(h).zfill(2): {"orders": 0, "sales": 0} for h in range(24)}
    
    for order in orders:
        try:
            created_at = order.get("created_at", "")
            if "T" in created_at:
                hour = created_at.split("T")[1][:2]
            else:
                hour = "00"
            
            if hour in hourly_data:
                hourly_data[hour]["orders"] += 1
                hourly_data[hour]["sales"] += order.get("total", 0)
        except:
            pass
    
    return {
        "date": date or datetime.now(timezone.utc).strftime("%Y-%m-%d"),
        "hourly": hourly_data
    }

# ==================== SMART REPORTS EXPORT ====================

@api_router.get("/smart-reports/export/excel")
async def export_smart_report_excel(
    report_type: str = "sales",  # sales, products, hourly
    period: str = "month",
    branch_id: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    """تصدير التقارير الذكية إلى Excel"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    
    tenant_id = get_user_tenant_id(current_user)
    
    wb = Workbook()
    ws = wb.active
    
    # التنسيق
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # تحديد الفترة
    now = datetime.now(timezone.utc)
    if period == "today":
        start = now.replace(hour=0, minute=0, second=0, microsecond=0)
    elif period == "week":
        start = now - timedelta(days=7)
    elif period == "month":
        start = now - timedelta(days=30)
    elif period == "year":
        start = now - timedelta(days=365)
    else:
        start = now.replace(hour=0, minute=0, second=0, microsecond=0)
    
    # Build query
    query = build_tenant_query(current_user)
    
    # فلترة الفرع
    user_branch_id = current_user.get("branch_id")
    user_role = current_user.get("role")
    if user_branch_id and user_role not in [UserRole.SUPER_ADMIN, UserRole.ADMIN, UserRole.MANAGER]:
        query["branch_id"] = user_branch_id
    elif branch_id:
        query["branch_id"] = branch_id
    
    query["created_at"] = {"$gte": start.isoformat()}
    
    if report_type == "sales":
        ws.title = "تقرير المبيعات الذكي"
        orders = await db.orders.find(query, {"_id": 0}).to_list(10000)
        
        # العنوان
        ws.merge_cells('A1:F1')
        ws['A1'] = f"تقرير المبيعات الذكي - {period}"
        ws['A1'].font = Font(bold=True, size=14)
        ws['A1'].alignment = Alignment(horizontal='center')
        
        # الرؤوس
        headers = ['#', 'رقم الطلب', 'التاريخ', 'النوع', 'طريقة الدفع', 'المبلغ']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')
            cell.border = thin_border
        
        order_types = {"dine_in": "محلي", "takeaway": "سفري", "delivery": "توصيل"}
        payment_methods = {"cash": "نقدي", "card": "بطاقة", "credit": "آجل"}
        
        row_num = 4
        total = 0
        for idx, order in enumerate(orders, 1):
            total += order.get("total", 0)
            data = [
                idx,
                order.get("order_number", ""),
                order.get("created_at", "")[:10],
                order_types.get(order.get("order_type"), order.get("order_type", "")),
                payment_methods.get(order.get("payment_method"), order.get("payment_method", "")),
                order.get("total", 0)
            ]
            for col, value in enumerate(data, 1):
                cell = ws.cell(row=row_num, column=col, value=value)
                cell.border = thin_border
                if col == 6:
                    cell.number_format = '#,##0'
            row_num += 1
        
        # الإجمالي
        ws.cell(row=row_num, column=5, value="الإجمالي:").font = Font(bold=True)
        ws.cell(row=row_num, column=6, value=total).font = Font(bold=True)
        ws.cell(row=row_num, column=6).number_format = '#,##0'
        
    elif report_type == "products":
        ws.title = "المنتجات الأكثر مبيعاً"
        orders = await db.orders.find(query, {"_id": 0}).to_list(10000)
        
        # حساب مبيعات المنتجات
        product_sales = {}
        for order in orders:
            for item in order.get("items", []):
                pid = item.get("product_id")
                name = item.get("name", "Unknown")
                if pid not in product_sales:
                    product_sales[pid] = {"name": name, "qty": 0, "revenue": 0}
                product_sales[pid]["qty"] += item.get("quantity", 0)
                product_sales[pid]["revenue"] += item.get("price", 0) * item.get("quantity", 0)
        
        # ترتيب حسب الكمية
        sorted_products = sorted(product_sales.values(), key=lambda x: x["qty"], reverse=True)[:20]
        
        ws.merge_cells('A1:D1')
        ws['A1'] = "المنتجات الأكثر مبيعاً"
        ws['A1'].font = Font(bold=True, size=14)
        ws['A1'].alignment = Alignment(horizontal='center')
        
        headers = ['#', 'المنتج', 'الكمية', 'الإيرادات']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')
            cell.border = thin_border
        
        row_num = 4
        for idx, prod in enumerate(sorted_products, 1):
            data = [idx, prod["name"], prod["qty"], prod["revenue"]]
            for col, value in enumerate(data, 1):
                cell = ws.cell(row=row_num, column=col, value=value)
                cell.border = thin_border
                if col == 4:
                    cell.number_format = '#,##0'
            row_num += 1
    
    elif report_type == "hourly":
        ws.title = "التقرير الساعي"
        today = datetime.now(timezone.utc).strftime("%Y-%m-%d")
        query["created_at"] = {"$regex": f"^{today}"}
        
        orders = await db.orders.find(query, {"_id": 0, "created_at": 1, "total": 1}).to_list(10000)
        
        hourly_data = {str(h).zfill(2): {"orders": 0, "sales": 0} for h in range(24)}
        for order in orders:
            try:
                created_at = order.get("created_at", "")
                if "T" in created_at:
                    hour = created_at.split("T")[1][:2]
                else:
                    hour = "00"
                if hour in hourly_data:
                    hourly_data[hour]["orders"] += 1
                    hourly_data[hour]["sales"] += order.get("total", 0)
            except:
                pass
        
        ws.merge_cells('A1:D1')
        ws['A1'] = f"التقرير الساعي - {today}"
        ws['A1'].font = Font(bold=True, size=14)
        ws['A1'].alignment = Alignment(horizontal='center')
        
        headers = ['الساعة', 'عدد الطلبات', 'المبيعات']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')
            cell.border = thin_border
        
        row_num = 4
        for hour in sorted(hourly_data.keys()):
            data = [f"{hour}:00", hourly_data[hour]["orders"], hourly_data[hour]["sales"]]
            for col, value in enumerate(data, 1):
                cell = ws.cell(row=row_num, column=col, value=value)
                cell.border = thin_border
                if col == 3:
                    cell.number_format = '#,##0'
            row_num += 1
    
    # ضبط عرض الأعمدة
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        ws.column_dimensions[column].width = max(max_length + 2, 12)
    
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=smart_report_{report_type}_{period}.xlsx"}
    )

@api_router.get("/smart-reports/export/pdf")
async def export_smart_report_pdf(
    report_type: str = "sales",
    period: str = "month",
    branch_id: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    """تصدير التقارير الذكية إلى PDF"""
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.lib.units import cm
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    
    tenant_id = get_user_tenant_id(current_user)
    
    # تحديد الفترة
    now = datetime.now(timezone.utc)
    if period == "today":
        start = now.replace(hour=0, minute=0, second=0, microsecond=0)
    elif period == "week":
        start = now - timedelta(days=7)
    elif period == "month":
        start = now - timedelta(days=30)
    elif period == "year":
        start = now - timedelta(days=365)
    else:
        start = now.replace(hour=0, minute=0, second=0, microsecond=0)
    
    # Build query
    query = build_tenant_query(current_user)
    
    user_branch_id = current_user.get("branch_id")
    user_role = current_user.get("role")
    if user_branch_id and user_role not in [UserRole.SUPER_ADMIN, UserRole.ADMIN, UserRole.MANAGER]:
        query["branch_id"] = user_branch_id
    elif branch_id:
        query["branch_id"] = branch_id
    
    query["created_at"] = {"$gte": start.isoformat()}
    
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(A4), rightMargin=1*cm, leftMargin=1*cm, topMargin=1*cm, bottomMargin=1*cm)
    
    styles = getSampleStyleSheet()
    elements = []
    
    title_style = styles['Heading1']
    title_style.alignment = 1
    
    headers = []
    data_rows = []
    totals_row = None
    title_text = ""
    
    if report_type == "sales":
        title_text = f"تقرير المبيعات الذكي - {period}"
        orders = await db.orders.find(query, {"_id": 0}).to_list(10000)
        
        headers = ["#", "رقم الطلب", "التاريخ", "النوع", "طريقة الدفع", "المبلغ"]
        order_types = {"dine_in": "محلي", "takeaway": "سفري", "delivery": "توصيل"}
        payment_methods = {"cash": "نقدي", "card": "بطاقة", "credit": "آجل"}
        
        total = 0
        for idx, order in enumerate(orders[:100], 1):  # Limit to 100 for PDF
            total += order.get("total", 0)
            data_rows.append([
                str(idx),
                order.get("order_number", ""),
                order.get("created_at", "")[:10],
                order_types.get(order.get("order_type"), order.get("order_type", "")),
                payment_methods.get(order.get("payment_method"), order.get("payment_method", "")),
                f"{order.get('total', 0):,.0f}"
            ])
        
        totals_row = ["", "", "", "", "الإجمالي:", f"{total:,.0f}"]
        
    elif report_type == "products":
        title_text = "المنتجات الأكثر مبيعاً"
        orders = await db.orders.find(query, {"_id": 0}).to_list(10000)
        
        product_sales = {}
        for order in orders:
            for item in order.get("items", []):
                pid = item.get("product_id")
                name = item.get("name", "Unknown")
                if pid not in product_sales:
                    product_sales[pid] = {"name": name, "qty": 0, "revenue": 0}
                product_sales[pid]["qty"] += item.get("quantity", 0)
                product_sales[pid]["revenue"] += item.get("price", 0) * item.get("quantity", 0)
        
        sorted_products = sorted(product_sales.values(), key=lambda x: x["qty"], reverse=True)[:20]
        
        headers = ["#", "المنتج", "الكمية", "الإيرادات"]
        for idx, prod in enumerate(sorted_products, 1):
            data_rows.append([str(idx), prod["name"], str(prod["qty"]), f"{prod['revenue']:,.0f}"])
    
    elif report_type == "hourly":
        title_text = f"التقرير الساعي - {datetime.now(timezone.utc).strftime('%Y-%m-%d')}"
        today = datetime.now(timezone.utc).strftime("%Y-%m-%d")
        query["created_at"] = {"$regex": f"^{today}"}
        
        orders = await db.orders.find(query, {"_id": 0, "created_at": 1, "total": 1}).to_list(10000)
        
        hourly_data = {str(h).zfill(2): {"orders": 0, "sales": 0} for h in range(24)}
        for order in orders:
            try:
                created_at = order.get("created_at", "")
                if "T" in created_at:
                    hour = created_at.split("T")[1][:2]
                else:
                    hour = "00"
                if hour in hourly_data:
                    hourly_data[hour]["orders"] += 1
                    hourly_data[hour]["sales"] += order.get("total", 0)
            except:
                pass
        
        headers = ["الساعة", "عدد الطلبات", "المبيعات"]
        for hour in sorted(hourly_data.keys()):
            data_rows.append([f"{hour}:00", str(hourly_data[hour]["orders"]), f"{hourly_data[hour]['sales']:,.0f}"])
    
    # Build PDF
    elements.append(Paragraph(title_text, title_style))
    elements.append(Spacer(1, 20))
    
    if headers and data_rows:
        table_data = [headers] + data_rows
        if totals_row:
            table_data.append(totals_row)
        
        col_widths = [doc.width / len(headers)] * len(headers)
        table = Table(table_data, colWidths=col_widths)
        
        style = TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4472C4')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('FONTSIZE', (0, 1), (-1, -1), 9),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('ROWBACKGROUNDS', (0, 1), (-1, -2), [colors.white, colors.HexColor('#F5F5F5')]),
        ])
        
        if totals_row:
            style.add('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#E6E6E6'))
            style.add('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold')
        
        table.setStyle(style)
        elements.append(table)
    
    doc.build(elements)
    buffer.seek(0)
    
    return StreamingResponse(
        buffer,
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename=smart_report_{report_type}_{period}.pdf"}
    )

# ==================== CALL CENTER / CALLER ID ====================

class CallCenterConfig(BaseModel):
    enabled: bool = False
    provider: str = ""
    api_url: str = ""
    api_key: str = ""
    api_secret: str = ""
    webhook_secret: str = ""
    auto_popup: bool = True
    auto_save_new_callers: bool = True
    play_sound: bool = True

class IncomingCall(BaseModel):
    phone: str
    caller_name: Optional[str] = None
    call_id: Optional[str] = None
    direction: str = "incoming"
    timestamp: Optional[str] = None

# Store active calls in memory (in production, use Redis)
active_calls = {}

@api_router.post("/callcenter/config")
async def save_callcenter_config(config: CallCenterConfig, current_user: dict = Depends(get_current_user)):
    """حفظ إعدادات الكول سنتر"""
    tenant_id = get_user_tenant_id(current_user)
    
    config_doc = {
        "tenant_id": tenant_id,
        **config.dict(),
        "updated_at": datetime.now(timezone.utc).isoformat(),
        "updated_by": current_user["id"]
    }
    
    await db.callcenter_config.update_one(
        {"tenant_id": tenant_id},
        {"$set": config_doc},
        upsert=True
    )
    
    return {"message": "تم حفظ إعدادات الكول سنتر"}

@api_router.get("/callcenter/config")
async def get_callcenter_config(current_user: dict = Depends(get_current_user)):
    """جلب إعدادات الكول سنتر"""
    tenant_id = get_user_tenant_id(current_user)
    
    config = await db.callcenter_config.find_one({"tenant_id": tenant_id}, {"_id": 0})
    
    if not config:
        return CallCenterConfig().dict()
    
    return config

@api_router.post("/callcenter/test")
async def test_callcenter_connection(config: CallCenterConfig, current_user: dict = Depends(get_current_user)):
    """اختبار اتصال الكول سنتر"""
    
    # في الإنتاج، يجب اختبار الاتصال الفعلي مع المزود
    # حالياً نرجع نجاح للمحاكاة
    
    if not config.provider:
        raise HTTPException(status_code=400, detail="يرجى اختيار مزود الخدمة")
    
    if not config.api_url and config.provider not in ["custom"]:
        raise HTTPException(status_code=400, detail="يرجى إدخال رابط API")
    
    # محاكاة اختبار الاتصال
    return {"success": True, "message": f"تم الاتصال بـ {config.provider} بنجاح"}

@api_router.post("/callcenter/webhook")
async def callcenter_webhook(request: Request):
    """Webhook لاستقبال المكالمات من نظام الكول سنتر"""
    
    try:
        body = await request.json()
    except:
        body = {}
    
    # استخراج رقم المتصل من البيانات (يختلف حسب المزود)
    phone = body.get("phone") or body.get("caller_id") or body.get("from") or body.get("callerNumber")
    caller_name = body.get("caller_name") or body.get("name") or body.get("callerName")
    call_id = body.get("call_id") or body.get("callId") or body.get("id") or str(uuid.uuid4())
    direction = body.get("direction") or body.get("type") or "incoming"
    
    if not phone:
        return {"status": "error", "message": "No phone number provided"}
    
    # تنظيف رقم الهاتف
    phone = phone.replace(" ", "").replace("-", "").replace("+", "")
    if phone.startswith("964"):
        phone = "0" + phone[3:]
    
    # البحث عن العميل
    customer = await db.customers.find_one(
        {"$or": [{"phone": phone}, {"phone2": phone}]},
        {"_id": 0}
    )
    
    # آخر طلب للعميل
    last_order = None
    if customer:
        last_order = await db.orders.find_one(
            {"customer_phone": phone},
            {"_id": 0},
            sort=[("created_at", -1)]
        )
    
    # تخزين المكالمة النشطة
    call_data = {
        "call_id": call_id,
        "phone": phone,
        "caller_name": caller_name or (customer["name"] if customer else "New Customer"),
        "direction": direction,
        "customer": customer,
        "last_order": last_order,
        "is_new_customer": customer is None,
        "timestamp": datetime.now(timezone.utc).isoformat(),
        "status": "ringing"
    }
    
    active_calls[call_id] = call_data
    
    # إذا كان عميل جديد وتم تفعيل الحفظ التلقائي
    # سيتم الحفظ عند إنشاء الطلب
    
    return {
        "status": "success",
        "call_id": call_id,
        "customer_found": customer is not None,
        "customer": customer,
        "last_order": last_order
    }

@api_router.post("/callcenter/simulate")
async def simulate_incoming_call(data: dict, current_user: dict = Depends(get_current_user)):
    """محاكاة مكالمة واردة للاختبار"""
    
    phone = data.get("phone", "07801234567")
    tenant_id = get_user_tenant_id(current_user)
    
    # تنظيف رقم الهاتف
    phone = phone.replace(" ", "").replace("-", "").replace("+", "")
    
    # البحث عن العميل مع مراعاة tenant_id
    customer_query = {"$or": [{"phone": phone}, {"phone2": phone}]}
    if tenant_id:
        customer_query["tenant_id"] = tenant_id
    else:
        customer_query["$or"] = [
            {"$and": [{"phone": phone}, {"$or": [{"tenant_id": None}, {"tenant_id": {"$exists": False}}]}]},
            {"$and": [{"phone2": phone}, {"$or": [{"tenant_id": None}, {"tenant_id": {"$exists": False}}]}]}
        ]
    
    customer = await db.customers.find_one(customer_query, {"_id": 0})
    
    # آخر طلب للعميل
    last_order = None
    if customer:
        order_query = {"customer_phone": phone}
        if tenant_id:
            order_query["tenant_id"] = tenant_id
        else:
            order_query["$or"] = [{"tenant_id": None}, {"tenant_id": {"$exists": False}}]
        
        last_order = await db.orders.find_one(
            order_query,
            {"_id": 0},
            sort=[("created_at", -1)]
        )
    
    call_id = str(uuid.uuid4())
    call_data = {
        "call_id": call_id,
        "phone": phone,
        "caller_name": customer["name"] if customer else "New Customer",
        "direction": "incoming",
        "customer": customer,
        "last_order": last_order,
        "is_new_customer": customer is None,
        "timestamp": datetime.now(timezone.utc).isoformat(),
        "status": "ringing",
        "simulated": True,
        "tenant_id": tenant_id
    }
    
    active_calls[call_id] = call_data
    
    return call_data

@api_router.get("/callcenter/active-calls")
async def get_active_calls(current_user: dict = Depends(get_current_user)):
    """جلب المكالمات النشطة للمستخدم الحالي"""
    tenant_id = get_user_tenant_id(current_user)
    
    # فلترة المكالمات حسب tenant_id
    if tenant_id:
        filtered_calls = [c for c in active_calls.values() if c.get("tenant_id") == tenant_id]
    else:
        # النظام الرئيسي يرى المكالمات بدون tenant_id
        filtered_calls = [c for c in active_calls.values() if not c.get("tenant_id")]
    
    return filtered_calls

@api_router.post("/callcenter/calls/{call_id}/answer")
async def answer_call(call_id: str, current_user: dict = Depends(get_current_user)):
    """الرد على المكالمة"""
    if call_id in active_calls:
        active_calls[call_id]["status"] = "answered"
        active_calls[call_id]["answered_by"] = current_user["id"]
        active_calls[call_id]["answered_at"] = datetime.now(timezone.utc).isoformat()
        return active_calls[call_id]
    raise HTTPException(status_code=404, detail="المكالمة غير موجودة")

@api_router.post("/callcenter/calls/{call_id}/end")
async def end_call(call_id: str, current_user: dict = Depends(get_current_user)):
    """إنهاء المكالمة"""
    if call_id in active_calls:
        call_data = active_calls.pop(call_id)
        call_data["status"] = "ended"
        call_data["ended_at"] = datetime.now(timezone.utc).isoformat()
        
        # حفظ سجل المكالمة
        call_log = {k: v for k, v in call_data.items() if k != '_id'}
        call_log["id"] = call_id
        await db.call_logs.insert_one(call_log)
        
        return {"message": "تم إنهاء المكالمة", "call": call_data}
    raise HTTPException(status_code=404, detail="المكالمة غير موجودة")

@api_router.get("/callcenter/call-logs")
async def get_call_logs(
    limit: int = 50,
    skip: int = 0,
    current_user: dict = Depends(get_current_user)
):
    """سجل المكالمات"""
    tenant_id = get_user_tenant_id(current_user)
    
    query = {}
    if tenant_id:
        query["tenant_id"] = tenant_id
    
    logs = await db.call_logs.find(query, {"_id": 0}).sort("timestamp", -1).skip(skip).limit(limit).to_list(limit)
    total = await db.call_logs.count_documents(query)
    
    return {
        "logs": logs,
        "total": total
    }

# ==================== SEED DATA ====================

@api_router.post("/seed")
async def seed_data():
    # Check if already seeded
    existing = await db.users.find_one({"email": "admin@maestroegp.com"})
    if existing:
        return {"message": "البيانات موجودة بالفعل"}
    
    # Create admin user
    admin_doc = {
        "id": str(uuid.uuid4()),
        "username": "admin",
        "email": "admin@maestroegp.com",
        "password": hash_password("admin123"),
        "full_name": "مدير النظام",
        "role": UserRole.ADMIN,
        "branch_id": None,
        "permissions": ["all"],
        "is_active": True,
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    await db.users.insert_one(admin_doc)
    
    # Create default branch
    branch_doc = {
        "id": str(uuid.uuid4()),
        "name": "الفرع الرئيسي",
        "address": "بغداد - الكرادة",
        "phone": "+964 770 123 4567",
        "email": "main@maestroegp.com",
        "is_active": True,
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    await db.branches.insert_one(branch_doc)
    branch_id = branch_doc["id"]
    
    # Create categories
    categories = [
        {"name": "برغر", "name_en": "Burgers", "icon": "Beef", "color": "#EF4444", "sort_order": 1},
        {"name": "بيتزا", "name_en": "Pizza", "icon": "Pizza", "color": "#F59E0B", "sort_order": 2},
        {"name": "مشروبات", "name_en": "Drinks", "icon": "Coffee", "color": "#8B5CF6", "sort_order": 3},
        {"name": "حلويات", "name_en": "Desserts", "icon": "Cake", "color": "#EC4899", "sort_order": 4},
        {"name": "سلطات", "name_en": "Salads", "icon": "Salad", "color": "#10B981", "sort_order": 5},
    ]
    
    cat_ids = {}
    for cat in categories:
        cat_doc = {"id": str(uuid.uuid4()), **cat, "is_active": True}
        await db.categories.insert_one(cat_doc)
        cat_ids[cat["name_en"]] = cat_doc["id"]
    
    # Create products with costs
    products = [
        {"name": "برغر كلاسيك", "name_en": "Classic Burger", "category_id": cat_ids["Burgers"], "price": 12000, "cost": 4000, "operating_cost": 1000, "image": "https://images.pexels.com/photos/18796078/pexels-photo-18796078.jpeg"},
        {"name": "برغر دبل", "name_en": "Double Burger", "category_id": cat_ids["Burgers"], "price": 18000, "cost": 7000, "operating_cost": 1500, "image": "https://images.pexels.com/photos/5672397/pexels-photo-5672397.jpeg"},
        {"name": "بيتزا مارغريتا", "name_en": "Margherita Pizza", "category_id": cat_ids["Pizza"], "price": 15000, "cost": 5000, "operating_cost": 1200, "image": "https://images.pexels.com/photos/35532821/pexels-photo-35532821.jpeg"},
        {"name": "بيتزا خضار", "name_en": "Veggie Pizza", "category_id": cat_ids["Pizza"], "price": 14000, "cost": 4500, "operating_cost": 1100, "image": "https://images.pexels.com/photos/34956178/pexels-photo-34956178.jpeg"},
        {"name": "قهوة عربية", "name_en": "Arabic Coffee", "category_id": cat_ids["Drinks"], "price": 3000, "cost": 500, "operating_cost": 200, "image": "https://images.pexels.com/photos/29799615/pexels-photo-29799615.jpeg"},
        {"name": "لاتيه", "name_en": "Latte", "category_id": cat_ids["Drinks"], "price": 5000, "cost": 1200, "operating_cost": 300, "image": "https://images.pexels.com/photos/15800375/pexels-photo-15800375.jpeg"},
        {"name": "كيكة شوكولاتة", "name_en": "Chocolate Cake", "category_id": cat_ids["Desserts"], "price": 8000, "cost": 2500, "operating_cost": 500, "image": "https://images.pexels.com/photos/29538417/pexels-photo-29538417.jpeg"},
        {"name": "تشيز كيك", "name_en": "Cheesecake", "category_id": cat_ids["Desserts"], "price": 9000, "cost": 3000, "operating_cost": 500, "image": "https://images.pexels.com/photos/15564368/pexels-photo-15564368.jpeg"},
    ]
    
    for prod in products:
        profit = prod["price"] - prod["cost"] - prod["operating_cost"]
        prod_doc = {"id": str(uuid.uuid4()), **prod, "profit": profit, "is_available": True, "ingredients": []}
        await db.products.insert_one(prod_doc)
    
    # Create tables
    for i in range(1, 11):
        table_doc = {
            "id": str(uuid.uuid4()),
            "number": i,
            "capacity": 4 if i <= 6 else 6,
            "branch_id": branch_id,
            "section": "داخلي" if i <= 6 else "خارجي",
            "status": "available",
            "current_order_id": None
        }
        await db.tables.insert_one(table_doc)
    
    # Create cashier user
    cashier_doc = {
        "id": str(uuid.uuid4()),
        "username": "cashier1",
        "email": "cashier@maestroegp.com",
        "password": hash_password("cashier123"),
        "full_name": "أحمد الكاشير",
        "role": UserRole.CASHIER,
        "branch_id": branch_id,
        "permissions": ["pos", "orders"],
        "is_active": True,
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    await db.users.insert_one(cashier_doc)
    
    # Set default currencies
    currencies = [
        {"code": "IQD", "name": "دينار عراقي", "symbol": "د.ع", "exchange_rate": 1.0},
        {"code": "USD", "name": "دولار أمريكي", "symbol": "$", "exchange_rate": 1460.0},
    ]
    await db.settings.insert_one({"type": "currencies", "value": currencies})
    
    # Set default delivery app settings
    delivery_apps = [
        {"app_id": "toters", "name": "توترز", "name_en": "Toters", "commission_type": "percentage", "commission_rate": 15, "is_active": True, "payment_terms": "weekly"},
        {"app_id": "talabat", "name": "طلبات", "name_en": "Talabat", "commission_type": "percentage", "commission_rate": 18, "is_active": True, "payment_terms": "weekly"},
        {"app_id": "baly", "name": "بالي", "name_en": "Baly", "commission_type": "percentage", "commission_rate": 12, "is_active": True, "payment_terms": "weekly"},
        {"app_id": "alsaree3", "name": "عالسريع", "name_en": "Al-Sari3", "commission_type": "percentage", "commission_rate": 10, "is_active": True, "payment_terms": "weekly"},
        {"app_id": "talabati", "name": "طلباتي", "name_en": "Talabati", "commission_type": "percentage", "commission_rate": 14, "is_active": True, "payment_terms": "weekly"},
    ]
    for app in delivery_apps:
        await db.delivery_app_settings.insert_one(app)
    
    return {"message": "تم إنشاء البيانات الأولية بنجاح"}

# ==================== ROOT ====================

@api_router.get("/")
async def root():
    return {"message": "Maestro EGP API", "version": "2.0.0"}

# ==================== BIOMETRIC DEVICE ROUTES ====================
# تكامل أجهزة البصمة ZKTeco

class BiometricDeviceCreate(BaseModel):
    name: str
    ip_address: str
    port: int = 4370
    branch_id: str
    device_type: str = "fingerprint"  # fingerprint, face, card

class ZKTecoPushData(BaseModel):
    AuthToken: Optional[str] = None
    OperationID: Optional[str] = None
    CommandName: Optional[str] = None
    VerifyType: Optional[str] = None
    PIN: Optional[str] = None
    DateTime: Optional[str] = None
    DeviceSN: Optional[str] = None
    Status: Optional[int] = None

@api_router.get("/biometric/devices")
async def list_biometric_devices(branch_id: Optional[str] = None, current_user: dict = Depends(get_current_user)):
    """قائمة أجهزة البصمة"""
    query = {"tenant_id": current_user.get("tenant_id")}
    if branch_id:
        query["branch_id"] = branch_id
    
    devices = await db.biometric_devices.find(query, {"_id": 0}).to_list(100)
    return devices

@api_router.post("/biometric/devices")
async def create_biometric_device(device: BiometricDeviceCreate, current_user: dict = Depends(get_current_user)):
    """إضافة جهاز بصمة جديد"""
    new_device = {
        "id": str(uuid.uuid4()),
        "name": device.name,
        "ip_address": device.ip_address,
        "port": device.port,
        "branch_id": device.branch_id,
        "device_type": device.device_type,
        "tenant_id": current_user.get("tenant_id"),
        "is_active": True,
        "last_sync": None,
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    
    await db.biometric_devices.insert_one(new_device)
    if "_id" in new_device:
        del new_device["_id"]
    
    return {"message": "تم إضافة الجهاز بنجاح", "device": new_device}

@api_router.post("/biometric/devices/{device_id}/test")
async def test_biometric_connection(device_id: str, current_user: dict = Depends(get_current_user)):
    """اختبار الاتصال بجهاز البصمة"""
    device = await db.biometric_devices.find_one({
        "id": device_id, 
        "tenant_id": current_user.get("tenant_id")
    }, {"_id": 0})
    
    if not device:
        raise HTTPException(status_code=404, detail="الجهاز غير موجود")
    
    # محاولة الاتصال بالجهاز
    try:
        from zk import ZK
        zk = ZK(device["ip_address"], port=device["port"], timeout=5)
        conn = zk.connect()
        
        if conn:
            info = {
                "serial_number": zk.get_serialnumber() if hasattr(zk, 'get_serialnumber') else "N/A",
                "users_count": len(zk.get_users()) if hasattr(zk, 'get_users') else 0
            }
            zk.disconnect()
            
            return {"success": True, "message": "تم الاتصال بنجاح", "device_info": info}
        else:
            return {"success": False, "message": "فشل الاتصال بالجهاز"}
    except ImportError:
        # المكتبة غير مثبتة - وضع المحاكاة
        return {
            "success": True, 
            "message": "وضع المحاكاة - مكتبة pyzk غير مثبتة",
            "device_info": {"serial_number": "MOCK-001", "users_count": 0}
        }
    except Exception as e:
        return {"success": False, "message": f"خطأ: {str(e)}"}

@api_router.post("/biometric/devices/{device_id}/sync")
async def sync_biometric_attendance(device_id: str, current_user: dict = Depends(get_current_user)):
    """مزامنة سجلات الحضور من جهاز البصمة"""
    device = await db.biometric_devices.find_one({
        "id": device_id,
        "tenant_id": current_user.get("tenant_id")
    }, {"_id": 0})
    
    if not device:
        raise HTTPException(status_code=404, detail="الجهاز غير موجود")
    
    synced_records = []
    
    try:
        from zk import ZK
        zk = ZK(device["ip_address"], port=device["port"], timeout=10)
        conn = zk.connect()
        
        if conn:
            attendance = zk.get_attendance()
            
            for record in attendance:
                att_record = {
                    "id": str(uuid.uuid4()),
                    "device_id": device_id,
                    "employee_code": str(record.user_id),
                    "punch_time": record.timestamp.isoformat() if record.timestamp else None,
                    "punch_type": "in" if record.status == 0 else "out",
                    "verify_type": "fingerprint",
                    "tenant_id": current_user.get("tenant_id"),
                    "synced_at": datetime.now(timezone.utc).isoformat()
                }
                synced_records.append(att_record)
            
            zk.disconnect()
            
            # حفظ السجلات في قاعدة البيانات
            if synced_records:
                await db.biometric_attendance.insert_many(synced_records)
            
            # تحديث وقت آخر مزامنة
            await db.biometric_devices.update_one(
                {"id": device_id},
                {"$set": {"last_sync": datetime.now(timezone.utc).isoformat()}}
            )
    except ImportError:
        # وضع المحاكاة
        pass
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"فشل المزامنة: {str(e)}")
    
    return {
        "message": "تمت المزامنة بنجاح",
        "records_count": len(synced_records)
    }

@api_router.post("/biometric/push")
async def receive_biometric_push(request: Request):
    """
    استقبال بيانات الحضور من أجهزة ZKTeco (Push SDK)
    يجب تكوين الجهاز لإرسال البيانات لهذا الـ endpoint
    """
    try:
        data = await request.json()
        payload = ZKTecoPushData(**data)
        
        if payload.PIN:
            # البحث عن الموظف بالكود
            employee = await db.employees.find_one({"code": payload.PIN}, {"_id": 0})
            
            punch_type = "in" if payload.Status == 0 else "out"
            
            attendance_record = {
                "id": str(uuid.uuid4()),
                "employee_id": employee["id"] if employee else None,
                "employee_code": payload.PIN,
                "punch_time": payload.DateTime or datetime.now(timezone.utc).isoformat(),
                "punch_type": punch_type,
                "device_serial": payload.DeviceSN,
                "verify_type": payload.VerifyType or "fingerprint",
                "created_at": datetime.now(timezone.utc).isoformat()
            }
            
            await db.biometric_attendance.insert_one(attendance_record)
            
            # تحديث سجل الحضور اليومي للموظف
            if employee:
                today = datetime.now(timezone.utc).strftime("%Y-%m-%d")
                
                if punch_type == "in":
                    await db.attendance.update_one(
                        {"employee_id": employee["id"], "date": today},
                        {"$set": {"check_in": payload.DateTime, "updated_at": datetime.now(timezone.utc).isoformat()}},
                        upsert=True
                    )
                else:
                    await db.attendance.update_one(
                        {"employee_id": employee["id"], "date": today},
                        {"$set": {"check_out": payload.DateTime, "updated_at": datetime.now(timezone.utc).isoformat()}}
                    )
            
            logger.info(f"Biometric punch: {payload.PIN} - {punch_type}")
            
            return {"status": "received", "OperationID": payload.OperationID}
        
        return {"status": "no_data"}
    except Exception as e:
        logger.error(f"Biometric push error: {e}")
        return {"status": "error", "message": str(e)}

@api_router.delete("/biometric/devices/{device_id}")
async def delete_biometric_device(device_id: str, current_user: dict = Depends(get_current_user)):
    """حذف جهاز بصمة"""
    result = await db.biometric_devices.delete_one({
        "id": device_id,
        "tenant_id": current_user.get("tenant_id")
    })
    
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="الجهاز غير موجود")
    
    return {"message": "تم حذف الجهاز بنجاح"}

@api_router.get("/biometric/attendance")
async def get_biometric_attendance(
    branch_id: Optional[str] = None,
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    """سجلات الحضور من أجهزة البصمة"""
    query = {"tenant_id": current_user.get("tenant_id")} if current_user.get("tenant_id") else {}
    
    if start_date:
        query["punch_time"] = {"$gte": start_date}
    if end_date:
        if "punch_time" in query:
            query["punch_time"]["$lte"] = end_date
        else:
            query["punch_time"] = {"$lte": end_date}
    
    records = await db.biometric_attendance.find(query, {"_id": 0}).sort("punch_time", -1).to_list(500)
    return records

# ==================== LOYALTY PROGRAM ROUTES ====================

class LoyaltySettingsUpdate(BaseModel):
    is_enabled: bool = True
    points_per_currency: float = 1.0
    currency_per_point: float = 0.01
    min_redeem_points: int = 100
    max_redeem_percent: float = 50
    points_expiry_days: int = 365
    welcome_bonus: int = 50
    birthday_bonus: int = 100
    referral_bonus: int = 200
    tiers: List[Dict[str, Any]] = []

class LoyaltyMemberCreate(BaseModel):
    customer_id: str
    customer_name: str
    phone: str
    email: Optional[str] = None
    birthday: Optional[str] = None
    referred_by: Optional[str] = None

class EarnPointsRequest(BaseModel):
    member_id: str
    order_id: str
    order_total: float

class RedeemPointsRequest(BaseModel):
    member_id: str
    points_to_redeem: int
    order_id: str

DEFAULT_LOYALTY_TIERS = [
    {"name": "برونزي", "name_en": "Bronze", "min_points": 0, "discount_percent": 0, "points_multiplier": 1.0, "color": "#CD7F32"},
    {"name": "فضي", "name_en": "Silver", "min_points": 500, "discount_percent": 5, "points_multiplier": 1.25, "color": "#C0C0C0"},
    {"name": "ذهبي", "name_en": "Gold", "min_points": 1500, "discount_percent": 10, "points_multiplier": 1.5, "color": "#FFD700"},
    {"name": "بلاتيني", "name_en": "Platinum", "min_points": 5000, "discount_percent": 15, "points_multiplier": 2.0, "color": "#E5E4E2"}
]

@api_router.get("/loyalty/settings")
async def get_loyalty_settings(current_user: dict = Depends(get_current_user)):
    """جلب إعدادات برنامج الولاء"""
    settings = await db.loyalty_settings.find_one(
        {"tenant_id": current_user.get("tenant_id")}, 
        {"_id": 0}
    )
    
    if not settings:
        return {
            "is_enabled": True,
            "points_per_currency": 1.0,
            "currency_per_point": 0.01,
            "min_redeem_points": 100,
            "max_redeem_percent": 50,
            "points_expiry_days": 365,
            "welcome_bonus": 50,
            "birthday_bonus": 100,
            "referral_bonus": 200,
            "tiers": DEFAULT_LOYALTY_TIERS
        }
    return settings

@api_router.put("/loyalty/settings")
async def update_loyalty_settings(settings: LoyaltySettingsUpdate, current_user: dict = Depends(get_current_user)):
    """تحديث إعدادات برنامج الولاء"""
    await db.loyalty_settings.update_one(
        {"tenant_id": current_user.get("tenant_id")},
        {"$set": {**settings.model_dump(), "tenant_id": current_user.get("tenant_id")}},
        upsert=True
    )
    return {"message": "تم تحديث الإعدادات"}

@api_router.get("/loyalty/members")
async def get_loyalty_members(current_user: dict = Depends(get_current_user)):
    """قائمة أعضاء برنامج الولاء"""
    members = await db.loyalty_members.find(
        {"tenant_id": current_user.get("tenant_id")},
        {"_id": 0}
    ).sort("total_points", -1).to_list(500)
    return members

@api_router.post("/loyalty/members")
async def create_loyalty_member(member: LoyaltyMemberCreate, current_user: dict = Depends(get_current_user)):
    """إضافة عضو جديد"""
    # التحقق من عدم وجود العضو
    existing = await db.loyalty_members.find_one({
        "phone": member.phone,
        "tenant_id": current_user.get("tenant_id")
    })
    if existing:
        raise HTTPException(status_code=400, detail="العضو موجود مسبقاً")
    
    # جلب إعدادات الولاء
    settings = await db.loyalty_settings.find_one({"tenant_id": current_user.get("tenant_id")})
    welcome_bonus = settings.get("welcome_bonus", 50) if settings else 50
    
    new_member = {
        "id": str(uuid.uuid4()),
        "customer_id": member.customer_id,
        "customer_name": member.customer_name,
        "phone": member.phone,
        "email": member.email,
        "total_points": welcome_bonus,
        "available_points": welcome_bonus,
        "redeemed_points": 0,
        "current_tier": "bronze",
        "lifetime_spending": 0,
        "total_orders": 0,
        "join_date": datetime.now(timezone.utc).isoformat(),
        "birthday": member.birthday,
        "referral_code": str(uuid.uuid4())[:8].upper(),
        "referred_by": member.referred_by,
        "tenant_id": current_user.get("tenant_id")
    }
    
    await db.loyalty_members.insert_one(new_member)
    
    # تسجيل نقاط الترحيب
    if welcome_bonus > 0:
        transaction = {
            "id": str(uuid.uuid4()),
            "member_id": new_member["id"],
            "transaction_type": "bonus",
            "points": welcome_bonus,
            "description": "نقاط الترحيب",
            "balance_after": welcome_bonus,
            "created_at": datetime.now(timezone.utc).isoformat(),
            "tenant_id": current_user.get("tenant_id")
        }
        await db.loyalty_transactions.insert_one(transaction)
    
    # مكافأة الإحالة
    if member.referred_by:
        referrer = await db.loyalty_members.find_one({"referral_code": member.referred_by})
        if referrer:
            referral_bonus = settings.get("referral_bonus", 200) if settings else 200
            await db.loyalty_members.update_one(
                {"id": referrer["id"]},
                {"$inc": {"total_points": referral_bonus, "available_points": referral_bonus}}
            )
            ref_transaction = {
                "id": str(uuid.uuid4()),
                "member_id": referrer["id"],
                "transaction_type": "bonus",
                "points": referral_bonus,
                "description": f"مكافأة إحالة - {member.customer_name}",
                "balance_after": referrer.get("available_points", 0) + referral_bonus,
                "created_at": datetime.now(timezone.utc).isoformat(),
                "tenant_id": current_user.get("tenant_id")
            }
            await db.loyalty_transactions.insert_one(ref_transaction)
    
    if "_id" in new_member:
        del new_member["_id"]
    
    return {"message": "تم إضافة العضو", "member": new_member}

@api_router.get("/loyalty/members/{member_id}")
async def get_loyalty_member(member_id: str, current_user: dict = Depends(get_current_user)):
    """تفاصيل عضو"""
    member = await db.loyalty_members.find_one(
        {"id": member_id, "tenant_id": current_user.get("tenant_id")},
        {"_id": 0}
    )
    if not member:
        raise HTTPException(status_code=404, detail="العضو غير موجود")
    return member

@api_router.post("/loyalty/earn")
async def earn_points(request: EarnPointsRequest, current_user: dict = Depends(get_current_user)):
    """كسب نقاط من طلب"""
    member = await db.loyalty_members.find_one({"id": request.member_id, "tenant_id": current_user.get("tenant_id")})
    if not member:
        raise HTTPException(status_code=404, detail="العضو غير موجود")
    
    settings = await db.loyalty_settings.find_one({"tenant_id": current_user.get("tenant_id")})
    points_per_currency = settings.get("points_per_currency", 1.0) if settings else 1.0
    tiers = settings.get("tiers", DEFAULT_LOYALTY_TIERS) if settings else DEFAULT_LOYALTY_TIERS
    
    # حساب مضاعف المستوى
    multiplier = 1.0
    for tier in tiers:
        if tier.get("name_en", "").lower() == member.get("current_tier", "bronze").lower():
            multiplier = tier.get("points_multiplier", 1.0)
            break
    
    earned_points = int(request.order_total * points_per_currency * multiplier)
    new_total = member.get("total_points", 0) + earned_points
    new_available = member.get("available_points", 0) + earned_points
    
    # تحديد المستوى الجديد
    new_tier = "bronze"
    for tier in sorted(tiers, key=lambda x: x.get("min_points", 0), reverse=True):
        if new_total >= tier.get("min_points", 0):
            new_tier = tier.get("name_en", "bronze").lower()
            break
    
    await db.loyalty_members.update_one(
        {"id": request.member_id},
        {"$set": {
            "total_points": new_total,
            "available_points": new_available,
            "current_tier": new_tier,
            "lifetime_spending": member.get("lifetime_spending", 0) + request.order_total,
            "total_orders": member.get("total_orders", 0) + 1
        }}
    )
    
    transaction = {
        "id": str(uuid.uuid4()),
        "member_id": request.member_id,
        "order_id": request.order_id,
        "transaction_type": "earn",
        "points": earned_points,
        "description": f"طلب #{request.order_id[:8]}",
        "balance_after": new_available,
        "created_at": datetime.now(timezone.utc).isoformat(),
        "tenant_id": current_user.get("tenant_id")
    }
    await db.loyalty_transactions.insert_one(transaction)
    
    return {"earned_points": earned_points, "new_balance": new_available, "new_tier": new_tier}

@api_router.post("/loyalty/redeem")
async def redeem_points(request: RedeemPointsRequest, current_user: dict = Depends(get_current_user)):
    """استبدال نقاط"""
    member = await db.loyalty_members.find_one({"id": request.member_id, "tenant_id": current_user.get("tenant_id")})
    if not member:
        raise HTTPException(status_code=404, detail="العضو غير موجود")
    
    if member.get("available_points", 0) < request.points_to_redeem:
        raise HTTPException(status_code=400, detail="رصيد النقاط غير كافي")
    
    settings = await db.loyalty_settings.find_one({"tenant_id": current_user.get("tenant_id")})
    min_redeem = settings.get("min_redeem_points", 100) if settings else 100
    currency_per_point = settings.get("currency_per_point", 0.01) if settings else 0.01
    
    if request.points_to_redeem < min_redeem:
        raise HTTPException(status_code=400, detail=f"الحد الأدنى للاستبدال {min_redeem} نقطة")
    
    discount_value = request.points_to_redeem * currency_per_point
    new_available = member.get("available_points", 0) - request.points_to_redeem
    new_redeemed = member.get("redeemed_points", 0) + request.points_to_redeem
    
    await db.loyalty_members.update_one(
        {"id": request.member_id},
        {"$set": {"available_points": new_available, "redeemed_points": new_redeemed}}
    )
    
    transaction = {
        "id": str(uuid.uuid4()),
        "member_id": request.member_id,
        "order_id": request.order_id,
        "transaction_type": "redeem",
        "points": -request.points_to_redeem,
        "description": f"استبدال في طلب #{request.order_id[:8]}",
        "balance_after": new_available,
        "created_at": datetime.now(timezone.utc).isoformat(),
        "tenant_id": current_user.get("tenant_id")
    }
    await db.loyalty_transactions.insert_one(transaction)
    
    return {"discount_value": discount_value, "new_balance": new_available}

@api_router.get("/loyalty/transactions/{member_id}")
async def get_member_transactions(member_id: str, current_user: dict = Depends(get_current_user)):
    """سجل معاملات العضو"""
    transactions = await db.loyalty_transactions.find(
        {"member_id": member_id, "tenant_id": current_user.get("tenant_id")},
        {"_id": 0}
    ).sort("created_at", -1).to_list(100)
    return transactions

# ==================== CUSTOMER REVIEWS ====================

@api_router.get("/customer-reviews")
async def get_customer_reviews(current_user: dict = Depends(get_current_user)):
    """جلب تقييمات العملاء"""
    tenant_id = current_user.get("tenant_id")
    query = {"tenant_id": tenant_id} if tenant_id else {}
    
    reviews = await db.customer_reviews.find(
        query,
        {"_id": 0}
    ).sort("created_at", -1).to_list(100)
    
    return reviews

@api_router.post("/customer-reviews")
async def create_customer_review(review: Dict[str, Any]):
    """إضافة تقييم من العميل (بدون توثيق - للعملاء)"""
    review_doc = {
        "id": str(uuid.uuid4()),
        "order_id": review.get("order_id"),
        "order_number": review.get("order_number"),
        "customer_name": review.get("customer_name"),
        "customer_phone": review.get("customer_phone"),
        "rating": review.get("rating", 5),
        "comment": review.get("comment", ""),
        "food_rating": review.get("food_rating"),
        "service_rating": review.get("service_rating"),
        "speed_rating": review.get("speed_rating"),
        "tenant_id": review.get("tenant_id"),
        "branch_id": review.get("branch_id"),
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    
    await db.customer_reviews.insert_one(review_doc)
    review_doc.pop("_id", None)
    
    return review_doc

# ==================== RECIPES & RAW MATERIALS ROUTES ====================

class RawMaterialCreate(BaseModel):
    name: str
    name_en: Optional[str] = None
    unit: str
    unit_cost: float
    current_stock: float = 0
    min_stock: float = 0
    max_stock: float = 0
    category: str = "general"
    branch_id: Optional[str] = None

class RecipeCreate(BaseModel):
    product_id: str
    ingredients: List[Dict[str, Any]]
    labor_cost: float = 0
    overhead_cost: float = 0
    portions: int = 1
    preparation_time: int = 0
    instructions: Optional[str] = None

MATERIAL_CATEGORIES = [
    {"id": "meat", "name": "لحوم ودواجن"},
    {"id": "seafood", "name": "مأكولات بحرية"},
    {"id": "vegetables", "name": "خضروات"},
    {"id": "fruits", "name": "فواكه"},
    {"id": "dairy", "name": "ألبان وبيض"},
    {"id": "grains", "name": "حبوب ونشويات"},
    {"id": "spices", "name": "توابل وبهارات"},
    {"id": "oils", "name": "زيوت ودهون"},
    {"id": "beverages", "name": "مشروبات"},
    {"id": "packaging", "name": "تغليف"},
    {"id": "general", "name": "عام"}
]

@api_router.get("/recipes/categories")
async def get_material_categories():
    """تصنيفات المواد الخام"""
    return MATERIAL_CATEGORIES

@api_router.get("/recipes/materials")
async def get_raw_materials(category: Optional[str] = None, current_user: dict = Depends(get_current_user)):
    """قائمة المواد الخام"""
    query = {"tenant_id": current_user.get("tenant_id")}
    if category:
        query["category"] = category
    
    materials = await db.raw_materials.find(query, {"_id": 0}).to_list(500)
    return materials

@api_router.post("/recipes/materials")
async def create_raw_material(material: RawMaterialCreate, current_user: dict = Depends(get_current_user)):
    """إضافة مادة خام"""
    new_material = {
        "id": str(uuid.uuid4()),
        **material.model_dump(),
        "tenant_id": current_user.get("tenant_id"),
        "is_active": True,
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    
    await db.raw_materials.insert_one(new_material)
    if "_id" in new_material:
        del new_material["_id"]
    
    return {"message": "تم إضافة المادة", "material": new_material}

@api_router.put("/recipes/materials/{material_id}")
async def update_raw_material(material_id: str, material: RawMaterialCreate, current_user: dict = Depends(get_current_user)):
    """تحديث مادة خام"""
    await db.raw_materials.update_one(
        {"id": material_id, "tenant_id": current_user.get("tenant_id")},
        {"$set": {**material.model_dump(), "updated_at": datetime.now(timezone.utc).isoformat()}}
    )
    return {"message": "تم التحديث"}

@api_router.delete("/recipes/materials/{material_id}")
async def delete_raw_material(material_id: str, current_user: dict = Depends(get_current_user)):
    """حذف مادة خام"""
    await db.raw_materials.delete_one({"id": material_id, "tenant_id": current_user.get("tenant_id")})
    return {"message": "تم الحذف"}

@api_router.get("/recipes")
async def get_recipes(current_user: dict = Depends(get_current_user)):
    """قائمة الوصفات"""
    recipes = await db.recipes.find({"tenant_id": current_user.get("tenant_id")}, {"_id": 0}).to_list(500)
    return recipes

@api_router.post("/recipes")
async def create_recipe(recipe: RecipeCreate, current_user: dict = Depends(get_current_user)):
    """إنشاء وصفة"""
    # جلب معلومات المنتج
    product = await db.products.find_one({"id": recipe.product_id}, {"_id": 0})
    if not product:
        raise HTTPException(status_code=404, detail="المنتج غير موجود")
    
    # جلب المواد الخام
    material_ids = [ing.get("material_id") for ing in recipe.ingredients]
    materials = await db.raw_materials.find({"id": {"$in": material_ids}}, {"_id": 0}).to_list(100)
    materials_dict = {m["id"]: m for m in materials}
    
    # حساب التكلفة
    total_cost = 0
    ingredients_list = []
    
    for ing in recipe.ingredients:
        mat = materials_dict.get(ing.get("material_id"))
        if mat:
            ing_cost = ing.get("quantity", 0) * mat.get("unit_cost", 0)
            total_cost += ing_cost
            ingredients_list.append({
                "material_id": mat["id"],
                "material_name": mat["name"],
                "quantity": ing.get("quantity", 0),
                "unit": mat["unit"],
                "unit_cost": mat["unit_cost"],
                "total_cost": round(ing_cost, 3)
            })
    
    final_cost = total_cost + recipe.labor_cost + recipe.overhead_cost
    selling_price = product.get("price", 0)
    profit_margin = ((selling_price - final_cost) / selling_price * 100) if selling_price > 0 else 0
    
    new_recipe = {
        "id": str(uuid.uuid4()),
        "product_id": recipe.product_id,
        "product_name": product.get("name", ""),
        "ingredients": ingredients_list,
        "total_cost": round(total_cost, 3),
        "labor_cost": recipe.labor_cost,
        "overhead_cost": recipe.overhead_cost,
        "final_cost": round(final_cost, 3),
        "selling_price": selling_price,
        "profit_margin": round(profit_margin, 2),
        "portions": recipe.portions,
        "preparation_time": recipe.preparation_time,
        "instructions": recipe.instructions,
        "tenant_id": current_user.get("tenant_id"),
        "is_active": True,
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    
    await db.recipes.insert_one(new_recipe)
    if "_id" in new_recipe:
        del new_recipe["_id"]
    
    return {"message": "تم إنشاء الوصفة", "recipe": new_recipe}

@api_router.get("/recipes/{recipe_id}")
async def get_recipe(recipe_id: str, current_user: dict = Depends(get_current_user)):
    """تفاصيل وصفة"""
    recipe = await db.recipes.find_one({"id": recipe_id, "tenant_id": current_user.get("tenant_id")}, {"_id": 0})
    if not recipe:
        raise HTTPException(status_code=404, detail="الوصفة غير موجودة")
    return recipe

@api_router.delete("/recipes/{recipe_id}")
async def delete_recipe(recipe_id: str, current_user: dict = Depends(get_current_user)):
    """حذف وصفة"""
    await db.recipes.delete_one({"id": recipe_id, "tenant_id": current_user.get("tenant_id")})
    return {"message": "تم الحذف"}

@api_router.get("/recipes/alerts/low-stock")
async def get_low_stock_alerts(current_user: dict = Depends(get_current_user)):
    """تنبيهات المخزون المنخفض"""
    materials = await db.raw_materials.find({
        "tenant_id": current_user.get("tenant_id"),
        "is_active": True,
        "$expr": {"$lte": ["$current_stock", "$min_stock"]}
    }, {"_id": 0}).to_list(100)
    
    alerts = []
    for mat in materials:
        severity = "critical" if mat.get("current_stock", 0) == 0 else "warning"
        alerts.append({
            "material_id": mat["id"],
            "material_name": mat["name"],
            "current_stock": mat.get("current_stock", 0),
            "min_stock": mat.get("min_stock", 0),
            "unit": mat.get("unit", ""),
            "severity": severity
        })
    
    return alerts

# ==================== INVOICE & PRINTING ROUTES ====================

class PrinterCreate(BaseModel):
    name: str
    printer_type: str = "thermal"
    paper_width: int = 80
    connection_type: str = "network"
    ip_address: Optional[str] = None
    port: int = 9100
    branch_id: str
    is_default: bool = False

class InvoiceTemplateCreate(BaseModel):
    name: str
    template_type: str = "receipt"
    show_logo: bool = True
    logo_url: Optional[str] = None
    business_name: str = ""
    business_name_en: Optional[str] = None
    address: Optional[str] = None
    phone: Optional[str] = None
    tax_number: Optional[str] = None
    footer_text: Optional[str] = None
    footer_text_en: Optional[str] = None
    show_qr_code: bool = False
    paper_width: int = 80
    branch_id: Optional[str] = None
    is_default: bool = False

@api_router.get("/invoices/printers")
async def get_printers(current_user: dict = Depends(get_current_user)):
    """قائمة الطابعات"""
    printers = await db.printers.find({"tenant_id": current_user.get("tenant_id")}, {"_id": 0}).to_list(50)
    return printers

@api_router.post("/invoices/printers")
async def create_printer(printer: PrinterCreate, current_user: dict = Depends(get_current_user)):
    """إضافة طابعة"""
    new_printer = {
        "id": str(uuid.uuid4()),
        **printer.model_dump(),
        "tenant_id": current_user.get("tenant_id"),
        "is_active": True,
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    
    await db.printers.insert_one(new_printer)
    if "_id" in new_printer:
        del new_printer["_id"]
    
    return {"message": "تم إضافة الطابعة", "printer": new_printer}

@api_router.delete("/invoices/printers/{printer_id}")
async def delete_printer(printer_id: str, current_user: dict = Depends(get_current_user)):
    """حذف طابعة"""
    await db.printers.delete_one({"id": printer_id, "tenant_id": current_user.get("tenant_id")})
    return {"message": "تم الحذف"}

@api_router.get("/invoices/templates")
async def get_invoice_templates(current_user: dict = Depends(get_current_user)):
    """قائمة قوالب الفواتير"""
    templates = await db.invoice_templates.find({"tenant_id": current_user.get("tenant_id")}, {"_id": 0}).to_list(50)
    return templates

@api_router.post("/invoices/templates")
async def create_invoice_template(template: InvoiceTemplateCreate, current_user: dict = Depends(get_current_user)):
    """إنشاء قالب فاتورة"""
    new_template = {
        "id": str(uuid.uuid4()),
        **template.model_dump(),
        "tenant_id": current_user.get("tenant_id"),
        "is_active": True,
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    
    await db.invoice_templates.insert_one(new_template)
    if "_id" in new_template:
        del new_template["_id"]
    
    return {"message": "تم إنشاء القالب", "template": new_template}

@api_router.put("/invoices/templates/{template_id}")
async def update_invoice_template(template_id: str, template: InvoiceTemplateCreate, current_user: dict = Depends(get_current_user)):
    """تحديث قالب فاتورة"""
    await db.invoice_templates.update_one(
        {"id": template_id, "tenant_id": current_user.get("tenant_id")},
        {"$set": {**template.model_dump(), "updated_at": datetime.now(timezone.utc).isoformat()}}
    )
    return {"message": "تم التحديث"}

@api_router.delete("/invoices/templates/{template_id}")
async def delete_invoice_template(template_id: str, current_user: dict = Depends(get_current_user)):
    """حذف قالب"""
    await db.invoice_templates.delete_one({"id": template_id, "tenant_id": current_user.get("tenant_id")})
    return {"message": "تم الحذف"}

@api_router.post("/invoices/print/{order_id}")
async def print_invoice(order_id: str, print_type: str = "customer", printer_id: Optional[str] = None, current_user: dict = Depends(get_current_user)):
    """طباعة فاتورة مع تطبيق صلاحيات الطابعة"""
    query = build_tenant_query(current_user, {"id": order_id})
    order = await db.orders.find_one(query, {"_id": 0})
    if not order:
        raise HTTPException(status_code=404, detail="الطلب غير موجود")
    
    # جلب إعدادات الطابعة إذا تم تحديدها
    printer_settings = None
    if printer_id:
        printer_query = build_tenant_query(current_user, {"id": printer_id})
        printer_settings = await db.printers.find_one(printer_query, {"_id": 0})
    
    # جلب قالب الفاتورة
    template = await db.invoice_templates.find_one({
        "tenant_id": current_user.get("tenant_id"),
        "template_type": "receipt" if print_type == "customer" else print_type,
        "is_default": True
    }, {"_id": 0})
    
    if not template:
        template = {
            "business_name": "المطعم",
            "show_logo": True,
            "footer_text": "شكراً لزيارتكم",
            "paper_width": 80
        }
    
    # تحضير الأصناف بناءً على صلاحيات الطابعة
    items = order.get("items", [])
    show_prices = True
    print_mode = "full_receipt"
    print_individual_items = False
    
    if printer_settings:
        show_prices = printer_settings.get("show_prices", True)
        print_mode = printer_settings.get("print_mode", "full_receipt")
        print_individual_items = printer_settings.get("print_individual_items", False)
    
    # تطبيق صلاحيات الطابعة على الأصناف
    processed_items = []
    for item in items:
        processed_item = {
            "name": item.get("product_name") or item.get("name"),
            "name_en": item.get("product_name_en") or item.get("name_en"),
            "quantity": item.get("quantity", 1),
            "notes": item.get("notes"),
        }
        
        # إضافة الأسعار فقط إذا كان مسموحاً
        if show_prices:
            processed_item["price"] = item.get("price", 0)
            processed_item["total"] = item.get("price", 0) * item.get("quantity", 1)
        
        processed_items.append(processed_item)
    
    # تحضير بيانات الطلب للطباعة
    print_data = {
        "order_number": order.get("order_number", order["id"][:8]),
        "date": datetime.fromisoformat(order.get("created_at", datetime.now(timezone.utc).isoformat())).strftime("%Y-%m-%d %H:%M"),
        "table_number": order.get("table_number"),
        "customer_name": order.get("customer_name"),
        "items": processed_items,
        "order_type": order.get("order_type", "dine_in"),
        "notes": order.get("notes"),
        "print_mode": print_mode,
        "show_prices": show_prices,
        "print_individual_items": print_individual_items
    }
    
    # إضافة المعلومات المالية فقط للفاتورة الكاملة وإذا كان عرض الأسعار مسموحاً
    if print_mode == "full_receipt" and show_prices:
        print_data["subtotal"] = order.get("subtotal", 0)
        print_data["discount"] = order.get("discount", 0)
        print_data["tax"] = order.get("tax", 0)
        print_data["total"] = order.get("total", 0)
        print_data["payment_method"] = order.get("payment_method", "cash")
    
    # إذا كانت طباعة كل صنف على حدة، نجهز مصفوفة من الطباعات
    print_jobs = []
    if print_individual_items:
        for item in processed_items:
            job = {
                "order_number": print_data["order_number"],
                "date": print_data["date"],
                "table_number": print_data["table_number"],
                "items": [item],
                "notes": print_data["notes"],
                "is_individual": True
            }
            print_jobs.append(job)
    else:
        print_jobs = [print_data]
    
    return {
        "message": "جاهز للطباعة",
        "print_data": print_data,
        "print_jobs": print_jobs,
        "template": template,
        "printer_settings": {
            "print_mode": print_mode,
            "show_prices": show_prices,
            "print_individual_items": print_individual_items
        }
    }

@api_router.get("/invoices/auto-print/{order_id}")
async def get_auto_print_data(order_id: str, branch_id: Optional[str] = None, current_user: dict = Depends(get_current_user)):
    """
    جلب بيانات الطباعة التلقائية لكل الطابعات النشطة
    يُستخدم عند إنشاء طلب جديد لإرسال بيانات الطباعة لكل طابعة
    """
    query = build_tenant_query(current_user, {"id": order_id})
    order = await db.orders.find_one(query, {"_id": 0})
    if not order:
        raise HTTPException(status_code=404, detail="الطلب غير موجود")
    
    # جلب جميع الطابعات المفعّلة للطباعة التلقائية
    printer_query = build_tenant_query(current_user, {
        "is_active": True,
        "auto_print_on_order": True
    })
    if branch_id:
        printer_query["branch_id"] = branch_id
    elif order.get("branch_id"):
        printer_query["branch_id"] = order.get("branch_id")
    
    printers = await db.printers.find(printer_query, {"_id": 0}).to_list(50)
    
    if not printers:
        return {"message": "لا توجد طابعات مفعّلة للطباعة التلقائية", "printers": []}
    
    # تجهيز بيانات الطباعة لكل طابعة
    print_results = []
    
    for printer in printers:
        show_prices = printer.get("show_prices", True)
        print_mode = printer.get("print_mode", "full_receipt")
        print_individual_items = printer.get("print_individual_items", False)
        
        # تحضير الأصناف بناءً على صلاحيات الطابعة
        processed_items = []
        for item in order.get("items", []):
            processed_item = {
                "name": item.get("product_name") or item.get("name"),
                "name_en": item.get("product_name_en") or item.get("name_en"),
                "quantity": item.get("quantity", 1),
                "notes": item.get("notes"),
            }
            
            if show_prices:
                processed_item["price"] = item.get("price", 0)
                processed_item["total"] = item.get("price", 0) * item.get("quantity", 1)
            
            processed_items.append(processed_item)
        
        # تحضير بيانات الطلب
        print_data = {
            "order_number": order.get("order_number", order["id"][:8]),
            "date": datetime.fromisoformat(order.get("created_at", datetime.now(timezone.utc).isoformat())).strftime("%Y-%m-%d %H:%M"),
            "table_number": order.get("table_number"),
            "customer_name": order.get("customer_name"),
            "items": processed_items,
            "order_type": order.get("order_type", "dine_in"),
            "notes": order.get("notes"),
        }
        
        # إضافة المعلومات المالية للفاتورة الكاملة فقط
        if print_mode == "full_receipt" and show_prices:
            print_data["subtotal"] = order.get("subtotal", 0)
            print_data["discount"] = order.get("discount", 0)
            print_data["tax"] = order.get("tax", 0)
            print_data["total"] = order.get("total", 0)
            print_data["payment_method"] = order.get("payment_method", "cash")
        
        # تجهيز الطباعات
        print_jobs = []
        if print_individual_items:
            for item in processed_items:
                job = {
                    "order_number": print_data["order_number"],
                    "date": print_data["date"],
                    "table_number": print_data["table_number"],
                    "items": [item],
                    "notes": print_data["notes"],
                    "is_individual": True
                }
                print_jobs.append(job)
        else:
            print_jobs = [print_data]
        
        print_results.append({
            "printer": {
                "id": printer.get("id"),
                "name": printer.get("name"),
                "ip_address": printer.get("ip_address"),
                "port": printer.get("port", 9100),
                "printer_type": printer.get("printer_type")
            },
            "settings": {
                "print_mode": print_mode,
                "show_prices": show_prices,
                "print_individual_items": print_individual_items
            },
            "print_jobs": print_jobs
        })
    
    return {
        "message": "بيانات الطباعة جاهزة",
        "order_id": order_id,
        "printers_count": len(print_results),
        "printers": print_results
    }

# ==================== PUSH NOTIFICATIONS ROUTES ====================

class FCMTokenCreate(BaseModel):
    user_id: str
    user_type: str
    token: str
    device_type: str = "web"
    device_id: Optional[str] = None
    branch_id: Optional[str] = None

class SendNotificationRequest(BaseModel):
    target_type: str  # user, role, branch, all
    target_id: Optional[str] = None
    title: str
    body: str
    data: Dict[str, Any] = {}

@api_router.post("/notifications/fcm/register")
async def register_fcm_token(token_data: FCMTokenCreate, current_user: dict = Depends(get_current_user)):
    """تسجيل FCM Token"""
    existing = await db.fcm_tokens.find_one({"token": token_data.token})
    
    if existing:
        await db.fcm_tokens.update_one(
            {"token": token_data.token},
            {"$set": {
                "user_id": token_data.user_id,
                "user_type": token_data.user_type,
                "updated_at": datetime.now(timezone.utc).isoformat()
            }}
        )
    else:
        new_token = {
            "id": str(uuid.uuid4()),
            **token_data.model_dump(),
            "tenant_id": current_user.get("tenant_id"),
            "is_active": True,
            "created_at": datetime.now(timezone.utc).isoformat()
        }
        await db.fcm_tokens.insert_one(new_token)
    
    return {"message": "تم التسجيل"}

@api_router.delete("/notifications/fcm/unregister")
async def unregister_fcm_token(token: str, current_user: dict = Depends(get_current_user)):
    """إلغاء تسجيل FCM Token"""
    await db.fcm_tokens.delete_one({"token": token})
    return {"message": "تم الإلغاء"}

@api_router.post("/notifications/send")
async def send_notification(request: SendNotificationRequest, background_tasks: BackgroundTasks, current_user: dict = Depends(get_current_user)):
    """إرسال إشعار"""
    # جلب Tokens المستهدفة
    query = {"tenant_id": current_user.get("tenant_id"), "is_active": True}
    
    if request.target_type == "user":
        query["user_id"] = request.target_id
    elif request.target_type == "role":
        query["user_type"] = request.target_id
    elif request.target_type == "branch":
        query["branch_id"] = request.target_id
    
    tokens = await db.fcm_tokens.find(query, {"token": 1, "_id": 0}).to_list(1000)
    token_list = [t["token"] for t in tokens]
    
    if not token_list:
        return {"message": "لا توجد أجهزة مسجلة", "sent": 0}
    
    # TODO: إرسال عبر Firebase
    # في الوقت الحالي نسجل الإشعار فقط
    
    notification_log = {
        "id": str(uuid.uuid4()),
        "target_type": request.target_type,
        "target_id": request.target_id,
        "title": request.title,
        "body": request.body,
        "data": request.data,
        "sent_count": len(token_list),
        "status": "sent",
        "tenant_id": current_user.get("tenant_id"),
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    await db.notification_logs.insert_one(notification_log)
    
    return {"message": "تم الإرسال", "sent": len(token_list)}

@api_router.get("/notifications/logs")
async def get_notification_logs(current_user: dict = Depends(get_current_user)):
    """سجل الإشعارات"""
    logs = await db.notification_logs.find(
        {"tenant_id": current_user.get("tenant_id")},
        {"_id": 0}
    ).sort("created_at", -1).to_list(100)
    return logs

# Helper function to send notification on new order (called from order creation)
async def notify_new_order(order: dict, tenant_id: str):
    """إرسال إشعار طلب جديد"""
    try:
        # جلب tokens السائقين والموظفين
        tokens = await db.fcm_tokens.find({
            "tenant_id": tenant_id,
            "user_type": {"$in": ["driver", "admin", "staff"]},
            "is_active": True
        }).to_list(100)
        
        if tokens:
            # TODO: إرسال عبر Firebase
            logger.info(f"Would send notification for order {order.get('id')} to {len(tokens)} devices")
    except Exception as e:
        logger.error(f"Failed to send notification: {e}")


# ==================== COMPREHENSIVE DASHBOARD STATS ====================

@api_router.get("/dashboard/stats")
async def get_dashboard_stats(
    branch_id: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    """إحصائيات Dashboard الشاملة - يومي/أسبوعي/شهري/إجمالي"""
    tenant_id = get_user_tenant_id(current_user)
    
    # تحديد الفلتر الأساسي
    base_query = {"status": {"$ne": OrderStatus.CANCELLED}}
    
    if tenant_id:
        base_query["tenant_id"] = tenant_id
    else:
        base_query["$or"] = [{"tenant_id": {"$exists": False}}, {"tenant_id": None}]
    
    # فلترة الفرع
    user_branch_id = current_user.get("branch_id")
    user_role = current_user.get("role")
    
    if user_branch_id and user_role not in [UserRole.SUPER_ADMIN, UserRole.ADMIN, UserRole.MANAGER]:
        base_query["branch_id"] = user_branch_id
    elif branch_id:
        base_query["branch_id"] = branch_id
    
    # حساب التواريخ
    now = datetime.now(timezone.utc)
    today = now.strftime('%Y-%m-%d')
    week_ago = (now - timedelta(days=7)).strftime('%Y-%m-%d')
    month_ago = (now - timedelta(days=30)).strftime('%Y-%m-%d')
    
    # استعلامات الفترات المختلفة
    async def get_period_stats(start_date: Optional[str] = None):
        query = base_query.copy()
        if start_date:
            query["created_at"] = {"$gte": start_date}
        
        orders = await db.orders.find(query, {"_id": 0, "total": 1, "total_cost": 1, "profit": 1, "payment_method": 1}).to_list(10000)
        
        total_sales = sum(o.get("total", 0) for o in orders)
        total_orders = len(orders)
        avg_order = total_sales / total_orders if total_orders > 0 else 0
        total_profit = sum(o.get("profit", 0) for o in orders)
        
        by_payment = {}
        for o in orders:
            pm = o.get("payment_method", "cash")
            by_payment[pm] = by_payment.get(pm, 0) + o.get("total", 0)
        
        return {
            "total_sales": total_sales,
            "total_orders": total_orders,
            "average_order_value": avg_order,
            "total_profit": total_profit,
            "by_payment_method": by_payment
        }
    
    # جلب جميع الإحصائيات بالتوازي
    today_stats, week_stats, month_stats, all_stats = await asyncio.gather(
        get_period_stats(today),
        get_period_stats(week_ago),
        get_period_stats(month_ago),
        get_period_stats(None)
    )
    
    # جلب آخر الطلبات
    recent_query = base_query.copy()
    recent_orders = await db.orders.find(recent_query, {"_id": 0}).sort("created_at", -1).limit(10).to_list(10)
    
    # جلب معلومات الوردية الحالية
    shift_query = {"status": "open"}
    if branch_id:
        shift_query["branch_id"] = branch_id
    current_shift = await db.shifts.find_one(shift_query, {"_id": 0})
    
    # جلب الطلبات المعلقة
    pending_query = base_query.copy()
    pending_query["status"] = {"$in": ["pending", "preparing", "ready"]}
    pending_count = await db.orders.count_documents(pending_query)
    
    return {
        "today": today_stats,
        "week": week_stats,
        "month": month_stats,
        "all_time": all_stats,
        "recent_orders": recent_orders,
        "current_shift": current_shift,
        "pending_orders_count": pending_count,
        "current_date": today,
        "server_time": now.isoformat()
    }


# ==================== AUTO DAY CLOSE SYSTEM (نظام الترحيل التلقائي) ====================

class DayCloseRequest(BaseModel):
    force: bool = False  # إغلاق إجباري حتى مع وجود طلبات معلقة
    notes: Optional[str] = None

class DayCloseResponse(BaseModel):
    success: bool
    message: str
    day_summary: Optional[Dict] = None
    pending_orders: Optional[List] = None
    shifts_closed: int = 0

@api_router.get("/day-management/status")
async def get_day_status(
    branch_id: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    """حالة اليوم الحالي - هل يوجد ورديات مفتوحة، طلبات معلقة، إلخ"""
    tenant_id = get_user_tenant_id(current_user)
    
    base_query = {}
    if tenant_id:
        base_query["tenant_id"] = tenant_id
    
    if branch_id:
        base_query["branch_id"] = branch_id
    
    # الورديات المفتوحة
    shift_query = {**base_query, "status": "open"}
    open_shifts = await db.shifts.find(shift_query, {"_id": 0}).to_list(100)
    
    # الطلبات المعلقة
    pending_query = {**base_query, "status": {"$in": ["pending", "preparing", "ready"]}}
    pending_orders = await db.orders.find(pending_query, {"_id": 0, "id": 1, "order_number": 1, "status": 1, "total": 1, "created_at": 1}).to_list(100)
    
    # آخر إغلاق يومي
    last_close = await db.day_closures.find_one(base_query, sort=[("closed_at", -1)])
    if last_close:
        last_close.pop("_id", None)
    
    # حساب عمر الوردية (بالساعات)
    oldest_shift_hours = 0
    if open_shifts:
        for shift in open_shifts:
            started = datetime.fromisoformat(shift["started_at"].replace("Z", "+00:00"))
            hours = (datetime.now(timezone.utc) - started).total_seconds() / 3600
            if hours > oldest_shift_hours:
                oldest_shift_hours = hours
    
    return {
        "open_shifts": open_shifts,
        "open_shifts_count": len(open_shifts),
        "pending_orders": pending_orders,
        "pending_orders_count": len(pending_orders),
        "last_day_close": last_close,
        "oldest_shift_hours": round(oldest_shift_hours, 1),
        "should_close": oldest_shift_hours >= 24,  # إشعار إذا مر 24 ساعة
        "can_close": len(pending_orders) == 0  # يمكن الإغلاق إذا لا توجد طلبات معلقة
    }

@api_router.post("/day-management/close", response_model=DayCloseResponse)
async def close_day(
    request: DayCloseRequest,
    branch_id: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    """إغلاق اليوم وترحيل البيانات"""
    tenant_id = get_user_tenant_id(current_user)
    
    base_query = {}
    if tenant_id:
        base_query["tenant_id"] = tenant_id
    
    if branch_id:
        base_query["branch_id"] = branch_id
    
    # التحقق من الطلبات المعلقة
    pending_query = {**base_query, "status": {"$in": ["pending", "preparing", "ready"]}}
    pending_orders = await db.orders.find(pending_query, {"_id": 0}).to_list(100)
    
    if pending_orders and not request.force:
        return DayCloseResponse(
            success=False,
            message="يوجد طلبات معلقة يجب إغلاقها أولاً",
            pending_orders=pending_orders
        )
    
    # إغلاق جميع الورديات المفتوحة
    shift_query = {**base_query, "status": "open"}
    open_shifts = await db.shifts.find(shift_query, {"_id": 0}).to_list(100)
    
    shifts_closed = 0
    total_day_sales = 0
    total_day_profit = 0
    total_day_expenses = 0
    
    for shift in open_shifts:
        # حساب إحصائيات الوردية
        orders = await db.orders.find({
            "shift_id": shift["id"],
            "status": {"$ne": OrderStatus.CANCELLED}
        }).to_list(1000)
        
        shift_sales = sum(o.get("total", 0) for o in orders)
        shift_profit = sum(o.get("profit", 0) for o in orders)
        
        expenses = await db.expenses.find({
            "branch_id": shift.get("branch_id"),
            "created_at": {"$gte": shift["started_at"]}
        }).to_list(100)
        shift_expenses = sum(e.get("amount", 0) for e in expenses)
        
        # تحديث الوردية كمغلقة
        await db.shifts.update_one(
            {"id": shift["id"]},
            {"$set": {
                "status": "closed",
                "ended_at": datetime.now(timezone.utc).isoformat(),
                "closed_by": current_user.get("full_name", current_user.get("username")),
                "closed_by_id": current_user.get("id"),
                "auto_closed": not request.force,
                "total_sales": shift_sales,
                "total_expenses": shift_expenses,
                "net_profit": shift_profit - shift_expenses,
                "notes": request.notes or "إغلاق يومي تلقائي"
            }}
        )
        
        shifts_closed += 1
        total_day_sales += shift_sales
        total_day_profit += shift_profit
        total_day_expenses += shift_expenses
    
    # إذا كان هناك طلبات معلقة مع force=True، نغلقها كملغية
    if pending_orders and request.force:
        for order in pending_orders:
            await db.orders.update_one(
                {"id": order["id"]},
                {"$set": {
                    "status": OrderStatus.CANCELLED,
                    "cancelled_reason": "إغلاق يومي إجباري",
                    "cancelled_at": datetime.now(timezone.utc).isoformat(),
                    "cancelled_by": current_user.get("full_name")
                }}
            )
    
    # إنشاء سجل إغلاق اليوم
    day_summary = {
        "id": str(uuid.uuid4()),
        "date": datetime.now(timezone.utc).strftime('%Y-%m-%d'),
        "closed_at": datetime.now(timezone.utc).isoformat(),
        "closed_by": current_user.get("full_name", current_user.get("username")),
        "closed_by_id": current_user.get("id"),
        "branch_id": branch_id,
        "tenant_id": tenant_id,
        "shifts_closed": shifts_closed,
        "total_sales": total_day_sales,
        "total_profit": total_day_profit,
        "total_expenses": total_day_expenses,
        "net_profit": total_day_profit - total_day_expenses,
        "forced_close": request.force,
        "pending_orders_cancelled": len(pending_orders) if request.force else 0,
        "notes": request.notes
    }
    
    await db.day_closures.insert_one(day_summary)
    day_summary.pop("_id", None)
    
    return DayCloseResponse(
        success=True,
        message=f"تم إغلاق اليوم بنجاح - {shifts_closed} وردية",
        day_summary=day_summary,
        shifts_closed=shifts_closed
    )

@api_router.get("/day-management/history")
async def get_day_close_history(
    branch_id: Optional[str] = None,
    limit: int = 30,
    current_user: dict = Depends(get_current_user)
):
    """سجل إغلاقات الأيام السابقة"""
    tenant_id = get_user_tenant_id(current_user)
    
    query = {}
    if tenant_id:
        query["tenant_id"] = tenant_id
    if branch_id:
        query["branch_id"] = branch_id
    
    closures = await db.day_closures.find(query, {"_id": 0}).sort("closed_at", -1).limit(limit).to_list(limit)
    return closures


# ==================== AUTO DAY CLOSE SCHEDULER (المجدول التلقائي) ====================

async def auto_close_old_shifts():
    """إغلاق تلقائي للورديات القديمة (أكثر من 24 ساعة)"""
    try:
        cutoff = (datetime.now(timezone.utc) - timedelta(hours=24)).isoformat()
        
        # البحث عن الورديات المفتوحة لأكثر من 24 ساعة
        old_shifts = await db.shifts.find({
            "status": "open",
            "started_at": {"$lt": cutoff}
        }).to_list(100)
        
        for shift in old_shifts:
            # حساب الإحصائيات
            orders = await db.orders.find({
                "shift_id": shift["id"],
                "status": {"$ne": OrderStatus.CANCELLED}
            }).to_list(1000)
            
            total_sales = sum(o.get("total", 0) for o in orders)
            total_profit = sum(o.get("profit", 0) for o in orders)
            
            # إغلاق الوردية تلقائياً
            await db.shifts.update_one(
                {"id": shift["id"]},
                {"$set": {
                    "status": "closed",
                    "ended_at": datetime.now(timezone.utc).isoformat(),
                    "auto_closed": True,
                    "total_sales": total_sales,
                    "net_profit": total_profit,
                    "notes": "إغلاق تلقائي بعد 24 ساعة"
                }}
            )
            
            logger.info(f"Auto-closed shift {shift['id']} after 24 hours")
        
        if old_shifts:
            logger.info(f"Auto-closed {len(old_shifts)} old shifts")
            
    except Exception as e:
        logger.error(f"Error in auto_close_old_shifts: {e}")


# ==================== DAILY REPORT EMAIL (تقرير يومي بالبريد) ====================

class DailyReportEmailRequest(BaseModel):
    recipient_emails: List[str] = []  # قائمة البريد للإرسال
    include_all_branches: bool = True

@api_router.post("/day-management/send-report")
async def send_daily_report_email(
    request: DailyReportEmailRequest,
    branch_id: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    """إرسال التقرير اليومي عبر البريد الإلكتروني"""
    if current_user["role"] not in [UserRole.ADMIN, UserRole.MANAGER, UserRole.SUPER_ADMIN]:
        raise HTTPException(status_code=403, detail="غير مصرح")
    
    tenant_id = get_user_tenant_id(current_user)
    today = datetime.now(timezone.utc).strftime('%Y-%m-%d')
    
    # جمع بيانات التقرير
    base_query = {"status": {"$ne": OrderStatus.CANCELLED}}
    if tenant_id:
        base_query["tenant_id"] = tenant_id
    
    # جلب جميع الفروع
    branch_query = {}
    if tenant_id:
        branch_query["tenant_id"] = tenant_id
    branches = await db.branches.find(branch_query, {"_id": 0}).to_list(100)
    
    branches_data = []
    total_sales = 0
    total_orders = 0
    total_expenses = 0
    total_profit = 0
    
    for branch in branches:
        if branch_id and branch["id"] != branch_id:
            continue
        
        # طلبات الفرع اليوم
        orders_query = {**base_query, "branch_id": branch["id"], "created_at": {"$gte": today}}
        orders = await db.orders.find(orders_query, {"_id": 0, "total": 1, "profit": 1}).to_list(1000)
        
        # مصاريف الفرع اليوم
        expenses_query = {"branch_id": branch["id"], "date": {"$gte": today}}
        if tenant_id:
            expenses_query["tenant_id"] = tenant_id
        expenses = await db.expenses.find(expenses_query, {"_id": 0, "amount": 1}).to_list(1000)
        
        branch_sales = sum(o.get("total", 0) for o in orders)
        branch_expenses = sum(e.get("amount", 0) for e in expenses)
        branch_profit = sum(o.get("profit", 0) for o in orders) - branch_expenses
        
        branches_data.append({
            "name": branch["name"],
            "orders": len(orders),
            "sales": branch_sales,
            "expenses": branch_expenses,
            "profit": branch_profit
        })
        
        total_sales += branch_sales
        total_orders += len(orders)
        total_expenses += branch_expenses
        total_profit += branch_profit
    
    # عدد الورديات المغلقة اليوم
    shifts_closed = await db.day_closures.count_documents({
        "closed_at": {"$gte": today},
        **({"tenant_id": tenant_id} if tenant_id else {})
    })
    
    # عدد الطلبات الملغية
    cancelled_query = {**base_query, "status": OrderStatus.CANCELLED, "created_at": {"$gte": today}}
    cancelled_orders = await db.orders.count_documents(cancelled_query)
    
    report_data = {
        "branches": branches_data,
        "total_sales": total_sales,
        "total_orders": total_orders,
        "total_expenses": total_expenses,
        "net_profit": total_profit,
        "shifts_closed": shifts_closed,
        "cancelled_orders": cancelled_orders
    }
    
    # تحديد قائمة المستلمين
    recipient_emails = request.recipient_emails
    if not recipient_emails:
        # استخدام بريد المستخدم الحالي كافتراضي
        recipient_emails = [current_user.get("email")]
    
    # إرسال التقرير
    try:
        from services.email_service import send_daily_report
        result = await send_daily_report(tenant_id, report_data, recipient_emails)
        return {
            "success": True,
            "message": f"تم إرسال التقرير إلى {result['success']} مستلم",
            "report_data": report_data,
            "email_results": result
        }
    except ImportError:
        # إذا لم تكن خدمة البريد متاحة، نرجع البيانات فقط
        return {
            "success": False,
            "message": "خدمة البريد غير متاحة حالياً",
            "report_data": report_data
        }
    except Exception as e:
        logger.error(f"Failed to send daily report: {e}")
        return {
            "success": False,
            "message": str(e),
            "report_data": report_data
        }

@api_router.get("/day-management/report-preview")
async def get_daily_report_preview(
    branch_id: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    """معاينة التقرير اليومي قبل الإرسال"""
    tenant_id = get_user_tenant_id(current_user)
    today = datetime.now(timezone.utc).strftime('%Y-%m-%d')
    
    base_query = {"status": {"$ne": OrderStatus.CANCELLED}, "created_at": {"$gte": today}}
    if tenant_id:
        base_query["tenant_id"] = tenant_id
    if branch_id:
        base_query["branch_id"] = branch_id
    
    # جلب إحصائيات اليوم
    orders = await db.orders.find(base_query, {"_id": 0}).to_list(1000)
    
    expenses_query = {"date": {"$gte": today}}
    if tenant_id:
        expenses_query["tenant_id"] = tenant_id
    if branch_id:
        expenses_query["branch_id"] = branch_id
    expenses = await db.expenses.find(expenses_query, {"_id": 0}).to_list(1000)
    
    total_sales = sum(o.get("total", 0) for o in orders)
    total_profit = sum(o.get("profit", 0) for o in orders)
    total_expenses = sum(e.get("amount", 0) for e in expenses)
    
    return {
        "date": today,
        "total_sales": total_sales,
        "total_orders": len(orders),
        "total_expenses": total_expenses,
        "gross_profit": total_profit,
        "net_profit": total_profit - total_expenses,
        "average_order_value": total_sales / len(orders) if orders else 0
    }

# إضافة مهمة الإغلاق التلقائي عند بدء التطبيق
@app.on_event("startup")
async def start_auto_close_scheduler():
    """بدء مجدول الإغلاق التلقائي"""
    async def scheduler():
        while True:
            await asyncio.sleep(3600)  # كل ساعة
            await auto_close_old_shifts()
    
    asyncio.create_task(scheduler())
    logger.info("✅ Auto day close scheduler started")

# إضافة indexes عند بدء التطبيق
@app.on_event("startup")
async def setup_database_indexes():
    """إعداد indexes لتحسين الأداء"""
    try:
        from services.reliability_service import create_database_indexes
        await create_database_indexes(db)
    except Exception as e:
        logger.error(f"Failed to create indexes: {e}")


# ==================== SYSTEM HEALTH & RELIABILITY APIS ====================

@api_router.get("/system/health")
async def health_check():
    """فحص صحة النظام - لا يحتاج توثيق"""
    try:
        from services.reliability_service import SystemHealth
        return await SystemHealth.full_health_check(db)
    except Exception as e:
        return {"status": "error", "message": str(e)}

@api_router.get("/system/stats")
async def get_system_stats(current_user: dict = Depends(get_current_user)):
    """إحصائيات النظام"""
    if current_user["role"] not in [UserRole.ADMIN, UserRole.SUPER_ADMIN]:
        raise HTTPException(status_code=403, detail="غير مصرح")
    
    try:
        from services.reliability_service import get_database_stats
        db_stats = await get_database_stats(db)
        
        # إحصائيات إضافية
        tenant_id = get_user_tenant_id(current_user)
        query = {}
        if tenant_id:
            query["tenant_id"] = tenant_id
        
        stats = {
            "database": db_stats,
            "business": {
                "total_orders": await db.orders.count_documents(query),
                "total_products": await db.products.count_documents(query),
                "total_customers": await db.customers.count_documents(query) if "customers" in await db.list_collection_names() else 0,
                "total_employees": await db.employees.count_documents(query),
                "total_branches": await db.branches.count_documents(query),
                "active_shifts": await db.shifts.count_documents({**query, "status": "open"})
            },
            "capacity": {
                "orders_limit": 1000000,
                "products_limit": 100000,
                "users_limit": 10000,
                "status": "healthy"
            }
        }
        
        # تحديد حالة السعة
        if stats["business"]["total_orders"] > 500000:
            stats["capacity"]["status"] = "warning"
        if stats["business"]["total_orders"] > 900000:
            stats["capacity"]["status"] = "critical"
        
        return stats
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@api_router.post("/system/backup")
async def create_backup(current_user: dict = Depends(get_current_user)):
    """إنشاء نسخة احتياطية"""
    if current_user["role"] not in [UserRole.SUPER_ADMIN]:
        raise HTTPException(status_code=403, detail="غير مصرح - Super Admin فقط")
    
    try:
        from services.reliability_service import full_backup
        result = await full_backup(db)
        return {
            "success": True,
            "message": f"تم النسخ الاحتياطي: {len(result['success'])} مجموعة",
            "details": result
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@api_router.get("/system/backup/list")
async def list_backups(current_user: dict = Depends(get_current_user)):
    """قائمة النسخ الاحتياطية"""
    if current_user["role"] not in [UserRole.ADMIN, UserRole.SUPER_ADMIN]:
        raise HTTPException(status_code=403, detail="غير مصرح")
    
    import os
    backup_path = "/app/backups"
    
    if not os.path.exists(backup_path):
        return {"backups": [], "message": "لا توجد نسخ احتياطية"}
    
    backups = []
    for f in os.listdir(backup_path):
        if f.endswith('.json'):
            file_path = os.path.join(backup_path, f)
            stat = os.stat(file_path)
            backups.append({
                "filename": f,
                "size_mb": round(stat.st_size / (1024*1024), 2),
                "created_at": datetime.fromtimestamp(stat.st_mtime).isoformat()
            })
    
    backups.sort(key=lambda x: x["created_at"], reverse=True)
    return {"backups": backups[:50]}  # آخر 50 نسخة


# ==================== CUSTOMER MENU APP APIs ====================

class CustomerRegister(BaseModel):
    name: str
    phone: str
    email: Optional[str] = None
    address: Optional[str] = None
    password: Optional[str] = None

class CustomerLogin(BaseModel):
    phone: str
    password: Optional[str] = None

class CustomerOrderItem(BaseModel):
    product_id: str
    quantity: int
    notes: Optional[str] = None

class DeliveryLocation(BaseModel):
    lat: float
    lng: float

class CustomerOrderCreate(BaseModel):
    items: List[CustomerOrderItem]
    delivery_address: str
    delivery_notes: Optional[str] = None
    delivery_location: Optional[DeliveryLocation] = None
    payment_method: str = "cash"
    customer_name: Optional[str] = None
    customer_phone: Optional[str] = None
    branch_id: Optional[str] = None

def generate_menu_slug(name: str) -> str:
    """إنشاء slug من اسم المطعم"""
    import re
    slug = name.lower().replace(" ", "-").replace("_", "-")
    slug = re.sub(r'[^a-z0-9\-]', '', slug)
    return slug or "menu"

def hash_customer_password(password: str) -> str:
    """تشفير كلمة مرور العميل"""
    import hashlib
    return hashlib.sha256(password.encode()).hexdigest()


@api_router.get("/customer/restaurants")
async def get_customer_restaurants():
    """جلب قائمة المطاعم المتاحة للعملاء"""
    # جلب جميع المستأجرين النشطين الذين لديهم menu_slug
    tenants = await db.tenants.find(
        {"menu_slug": {"$ne": None, "$exists": True}},
        {"_id": 0}
    ).to_list(100)
    
    restaurants = []
    for tenant in tenants:
        # التحقق من أن ميزة قائمة الطعام مفعلة
        enabled_features = tenant.get("enabled_features", {})
        if enabled_features.get("showCustomerMenu") == False:
            continue  # تخطي هذا المطعم
        
        # جلب إعدادات المطعم
        settings = await db.settings.find_one(
            {"tenant_id": tenant.get("id"), "type": "restaurant"},
            {"_id": 0}
        )
        
        # جلب عدد الفروع
        branches_count = await db.branches.count_documents({"tenant_id": tenant.get("id")})
        
        restaurant_data = settings.get("value", {}) if settings else {}
        
        restaurants.append({
            "id": tenant.get("id"),
            "name": restaurant_data.get("name") or tenant.get("name", "مطعم"),
            "menu_slug": tenant.get("menu_slug"),
            "logo": restaurant_data.get("logo"),
            "description": restaurant_data.get("description"),
            "address": restaurant_data.get("address"),
            "branches_count": branches_count
        })
    
    return restaurants


@api_router.get("/customer/menu/{tenant_id}")
async def get_customer_menu(tenant_id: str):
    """جلب قائمة الطعام للعملاء - بدون توثيق"""
    # البحث عن tenant
    tenant = await db.tenants.find_one(
        {"$or": [{"id": tenant_id}, {"menu_slug": tenant_id}]},
        {"_id": 0}
    )
    
    if not tenant:
        # إذا لم يوجد tenant، نستخدم tenant_id كـ query للمنتجات
        # هذا للتوافق مع الأنظمة التي لا تستخدم tenants
        tenant = {"id": tenant_id, "name": "المطعم"}
    
    # التحقق من أن ميزة قائمة الطعام مفعلة
    enabled_features = tenant.get("enabled_features", {})
    if enabled_features.get("showCustomerMenu") == False:
        raise HTTPException(status_code=403, detail="قائمة الطعام غير متاحة لهذا المطعم")
    
    tid = tenant.get("id", tenant_id)
    
    # جلب الفئات - فقط للعميل المحدد
    categories = await db.categories.find(
        {"tenant_id": tid},
        {"_id": 0}
    ).sort("sort_order", 1).to_list(100)
    
    # جلب المنتجات - فقط للعميل المحدد
    products = await db.products.find(
        {"tenant_id": tid, "is_available": {"$ne": False}},
        {"_id": 0}
    ).to_list(500)
    
    # جلب الفروع - فقط للعميل المحدد مع إخفاء الفروع الافتراضية
    default_branch_names = ["الفرع الرئيسي", "Main Branch", "الفرع الثاني", "فرع المالك الرئيسي"]
    branches = await db.branches.find(
        {
            "tenant_id": tid, 
            "is_active": {"$ne": False},
            "name": {"$nin": default_branch_names}
        },
        {"_id": 0}
    ).to_list(50)
    
    # إذا لم توجد فروع حقيقية، لا نُنشئ فرع افتراضي للعملاء
    # بل نعرض رسالة أنه لا توجد فروع متاحة
    
    # جلب الإعدادات - للحصول على الشعار والاسم
    settings = await db.tenant_settings.find_one({"tenant_id": tid}, {"_id": 0}) or {}
    
    # جلب إعدادات المطعم الرئيسية (fallback)
    main_settings = await db.settings.find_one({"tenant_id": tid}, {"_id": 0}) or {}
    
    # تحديد الشعار والاسم - الأولوية: tenant -> settings -> main_settings
    # ملاحظة: في جدول tenants الحقل هو logo_url وليس logo
    restaurant_logo = tenant.get("logo_url") or tenant.get("logo") or settings.get("restaurant_logo") or main_settings.get("restaurant_logo", "")
    restaurant_name = tenant.get("name") or tenant.get("name_ar") or settings.get("restaurant_name") or main_settings.get("restaurant_name", "المطعم")
    
    return {
        "restaurant": {
            "id": tid,
            "name": restaurant_name,
            "logo": restaurant_logo,
            "description": tenant.get("description", ""),
            "phone": tenant.get("phone", ""),
            "address": tenant.get("address", ""),
            "delivery_fee": settings.get("delivery_fee", 0),
            "min_order": settings.get("min_order", 0),
            "payment_methods": settings.get("payment_methods", ["cash"]),
            "menu_slug": tenant.get("menu_slug", tid)
        },
        "categories": categories,
        "products": products,
        "branches": branches
    }

@api_router.post("/customer/auth/register/{tenant_id}")
async def register_customer(tenant_id: str, data: CustomerRegister):
    """تسجيل عميل جديد"""
    # التحقق من عدم وجود العميل
    existing = await db.customers.find_one({
        "phone": data.phone,
        "$or": [{"tenant_id": tenant_id}, {"tenant_id": {"$exists": False}}]
    })
    
    if existing:
        raise HTTPException(status_code=400, detail="رقم الهاتف مسجل بالفعل")
    
    customer = {
        "id": str(uuid.uuid4()),
        "name": data.name,
        "phone": data.phone,
        "email": data.email,
        "address": data.address,
        "password": hash_customer_password(data.password) if data.password else None,
        "tenant_id": tenant_id,
        "total_orders": 0,
        "total_spent": 0,
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    
    await db.customers.insert_one(customer)
    customer.pop("_id", None)
    customer.pop("password", None)
    
    # إنشاء token
    import secrets
    token = secrets.token_urlsafe(32)
    await db.customer_tokens.insert_one({
        "token": token,
        "customer_id": customer["id"],
        "tenant_id": tenant_id,
        "created_at": datetime.now(timezone.utc).isoformat()
    })
    
    return {"customer": customer, "token": token}

@api_router.post("/customer/auth/login/{tenant_id}")
async def login_customer(tenant_id: str, data: CustomerLogin):
    """تسجيل دخول العميل"""
    customer = await db.customers.find_one({
        "phone": data.phone,
        "$or": [{"tenant_id": tenant_id}, {"tenant_id": {"$exists": False}}]
    }, {"_id": 0})
    
    if not customer:
        raise HTTPException(status_code=401, detail="رقم الهاتف غير مسجل")
    
    # التحقق من كلمة المرور إذا كانت موجودة
    if customer.get("password") and data.password:
        if hash_customer_password(data.password) != customer["password"]:
            raise HTTPException(status_code=401, detail="كلمة المرور غير صحيحة")
    
    # إنشاء token
    import secrets
    token = secrets.token_urlsafe(32)
    await db.customer_tokens.insert_one({
        "token": token,
        "customer_id": customer["id"],
        "tenant_id": tenant_id,
        "created_at": datetime.now(timezone.utc).isoformat()
    })
    
    customer.pop("password", None)
    return {"customer": customer, "token": token}

async def get_customer_from_token(token: str):
    """جلب العميل من token"""
    if not token:
        return None
    
    token_doc = await db.customer_tokens.find_one({"token": token})
    if not token_doc:
        return None
    
    customer = await db.customers.find_one(
        {"id": token_doc["customer_id"]},
        {"_id": 0, "password": 0}
    )
    return customer

@api_router.post("/customer/order/{tenant_id}")
async def create_customer_order(
    tenant_id: str,
    order: CustomerOrderCreate,
    customer_token: Optional[str] = None
):
    """إنشاء طلب من تطبيق العميل"""
    # Resolve menu_slug to actual tenant_id if needed
    tenant = await db.tenants.find_one(
        {"$or": [{"id": tenant_id}, {"menu_slug": tenant_id}]},
        {"_id": 0, "id": 1}
    )
    if tenant:
        tenant_id = tenant["id"]
    
    # جلب العميل
    customer = None
    if customer_token:
        customer = await get_customer_from_token(customer_token)
    
    # حساب المجموع
    total = 0
    total_cost = 0
    order_items = []
    
    for item in order.items:
        product = await db.products.find_one({"id": item.product_id}, {"_id": 0})
        if not product:
            raise HTTPException(status_code=400, detail=f"المنتج غير موجود: {item.product_id}")
        
        item_total = product.get("price", 0) * item.quantity
        item_cost = product.get("cost", 0) * item.quantity
        
        order_items.append({
            "product_id": item.product_id,
            "product_name": product.get("name"),
            "name": product.get("name"),
            "name_en": product.get("name_en"),
            "price": product.get("price", 0),
            "quantity": item.quantity,
            "total": item_total,
            "notes": item.notes
        })
        
        total += item_total
        total_cost += item_cost
    
    # جلب رسوم التوصيل
    settings = await db.tenant_settings.find_one({"tenant_id": tenant_id}) or {}
    delivery_fee = settings.get("delivery_fee", 0)
    
    # إنشاء رقم الطلب
    today = datetime.now(timezone.utc).strftime('%Y%m%d')
    count = await db.orders.count_documents({"created_at": {"$gte": datetime.now(timezone.utc).strftime('%Y-%m-%d')}})
    order_number = int(f"{today[-4:]}{count + 1:04d}")
    
    # تحديد الفرع
    branch_id = order.branch_id
    if not branch_id:
        # استخدام أول فرع نشط
        branch = await db.branches.find_one({"is_active": {"$ne": False}}, {"id": 1})
        branch_id = branch["id"] if branch else None
    
    # إنشاء الطلب
    order_doc = {
        "id": str(uuid.uuid4()),
        "order_number": order_number,
        "tenant_id": tenant_id,
        "branch_id": branch_id,
        "customer_id": customer["id"] if customer else None,
        "customer_name": customer["name"] if customer else order.customer_name,
        "customer_phone": customer["phone"] if customer else order.customer_phone,
        "delivery_address": order.delivery_address,
        "delivery_notes": order.delivery_notes,
        "delivery_location": order.delivery_location.model_dump() if hasattr(order, 'delivery_location') and order.delivery_location else None,
        "items": order_items,
        "subtotal": total,
        "delivery_fee": delivery_fee,
        "discount": 0,
        "tax": 0,
        "total": total + delivery_fee,
        "total_cost": total_cost,
        "profit": total - total_cost,
        "payment_method": order.payment_method,
        "payment_status": "paid" if order.payment_method in ["card", "zain_cash"] else "pending",
        "status": "pending",
        "order_type": "delivery",
        "source": "customer_app",
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    
    await db.orders.insert_one(order_doc)
    order_doc.pop("_id", None)
    
    # تحديث إحصائيات العميل
    if customer:
        await db.customers.update_one(
            {"id": customer["id"]},
            {"$inc": {"total_orders": 1, "total_spent": order_doc["total"]}}
        )
    
    return {
        "success": True,
        "message": "تم إنشاء الطلب بنجاح! سيتم التواصل معك قريباً",
        "order": order_doc
    }


@api_router.get("/customer/orders/history")
async def get_customer_order_history(
    tenant_id: str = None,
    phone: str = None
):
    """جلب سجل طلبات العميل بناءً على رقم الهاتف"""
    if not phone:
        return []
    
    query = {"customer_phone": phone}
    if tenant_id:
        # تحويل menu_slug إلى tenant_id إذا لزم الأمر
        tenant = await db.tenants.find_one({"menu_slug": tenant_id})
        if tenant:
            query["tenant_id"] = tenant.get("id")
        else:
            query["tenant_id"] = tenant_id
    
    orders = await db.orders.find(
        query,
        {"_id": 0, "id": 1, "order_number": 1, "items": 1, "total": 1, 
         "status": 1, "created_at": 1, "order_type": 1}
    ).sort("created_at", -1).limit(20).to_list(20)
    
    # إضافة تسميات الحالة بالعربية
    status_labels = {
        "pending": "قيد الانتظار",
        "preparing": "قيد التحضير",
        "ready": "جاهز",
        "out_for_delivery": "في الطريق",
        "delivered": "تم التوصيل",
        "completed": "مكتمل",
        "cancelled": "ملغي"
    }
    
    for order in orders:
        order["status_label"] = status_labels.get(order.get("status"), order.get("status"))
        order["items_count"] = len(order.get("items", []))
    
    return orders



# ==================== الطلبات المفضلة للزبائن ====================

class FavoriteItem(BaseModel):
    product_id: str
    product_name: str
    quantity: int
    price: float
    notes: str = ""

class AddFavoriteRequest(BaseModel):
    tenant_id: str = None
    phone: str
    name: str = None
    items: List[FavoriteItem]

@api_router.post("/customer/favorites/add")
async def add_to_favorites(request: AddFavoriteRequest):
    """إضافة طلب للمفضلة"""
    if not request.phone or not request.items:
        raise HTTPException(status_code=400, detail="رقم الهاتف والمنتجات مطلوبة")
    
    # التحقق من وجود المستأجر
    tenant = None
    if request.tenant_id:
        tenant = await db.tenants.find_one({"menu_slug": request.tenant_id})
        if not tenant:
            tenant = await db.tenants.find_one({"id": request.tenant_id})
    
    actual_tenant_id = tenant.get("id") if tenant else request.tenant_id
    
    favorite = {
        "id": str(uuid.uuid4()),
        "tenant_id": actual_tenant_id,
        "phone": request.phone,
        "name": request.name or f"طلبي المفضل #{datetime.now().strftime('%d/%m')}",
        "items": [item.dict() for item in request.items],
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    
    await db.customer_favorites.insert_one(favorite)
    
    # إرجاع بدون _id
    return {"message": "تمت الإضافة للمفضلة", "favorite": {k: v for k, v in favorite.items() if k != '_id'}}

@api_router.get("/customer/favorites")
async def get_favorites(
    tenant_id: str = None,
    phone: str = None
):
    """جلب الطلبات المفضلة للزبون"""
    if not phone:
        return []
    
    query = {"phone": phone}
    
    if tenant_id:
        tenant = await db.tenants.find_one({"menu_slug": tenant_id})
        if tenant:
            query["tenant_id"] = tenant.get("id")
        else:
            query["tenant_id"] = tenant_id
    
    favorites = await db.customer_favorites.find(
        query,
        {"_id": 0}
    ).sort("created_at", -1).limit(10).to_list(10)
    
    return favorites

@api_router.delete("/customer/favorites/{favorite_id}")
async def remove_from_favorites(
    favorite_id: str,
    phone: str = None
):
    """حذف طلب من المفضلة"""
    if not phone:
        raise HTTPException(status_code=400, detail="رقم الهاتف مطلوب")
    
    result = await db.customer_favorites.delete_one({
        "id": favorite_id,
        "phone": phone
    })
    
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="لم يتم العثور على الطلب المفضل")
    
    return {"message": "تم الحذف من المفضلة"}


# ==================== نظام تقييم الطلبات ====================

class OrderRating(BaseModel):
    order_id: str
    tenant_id: str
    phone: str
    rating: int  # 1-5 نجوم
    comment: Optional[str] = None
    food_quality: Optional[int] = None  # جودة الطعام 1-5
    delivery_speed: Optional[int] = None  # سرعة التوصيل 1-5
    service_quality: Optional[int] = None  # جودة الخدمة 1-5

@api_router.post("/customer/rate-order")
async def rate_order(rating: OrderRating):
    """تقييم طلب من الزبون"""
    if rating.rating < 1 or rating.rating > 5:
        raise HTTPException(status_code=400, detail="التقييم يجب أن يكون من 1 إلى 5")
    
    # التحقق من وجود الطلب
    order = await db.orders.find_one({"id": rating.order_id})
    if not order:
        raise HTTPException(status_code=404, detail="الطلب غير موجود")
    
    # التحقق من أن الطلب مكتمل
    if order.get("status") not in ["delivered", "completed"]:
        raise HTTPException(status_code=400, detail="لا يمكن تقييم طلب غير مكتمل")
    
    # التحقق من أن الرقم مطابق
    if order.get("customer_phone") != rating.phone:
        raise HTTPException(status_code=403, detail="رقم الهاتف غير مطابق")
    
    # التحقق من عدم وجود تقييم سابق
    existing = await db.order_ratings.find_one({"order_id": rating.order_id})
    if existing:
        raise HTTPException(status_code=400, detail="تم تقييم هذا الطلب مسبقاً")
    
    # حفظ التقييم
    rating_doc = {
        "id": str(uuid.uuid4()),
        **rating.model_dump(),
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    
    await db.order_ratings.insert_one(rating_doc)
    rating_doc.pop("_id", None)
    
    # تحديث متوسط تقييم الفرع
    await update_branch_rating(order.get("branch_id"))
    
    return {"message": "شكراً لتقييمك! ⭐", "rating": rating_doc}

async def update_branch_rating(branch_id: str):
    """تحديث متوسط تقييم الفرع"""
    if not branch_id:
        return
    
    pipeline = [
        {"$lookup": {
            "from": "orders",
            "localField": "order_id",
            "foreignField": "id",
            "as": "order"
        }},
        {"$unwind": "$order"},
        {"$match": {"order.branch_id": branch_id}},
        {"$group": {
            "_id": None,
            "avg_rating": {"$avg": "$rating"},
            "total_ratings": {"$sum": 1},
            "avg_food": {"$avg": "$food_quality"},
            "avg_delivery": {"$avg": "$delivery_speed"},
            "avg_service": {"$avg": "$service_quality"}
        }}
    ]
    
    result = await db.order_ratings.aggregate(pipeline).to_list(1)
    
    if result:
        await db.branches.update_one(
            {"id": branch_id},
            {"$set": {
                "rating": round(result[0].get("avg_rating", 0), 1),
                "total_ratings": result[0].get("total_ratings", 0),
                "food_rating": round(result[0].get("avg_food", 0) or 0, 1),
                "delivery_rating": round(result[0].get("avg_delivery", 0) or 0, 1),
                "service_rating": round(result[0].get("avg_service", 0) or 0, 1)
            }}
        )

@api_router.get("/customer/order-rating/{order_id}")
async def get_order_rating(order_id: str, phone: str = None):
    """جلب تقييم طلب معين"""
    rating = await db.order_ratings.find_one({"order_id": order_id}, {"_id": 0})
    
    if not rating:
        return {"can_rate": True, "rating": None}
    
    return {"can_rate": False, "rating": rating}

@api_router.get("/ratings/branch/{branch_id}")
async def get_branch_ratings(branch_id: str, limit: int = 20, current_user: dict = Depends(get_current_user)):
    """جلب تقييمات فرع معين"""
    pipeline = [
        {"$lookup": {
            "from": "orders",
            "localField": "order_id",
            "foreignField": "id",
            "as": "order"
        }},
        {"$unwind": "$order"},
        {"$match": {"order.branch_id": branch_id}},
        {"$sort": {"created_at": -1}},
        {"$limit": limit},
        {"$project": {
            "_id": 0,
            "id": 1,
            "rating": 1,
            "comment": 1,
            "food_quality": 1,
            "delivery_speed": 1,
            "service_quality": 1,
            "created_at": 1,
            "customer_name": "$order.customer_name"
        }}
    ]
    
    ratings = await db.order_ratings.aggregate(pipeline).to_list(limit)
    
    # إحصائيات
    branch = await db.branches.find_one({"id": branch_id}, {"_id": 0, "rating": 1, "total_ratings": 1, "food_rating": 1, "delivery_rating": 1, "service_rating": 1})
    
    return {
        "ratings": ratings,
        "stats": branch or {}
    }

@api_router.get("/ratings/tenant-summary")
async def get_tenant_ratings_summary(current_user: dict = Depends(get_current_user)):
    """ملخص تقييمات العميل (المطعم)"""
    tenant_id = get_user_tenant_id(current_user)
    
    pipeline = [
        {"$lookup": {
            "from": "orders",
            "localField": "order_id",
            "foreignField": "id",
            "as": "order"
        }},
        {"$unwind": "$order"},
        {"$match": {"order.tenant_id": tenant_id}},
        {"$group": {
            "_id": None,
            "avg_rating": {"$avg": "$rating"},
            "total_ratings": {"$sum": 1},
            "five_stars": {"$sum": {"$cond": [{"$eq": ["$rating", 5]}, 1, 0]}},
            "four_stars": {"$sum": {"$cond": [{"$eq": ["$rating", 4]}, 1, 0]}},
            "three_stars": {"$sum": {"$cond": [{"$eq": ["$rating", 3]}, 1, 0]}},
            "two_stars": {"$sum": {"$cond": [{"$eq": ["$rating", 2]}, 1, 0]}},
            "one_star": {"$sum": {"$cond": [{"$eq": ["$rating", 1]}, 1, 0]}},
            "avg_food": {"$avg": "$food_quality"},
            "avg_delivery": {"$avg": "$delivery_speed"},
            "avg_service": {"$avg": "$service_quality"}
        }}
    ]
    
    result = await db.order_ratings.aggregate(pipeline).to_list(1)
    
    if not result:
        return {
            "avg_rating": 0,
            "total_ratings": 0,
            "distribution": {"5": 0, "4": 0, "3": 0, "2": 0, "1": 0},
            "categories": {"food": 0, "delivery": 0, "service": 0}
        }
    
    data = result[0]
    return {
        "avg_rating": round(data.get("avg_rating", 0), 1),
        "total_ratings": data.get("total_ratings", 0),
        "distribution": {
            "5": data.get("five_stars", 0),
            "4": data.get("four_stars", 0),
            "3": data.get("three_stars", 0),
            "2": data.get("two_stars", 0),
            "1": data.get("one_star", 0)
        },
        "categories": {
            "food": round(data.get("avg_food", 0) or 0, 1),
            "delivery": round(data.get("avg_delivery", 0) or 0, 1),
            "service": round(data.get("avg_service", 0) or 0, 1)
        }
    }

@api_router.get("/super-admin/ratings-overview")
async def get_super_admin_ratings_overview(current_user: dict = Depends(verify_super_admin)):
    """ملخص تقييمات جميع العملاء للمالك"""
    pipeline = [
        {"$lookup": {
            "from": "orders",
            "localField": "order_id",
            "foreignField": "id",
            "as": "order"
        }},
        {"$unwind": "$order"},
        {"$group": {
            "_id": "$order.tenant_id",
            "avg_rating": {"$avg": "$rating"},
            "total_ratings": {"$sum": 1}
        }}
    ]
    
    tenant_ratings = await db.order_ratings.aggregate(pipeline).to_list(100)
    
    # جلب أسماء العملاء
    result = []
    for tr in tenant_ratings:
        tenant = await db.tenants.find_one({"id": tr["_id"]}, {"_id": 0, "name": 1})
        result.append({
            "tenant_id": tr["_id"],
            "tenant_name": tenant.get("name") if tenant else "Unknown",
            "avg_rating": round(tr.get("avg_rating", 0), 1),
            "total_ratings": tr.get("total_ratings", 0)
        })
    
    # الترتيب حسب التقييم
    result.sort(key=lambda x: x["avg_rating"], reverse=True)
    
    # الإحصائيات العامة
    total_pipeline = [
        {"$group": {
            "_id": None,
            "avg_rating": {"$avg": "$rating"},
            "total_ratings": {"$sum": 1}
        }}
    ]
    total = await db.order_ratings.aggregate(total_pipeline).to_list(1)
    
    return {
        "overall": {
            "avg_rating": round(total[0].get("avg_rating", 0), 1) if total else 0,
            "total_ratings": total[0].get("total_ratings", 0) if total else 0
        },
        "tenants": result
    }


@api_router.get("/customer/orders/{tenant_id}")
async def get_customer_orders(tenant_id: str, customer_token: str):
    """جلب طلبات العميل"""
    customer = await get_customer_from_token(customer_token)
    if not customer:
        raise HTTPException(status_code=401, detail="غير مصرح")
    
    orders = await db.orders.find(
        {"customer_id": customer["id"]},
        {"_id": 0}
    ).sort("created_at", -1).limit(50).to_list(50)
    
    return orders

@api_router.get("/customer/order/{tenant_id}/{order_id}")
async def track_customer_order(tenant_id: str, order_id: str):
    """تتبع حالة الطلب"""
    order = await db.orders.find_one({"id": order_id}, {"_id": 0})
    
    if not order:
        raise HTTPException(status_code=404, detail="الطلب غير موجود")
    
    # جلب معلومات السائق مع الموقع
    driver_info = None
    if order.get("driver_id"):
        driver = await db.drivers.find_one(
            {"id": order["driver_id"]},
            {"_id": 0, "id": 1, "name": 1, "phone": 1, "photo": 1, "current_location": 1, "last_location_update": 1}
        )
        if driver:
            driver_info = driver
    
    # مراحل الطلب
    status_labels = {
        "pending": "قيد الانتظار",
        "preparing": "قيد التحضير",
        "ready": "جاهز للتوصيل",
        "out_for_delivery": "السائق في الطريق",
        "delivered": "تم التسليم",
        "cancelled": "ملغي"
    }
    
    current_status_index = ["pending", "preparing", "ready", "out_for_delivery", "delivered"].index(order["status"]) if order["status"] in ["pending", "preparing", "ready", "out_for_delivery", "delivered"] else 0
    
    timeline = []
    for i, status in enumerate(["pending", "preparing", "ready", "out_for_delivery", "delivered"]):
        timeline.append({
            "status": status,
            "label": status_labels.get(status, status),
            "completed": i <= current_status_index
        })
    
    return {
        "order": order,
        "driver": driver_info,
        "status_label": status_labels.get(order["status"], order["status"]),
        "timeline": timeline
    }

@api_router.get("/customer/menu-link")
async def get_menu_link(request: Request, current_user: dict = Depends(get_current_user)):
    """جلب رابط القائمة للمستخدم"""
    tenant_id = get_user_tenant_id(current_user) or "default"
    
    # التحقق من وجود tenant أو إنشائه
    tenant = await db.tenants.find_one({"id": tenant_id})
    
    # التحقق من صلاحية قائمة الطعام للعملاء
    if tenant:
        enabled_features = tenant.get("enabled_features", {})
        if enabled_features.get("showCustomerMenu") == False:
            raise HTTPException(status_code=403, detail="قائمة الطعام للعملاء غير مفعلة")
    
    if not tenant:
        # جلب اسم المطعم من الإعدادات
        restaurant_settings = await db.settings.find_one({"tenant_id": tenant_id, "type": "restaurant"})
        restaurant_name = None
        if restaurant_settings:
            restaurant_name = restaurant_settings.get("value", {}).get("name")
        
        # إذا لم يوجد اسم، استخدم اسم المستخدم أو البريد الإلكتروني
        if not restaurant_name:
            restaurant_name = current_user.get("restaurant_name") or current_user.get("full_name") or current_user.get("email", "").split("@")[0]
        
        # إنشاء tenant جديد
        tenant = {
            "id": tenant_id,
            "name": restaurant_name,
            "menu_slug": generate_menu_slug(restaurant_name or tenant_id),
            "created_at": datetime.now(timezone.utc).isoformat()
        }
        await db.tenants.insert_one(tenant)
    
    # استخدام نفس النطاق الذي جاء منه الطلب
    # هذا يضمن أن رابط القائمة يعمل على أي نسخة (preview أو production)
    origin = request.headers.get("origin") or request.headers.get("referer", "")
    if origin:
        # استخراج النطاق من origin
        from urllib.parse import urlparse
        parsed = urlparse(origin)
        base_url = f"{parsed.scheme}://{parsed.netloc}"
    else:
        # fallback للـ environment variable
        base_url = os.environ.get('REACT_APP_BACKEND_URL', 'https://bistromate-2.preview.emergentagent.com')
    
    menu_url = f"{base_url}/menu/{tenant.get('menu_slug', tenant_id)}"
    
    return {
        "menu_url": menu_url,
        "tenant_id": tenant_id,
        "menu_slug": tenant.get("menu_slug", tenant_id)
    }


# ==================== DRIVER TRACKING ROUTES ====================

class DriverLocation(BaseModel):
    latitude: float
    longitude: float

@api_router.get("/drivers")
async def get_drivers(current_user: dict = Depends(get_current_user)):
    """جلب قائمة السائقين"""
    tenant_id = get_user_tenant_id(current_user)
    
    drivers = await db.drivers.find(
        {"tenant_id": tenant_id},
        {"_id": 0}
    ).sort("name", 1).to_list(100)
    
    return drivers

class DriverCreateRequest(BaseModel):
    name: str
    phone: str
    branch_id: Optional[str] = None
    pin: str = "1234"

@api_router.post("/drivers")
async def create_driver(
    driver_data: DriverCreateRequest,
    current_user: dict = Depends(get_current_user)
):
    """إنشاء سائق جديد"""
    if current_user["role"] not in [UserRole.ADMIN, UserRole.MANAGER]:
        raise HTTPException(status_code=403, detail="غير مصرح")
    
    tenant_id = get_user_tenant_id(current_user)
    
    # Debug logging
    logger.info(f"Creating driver with PIN: {driver_data.pin}")
    
    driver = {
        "id": str(uuid.uuid4()),
        "tenant_id": tenant_id,
        "branch_id": driver_data.branch_id,
        "name": driver_data.name,
        "phone": driver_data.phone,
        "pin": driver_data.pin,  # الرمز السري للسائق
        "is_active": True,
        "is_available": True,
        "current_location": None,
        "last_location_update": None,
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    
    logger.info(f"Driver dict before insert: pin={driver.get('pin')}")
    
    await db.drivers.insert_one(driver)
    driver.pop("_id", None)
    driver.pop("pin", None)  # لا ترجع PIN في الاستجابة
    
    return {"message": "تم إضافة السائق", "driver": driver}

class DriverUpdateRequest(BaseModel):
    name: Optional[str] = None
    phone: Optional[str] = None
    pin: Optional[str] = None
    is_active: Optional[bool] = None
    is_available: Optional[bool] = None
    branch_id: Optional[str] = None

@api_router.put("/drivers/{driver_id}")
async def update_driver(
    driver_id: str,
    driver_data: DriverUpdateRequest,
    current_user: dict = Depends(get_current_user)
):
    """تحديث بيانات سائق"""
    if current_user["role"] not in [UserRole.ADMIN, UserRole.MANAGER]:
        raise HTTPException(status_code=403, detail="غير مصرح")
    
    tenant_id = get_user_tenant_id(current_user)
    
    update_data = {"updated_at": datetime.now(timezone.utc).isoformat()}
    if driver_data.name: update_data["name"] = driver_data.name
    if driver_data.phone: update_data["phone"] = driver_data.phone
    if driver_data.pin: update_data["pin"] = driver_data.pin  # تحديث الرمز السري
    if driver_data.is_active is not None: update_data["is_active"] = driver_data.is_active
    if driver_data.is_available is not None: update_data["is_available"] = driver_data.is_available
    if driver_data.branch_id: update_data["branch_id"] = driver_data.branch_id
    
    result = await db.drivers.update_one(
        {"id": driver_id, "tenant_id": tenant_id},
        {"$set": update_data}
    )
    
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="السائق غير موجود")
    
    return {"message": "تم تحديث بيانات السائق"}

@api_router.delete("/drivers/{driver_id}")
async def delete_driver(driver_id: str, current_user: dict = Depends(get_current_user)):
    """حذف سائق"""
    if current_user["role"] not in [UserRole.ADMIN, UserRole.MANAGER]:
        raise HTTPException(status_code=403, detail="غير مصرح")
    
    tenant_id = get_user_tenant_id(current_user)
    
    result = await db.drivers.delete_one({"id": driver_id, "tenant_id": tenant_id})
    
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="السائق غير موجود")
    
    return {"message": "تم حذف السائق"}

@api_router.post("/drivers/{driver_id}/location")
async def update_driver_location(
    driver_id: str,
    location: DriverLocation,
    current_user: dict = Depends(get_current_user)
):
    """تحديث موقع السائق (يستخدمها تطبيق السائق)"""
    tenant_id = get_user_tenant_id(current_user)
    
    result = await db.drivers.update_one(
        {"id": driver_id, "tenant_id": tenant_id},
        {"$set": {
            "current_location": {
                "latitude": location.latitude,
                "longitude": location.longitude
            },
            "last_location_update": datetime.now(timezone.utc).isoformat()
        }}
    )
    
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="السائق غير موجود")
    
    return {"message": "تم تحديث الموقع"}

@api_router.get("/drivers/{driver_id}/location")
async def get_driver_location(driver_id: str):
    """جلب موقع السائق (للزبون)"""
    driver = await db.drivers.find_one(
        {"id": driver_id},
        {"_id": 0, "current_location": 1, "last_location_update": 1, "name": 1, "phone": 1}
    )
    
    if not driver:
        raise HTTPException(status_code=404, detail="السائق غير موجود")
    
    return driver

@api_router.post("/orders/{order_id}/assign-driver")
async def assign_driver_to_order(
    order_id: str,
    driver_id: str,
    current_user: dict = Depends(get_current_user)
):
    """تخصيص سائق للطلب"""
    if current_user["role"] not in [UserRole.ADMIN, UserRole.MANAGER, UserRole.SUPERVISOR, UserRole.CASHIER]:
        raise HTTPException(status_code=403, detail="غير مصرح")
    
    tenant_id = get_user_tenant_id(current_user)
    
    # التحقق من وجود السائق
    driver = await db.drivers.find_one({"id": driver_id, "tenant_id": tenant_id})
    if not driver:
        raise HTTPException(status_code=404, detail="السائق غير موجود")
    
    # تحديث الطلب
    result = await db.orders.update_one(
        {"id": order_id, "tenant_id": tenant_id},
        {"$set": {
            "driver_id": driver_id,
            "driver_assigned_at": datetime.now(timezone.utc).isoformat(),
            "status": "out_for_delivery"
        }}
    )
    
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="الطلب غير موجود")
    
    # تحديث حالة السائق
    await db.drivers.update_one(
        {"id": driver_id},
        {"$set": {"is_available": False}}
    )
    
    return {
        "message": "تم تخصيص السائق للطلب",
        "driver": {
            "id": driver["id"],
            "name": driver["name"],
            "phone": driver["phone"]
        }
    }


# ==================== DRIVER APP ROUTES ====================

@api_router.post("/driver/login")
async def driver_login(phone: str, pin: str):
    """تسجيل دخول السائق برقم الهاتف والرمز السري"""
    driver = await db.drivers.find_one({"phone": phone}, {"_id": 0})
    
    if not driver:
        raise HTTPException(status_code=404, detail="رقم الهاتف غير مسجل كسائق")
    
    # التحقق من الرمز السري
    if driver.get("pin", "1234") != pin:
        raise HTTPException(status_code=401, detail="الرمز السري غير صحيح")
    
    if not driver.get("is_active", True):
        raise HTTPException(status_code=403, detail="حساب السائق غير مفعل")
    
    # إزالة PIN من الاستجابة لأسباب أمنية
    driver_response = {k: v for k, v in driver.items() if k != "pin"}
    
    return {"driver": driver_response, "message": "تم تسجيل الدخول بنجاح"}

@api_router.get("/driver/orders")
async def get_driver_orders(driver_id: str):
    """جلب الطلبات المسندة للسائق"""
    orders = await db.orders.find(
        {
            "driver_id": driver_id,
            "status": {"$in": ["ready", "out_for_delivery"]}
        },
        {"_id": 0}
    ).sort("created_at", -1).to_list(50)
    
    # إضافة status_label
    status_labels = {
        'ready': 'جاهز للتوصيل',
        'out_for_delivery': 'في الطريق'
    }
    
    for order in orders:
        order['status_label'] = status_labels.get(order.get('status'), order.get('status'))
    
    return orders

@api_router.put("/orders/{order_id}/status")
async def update_order_status(order_id: str, status: str):
    """تحديث حالة الطلب (للسائق)"""
    valid_statuses = ['pending', 'preparing', 'ready', 'out_for_delivery', 'delivered', 'cancelled']
    if status not in valid_statuses:
        raise HTTPException(status_code=400, detail="حالة غير صحيحة")
    
    order = await db.orders.find_one({"id": order_id})
    if not order:
        raise HTTPException(status_code=404, detail="الطلب غير موجود")
    
    update_data = {
        "status": status,
        "updated_at": datetime.now(timezone.utc).isoformat()
    }
    
    if status == 'delivered':
        update_data["delivered_at"] = datetime.now(timezone.utc).isoformat()
        # تحرير السائق
        if order.get("driver_id"):
            await db.drivers.update_one(
                {"id": order["driver_id"]},
                {"$set": {"is_available": True}}
            )
    
    await db.orders.update_one({"id": order_id}, {"$set": update_data})
    
    # إرسال إشعار Push للعميل عند تغيير حالة الطلب
    await notify_order_status_change(order_id, status)
    
    return {"message": "تم تحديث حالة الطلب", "status": status}


@api_router.get("/customer/order-driver/{order_id}")
async def get_order_driver_info(order_id: str, phone: str = None):
    """جلب معلومات سائق الطلب للزبون"""
    # جلب الطلب
    order = await db.orders.find_one({"id": order_id})
    if not order:
        raise HTTPException(status_code=404, detail="الطلب غير موجود")
    
    # التحقق من رقم الهاتف
    if phone and order.get("customer_phone") != phone:
        raise HTTPException(status_code=403, detail="غير مصرح بالوصول لهذا الطلب")
    
    if not order.get("driver_id"):
        return {"driver": None, "message": "لم يتم تخصيص سائق بعد"}
    
    # جلب معلومات السائق
    driver = await db.drivers.find_one(
        {"id": order["driver_id"]},
        {"_id": 0, "id": 1, "name": 1, "phone": 1, "photo": 1, "current_location": 1, "last_location_update": 1}
    )
    
    if not driver:
        return {"driver": None, "message": "السائق غير متاح"}
    
    # إضافة موقع التوصيل
    delivery_location = order.get("delivery_location")
    
    return {
        "driver": driver,
        "delivery_location": delivery_location,
        "order_status": order.get("status")
    }


# ==================== PUSH NOTIFICATIONS ROUTES ====================

class PushSubscription(BaseModel):
    endpoint: str
    keys: dict
    phone: Optional[str] = None
    user_type: str = "customer"  # customer, driver, admin

# ==================== DRIVER APP ROUTES (بدون مصادقة JWT) ====================

class DriverLocationUpdate(BaseModel):
    latitude: float
    longitude: float

@api_router.post("/driver/update-location")
async def driver_update_location(driver_id: str, location: DriverLocationUpdate):
    """تحديث موقع السائق - للاستخدام من تطبيق السائق (بدون JWT)"""
    result = await db.drivers.update_one(
        {"id": driver_id},
        {"$set": {
            "current_location": {
                "latitude": location.latitude,
                "longitude": location.longitude
            },
            "last_location_update": datetime.now(timezone.utc).isoformat()
        }}
    )
    
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="السائق غير موجود")
    
    return {"message": "تم تحديث الموقع", "success": True}

@api_router.put("/driver/orders/{order_id}/status")
async def driver_update_order_status(order_id: str, status: str, driver_id: str):
    """تحديث حالة الطلب من تطبيق السائق (بدون JWT)"""
    valid_statuses = ['out_for_delivery', 'delivered']
    if status not in valid_statuses:
        raise HTTPException(status_code=400, detail="حالة غير صحيحة للسائق")
    
    # التحقق من أن الطلب مُسند لهذا السائق
    order = await db.orders.find_one({"id": order_id})
    if not order:
        raise HTTPException(status_code=404, detail="الطلب غير موجود")
    
    if order.get("driver_id") != driver_id:
        raise HTTPException(status_code=403, detail="هذا الطلب غير مُسند لك")
    
    update_data = {
        "status": status,
        "updated_at": datetime.now(timezone.utc).isoformat()
    }
    
    if status == 'delivered':
        update_data["delivered_at"] = datetime.now(timezone.utc).isoformat()
        # تحرير السائق
        await db.drivers.update_one(
            {"id": driver_id},
            {"$set": {"is_available": True, "current_order_id": None}}
        )
    
    await db.orders.update_one({"id": order_id}, {"$set": update_data})
    
    # إرسال إشعار Push للعميل
    await notify_order_status_change(order_id, status)
    
    return {"message": "تم تحديث حالة الطلب", "status": status}

@api_router.get("/driver/order-driver-info/{order_id}")
async def get_driver_info_for_customer(order_id: str):
    """جلب معلومات السائق وموقعه للزبون - بدون مصادقة"""
    order = await db.orders.find_one({"id": order_id}, {"_id": 0})
    if not order:
        raise HTTPException(status_code=404, detail="الطلب غير موجود")
    
    driver_id = order.get("driver_id")
    if not driver_id:
        return {
            "driver": None,
            "message": "لم يتم تخصيص سائق بعد",
            "order_status": order.get("status")
        }
    
    driver = await db.drivers.find_one(
        {"id": driver_id},
        {"_id": 0, "id": 1, "name": 1, "phone": 1, "current_location": 1, "last_location_update": 1}
    )
    
    return {
        "driver": driver,
        "delivery_location": order.get("delivery_location"),
        "order_status": order.get("status"),
        "delivery_address": order.get("delivery_address")
    }

# ==================== PUSH NOTIFICATIONS ROUTES ====================

@api_router.post("/push/subscribe")
async def subscribe_push(subscription: PushSubscription):
    """تسجيل اشتراك في إشعارات Push"""
    sub_doc = {
        "id": str(uuid.uuid4()),
        "endpoint": subscription.endpoint,
        "keys": subscription.keys,
        "phone": subscription.phone,
        "user_type": subscription.user_type,
        "created_at": datetime.now(timezone.utc).isoformat(),
        "is_active": True
    }
    
    # تحديث أو إنشاء الاشتراك
    await db.push_subscriptions.update_one(
        {"endpoint": subscription.endpoint},
        {"$set": sub_doc},
        upsert=True
    )
    
    return {"message": "تم تسجيل الاشتراك بنجاح"}

@api_router.delete("/push/unsubscribe")
async def unsubscribe_push(endpoint: str):
    """إلغاء اشتراك في إشعارات Push"""
    await db.push_subscriptions.delete_one({"endpoint": endpoint})
    return {"message": "تم إلغاء الاشتراك"}

async def send_push_notification(phone: str, title: str, body: str, data: dict = None, user_type: str = None):
    """إرسال إشعار Push للمستخدم"""
    try:
        from pywebpush import webpush, WebPushException
        
        # جلب VAPID keys من البيئة (يجب توليدها مسبقاً)
        # لأغراض العرض، سنستخدم طريقة بديلة
        
        query = {"is_active": True}
        if phone:
            query["phone"] = phone
        if user_type:
            query["user_type"] = user_type
        
        subscriptions = await db.push_subscriptions.find(query).to_list(100)
        
        notification_data = {
            "title": title,
            "body": body,
            "data": data or {},
            "icon": "/icons/admin-icon-192.png"
        }
        
        # هنا يجب استخدام webpush library لإرسال الإشعارات الفعلية
        # لكن لأن هذا يتطلب VAPID keys، سنسجل الإشعار في قاعدة البيانات
        
        notification_log = {
            "id": str(uuid.uuid4()),
            "phone": phone,
            "user_type": user_type,
            "title": title,
            "body": body,
            "data": data,
            "sent_at": datetime.now(timezone.utc).isoformat(),
            "subscriptions_count": len(subscriptions)
        }
        
        await db.notification_logs.insert_one(notification_log)
        
        return True
        
    except Exception as e:
        logger.error(f"Push notification error: {str(e)}")
        return False

@api_router.post("/push/test")
async def test_push_notification(phone: str, message: str = "هذا إشعار تجريبي"):
    """إرسال إشعار تجريبي"""
    await send_push_notification(
        phone=phone,
        title="Maestro EGP",
        body=message,
        data={"type": "test"}
    )
    return {"message": "تم إرسال الإشعار"}

@api_router.get("/notifications/{phone}")
async def get_notifications(phone: str, limit: int = 20):
    """جلب سجل الإشعارات للمستخدم"""
    notifications = await db.notification_logs.find(
        {"phone": phone},
        {"_id": 0}
    ).sort("sent_at", -1).limit(limit).to_list(limit)
    
    return notifications

# دالة لإرسال إشعار عند تغير حالة الطلب
async def notify_order_status_change(order_id: str, new_status: str):
    """إرسال إشعار للعميل عند تغير حالة الطلب"""
    order = await db.orders.find_one({"id": order_id})
    if not order:
        return
    
    status_messages = {
        'preparing': ('جاري تحضير طلبك! 👨‍🍳', 'طلبك قيد التحضير الآن'),
        'ready': ('طلبك جاهز! ✅', 'طلبك جاهز للتوصيل'),
        'out_for_delivery': ('السائق في الطريق! 🚚', 'السائق في طريقه إليك'),
        'delivered': ('تم التسليم! 🎉', 'استمتع بوجبتك! لا تنسى تقييم الطلب')
    }
    
    if new_status in status_messages:
        title, body = status_messages[new_status]
        await send_push_notification(
            phone=order.get("customer_phone"),
            title=title,
            body=body,
            data={
                "type": "order_status",
                "order_id": order_id,
                "status": new_status,
                "url": f"/menu/{order.get('tenant_id')}"
            },
            user_type="customer"
        )


# ==================== ADDRESS AUTOCOMPLETE ROUTES ====================

@api_router.get("/geocode/reverse")
async def reverse_geocode(lat: float, lng: float):
    """تحويل إحداثيات لعنوان (Reverse Geocoding)"""
    try:
        import httpx
        
        # استخدام Nominatim API المجاني
        url = f"https://nominatim.openstreetmap.org/reverse?lat={lat}&lon={lng}&format=json&accept-language=ar"
        
        async with httpx.AsyncClient() as client:
            response = await client.get(url, headers={"User-Agent": "MaestroEGP/1.0"})
            
            if response.status_code == 200:
                data = response.json()
                
                address = data.get("display_name", "")
                address_parts = data.get("address", {})
                
                return {
                    "address": address,
                    "street": address_parts.get("road", ""),
                    "neighbourhood": address_parts.get("neighbourhood", address_parts.get("suburb", "")),
                    "city": address_parts.get("city", address_parts.get("town", address_parts.get("village", ""))),
                    "country": address_parts.get("country", ""),
                    "lat": lat,
                    "lng": lng
                }
            else:
                return {"address": "", "lat": lat, "lng": lng}
                
    except Exception as e:
        logger.error(f"Reverse geocoding error: {str(e)}")
        return {"address": "", "lat": lat, "lng": lng}

@api_router.get("/geocode/search")
async def search_address(query: str, lat: Optional[float] = None, lng: Optional[float] = None):
    """البحث عن عنوان (Address Autocomplete)"""
    try:
        import httpx
        
        # استخدام Nominatim API للبحث
        url = f"https://nominatim.openstreetmap.org/search?q={query}&format=json&limit=5&accept-language=ar"
        
        # إضافة تفضيل للموقع الحالي إذا متاح
        if lat and lng:
            url += f"&viewbox={lng-0.5},{lat-0.5},{lng+0.5},{lat+0.5}&bounded=0"
        
        async with httpx.AsyncClient() as client:
            response = await client.get(url, headers={"User-Agent": "MaestroEGP/1.0"})
            
            if response.status_code == 200:
                data = response.json()
                
                results = []
                for item in data:
                    results.append({
                        "address": item.get("display_name", ""),
                        "lat": float(item.get("lat", 0)),
                        "lng": float(item.get("lon", 0)),
                        "type": item.get("type", "")
                    })
                
                return {"results": results}
            else:
                return {"results": []}
                
    except Exception as e:
        logger.error(f"Address search error: {str(e)}")
        return {"results": []}


# ==================== PAYMENT ROUTES ====================

@api_router.post("/payments/create-checkout/{tenant_id}")
async def create_payment_checkout(
    tenant_id: str,
    request: Request,
    order_id: str,
    amount: float,
    customer_phone: Optional[str] = None,
    save_card: bool = False
):
    """إنشاء جلسة دفع Stripe"""
    try:
        from emergentintegrations.payments.stripe.checkout import (
            StripeCheckout, 
            CheckoutSessionRequest,
            CheckoutSessionResponse
        )
        
        api_key = os.environ.get('STRIPE_API_KEY')
        if not api_key:
            raise HTTPException(status_code=500, detail="Stripe not configured")
        
        # إعداد URLs
        host_url = str(request.base_url).rstrip('/')
        webhook_url = f"{host_url}/api/webhook/stripe"
        
        # استخدام الـ origin من الـ referer
        referer = request.headers.get('referer', '')
        if referer:
            from urllib.parse import urlparse
            parsed = urlparse(referer)
            frontend_url = f"{parsed.scheme}://{parsed.netloc}"
        else:
            frontend_url = host_url
        
        success_url = f"{frontend_url}/menu/{tenant_id}?payment_success=true&session_id={{CHECKOUT_SESSION_ID}}"
        cancel_url = f"{frontend_url}/menu/{tenant_id}?payment_cancelled=true"
        
        stripe_checkout = StripeCheckout(api_key=api_key, webhook_url=webhook_url)
        
        # إنشاء طلب الدفع
        checkout_request = CheckoutSessionRequest(
            amount=float(amount),
            currency="usd",
            success_url=success_url,
            cancel_url=cancel_url,
            metadata={
                "order_id": order_id,
                "tenant_id": tenant_id,
                "customer_phone": customer_phone or "",
                "save_card": str(save_card)
            },
            payment_methods=["card"]
        )
        
        session: CheckoutSessionResponse = await stripe_checkout.create_checkout_session(checkout_request)
        
        # حفظ معاملة الدفع
        transaction = {
            "id": str(uuid.uuid4()),
            "session_id": session.session_id,
            "order_id": order_id,
            "tenant_id": tenant_id,
            "amount": amount,
            "currency": "usd",
            "customer_phone": customer_phone,
            "payment_status": "pending",
            "status": "initiated",
            "save_card": save_card,
            "created_at": datetime.now(timezone.utc).isoformat()
        }
        
        await db.payment_transactions.insert_one(transaction)
        
        return {
            "success": True,
            "checkout_url": session.url,
            "session_id": session.session_id
        }
        
    except ImportError:
        raise HTTPException(status_code=500, detail="Payment library not installed")
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Payment error: {str(e)}")

@api_router.get("/payments/status/{session_id}")
async def get_payment_status(session_id: str):
    """التحقق من حالة الدفع"""
    try:
        from emergentintegrations.payments.stripe.checkout import StripeCheckout
        
        api_key = os.environ.get('STRIPE_API_KEY')
        if not api_key:
            raise HTTPException(status_code=500, detail="Stripe not configured")
        
        stripe_checkout = StripeCheckout(api_key=api_key, webhook_url="")
        status = await stripe_checkout.get_checkout_status(session_id)
        
        # تحديث حالة المعاملة في قاعدة البيانات
        transaction = await db.payment_transactions.find_one({"session_id": session_id})
        
        if transaction:
            new_status = "completed" if status.payment_status == "paid" else status.payment_status
            
            # تجنب التحديث المتكرر
            if transaction.get("payment_status") != new_status:
                await db.payment_transactions.update_one(
                    {"session_id": session_id},
                    {
                        "$set": {
                            "payment_status": new_status,
                            "status": status.status,
                            "updated_at": datetime.now(timezone.utc).isoformat()
                        }
                    }
                )
                
                # تحديث حالة الطلب إذا تم الدفع
                if status.payment_status == "paid" and transaction.get("order_id"):
                    await db.orders.update_one(
                        {"id": transaction["order_id"]},
                        {
                            "$set": {
                                "payment_status": "paid",
                                "paid_at": datetime.now(timezone.utc).isoformat()
                            }
                        }
                    )
        
        return {
            "status": status.status,
            "payment_status": status.payment_status,
            "amount_total": status.amount_total / 100,
            "currency": status.currency,
            "order_id": status.metadata.get("order_id") if status.metadata else None
        }
        
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Status check error: {str(e)}")

@api_router.post("/webhook/stripe")
async def stripe_webhook(request: Request):
    """معالجة webhook من Stripe"""
    try:
        from emergentintegrations.payments.stripe.checkout import StripeCheckout
        
        api_key = os.environ.get('STRIPE_API_KEY')
        if not api_key:
            return {"error": "Stripe not configured"}
        
        body = await request.body()
        signature = request.headers.get("Stripe-Signature")
        
        stripe_checkout = StripeCheckout(api_key=api_key, webhook_url="")
        webhook_response = await stripe_checkout.handle_webhook(body, signature)
        
        if webhook_response.payment_status == "paid":
            # تحديث المعاملة
            await db.payment_transactions.update_one(
                {"session_id": webhook_response.session_id},
                {
                    "$set": {
                        "payment_status": "completed",
                        "status": "complete",
                        "webhook_event_id": webhook_response.event_id,
                        "updated_at": datetime.now(timezone.utc).isoformat()
                    }
                }
            )
            
            # تحديث الطلب
            transaction = await db.payment_transactions.find_one({"session_id": webhook_response.session_id})
            if transaction and transaction.get("order_id"):
                await db.orders.update_one(
                    {"id": transaction["order_id"]},
                    {
                        "$set": {
                            "payment_status": "paid",
                            "paid_at": datetime.now(timezone.utc).isoformat()
                        }
                    }
                )
        
        return {"received": True, "event_type": webhook_response.event_type}
        
    except Exception as e:
        return {"error": str(e)}


# ==================== PAYMENT SETTINGS APIs ====================

class PaymentSettingsUpdate(BaseModel):
    stripe_enabled: Optional[bool] = True
    stripe_publishable_key: Optional[str] = None
    stripe_secret_key: Optional[str] = None
    stripe_currency: Optional[str] = "USD"
    stripe_mode: Optional[str] = "test"  # test or live
    zaincash_enabled: Optional[bool] = True
    zaincash_phone: Optional[str] = None
    zaincash_name: Optional[str] = None
    zaincash_qr_image: Optional[str] = None
    cash_enabled: Optional[bool] = True
    delivery_fee: Optional[int] = 5000
    min_order_amount: Optional[int] = 10000

@api_router.get("/payment-settings")
async def get_payment_settings(current_user: dict = Depends(get_current_user)):
    """جلب إعدادات الدفع"""
    tenant_id = get_user_tenant_id(current_user) or "default"
    
    settings = await db.payment_settings.find_one(
        {"tenant_id": tenant_id},
        {"_id": 0, "stripe_secret_key": 0}  # إخفاء المفتاح السري
    )
    
    if not settings:
        settings = {
            "tenant_id": tenant_id,
            "stripe_enabled": True,
            "stripe_publishable_key": "",
            "stripe_currency": "USD",
            "stripe_mode": "test",
            "zaincash_enabled": True,
            "zaincash_phone": "",
            "zaincash_name": "",
            "zaincash_qr_image": "",
            "cash_enabled": True,
            "delivery_fee": 5000,
            "min_order_amount": 10000
        }
    
    # إخفاء المفتاح السري (عرض فقط أنه موجود أو لا)
    settings["stripe_secret_key_set"] = bool(settings.get("stripe_secret_key"))
    
    return settings

@api_router.post("/payment-settings")
async def update_payment_settings(
    settings: PaymentSettingsUpdate,
    current_user: dict = Depends(get_current_user)
):
    """تحديث إعدادات الدفع"""
    if not has_role(current_user, ['admin', 'owner']):
        raise HTTPException(status_code=403, detail="غير مصرح لك بتعديل الإعدادات")
    
    tenant_id = get_user_tenant_id(current_user) or "default"
    
    update_data = {
        "tenant_id": tenant_id,
        "updated_at": datetime.now(timezone.utc).isoformat(),
        "updated_by": current_user.get("id")
    }
    
    # إضافة الحقول غير الفارغة فقط
    settings_dict = settings.dict(exclude_unset=True, exclude_none=True)
    update_data.update(settings_dict)
    
    # تشفير المفتاح السري (في الإنتاج يجب استخدام تشفير حقيقي)
    if "stripe_secret_key" in update_data and update_data["stripe_secret_key"]:
        # في الإنتاج: استخدم تشفير AES أو KMS
        # هنا نحتفظ به كما هو للتبسيط
        pass
    
    await db.payment_settings.update_one(
        {"tenant_id": tenant_id},
        {"$set": update_data},
        upsert=True
    )
    
    return {"success": True, "message": "تم حفظ الإعدادات بنجاح"}

@api_router.post("/payment-settings/zaincash-qr")
async def upload_zaincash_qr(
    file: UploadFile = File(...),
    current_user: dict = Depends(get_current_user)
):
    """رفع صورة QR Code لزين كاش"""
    if not has_role(current_user, ['admin', 'owner']):
        raise HTTPException(status_code=403, detail="غير مصرح لك")
    
    tenant_id = get_user_tenant_id(current_user) or "default"
    
    # حفظ الصورة
    file_ext = file.filename.split('.')[-1] if '.' in file.filename else 'png'
    filename = f"zaincash_qr_{tenant_id}.{file_ext}"
    file_path = UPLOAD_DIR / "payment" / filename
    
    (UPLOAD_DIR / "payment").mkdir(exist_ok=True)
    
    content = await file.read()
    async with aiofiles.open(file_path, 'wb') as f:
        await f.write(content)
    
    image_url = f"/uploads/payment/{filename}"
    
    # تحديث الإعدادات
    await db.payment_settings.update_one(
        {"tenant_id": tenant_id},
        {"$set": {"zaincash_qr_image": image_url}},
        upsert=True
    )
    
    return {"success": True, "image_url": image_url}


# ==================== REAL-TIME NOTIFICATIONS APIs ====================

# تخزين الإشعارات غير المقروءة في الذاكرة (للتبسيط)
# في الإنتاج: استخدم Redis أو WebSockets
pending_notifications = {}

@api_router.get("/notifications/pending-orders")
async def get_pending_order_notifications(
    branch_id: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    """جلب الطلبات الجديدة (للكاشير)"""
    tenant_id = get_user_tenant_id(current_user) or "default"
    
    # جلب الطلبات الجديدة من آخر 5 دقائق
    five_minutes_ago = (datetime.now(timezone.utc) - timedelta(minutes=5)).isoformat()
    
    query = {
        "status": "pending",
        "source": "customer_app",
        "created_at": {"$gte": five_minutes_ago}
    }
    
    if branch_id:
        query["branch_id"] = branch_id
    elif tenant_id != "default":
        query["tenant_id"] = tenant_id
    
    orders = await db.orders.find(
        query,
        {"_id": 0, "id": 1, "order_number": 1, "customer_name": 1, 
         "total": 1, "created_at": 1, "payment_method": 1, "items": 1}
    ).sort("created_at", -1).to_list(20)
    
    # تحديد الطلبات الجديدة (غير المشاهدة)
    user_id = current_user.get("id", "")
    viewed_key = f"{tenant_id}_{user_id}"
    viewed_orders = pending_notifications.get(viewed_key, set())
    
    new_orders = []
    for order in orders:
        order["is_new"] = order["id"] not in viewed_orders
        new_orders.append(order)
    
    return {
        "orders": new_orders,
        "new_count": sum(1 for o in new_orders if o.get("is_new")),
        "total_count": len(new_orders)
    }

@api_router.post("/notifications/mark-seen")
async def mark_orders_as_seen(
    order_ids: List[str] = Body(...),
    current_user: dict = Depends(get_current_user)
):
    """تحديد الطلبات كمشاهدة"""
    tenant_id = get_user_tenant_id(current_user) or "default"
    user_id = current_user.get("id", "")
    viewed_key = f"{tenant_id}_{user_id}"
    
    if viewed_key not in pending_notifications:
        pending_notifications[viewed_key] = set()
    
    pending_notifications[viewed_key].update(order_ids)
    
    return {"success": True, "marked_count": len(order_ids)}

@api_router.get("/notifications/delayed-orders")
async def get_delayed_orders(
    branch_id: Optional[str] = None,
    delay_minutes: int = 15,
    current_user: dict = Depends(get_current_user)
):
    """
    جلب الطلبات المتأخرة - الطلبات التي مر عليها أكثر من المدة المحددة
    delay_minutes: عدد الدقائق للاعتبار التأخير (افتراضي 15 دقيقة)
    """
    tenant_id = get_user_tenant_id(current_user)
    
    # حساب وقت الحد الأقصى للتأخير
    delay_threshold = (datetime.now(timezone.utc) - timedelta(minutes=delay_minutes)).isoformat()
    
    # البحث عن الطلبات المتأخرة (pending أو preparing لأكثر من المدة المحددة)
    query = {
        "status": {"$in": ["pending", "preparing"]},
        "created_at": {"$lt": delay_threshold}
    }
    
    if branch_id and branch_id != 'all':
        query["branch_id"] = branch_id
    
    if tenant_id:
        query["tenant_id"] = tenant_id
    
    # جلب الطلبات المتأخرة
    delayed_orders = await db.orders.find(query, {"_id": 0}).sort("created_at", 1).to_list(50)
    
    # حساب مدة التأخير لكل طلب
    now = datetime.now(timezone.utc)
    for order in delayed_orders:
        try:
            created_at = datetime.fromisoformat(order.get("created_at", "").replace("Z", "+00:00"))
            delay_duration = now - created_at
            order["delay_minutes"] = int(delay_duration.total_seconds() / 60)
            
            # تصنيف مستوى التأخير
            if order["delay_minutes"] >= 45:
                order["delay_level"] = "critical"  # حرج
            elif order["delay_minutes"] >= 30:
                order["delay_level"] = "high"  # عالي
            elif order["delay_minutes"] >= 15:
                order["delay_level"] = "medium"  # متوسط
            else:
                order["delay_level"] = "low"  # منخفض
        except:
            order["delay_minutes"] = 0
            order["delay_level"] = "unknown"
    
    # تصنيف حسب نوع الطلب
    delayed_by_type = {
        "dine_in": [o for o in delayed_orders if o.get("order_type") == "dine_in"],
        "takeaway": [o for o in delayed_orders if o.get("order_type") == "takeaway"],
        "delivery": [o for o in delayed_orders if o.get("order_type") == "delivery"],
    }
    
    # إحصائيات سريعة
    stats = {
        "total_delayed": len(delayed_orders),
        "critical_count": len([o for o in delayed_orders if o.get("delay_level") == "critical"]),
        "high_count": len([o for o in delayed_orders if o.get("delay_level") == "high"]),
        "medium_count": len([o for o in delayed_orders if o.get("delay_level") == "medium"]),
        "avg_delay_minutes": round(sum(o.get("delay_minutes", 0) for o in delayed_orders) / max(len(delayed_orders), 1), 1),
        "max_delay_minutes": max((o.get("delay_minutes", 0) for o in delayed_orders), default=0)
    }
    
    return {
        "delayed_orders": delayed_orders,
        "delayed_by_type": delayed_by_type,
        "stats": stats,
        "delay_threshold_minutes": delay_minutes
    }

@api_router.get("/notifications/sound-alert")
async def check_sound_alert(
    last_check: Optional[str] = None,
    branch_id: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    """التحقق من وجود طلبات جديدة تحتاج تنبيه صوتي"""
    tenant_id = get_user_tenant_id(current_user) or "default"
    
    # تحديد وقت آخر فحص
    if last_check:
        try:
            check_time = last_check
        except:
            check_time = (datetime.now(timezone.utc) - timedelta(seconds=30)).isoformat()
    else:
        check_time = (datetime.now(timezone.utc) - timedelta(seconds=30)).isoformat()
    
    query = {
        "status": "pending",
        "source": "customer_app",
        "created_at": {"$gt": check_time}
    }
    
    if branch_id:
        query["branch_id"] = branch_id
    elif tenant_id != "default":
        query["tenant_id"] = tenant_id
    
    new_orders_count = await db.orders.count_documents(query)
    
    return {
        "has_new_orders": new_orders_count > 0,
        "new_orders_count": new_orders_count,
        "check_time": datetime.now(timezone.utc).isoformat()
    }


# ==================== SECURE CARD DATA (Stripe handles this) ====================
# ملاحظة مهمة: بيانات البطاقة لا تُخزن في قاعدة البيانات أبداً
# Stripe يتعامل مع جميع بيانات البطاقة الحساسة
# نحن نخزن فقط: آخر 4 أرقام، نوع البطاقة، تاريخ الانتهاء (للعرض فقط)

@api_router.get("/payment-transactions")
async def get_payment_transactions(
    limit: int = 20,
    current_user: dict = Depends(get_current_user)
):
    """جلب سجل المعاملات المالية"""
    if not has_role(current_user, ['admin', 'owner']):
        raise HTTPException(status_code=403, detail="غير مصرح لك")
    
    tenant_id = get_user_tenant_id(current_user) or "default"
    
    transactions = await db.payment_transactions.find(
        {"tenant_id": tenant_id},
        {"_id": 0}
    ).sort("created_at", -1).limit(limit).to_list(limit)
    
    return {"transactions": transactions}


# Include reports routes (refactored) - PRIORITY over old routes
from routes.reports_routes import router as reports_router
app.include_router(reports_router, prefix="/api")

# Include drivers routes (refactored)
from routes.drivers_routes import router as drivers_router
app.include_router(drivers_router, prefix="/api")

# Include payroll routes (refactored)
from routes.payroll_routes import router as payroll_router
app.include_router(payroll_router, prefix="/api")

# Include shifts routes (refactored)
from routes.shifts_routes import router as shifts_router
app.include_router(shifts_router, prefix="/api")

# Include router and middleware
app.include_router(api_router)

# Include new inventory system routes
from routes.inventory_system import router as inventory_router
app.include_router(inventory_router)

# Middleware to prevent caching of API responses
@app.middleware("http")
async def add_no_cache_headers(request, call_next):
    response = await call_next(request)
    if request.url.path.startswith("/api"):
        response.headers["Cache-Control"] = "no-cache, no-store, must-revalidate"
        response.headers["Pragma"] = "no-cache"
        response.headers["Expires"] = "0"
    return response

app.add_middleware(
    CORSMiddleware,
    allow_credentials=True,
    allow_origins=os.environ.get('CORS_ORIGINS', '*').split(','),
    allow_methods=["*"],
    allow_headers=["*"],
)

@app.on_event("shutdown")
async def shutdown_db_client():
    client.close()
